"""
WBS 앱 뷰
"""
import io
import urllib.parse
from datetime import date, timedelta
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import Sum, F, ExpressionWrapper, FloatField
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated

from .models import WBSItem
from .serializers import WBSItemSerializer, WBSSummarySerializer


@login_required
def wbs_view(request):
    return render(request, 'wbs/index.html')


class WBSItemViewSet(viewsets.ModelViewSet):
    serializer_class   = WBSItemSerializer
    permission_classes = [IsAuthenticated]
    pagination_class   = None

    def perform_update(self, serializer):
        """PATCH 시 일정 변경 이력 자동 기록"""
        from .models import WBSChangeLog
        item = serializer.instance
        tracked = ['planned_start', 'planned_end', 'actual_start', 'actual_end', 'name', 'weight', 'phase']
        for f in tracked:
            if f in serializer.validated_data:
                old = str(getattr(item, f, '') or '')
                new = str(serializer.validated_data[f] or '')
                if old != new:
                    ct = 'schedule' if 'start' in f or 'end' in f else 'edit'
                    WBSChangeLog.objects.create(
                        item=item, change_type=ct, field_name=f,
                        old_value=old, new_value=new,
                        changed_by=self.request.user,
                    )
        serializer.save()

    def get_queryset(self):
        qs = WBSItem.objects.select_related(
            'parent', 'assignee', 'linked_template', 'linked_inspection'
        )
        project_id = self.request.query_params.get('project')
        if project_id:
            qs = qs.filter(project_id=project_id)
        phase = self.request.query_params.get('phase')
        if phase:
            qs = qs.filter(phase=phase)
        depth = self.request.query_params.get('depth')
        if depth:
            qs = qs.filter(depth=depth)
        return qs

    @action(detail=False, methods=['get'], url_path='gantt')
    def gantt(self, request):
        """Gantt 차트용 포맷 반환"""
        project_id = request.query_params.get('project')
        if not project_id:
            return Response({'error': 'project 파라미터 필요'}, status=400)

        items = WBSItem.objects.filter(project_id=project_id).order_by('seq')
        data = []
        for item in items:
            data.append({
                'id':            item.id,
                'code':          item.code,
                'text':          item.name,
                'depth':         item.depth,
                'parent':        item.parent_id,
                'start_date':    item.planned_start.strftime('%Y-%m-%d') if item.planned_start else None,
                'end_date':      item.planned_end.strftime('%Y-%m-%d')   if item.planned_end   else None,
                'actual_start':  item.actual_start.strftime('%Y-%m-%d') if item.actual_start else None,
                'actual_end':    item.actual_end.strftime('%Y-%m-%d')   if item.actual_end   else None,
                'progress':      item.progress / 100,
                'weight':        float(item.weight),
                'is_milestone':  item.is_milestone,
                'open':          item.depth <= 2,
            })
        return Response(data)

    @action(detail=False, methods=['get'], url_path='summary')
    def summary(self, request):
        """페이즈별 계획진척률 / 실적진척률 / 공정준수율 집계"""
        project_id = request.query_params.get('project')
        if not project_id:
            return Response({'error': 'project 파라미터 필요'}, status=400)

        today = date.today()
        phase_labels = {'plan': '계획', 'execute': '수행', 'close': '종료'}
        result = []

        for phase_key, phase_label in phase_labels.items():
            items = WBSItem.objects.filter(
                project_id=project_id, phase=phase_key
            ).exclude(progress_source='children')

            total_weight = sum(float(i.weight) for i in items) or 1

            # 계획진척률: 오늘 기준으로 완료되어야 할 항목의 가중 합
            planned = 0.0
            actual  = 0.0
            for i in items:
                w = float(i.weight)
                # 계획진척: 계획 종료일이 오늘 이전이면 100%, 아직 안 됐으면 0%
                if i.planned_end and i.planned_end <= today:
                    planned += w * 100
                elif i.planned_start and i.planned_start <= today and i.planned_end:
                    elapsed = (today - i.planned_start).days
                    total   = (i.planned_end - i.planned_start).days or 1
                    planned += w * min(elapsed / total * 100, 100)

                actual += w * i.progress

            planned_pct = round(planned / total_weight, 1)
            actual_pct  = round(actual  / total_weight, 1)
            compliance  = round(actual_pct / planned_pct * 100, 1) if planned_pct else 0

            result.append({
                'phase':            phase_key,
                'phase_display':    phase_label,
                'planned_progress': planned_pct,
                'actual_progress':  actual_pct,
                'compliance_rate':  compliance,
                'total_weight':     total_weight,
            })

        # 전체 합산
        all_items = WBSItem.objects.filter(
            project_id=project_id
        ).exclude(progress_source='children')
        tw = sum(float(i.weight) for i in all_items) or 1
        planned_all = 0.0
        actual_all  = 0.0
        for i in all_items:
            w = float(i.weight)
            if i.planned_end and i.planned_end <= today:
                planned_all += w * 100
            elif i.planned_start and i.planned_start <= today and i.planned_end:
                elapsed = (today - i.planned_start).days
                total   = (i.planned_end - i.planned_start).days or 1
                planned_all += w * min(elapsed / total * 100, 100)
            actual_all += w * i.progress

        planned_all_pct = round(planned_all / tw, 1)
        actual_all_pct  = round(actual_all  / tw, 1)
        result.append({
            'phase':            'total',
            'phase_display':    '전체',
            'planned_progress': planned_all_pct,
            'actual_progress':  actual_all_pct,
            'compliance_rate':  round(actual_all_pct / planned_all_pct * 100, 1) if planned_all_pct else 0,
            'total_weight':     tw,
        })

        return Response(result)

    @action(detail=False, methods=['get'], url_path='evm')
    def evm(self, request):
        """EVM(Earned Value Management) 지표 계산
        PV: 계획가치(Planned Value) — 오늘까지 완료 예정인 작업의 가중 비용
        EV: 획득가치(Earned Value) — 실제 완료된 작업의 가중 비용
        SPI: 일정성과지수(Schedule Performance Index) = EV / PV
        CPI: 원가성과지수(Cost Performance Index) = EV / AC (AC=EV 가정)
        """
        project_id = request.query_params.get('project')
        if not project_id:
            return Response({'error': 'project 파라미터 필요'}, status=400)

        today = date.today()
        items = list(WBSItem.objects.filter(
            project_id=project_id
        ).exclude(progress_source='children'))

        if not items:
            return Response({'pv': 0, 'ev': 0, 'spi': 0, 'phases': []})

        # 총 예산 = 가중치 합계 (정규화 기준 = 100%)
        bac = sum(float(i.weight) for i in items) or 1

        # PV: 오늘 기준 계획진척률 × 가중치
        pv = 0.0
        ev = 0.0
        for i in items:
            w = float(i.weight)
            # PV 계산
            if i.planned_end and i.planned_end <= today:
                pv += w
            elif i.planned_start and i.planned_start <= today and i.planned_end:
                elapsed = (today - i.planned_start).days
                total = (i.planned_end - i.planned_start).days or 1
                pv += w * min(elapsed / total, 1.0)
            # EV 계산
            ev += w * (i.progress / 100)

        pv_pct = round(pv / bac * 100, 1)
        ev_pct = round(ev / bac * 100, 1)
        spi = round(ev / pv, 3) if pv > 0 else 0
        sv = round(ev_pct - pv_pct, 1)  # Schedule Variance

        # 단계별 EVM
        phase_labels = {'plan': '계획', 'execute': '수행', 'close': '종료'}
        phases = []
        for pk, pl in phase_labels.items():
            pi = [i for i in items if i.phase == pk]
            if not pi:
                continue
            p_bac = sum(float(i.weight) for i in pi) or 1
            p_pv = 0.0
            p_ev = 0.0
            for i in pi:
                w = float(i.weight)
                if i.planned_end and i.planned_end <= today:
                    p_pv += w
                elif i.planned_start and i.planned_start <= today and i.planned_end:
                    elapsed = (today - i.planned_start).days
                    total = (i.planned_end - i.planned_start).days or 1
                    p_pv += w * min(elapsed / total, 1.0)
                p_ev += w * (i.progress / 100)
            p_spi = round(p_ev / p_pv, 3) if p_pv > 0 else 0
            phases.append({
                'phase': pk, 'label': pl,
                'pv': round(p_pv / p_bac * 100, 1),
                'ev': round(p_ev / p_bac * 100, 1),
                'spi': p_spi,
            })

        # SPI 판정
        if pv == 0:
            spi_status = 'good'
            spi_label = '사업 시작 전'
        elif spi >= 1.0:
            spi_status = 'good'
            spi_label = '정상 (일정 준수)'
        elif spi >= 0.9:
            spi_status = 'warning'
            spi_label = '주의 (소폭 지연)'
        else:
            spi_status = 'danger'
            spi_label = '위험 (일정 지연)'

        return Response({
            'bac': round(bac, 4),
            'pv': pv_pct,
            'ev': ev_pct,
            'sv': sv,
            'spi': spi,
            'spi_status': spi_status,
            'spi_label': spi_label,
            'phases': phases,
            'as_of': today.strftime('%Y-%m-%d'),
        })

    @action(detail=False, methods=['get'], url_path='history')
    def progress_history(self, request):
        """진척 이력 조회 — 주차별 진척 추이"""
        from .models import WBSProgressHistory
        project_id = request.query_params.get('project')
        if not project_id:
            return Response({'error': 'project 파라미터 필요'}, status=400)

        # 주차별 전체 평균 진척
        histories = WBSProgressHistory.objects.filter(
            item__project_id=project_id
        ).order_by('week_date')

        week_data = {}
        for h in histories:
            wd = h.week_date.strftime('%Y-%m-%d')
            if wd not in week_data:
                week_data[wd] = {'planned': [], 'actual': []}
            week_data[wd]['planned'].append(float(h.planned_progress))
            week_data[wd]['actual'].append(h.progress)

        labels = sorted(week_data.keys())
        planned = []
        actual = []
        for wd in labels:
            d = week_data[wd]
            planned.append(round(sum(d['planned']) / len(d['planned']), 1) if d['planned'] else 0)
            actual.append(round(sum(d['actual']) / len(d['actual']), 1) if d['actual'] else 0)

        return Response({
            'labels': [l[5:] for l in labels],  # MM-DD 형식
            'planned': planned,
            'actual': actual,
            'total_snapshots': histories.count(),
        })

    @action(detail=False, methods=['post'], url_path='snapshot')
    def take_snapshot(self, request):
        """진척 스냅샷 수동 생성 (관리자용)"""
        from .tasks import snapshot_wbs_progress
        result = snapshot_wbs_progress()
        return Response(result)

    # ── 기준선 관리 ─────────────────────────────────
    @action(detail=False, methods=['get', 'post'], url_path='baselines')
    def baselines(self, request):
        """기준선 목록 조회 / 신규 기준선 저장"""
        from .models import WBSBaseline, WBSBaselineItem

        project_id = request.query_params.get('project') or request.data.get('project')
        if not project_id:
            return Response({'error': 'project 파라미터 필요'}, status=400)

        if request.method == 'GET':
            bls = WBSBaseline.objects.filter(project_id=project_id).select_related('created_by')
            return Response([{
                'id': b.id, 'version': b.version, 'name': b.name,
                'description': b.description,
                'created_by': b.created_by.name if b.created_by else '',
                'created_at': b.created_at.strftime('%Y-%m-%d %H:%M'),
                'item_count': b.items.count(),
            } for b in bls])

        # POST: 기준선 저장
        name = request.data.get('name', '')
        desc = request.data.get('description', '')
        if not name:
            return Response({'error': '기준선명을 입력하세요'}, status=400)

        last = WBSBaseline.objects.filter(project_id=project_id).order_by('-version').first()
        version = (last.version + 1) if last else 1

        baseline = WBSBaseline.objects.create(
            project_id=project_id, version=version,
            name=name, description=desc, created_by=request.user
        )
        items = WBSItem.objects.filter(project_id=project_id)
        bl_items = [
            WBSBaselineItem(
                baseline=baseline, code=i.code, name=i.name, depth=i.depth,
                phase=i.phase, weight=i.weight,
                planned_start=i.planned_start, planned_end=i.planned_end,
                progress=i.progress,
            ) for i in items
        ]
        WBSBaselineItem.objects.bulk_create(bl_items)

        return Response({
            'ok': True, 'version': version, 'name': name,
            'item_count': len(bl_items),
        })

    @action(detail=False, methods=['get'], url_path='baseline-compare')
    def baseline_compare(self, request):
        """기준선 vs 현재 비교"""
        from .models import WBSBaseline, WBSBaselineItem

        project_id = request.query_params.get('project')
        baseline_id = request.query_params.get('baseline')
        if not project_id:
            return Response({'error': 'project 파라미터 필요'}, status=400)

        # 기준선 선택 (없으면 최초 기준선)
        if baseline_id:
            try:
                baseline = WBSBaseline.objects.get(id=baseline_id)
            except WBSBaseline.DoesNotExist:
                return Response({'error': '기준선 없음'}, status=404)
        else:
            baseline = WBSBaseline.objects.filter(project_id=project_id).order_by('version').first()
            if not baseline:
                return Response({'error': '저장된 기준선이 없습니다', 'baselines': []}, status=200)

        bl_items = {bi.code: bi for bi in baseline.items.all()}
        current_items = WBSItem.objects.filter(project_id=project_id).order_by('seq')

        rows = []
        for item in current_items:
            bl = bl_items.get(item.code)
            if not bl:
                rows.append({
                    'code': item.code, 'name': item.name, 'depth': item.depth,
                    'status': 'added', 'status_label': '신규',
                    'cur_start': str(item.planned_start or ''), 'cur_end': str(item.planned_end or ''),
                    'bl_start': '', 'bl_end': '',
                    'start_diff': 0, 'end_diff': 0,
                    'cur_progress': item.progress, 'bl_progress': 0,
                })
                continue

            start_diff = 0
            end_diff = 0
            changed = False
            if item.planned_start and bl.planned_start:
                start_diff = (item.planned_start - bl.planned_start).days
                if start_diff != 0: changed = True
            if item.planned_end and bl.planned_end:
                end_diff = (item.planned_end - bl.planned_end).days
                if end_diff != 0: changed = True

            rows.append({
                'code': item.code, 'name': item.name, 'depth': item.depth,
                'status': 'changed' if changed else 'same',
                'status_label': '변경' if changed else '동일',
                'cur_start': str(item.planned_start or ''),
                'cur_end': str(item.planned_end or ''),
                'bl_start': str(bl.planned_start or ''),
                'bl_end': str(bl.planned_end or ''),
                'start_diff': start_diff, 'end_diff': end_diff,
                'cur_progress': item.progress, 'bl_progress': bl.progress,
            })

        # 삭제된 항목 (기준선에 있으나 현재 없는)
        current_codes = {i.code for i in current_items}
        for code, bl in bl_items.items():
            if code not in current_codes:
                rows.append({
                    'code': code, 'name': bl.name, 'depth': bl.depth,
                    'status': 'deleted', 'status_label': '삭제',
                    'cur_start': '', 'cur_end': '',
                    'bl_start': str(bl.planned_start or ''), 'bl_end': str(bl.planned_end or ''),
                    'start_diff': 0, 'end_diff': 0,
                    'cur_progress': 0, 'bl_progress': bl.progress,
                })

        return Response({
            'baseline': {'id': baseline.id, 'version': baseline.version, 'name': baseline.name,
                         'created_at': baseline.created_at.strftime('%Y-%m-%d')},
            'rows': rows,
            'summary': {
                'total': len(rows),
                'same': sum(1 for r in rows if r['status'] == 'same'),
                'changed': sum(1 for r in rows if r['status'] == 'changed'),
                'added': sum(1 for r in rows if r['status'] == 'added'),
                'deleted': sum(1 for r in rows if r['status'] == 'deleted'),
            }
        })

    @action(detail=False, methods=['get'], url_path='change-log')
    def change_log(self, request):
        """변경 이력 조회"""
        from .models import WBSChangeLog
        project_id = request.query_params.get('project')
        if not project_id:
            return Response({'error': 'project 파라미터 필요'}, status=400)
        page = max(1, int(request.query_params.get('page', 1)))
        page_size = 30
        offset = (page - 1) * page_size

        qs = WBSChangeLog.objects.filter(item__project_id=project_id).select_related('item', 'changed_by').order_by('-changed_at')
        total = qs.count()
        rows = [{
            'code': c.item.code, 'name': c.item.name,
            'change_type': c.get_change_type_display(),
            'field': c.field_name, 'old': c.old_value[:50], 'new': c.new_value[:50],
            'reason': c.reason, 'by': c.changed_by.name if c.changed_by else '',
            'at': c.changed_at.strftime('%Y-%m-%d %H:%M'),
        } for c in qs[offset:offset + page_size]]
        return Response({'rows': rows, 'total': total, 'page': page,
                         'total_pages': (total + page_size - 1) // page_size})

    @action(detail=False, methods=['get'], url_path='s-curve')
    def s_curve(self, request):
        """S-Curve 데이터 — 주간 단위 계획 vs 실적 진척 추이"""
        project_id = request.query_params.get('project')
        if not project_id:
            return Response({'error': 'project 파라미터 필요'}, status=400)

        items = list(WBSItem.objects.filter(
            project_id=project_id
        ).exclude(progress_source='children'))
        if not items:
            return Response({'labels': [], 'planned': [], 'actual': []})

        # 전체 기간 산출
        starts = [i.planned_start for i in items if i.planned_start]
        ends = [i.planned_end for i in items if i.planned_end]
        if not starts or not ends:
            return Response({'labels': [], 'planned': [], 'actual': []})

        proj_start = min(starts)
        proj_end = max(ends)
        today = date.today()

        tw = sum(float(i.weight) for i in items) or 1

        # 주간 단위 포인트 생성
        labels = []
        planned_data = []
        actual_data = []

        cursor = proj_start
        week_num = 1
        while cursor <= proj_end:
            labels.append(f'{cursor.month}/{cursor.day}')
            # 계획 진척률 (이 날짜 기준)
            planned = 0.0
            for i in items:
                w = float(i.weight)
                if i.planned_end and i.planned_end <= cursor:
                    planned += w * 100
                elif i.planned_start and i.planned_start <= cursor and i.planned_end:
                    elapsed = (cursor - i.planned_start).days
                    total = (i.planned_end - i.planned_start).days or 1
                    planned += w * min(elapsed / total * 100, 100)
            planned_data.append(round(planned / tw, 1))

            # 실적 진척률 (미래는 None)
            if cursor <= today:
                actual = sum(float(i.weight) * i.progress for i in items)
                actual_data.append(round(actual / tw, 1))
            else:
                actual_data.append(None)

            cursor += timedelta(days=7)
            week_num += 1

        return Response({
            'labels': labels,
            'planned': planned_data,
            'actual': actual_data,
        })

    @action(detail=False, methods=['get'], url_path='delayed')
    def delayed(self, request):
        """지연 항목 목록 — planned_end < today & progress < 100%"""
        project_id = request.query_params.get('project')
        if not project_id:
            return Response({'error': 'project 파라미터 필요'}, status=400)

        today = date.today()
        items = WBSItem.objects.filter(
            project_id=project_id,
            planned_end__lt=today,
            progress__lt=100,
        ).exclude(progress_source='children').order_by('planned_end')

        data = []
        for i in items:
            delay_days = (today - i.planned_end).days
            data.append({
                'id': i.id,
                'code': i.code,
                'name': i.name,
                'progress': i.progress,
                'planned_end': i.planned_end.strftime('%Y-%m-%d'),
                'delay_days': delay_days,
                'assignee': i.assignee.name if i.assignee else '-',
            })
        return Response(data)

    @action(detail=False, methods=['get'], url_path='export')
    def export_excel(self, request):
        """WBS 전체 Excel 내보내기"""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        project_id = request.query_params.get('project')
        if not project_id:
            return Response({'error': 'project 파라미터 필요'}, status=400)

        items = WBSItem.objects.filter(project_id=project_id).select_related(
            'assignee', 'parent'
        ).order_by('seq')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'WBS 진척현황'

        headers = ['코드', '작업명', '단계', '깊이', '담당자', '가중치',
                   '계획시작', '계획종료', '실적시작', '실적종료',
                   '진척률(%)', '소스', '금주계획', '금주실적', '차주계획', '비고']
        ws.append(headers)
        hdr_fill = PatternFill('solid', fgColor='1F497D')
        hdr_font = Font(bold=True, color='FFFFFF', size=10)
        for cell in ws[1]:
            cell.font = hdr_font
            cell.fill = hdr_fill

        SRC_KO = {'manual': '수동', 'artifact': '산출물', 'inspection': '점검',
                  'incident': '장애', 'children': '하위합산'}
        PHASE_KO = {'plan': '계획', 'execute': '수행', 'close': '종료'}

        for i in items:
            ws.append([
                i.code, i.name, PHASE_KO.get(i.phase, i.phase), i.depth,
                i.assignee.name if i.assignee else '',
                float(i.weight),
                i.planned_start.strftime('%Y-%m-%d') if i.planned_start else '',
                i.planned_end.strftime('%Y-%m-%d') if i.planned_end else '',
                i.actual_start.strftime('%Y-%m-%d') if i.actual_start else '',
                i.actual_end.strftime('%Y-%m-%d') if i.actual_end else '',
                i.progress, SRC_KO.get(i.progress_source, i.progress_source),
                i.this_week_plan or '', i.this_week_actual or '',
                i.next_week_plan or '', i.notes or '',
            ])

        # depth 1 행 굵게
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            if row[3].value == 1:
                for cell in row:
                    cell.font = Font(bold=True)

        # 컬럼 너비
        widths = [8, 30, 6, 4, 8, 6, 10, 10, 10, 10, 8, 8, 20, 20, 20, 20]
        for ci, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = 'WBS_진척현황.xlsx'
        encoded = urllib.parse.quote(fname)
        resp = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = f"attachment; filename*=UTF-8''{encoded}"
        return resp

    @action(detail=False, methods=['post'], url_path='import')
    def import_excel(self, request):
        """WBS Excel 일괄 업로드 — 코드 기준 매칭(업데이트), 없으면 신규 생성"""
        import openpyxl
        from rest_framework.parsers import MultiPartParser
        from apps.audit.models import AuditProject

        project_id = request.query_params.get('project') or request.data.get('project')
        if not project_id:
            return Response({'error': 'project 파라미터 필요'}, status=400)
        try:
            project = AuditProject.objects.get(id=project_id)
        except AuditProject.DoesNotExist:
            return Response({'error': '프로젝트 없음'}, status=404)

        file = request.FILES.get('file')
        if not file:
            return Response({'error': '파일을 첨부하세요'}, status=400)

        try:
            wb = openpyxl.load_workbook(file, read_only=True)
            ws = wb.active
        except Exception as e:
            return Response({'error': f'Excel 파일 오류: {e}'}, status=400)

        # 헤더 매핑 (한글 → 필드명)
        HEADER_MAP = {
            '코드': 'code', '작업명': 'name', '단계': 'phase', '깊이': 'depth',
            '가중치': 'weight', '계획시작': 'planned_start', '계획종료': 'planned_end',
            '실적시작': 'actual_start', '실적종료': 'actual_end',
            '진척률(%)': 'progress', '진척률': 'progress',
            '소스': 'progress_source', '금주계획': 'this_week_plan',
            '금주실적': 'this_week_actual', '차주계획': 'next_week_plan', '비고': 'notes',
        }
        PHASE_MAP = {'계획': 'plan', '수행': 'execute', '종료': 'close'}
        SRC_MAP = {'수동': 'manual', '산출물': 'artifact', '점검': 'inspection',
                   '장애': 'incident', '하위합산': 'children'}

        # 헤더 읽기
        headers = []
        for cell in ws[1]:
            val = str(cell.value or '').strip()
            headers.append(HEADER_MAP.get(val, val))

        if 'code' not in headers:
            return Response({'error': '코드 컬럼이 필요합니다'}, status=400)

        created = 0
        updated = 0
        errors = []

        # 기존 항목 캐시
        existing = {i.code: i for i in WBSItem.objects.filter(project=project)}

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            row_dict = {}
            for ci, val in enumerate(row):
                if ci < len(headers):
                    row_dict[headers[ci]] = val

            code = str(row_dict.get('code', '')).strip()
            if not code:
                continue

            try:
                item = existing.get(code)
                is_new = item is None
                if is_new:
                    item = WBSItem(project=project, code=code)

                # 필드 업데이트
                if 'name' in row_dict and row_dict['name']:
                    item.name = str(row_dict['name']).strip()
                if 'phase' in row_dict and row_dict['phase']:
                    phase_val = str(row_dict['phase']).strip()
                    item.phase = PHASE_MAP.get(phase_val, phase_val)
                if 'depth' in row_dict and row_dict['depth']:
                    item.depth = int(row_dict['depth'])
                if 'weight' in row_dict and row_dict['weight'] is not None:
                    item.weight = float(row_dict['weight'])
                if 'progress' in row_dict and row_dict['progress'] is not None:
                    item.progress = int(row_dict['progress'])
                if 'progress_source' in row_dict and row_dict['progress_source']:
                    src_val = str(row_dict['progress_source']).strip()
                    item.progress_source = SRC_MAP.get(src_val, src_val)

                # 날짜 필드
                for f in ['planned_start', 'planned_end', 'actual_start', 'actual_end']:
                    if f in row_dict and row_dict[f]:
                        val = row_dict[f]
                        if hasattr(val, 'date'):
                            setattr(item, f, val.date() if hasattr(val, 'date') else val)
                        elif isinstance(val, str) and val.strip():
                            from datetime import datetime as dt
                            try:
                                setattr(item, f, dt.strptime(val.strip(), '%Y-%m-%d').date())
                            except ValueError:
                                pass
                        elif val is None or val == '':
                            setattr(item, f, None)

                # 텍스트 필드
                for f in ['this_week_plan', 'this_week_actual', 'next_week_plan', 'notes']:
                    if f in row_dict:
                        setattr(item, f, str(row_dict[f] or '').strip())

                # 부모 자동 매칭 (코드 기반: 1.2.3 → 부모 1.2)
                if '.' in code:
                    parent_code = code.rsplit('.', 1)[0]
                    parent = existing.get(parent_code) or WBSItem.objects.filter(
                        project=project, code=parent_code).first()
                    if parent:
                        item.parent = parent

                item.save()
                existing[code] = item

                if is_new:
                    created += 1
                else:
                    updated += 1
            except Exception as e:
                errors.append({'row': row_idx, 'code': code, 'error': str(e)})

        # 부모 진척률 재계산
        for item in WBSItem.objects.filter(project=project, progress_source='children'):
            item.recalculate_from_children()

        return Response({
            'created': created,
            'updated': updated,
            'errors': errors,
        })

    @action(detail=True, methods=['patch'], url_path='progress')
    def update_progress(self, request, pk=None):
        """진척률 수동 업데이트 (manual 소스 항목만)"""
        item = self.get_object()
        if item.progress_source != 'manual':
            return Response(
                {'error': f'이 항목은 {item.get_progress_source_display()} 소스로 자동 관리됩니다.'},
                status=400
            )
        val = request.data.get('progress')
        try:
            progress_int = int(val)
        except (TypeError, ValueError):
            return Response({'error': '0~100 사이 값 필요'}, status=400)
        if not (0 <= progress_int <= 100):
            return Response({'error': '0~100 사이 값 필요'}, status=400)

        # 변경 이력 기록
        from .models import WBSChangeLog
        old_progress = item.progress
        if old_progress != progress_int:
            WBSChangeLog.objects.create(
                item=item, change_type='progress', field_name='progress',
                old_value=str(old_progress), new_value=str(progress_int),
                changed_by=request.user,
            )
        for f in ['actual_start', 'actual_end', 'planned_start', 'planned_end']:
            if f in request.data:
                old_val = str(getattr(item, f, '') or '')
                new_val = str(request.data[f] or '')
                if old_val != new_val:
                    WBSChangeLog.objects.create(
                        item=item, change_type='schedule', field_name=f,
                        old_value=old_val, new_value=new_val,
                        changed_by=request.user,
                    )

        item.progress = progress_int
        fields = ['progress', 'updated_at']
        for f in ['this_week_plan', 'this_week_actual', 'next_week_plan', 'notes',
                  'actual_start', 'actual_end']:
            if f in request.data:
                setattr(item, f, request.data[f])
                fields.append(f)
        item.save(update_fields=fields)

        # 부모 버블업
        from apps.wbs.signals import _bubble_up
        _bubble_up(item)

        return Response(WBSItemSerializer(item).data)

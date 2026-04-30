import logging
from django.shortcuts import render

logger = logging.getLogger(__name__)
from django.contrib.auth.decorators import login_required
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.utils import timezone
from django.http import FileResponse, Http404
from django.db.models import Q
import os

@login_required
def reports_view(request):
    return render(request, 'reports/index.html')


@login_required
def labeling_view(request):
    return render(request, 'reports/labeling.html')


@login_required
def performance_report_view(request):
    """성과보고서 페이지 (주간/월간/분기/반기/연간, 교육지원청별)"""
    from django.utils import timezone as tz
    from apps.schools.models import SupportCenter
    now = tz.localdate()
    centers = SupportCenter.objects.filter(is_active=True).order_by('id')
    return render(request, 'reports/performance.html', {
        'current_year':  now.year,
        'current_month': now.month,
        'years':  list(range(2024, now.year + 2)),
        'months': list(range(1, 13)),
        'weeks':  list(range(1, 54)),
        'centers': centers,
    })


@login_required
def performance_report_data_api(request):
    """성과보고서 데이터 API — GET ?type=monthly&year=2026&month=4&center=all 등"""
    import json, calendar
    from django.http import JsonResponse
    from django.db.models import Count, Avg, Q
    from django.utils import timezone as tz
    from datetime import date, timedelta

    period_type = request.GET.get('type', 'monthly')   # weekly|monthly|quarterly|half|annual
    year  = int(request.GET.get('year',  tz.localdate().year))
    month = int(request.GET.get('month', tz.localdate().month))
    week  = int(request.GET.get('week',  1))   # 주간용 ISO week
    center_code = request.GET.get('center', 'all')     # 교육지원청 코드 또는 'all'

    # ── 기간 범위 산정 ────────────────────────────────────
    if period_type == 'weekly':
        d = date.fromisocalendar(year, week, 1)
        date_from = d
        date_to   = d + timedelta(days=6)
        label = f"{year}년 {week}주차 ({date_from.strftime('%m.%d')}~{date_to.strftime('%m.%d')})"
    elif period_type == 'monthly':
        date_from = date(year, month, 1)
        date_to   = date(year, month, calendar.monthrange(year, month)[1])
        label = f"{year}년 {month}월"
    elif period_type == 'quarterly':
        q = int(request.GET.get('quarter', ((month - 1) // 3) + 1))
        qm_start = (q - 1) * 3 + 1
        qm_end   = qm_start + 2
        date_from = date(year, qm_start, 1)
        date_to   = date(year, qm_end, calendar.monthrange(year, qm_end)[1])
        label = f"{year}년 {q}분기"
    elif period_type == 'half':
        h = int(request.GET.get('half', 1 if month <= 6 else 2))
        hm_start = 1 if h == 1 else 7
        hm_end   = 6 if h == 1 else 12
        date_from = date(year, hm_start, 1)
        date_to   = date(year, hm_end, calendar.monthrange(year, hm_end)[1])
        label = f"{year}년 {h}반기"
    else:  # annual
        date_from = date(year, 1, 1)
        date_to   = date(year, 12, 31)
        label = f"{year}년 연간"

    # ── 교육지원청 필터 준비 ──────────────────────────────
    from apps.schools.models import SupportCenter
    center_name = '전체'
    center_filter = {}            # Incident: school__support_center
    school_filter = {}            # School 직접 쿼리
    report_filter = {}            # Report: school__support_center
    ws_filter = {}                # WorkSchedule: school__support_center

    if center_code and center_code != 'all':
        sc = SupportCenter.objects.filter(code=center_code).first()
        if sc:
            center_name = sc.name
            center_filter = {'school__support_center': sc}
            school_filter = {'support_center': sc}
            report_filter = {'school__support_center': sc}
            ws_filter = {'school__support_center': sc}

    label_prefix = f"[{center_name}] " if center_code != 'all' else ''
    label = label_prefix + label

    from apps.incidents.models import Incident, SLAMonthly

    # ── 장애 현황 ─────────────────────────────────────────
    inc_qs = Incident.objects.filter(
        received_at__date__gte=date_from,
        received_at__date__lte=date_to,
        **center_filter,
    )
    total_inc     = inc_qs.count()
    completed_inc = inc_qs.filter(status='completed').count()

    # 장애 유형별
    by_type = list(
        inc_qs.values('fault_type').annotate(cnt=Count('id')).order_by('-cnt')[:6]
    )
    # 학교별 상위
    by_school = list(
        inc_qs.values('school__name').annotate(cnt=Count('id')).order_by('-cnt')[:5]
    )

    # ── 교육지원청별 크로스탭 데이터 ──────────────────────
    from apps.schools.models import School as SchoolModel

    all_centers = list(SupportCenter.objects.filter(is_active=True).order_by('id'))
    centers_info = [{'code': c.code, 'name': c.name} for c in all_centers]

    # 기간 내 전체 장애 (교육지원청 필터 무관)
    inc_all = Incident.objects.filter(
        received_at__date__gte=date_from,
        received_at__date__lte=date_to,
    )

    # 1) 교육지원청별 접수/완료/진행
    center_summary = []
    for sc in all_centers:
        sc_inc = inc_all.filter(school__support_center=sc)
        center_summary.append({
            'code': sc.code, 'name': sc.name,
            'total': sc_inc.count(),
            'completed': sc_inc.filter(status='completed').count(),
            'in_progress': sc_inc.exclude(status='completed').count(),
        })

    # 서비스 시작일 필터: NULL이거나 기간 내 시작된 학교만
    svc_q = Q(service_start_date__isnull=True) | Q(service_start_date__lte=date_to)

    # 2) 학제별 × 교육지원청 크로스탭
    from apps.schools.models import SchoolType
    school_types = list(
        SchoolType.objects.all().order_by('order').values_list('name', flat=True)
    )
    school_type_cross = []
    for st_name in school_types:
        row = {'name': st_name}
        row_total = 0
        for sc in all_centers:
            cnt = inc_all.filter(school__school_type__name=st_name, school__support_center=sc).count()
            row[sc.code] = cnt
            row_total += cnt
        row['total'] = row_total
        school_type_cross.append(row)

    # 3) 장애 대분류(category) × 교육지원청 크로스탭
    from apps.incidents.models import IncidentCategory
    fault_categories = list(
        IncidentCategory.objects.all().order_by('order').values_list('code', 'name')
    )
    fault_type_cross = []
    for cat_code, cat_name in fault_categories:
        row = {'name': cat_name}
        row_total = 0
        for sc in all_centers:
            cnt = inc_all.filter(category__code=cat_code, school__support_center=sc).count()
            row[sc.code] = cnt
            row_total += cnt
        row['total'] = row_total
        fault_type_cross.append(row)

    # 학제별 단순 집계 (기존 호환)
    by_school_type = list(
        inc_qs.values('school__school_type__name')
        .annotate(cnt=Count('id'))
        .order_by('school__school_type__order')
    )

    # ── SLA 현황 (월간만 저장되므로 기간 내 해당 월들 집계) ──
    sla_months = SLAMonthly.objects.filter(
        year__gte=date_from.year, year__lte=date_to.year
    ).filter(
        **({} if period_type == 'annual' else {})
    ).order_by('year', 'month')

    if period_type == 'monthly':
        sla_months = sla_months.filter(year=year, month=month)
    elif period_type == 'quarterly':
        sla_months = sla_months.filter(
            year=year, month__gte=date_from.month, month__lte=date_to.month
        )
    elif period_type == 'half':
        sla_months = sla_months.filter(
            year=year, month__gte=date_from.month, month__lte=date_to.month
        )

    sla_list = []
    for s in sla_months:
        sla_list.append({
            'year': s.year, 'month': s.month,
            'total_score': s.total_score,
            'grade': s.grade,
            'uptime_pct': s.uptime_pct,
            'inspection_pct': s.inspection_pct,
            'fault_count': s.fault_count,
            'avg_fault_min': s.avg_fault_min,
            'overtime_count': s.overtime_count,
            'recurrence_count': s.recurrence_count,
            'human_error_count': s.human_error_count,
            'security_count': s.security_count,
            'satisfaction_pct': s.satisfaction_pct,
        })

    avg_sla_score = (
        sum(s['total_score'] for s in sla_list if s['total_score']) / len(sla_list)
        if sla_list else None
    )

    # ── 정기점검 현황 ─────────────────────────────────────
    total_schools = SchoolModel.objects.filter(is_active=True, **school_filter).filter(svc_q).count()

    inspect_all = Report.objects.filter(
        template__report_type='regular',
        completed_at__date__gte=date_from,
        completed_at__date__lte=date_to,
        status='completed',
    )
    inspect_qs = inspect_all.filter(**report_filter) if report_filter else inspect_all
    inspect_completed = inspect_qs.values('school_id').distinct().count()
    inspect_by_school = list(
        inspect_qs.values('school__name')
        .annotate(cnt=Count('id'))
        .order_by('school__name')[:10]
    )

    # 4) 정기점검 × 교육지원청 크로스탭
    inspect_cross = []
    for sc in all_centers:
        sc_schools = SchoolModel.objects.filter(is_active=True, support_center=sc).filter(svc_q).count()
        sc_done = inspect_all.filter(school__support_center=sc).values('school_id').distinct().count()
        inspect_cross.append({
            'code': sc.code, 'name': sc.name,
            'total_schools': sc_schools,
            'completed': sc_done,
            'pct': round(sc_done / sc_schools * 100, 1) if sc_schools else 0,
        })

    # ── 인력 투입 현황 ────────────────────────────────────
    from apps.workforce.models import WorkSchedule
    ws_all = WorkSchedule.objects.filter(
        start_dt__date__gte=date_from,
        start_dt__date__lte=date_to,
    )
    ws_qs = ws_all.filter(**ws_filter) if ws_filter else ws_all
    ws_by_type = list(
        ws_qs.values('schedule_type__name')
        .annotate(cnt=Count('id'))
        .order_by('-cnt')
    )
    ws_by_worker = list(
        ws_qs.values('worker__name')
        .annotate(cnt=Count('id'))
        .order_by('-cnt')[:8]
    )
    ws_total = ws_qs.count()
    ws_completed = ws_qs.filter(status='completed').count()

    # 5) 인력투입 × 교육지원청 크로스탭
    workforce_cross = []
    for sc in all_centers:
        sc_ws = ws_all.filter(school__support_center=sc)
        workforce_cross.append({
            'code': sc.code, 'name': sc.name,
            'total': sc_ws.count(),
            'completed': sc_ws.filter(status='completed').count(),
        })

    return JsonResponse({
        'period_type': period_type,
        'center_code': center_code,
        'center_name': center_name,
        'label':       label,
        'date_from':   date_from.isoformat(),
        'date_to':     date_to.isoformat(),
        'centers':     centers_info,
        'incidents': {
            'total':     total_inc,
            'completed': completed_inc,
            'by_type':   by_type,
            'by_school': by_school,
            'by_center': center_summary,
            'by_school_type': by_school_type,
            'school_type_cross': school_type_cross,
            'fault_type_cross': fault_type_cross,
        },
        'sla': {
            'months':    sla_list,
            'avg_score': round(avg_sla_score, 2) if avg_sla_score else None,
        },
        'inspection': {
            'total_schools':    total_schools,
            'completed_schools': inspect_completed,
            'pct': round(inspect_completed / total_schools * 100, 1) if total_schools else 0,
            'by_school': inspect_by_school,
            'by_center': inspect_cross,
        },
        'workforce': {
            'total':     ws_total,
            'completed': ws_completed,
            'by_type':   ws_by_type,
            'by_worker': ws_by_worker,
            'by_center': workforce_cross,
        },
    })


@login_required
def export_performance_excel(request):
    """성과보고서 Excel 내보내기"""
    import io, calendar
    from datetime import date, timedelta
    from django.http import HttpResponse
    from django.db.models import Count, Q
    from django.utils import timezone as tz

    try:
        import openpyxl
        from openpyxl.styles import (Font, PatternFill, Alignment,
                                     Border, Side, numbers)
        from openpyxl.utils import get_column_letter
    except ImportError:
        from django.http import JsonResponse
        return JsonResponse({'error': 'openpyxl 패키지가 필요합니다.'}, status=500)

    period_type = request.GET.get('type', 'monthly')
    year  = int(request.GET.get('year',  tz.localdate().year))
    month = int(request.GET.get('month', tz.localdate().month))
    week  = int(request.GET.get('week', 1))
    center_code = request.GET.get('center', 'all')

    # ── 월간업무보고 엑셀 ──
    if period_type == 'monthly_work':
        return _export_monthly_work_excel(request, year, month, center_code, openpyxl)

    if period_type == 'weekly':
        d = date.fromisocalendar(year, week, 1)
        date_from, date_to = d, d + timedelta(days=6)
        label = f"{year}년 {week}주차"
    elif period_type == 'monthly':
        date_from = date(year, month, 1)
        date_to   = date(year, month, calendar.monthrange(year, month)[1])
        label = f"{year}년 {month}월"
    elif period_type == 'quarterly':
        q = int(request.GET.get('quarter', ((month - 1) // 3) + 1))
        qm_start = (q - 1) * 3 + 1
        date_from = date(year, qm_start, 1)
        date_to   = date(year, qm_start + 2,
                         calendar.monthrange(year, qm_start + 2)[1])
        label = f"{year}년 {q}분기"
    elif period_type == 'half':
        h = int(request.GET.get('half', 1 if month <= 6 else 2))
        hm_s, hm_e = (1, 6) if h == 1 else (7, 12)
        date_from = date(year, hm_s, 1)
        date_to   = date(year, hm_e, calendar.monthrange(year, hm_e)[1])
        label = f"{year}년 {h}반기"
    else:
        date_from = date(year, 1, 1)
        date_to   = date(year, 12, 31)
        label = f"{year}년 연간"

    from apps.incidents.models import Incident, SLAMonthly
    from apps.schools.models import School as SchoolModel, SupportCenter
    from apps.workforce.models import WorkSchedule

    # 교육지원청 필터
    center_filter = {}
    school_filter = {}
    report_filter = {}
    ws_filter = {}
    center_name = '전체'
    if center_code and center_code != 'all':
        sc = SupportCenter.objects.filter(code=center_code).first()
        if sc:
            center_name = sc.name
            center_filter = {'school__support_center': sc}
            school_filter = {'support_center': sc}
            report_filter = {'school__support_center': sc}
            ws_filter = {'school__support_center': sc}
    label_prefix = f"[{center_name}] " if center_code != 'all' else ''
    label = label_prefix + label

    inc_qs = Incident.objects.filter(
        received_at__date__gte=date_from, received_at__date__lte=date_to,
        **center_filter)
    total_inc     = inc_qs.count()
    completed_inc = inc_qs.filter(status='completed').count()
    by_type   = list(inc_qs.values('fault_type').annotate(cnt=Count('id')).order_by('-cnt')[:10])
    by_school = list(inc_qs.values('school__name').annotate(cnt=Count('id')).order_by('-cnt')[:10])

    sla_qs = SLAMonthly.objects.filter(
        year__gte=date_from.year, year__lte=date_to.year).order_by('year', 'month')
    if period_type == 'monthly':
        sla_qs = sla_qs.filter(year=year, month=month)
    elif period_type in ('quarterly', 'half'):
        sla_qs = sla_qs.filter(
            year=year, month__gte=date_from.month, month__lte=date_to.month)

    svc_q = Q(service_start_date__isnull=True) | Q(service_start_date__lte=date_to)
    total_schools = SchoolModel.objects.filter(is_active=True, **school_filter).filter(svc_q).count()
    inspect_qs    = Report.objects.filter(
        template__report_type='regular', status='completed',
        completed_at__date__gte=date_from, completed_at__date__lte=date_to,
        **report_filter)
    inspect_cnt   = inspect_qs.values('school_id').distinct().count()

    ws_qs      = WorkSchedule.objects.filter(
        start_dt__date__gte=date_from, start_dt__date__lte=date_to,
        **ws_filter)
    ws_by_type = list(ws_qs.values('schedule_type__name').annotate(cnt=Count('id')).order_by('-cnt'))
    ws_by_wkr  = list(ws_qs.values('worker__name').annotate(cnt=Count('id')).order_by('-cnt')[:10])

    # ── 워크북 생성 ────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '성과보고서'

    # 스타일
    hdr_fill  = PatternFill('solid', fgColor='1F497D')
    hdr_font  = Font(bold=True, color='FFFFFF', size=11)
    sec_fill  = PatternFill('solid', fgColor='D6E4F0')
    sec_font  = Font(bold=True, size=10)
    good_fill = PatternFill('solid', fgColor='C6EFCE')
    warn_fill = PatternFill('solid', fgColor='FFEB9C')
    bad_fill  = PatternFill('solid', fgColor='FFC7CE')
    ctr = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin'))

    def _set(row, col, val, font=None, fill=None, align=None, border=None):
        c = ws.cell(row=row, column=col, value=val)
        if font:   c.font   = font
        if fill:   c.fill   = fill
        if align:  c.alignment = align
        if border: c.border = border
        return c

    r = 1
    # 타이틀
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    _set(r, 1, f'서울시교육청 학교망 유지보수 용역 \u2014 {label} 성과보고서',
         Font(bold=True, size=14), hdr_fill, ctr)
    ws.row_dimensions[r].height = 30
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    _set(r, 1, f'보고 기간: {date_from} ~ {date_to}   /   작성일: {date.today()}',
         Font(size=10, italic=True), align=ctr)
    ws.row_dimensions[r].height = 20
    r += 2

    # ① 장애 현황
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    _set(r, 1, '1. 장애 현황', sec_font, sec_fill, ctr, thin)
    r += 1
    for h, col in [('유형','A'),('건수','B'),('비율','C')]:
        _set(r, ord(col)-64, h, hdr_font, hdr_fill, ctr, thin)
    r += 1
    for row in by_type:
        ft = row.get('fault_type') or '미분류'
        cnt = row['cnt']
        pct = f"{cnt/total_inc*100:.1f}%" if total_inc else '0%'
        for ci, v in [(1, ft),(2, cnt),(3, pct)]:
            _set(r, ci, v, border=thin, align=ctr if ci > 1 else None)
        r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=1)
    _set(r, 1, f'합계: {total_inc}건  (완료: {completed_inc}건)', Font(bold=True), border=thin)
    r += 2

    # ② SLA 현황
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    _set(r, 1, '2. SLA 종합 현황', sec_font, sec_fill, ctr, thin)
    r += 1
    sla_hdrs = ['기간','종합점수','등급','가동률','점검준수율','장애건수','평균조치(분)','반복장애']
    for ci, h in enumerate(sla_hdrs, 1):
        _set(r, ci, h, hdr_font, hdr_fill, ctr, thin)
    r += 1
    for s in sla_qs:
        row_fill = (good_fill if s.total_score and s.total_score >= 90
                    else warn_fill if s.total_score and s.total_score >= 80
                    else bad_fill if s.total_score else None)
        vals = [
            f"{s.year}.{str(s.month).zfill(2)}",
            round(s.total_score, 1) if s.total_score else '-',
            s.grade or '-',
            f"{s.uptime_pct:.3f}%" if s.uptime_pct else '-',
            f"{s.inspection_pct:.1f}%" if s.inspection_pct else '-',
            s.fault_count or 0,
            round(s.avg_fault_min) if s.avg_fault_min else '-',
            s.recurrence_count or 0,
        ]
        for ci, v in enumerate(vals, 1):
            c = _set(r, ci, v, border=thin, align=ctr)
            if row_fill: c.fill = row_fill
        r += 1
    if not sla_qs.exists():
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        _set(r, 1, '해당 기간 SLA 데이터 없음', align=ctr, border=thin)
        r += 1
    r += 1

    # ③ 정기점검 현황
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    _set(r, 1, '3. 정기점검 현황', sec_font, sec_fill, ctr, thin)
    r += 1
    pct_str = f"{inspect_cnt/total_schools*100:.1f}%" if total_schools else '0%'
    _set(r, 1, f'완료 학교: {inspect_cnt}개교 / 전체: {total_schools}개교 ({pct_str})',
         Font(bold=True), border=thin)
    r += 1
    if inspect_qs.exists():
        for h, ci in [('학교명', 1), ('점검 보고서 수', 2)]:
            _set(r, ci, h, hdr_font, hdr_fill, ctr, thin)
        r += 1
        for row in inspect_qs.values('school__name').annotate(cnt=Count('id')).order_by('school__name')[:15]:
            _set(r, 1, row['school__name'] or '-', border=thin)
            _set(r, 2, row['cnt'], border=thin, align=ctr)
            r += 1
    r += 1

    # ④ 인력 투입 현황
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    _set(r, 1, '4. 인력 투입 현황', sec_font, sec_fill, ctr, thin)
    r += 1
    for h, ci in [('업무유형', 1), ('투입건수', 2)]:
        _set(r, ci, h, hdr_font, hdr_fill, ctr, thin)
    r += 1
    for row in ws_by_type:
        _set(r, 1, row.get('schedule_type__name') or '미분류', border=thin)
        _set(r, 2, row['cnt'], border=thin, align=ctr)
        r += 1
    if not ws_by_type:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        _set(r, 1, '데이터 없음', align=ctr, border=thin)
        r += 1
    r += 1

    # 열 너비
    col_widths = [22, 12, 10, 14, 14, 10, 14, 12]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_label = label.replace(' ', '_').replace('/', '-')
    resp = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="performance_{safe_label}.xlsx"'
    return resp


def _export_monthly_work_excel(request, year, month, center_code, openpyxl):
    """월간업무보고 엑셀 내보내기 — 일자별 업무 현황"""
    import io, calendar
    from datetime import date, timedelta
    from collections import defaultdict
    from django.http import HttpResponse
    from django.db.models import Count
    from urllib.parse import quote
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    from apps.schools.models import SupportCenter
    from apps.incidents.models import Incident
    from apps.progress.models import SchoolInspection
    from .models import Report

    date_from = date(year, month, 1)
    date_to = date(year, month, calendar.monthrange(year, month)[1])

    center_filter = {}
    center_name = '전체'
    if center_code and center_code != 'all':
        sc = SupportCenter.objects.filter(code=center_code).first()
        if sc:
            center_name = sc.name
            center_filter = {'school__support_center': sc}

    # 데이터 수집 (monthly_work_report_api 와 동일 로직)
    daily_work = defaultdict(lambda: {'inspection': [], 'incident': [], 'switch': [], 'note': ''})

    # 정기점검
    regular_reports = Report.objects.filter(
        template__report_type='regular', status='completed', **center_filter,
    ).select_related('school')
    for r in regular_reports:
        inspect_date_str = (r.data or {}).get('inspect_date', '')
        if inspect_date_str:
            try:
                d = date.fromisoformat(inspect_date_str)
                if date_from <= d <= date_to:
                    daily_work[d.isoformat()]['inspection'].append(r.school.name)
            except (ValueError, TypeError):
                pass

    # 장애처리
    inc_qs = Incident.objects.filter(
        received_at__date__gte=date_from, received_at__date__lte=date_to,
        **center_filter,
    ).select_related('school')
    for inc in inc_qs:
        d = inc.received_at.date()
        if date_from <= d <= date_to:
            daily_work[d.isoformat()]['incident'].append(inc.school.name)

    # 스위치교체
    switch_reports = Report.objects.filter(
        template__report_type='switch_install', status='completed', **center_filter,
    ).select_related('school')
    for r in switch_reports:
        install_date_str = (r.data or {}).get('install_date', '')
        if install_date_str:
            try:
                d = date.fromisoformat(install_date_str)
                if date_from <= d <= date_to:
                    daily_work[d.isoformat()]['switch'].append(r.school.name)
            except (ValueError, TypeError):
                pass

    # 엑셀 생성
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '일자별 업무 현황'

    hdr_fill = PatternFill('solid', fgColor='1F497D')
    hdr_font = Font(bold=True, color='FFFFFF', size=10)
    ctr = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))
    WEEKDAY_NAMES = ['월', '화', '수', '목', '금', '토', '일']

    # 타이틀
    ws.merge_cells('A1:F1')
    c = ws.cell(row=1, column=1, value=f'{year}년 {month}월 월간업무보고 — 일자별 업무 현황 ({center_name})')
    c.font = Font(bold=True, size=13)
    c.alignment = ctr
    ws.row_dimensions[1].height = 28

    # 헤더
    headers = ['일자', '요일', '정기점검', '장애처리', '스위치교체', '비고']
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = ctr
        c.border = thin

    # 데이터
    row = 4
    current = date_from
    while current <= date_to:
        if current.weekday() < 5:
            key = current.isoformat()
            entry = daily_work.get(key, {'inspection': [], 'incident': [], 'switch': [], 'note': ''})
            insp = ', '.join(dict.fromkeys(entry['inspection']))
            inci = ', '.join(dict.fromkeys(entry['incident']))
            swit = ', '.join(dict.fromkeys(entry['switch']))

            vals = [current.day, WEEKDAY_NAMES[current.weekday()], insp, inci, swit, entry['note']]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=row, column=ci, value=v)
                c.border = thin
                c.alignment = ctr if ci <= 2 else Alignment(vertical='center', wrap_text=True)
            row += 1
        current += timedelta(days=1)

    # 열 너비
    col_widths = [8, 6, 30, 30, 30, 15]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f'월간업무보고_{center_name}_{year}년{month}월.xlsx'
    resp = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(fname)}"
    return resp


from .models import ReportTemplate, Report, ReportVersion, ReportSignature
from .serializers import (
    ReportTemplateSerializer, ReportListSerializer,
    ReportDetailSerializer, ReportCreateSerializer,
    ReportVersionSerializer, ReportSignatureSerializer
)
from core.permissions.roles import IsAdmin
from core.pagination import StandardPagination


class ReportTemplateViewSet(viewsets.ModelViewSet):
    """보고서 템플릿 관리"""
    serializer_class = ReportTemplateSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = ReportTemplate.objects.all()
        active_only = self.request.query_params.get('active')
        if active_only == '1':
            qs = qs.filter(is_active=True)
        report_type = self.request.query_params.get('type')
        if report_type:
            qs = qs.filter(report_type=report_type)
        return qs

    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [IsAdmin()]
        return [IsAuthenticated()]

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)


class ReportViewSet(viewsets.ModelViewSet):
    """보고서 CRUD"""
    permission_classes = [IsAuthenticated]
    pagination_class = StandardPagination
    http_method_names = ['get', 'post', 'patch', 'delete', 'head', 'options']

    def get_serializer_class(self):
        if self.action == 'create':
            return ReportCreateSerializer
        if self.action == 'retrieve':
            return ReportDetailSerializer
        return ReportListSerializer

    def get_queryset(self):
        user = self.request.user
        qs = Report.objects.select_related('school', 'template', 'incident', 'created_by')
        if user.role == 'worker':
            qs = qs.filter(created_by=user)
        school_id = self.request.query_params.get('school_id')
        if school_id:
            qs = qs.filter(school_id=school_id)
        incident_id = self.request.query_params.get('incident_id')
        if incident_id:
            qs = qs.filter(incident_id=incident_id)
        st = self.request.query_params.get('status')
        if st:
            qs = qs.filter(status=st)
        q = self.request.query_params.get('q')
        if q:
            qs = qs.filter(Q(title__icontains=q) | Q(school__name__icontains=q))
        return qs.order_by('-updated_at')

    def perform_create(self, serializer):
        """스위치 설치확인서: 해당 학교 출고등록 장비 검증 후 생성"""
        validated = serializer.validated_data
        template  = validated.get('template')
        data      = validated.get('data') or {}
        school    = validated.get('school')

        if template and template.report_type == 'switch_install':
            from rest_framework.exceptions import ValidationError
            from apps.assets.models import Asset

            # 신규 형식: devices 배열
            devices = data.get('devices') or []
            if devices:
                for i, device in enumerate(devices):
                    sn = (device.get('serial_number') or '').strip()
                    if not sn:
                        continue  # 빈 항목은 건너뜀
                    if not Asset.objects.filter(serial_number=sn).exists():
                        raise ValidationError({
                            'devices': f'장비 {i+1}: 장비 관리에 등록되지 않은 제조번호입니다: {sn}.'
                        })
            else:
                # 구형 단일 serial_number 형식 호환
                serial_number = data.get('serial_number', '').strip()
                if not serial_number:
                    raise ValidationError({'serial_number': '스위치 제조번호(S/N)를 입력하세요.'})
                if not Asset.objects.filter(serial_number=serial_number).exists():
                    raise ValidationError({
                        'serial_number': f'장비 관리에 등록되지 않은 제조번호입니다: {serial_number}.'
                    })
        # ── 정기점검 보고서: 장비 수량 자동 + 확인자 정보 자동 ──
        if template and template.report_type == 'regular':
            from apps.schools.models import SchoolContact, SchoolEquipment
            if school:
                cats = list(SchoolEquipment.objects.filter(
                    school=school
                ).values_list('category', flat=True))
                data['switch_count'] = sum(1 for c in cats if '스위치' in c and 'PoE' not in c)
                data['poe_count'] = sum(1 for c in cats if 'PoE' in c)
                data['ap_count'] = sum(1 for c in cats if 'AP' in c or '무선' in c)
                # 분기 자동 계산
                if not data.get('quarter'):
                    from django.utils import timezone
                    month = timezone.localtime(timezone.now()).month
                    # 사업 기간: 5~6월=2분기, 7~9월=3분기, 10~12월=4분기
                    if month <= 6:
                        data['quarter'] = '2'
                    elif month <= 9:
                        data['quarter'] = '3'
                    else:
                        data['quarter'] = '4'
                # 확인자: 학교 담당자(선생님) DB 우선
                if not data.get('signature_school') or not data['signature_school'].get('name'):
                    contact = SchoolContact.objects.filter(school=school).first()
                    if contact:
                        data.setdefault('signature_school', {})
                        data['signature_school']['org'] = school.name
                        data['signature_school']['name'] = contact.name or ''
                        data['signature_school']['phone'] = contact.phone or ''
                # 점검자: 로그인 사용자 정보
                user = self.request.user
                if not data.get('signature_itl') or not data['signature_itl'].get('name'):
                    data.setdefault('signature_itl', {})
                    data['signature_itl']['org'] = '세종아이티엘 컨소시엄'
                    data['signature_itl']['name'] = user.name or user.username
                    data['signature_itl']['phone'] = getattr(user, 'phone', '') or ''
                validated['data'] = data

        serializer.save(created_by=self.request.user)

    @action(detail=False, methods=['get'])
    def school_switch_assets(self, request):
        """스위치 설치확인서용 - 해당 학교의 출고등록 장비 목록"""
        from apps.assets.models import AssetOutbound
        school_id = request.query_params.get('school_id')
        center_id = request.query_params.get('center_id')
        if not school_id:
            return Response({'error': 'school_id 필요'}, status=400)

        qs = AssetOutbound.objects.filter(
            to_location_type='school',
            to_school_id=school_id,
        ).select_related('asset__asset_model', 'from_center').order_by('-outbound_date')

        if center_id:
            qs = qs.filter(from_center_id=center_id)

        # 장비별 최신 출고 기록만 (중복 제거)
        seen = set()
        assets = []
        for ob in qs:
            a = ob.asset
            if a.id not in seen:
                seen.add(a.id)
                am = a.asset_model
                cat_code = ''
                if am and am.category:
                    cat_code = am.category.code if hasattr(am.category, 'code') else str(am.category)
                assets.append({
                    'id':             a.id,
                    'serial_number':  a.serial_number,
                    'asset_tag':      a.asset_tag or '',
                    'model_name':     am.model_name if am else '',
                    'manufacturer':   am.manufacturer if am else '',
                    'category':       cat_code,
                    'outbound_date':  ob.outbound_date.isoformat(),
                    'status':         a.get_status_display(),
                })

        return Response({'assets': assets, 'count': len(assets)})

    @action(detail=False, methods=['get'])
    def school_installed_assets(self, request):
        """스위치 설치확인서용 - SchoolEquipment 기반 교체전 장비 목록"""
        from apps.schools.models import SchoolEquipment
        school_id = request.query_params.get('school_id')
        if not school_id:
            return Response({'error': 'school_id 필요'}, status=400)

        from django.db.models import Q
        qs = SchoolEquipment.objects.filter(
            Q(category__icontains='스위치') | Q(category__icontains='poe'),
            school_id=school_id,
        ).order_by('model_name', 'building', 'floor')

        assets = []
        buildings, floors, locations = set(), set(), set()
        for d in qs:
            assets.append({
                'id':               d.id,
                'model_name':       d.model_name or '',
                'manufacturer':     d.manufacturer or '',
                'building':         d.building or '',
                'floor':            d.floor or '',
                'install_location': d.install_location or '',
                'network_type':     d.network_type or '',
                'device_id':        d.device_id or '',
                'category':         d.category or '',
            })
            # 건물/층/설치위치 목록 — 스위치·PoE 장비에서만 수집 (AP 제외)
            if d.building:         buildings.add(d.building)
            if d.floor:            floors.add(d.floor)
            if d.install_location: locations.add(d.install_location)

        return Response({
            'assets':    assets,
            'count':     len(assets),
            'buildings': sorted(buildings),
            'floors':    sorted(floors),
            'locations': sorted(locations),
        })

    @action(detail=False, methods=['get'])
    def school_vsdx_data(self, request):
        """스위치 설치확인서용 - 학교 VSDX 건물/층/교실 데이터"""
        from apps.schools.models import SchoolBuilding, SchoolFloor, SchoolRoom
        school_id = request.query_params.get('school_id')
        if not school_id:
            return Response({'error': 'school_id 필요'}, status=400)
        buildings = []
        for b in SchoolBuilding.objects.filter(school_id=school_id).order_by('order', 'name'):
            floors = []
            for f in SchoolFloor.objects.filter(building=b).order_by('floor_num'):
                rooms = list(
                    SchoolRoom.objects.filter(floor=f)
                    .values('id', 'name', 'room_number', 'room_type')
                    .order_by('room_number', 'name')
                )
                floors.append({
                    'id': f.id,
                    'floor_num': f.floor_num,
                    'floor_name': f.floor_name or f'{f.floor_num}층',
                    'rooms': rooms,
                })
            buildings.append({'id': b.id, 'name': b.name, 'floors': floors})
        return Response({'buildings': buildings})

    @action(detail=False, methods=['get', 'post'])
    def school_contacts(self, request):
        """스위치 설치확인서용 - 학교 담당자 목록 / 새 담당자 등록"""
        from apps.schools.models import SchoolContact
        school_id = (request.query_params.get('school_id')
                     if request.method == 'GET'
                     else request.data.get('school_id'))
        if not school_id:
            return Response({'error': 'school_id 필요'}, status=400)
        if request.method == 'POST':
            name = request.data.get('name', '').strip()
            phone = request.data.get('phone', '').strip()
            position = request.data.get('position', '').strip()
            if not name:
                return Response({'error': '이름을 입력하세요.'}, status=400)
            contact = SchoolContact.objects.create(
                school_id=school_id, name=name, phone=phone, position=position
            )
            return Response({'id': contact.id, 'name': contact.name,
                             'phone': contact.phone, 'position': contact.position})
        contacts = list(
            SchoolContact.objects.filter(school_id=school_id)
            .values('id', 'name', 'phone', 'position', 'is_primary')
            .order_by('-is_primary', 'name')
        )
        return Response({'contacts': contacts})

    @action(detail=False, methods=['get'])
    def validate_serial(self, request):
        """스위치 설치확인서용 제조번호 실시간 검증"""
        from apps.assets.models import Asset
        from apps.assets.serializers import AssetListSerializer
        serial = request.query_params.get('serial', '').strip()
        if not serial:
            return Response({'valid': False, 'message': '제조번호를 입력하세요.'})
        asset = Asset.objects.filter(serial_number=serial).select_related(
            'asset_model', 'current_school'
        ).first()
        if not asset:
            return Response({'valid': False, 'message': f'등록되지 않은 제조번호: {serial}'})
        return Response({
            'valid':        True,
            'asset_tag':    asset.asset_tag,
            'model_name':   asset.asset_model.model_name if asset.asset_model else '',
            'status':       asset.get_status_display(),
            'school':       asset.current_school.name if asset.current_school else '창고',
            'message':      '확인됨',
        })

    @action(detail=True, methods=['post'])
    def save_version(self, request, pk=None):
        """현재 데이터를 버전으로 저장"""
        report = self.get_object()
        last_version = report.versions.order_by('-version').first()
        next_v = (last_version.version + 1) if last_version else 1
        note = request.data.get('note', '')
        data_snapshot = request.data.get('data', report.data)
        # data 업데이트
        report.data = data_snapshot
        report.save(update_fields=['data', 'updated_at'])
        version = ReportVersion.objects.create(
            report=report,
            version=next_v,
            data=data_snapshot,
            saved_by=request.user,
            note=note,
        )
        return Response(ReportVersionSerializer(version).data)

    @action(detail=True, methods=['post'])
    def complete(self, request, pk=None):
        """보고서 완료 처리"""
        report = self.get_object()
        if report.status == 'completed':
            return Response({'error': '이미 완료된 보고서입니다.'}, status=status.HTTP_400_BAD_REQUEST)
        report.status = 'completed'
        report.completed_at = timezone.now()
        report.save(update_fields=['status', 'completed_at'])
        # PDF 생성 비동기
        from .tasks import generate_report_pdf_task
        generate_report_pdf_task.delay(report.id)
        # 정기점검 보고서 → WBS 진척 자동 연동
        if report.template.report_type == 'regular':
            _sync_wbs_regular_inspect(report)
        # 스위치/AP 설치 확인서 → 학교 선생님 정보 자동 등록
        if report.template.report_type == 'switch_install':
            _auto_register_school_contact(report)
        # 근태기록부에 업무 자동 기록
        _record_attendance_work(report)
        return Response(ReportDetailSerializer(report).data)

    @action(detail=True, methods=['post'])
    def add_signature(self, request, pk=None):
        """전자서명 추가"""
        report = self.get_object()
        sig_data = request.data.get('signature_data')
        if not sig_data:
            return Response({'error': '서명 데이터가 필요합니다.'}, status=status.HTTP_400_BAD_REQUEST)
        sig = ReportSignature.objects.create(
            report=report,
            signer=request.user,
            signer_name=request.data.get('signer_name', request.user.name),
            role=request.data.get('role', ''),
            signature_data=sig_data,
        )
        return Response(ReportSignatureSerializer(sig).data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['get'])
    def download_pdf(self, request, pk=None):
        """PDF 다운로드"""
        report = self.get_object()
        if not report.pdf_path or not os.path.exists(report.pdf_path):
            return Response({'error': 'PDF가 아직 생성되지 않았습니다.'}, status=status.HTTP_404_NOT_FOUND)
        import os as _os
        dl_filename = _os.path.basename(report.pdf_path)
        return FileResponse(
            open(report.pdf_path, 'rb'),
            as_attachment=True,
            filename=dl_filename,
            content_type='application/pdf'
        )

    @action(detail=True, methods=['post'])
    def generate_pdf(self, request, pk=None):
        """PDF 즉시 생성 요청"""
        report = self.get_object()
        from .tasks import generate_report_pdf_task
        generate_report_pdf_task.delay(report.id)
        return Response({'message': 'PDF 생성 요청됨'})

    @action(detail=True, methods=['patch'])
    def save_data(self, request, pk=None):
        """수정 시마다 별도 새 Report 문서를 생성하고 원본에 수정이력(버전) 기록"""
        from django.db.models import Max
        from .models import ReportVersion
        origin = self.get_object()
        new_data = request.data.get('data')
        if new_data is None:
            return Response({'error': 'data 필드가 필요합니다.'}, status=status.HTTP_400_BAD_REQUEST)
        note = request.data.get('note', '').strip()

        # ── 원본에 수정이력 버전 기록 (처음 생성은 제외, 수정v1~부터)
        last_ver = origin.versions.aggregate(m=Max('version'))['m'] or 0
        ReportVersion.objects.create(
            report=origin,
            version=last_ver + 1,
            data=origin.data,   # 수정 전 스냅샷
            saved_by=request.user,
            note=note or f'수정 v{last_ver + 1}',
        )

        # ── 새 Report 문서 생성
        new_title = request.data.get('title', origin.title)
        new_report = Report.objects.create(
            template=origin.template,
            school=origin.school,
            incident=origin.incident,
            title=new_title,
            data=new_data,
            status='completed',
            created_by=request.user,
            completed_at=timezone.now(),
        )

        # PDF 생성 (비동기)
        from .tasks import generate_report_pdf_task
        generate_report_pdf_task.delay(new_report.id)

        return Response({'id': new_report.id, 'title': new_report.title, 'pdf_path': new_report.pdf_path})

    @action(detail=True, methods=['post'])
    def set_final(self, request, pk=None):
        """최종 확정: is_final=True 설정 + SchoolEquipment 자동 갱신 (switch_install 전용)"""
        report = self.get_object()

        # 같은 학교+템플릿의 기존 최종확정 해제
        Report.objects.filter(
            school=report.school,
            template=report.template,
            is_final=True,
        ).update(is_final=False)

        report.is_final = True
        report.save(update_fields=['is_final'])

        # switch_install 보고서만 SchoolEquipment 갱신
        if report.template.report_type == 'switch_install':
            _sync_school_equipment(report)

        return Response({'id': report.id, 'is_final': True})

    @action(detail=True, methods=['post'])
    def unset_final(self, request, pk=None):
        """최종 확정 해제"""
        report = self.get_object()
        report.is_final = False
        report.save(update_fields=['is_final'])
        return Response({'id': report.id, 'is_final': False})


    # ── 보고서 데이터 Excel/CSV 내보내기 ──────────────────────────────
    @action(detail=False, methods=['get'])
    def export(self, request):
        """서류 종류별 보고서 데이터 내보내기 (Excel / CSV)"""
        import io, csv as _csv
        from django.http import HttpResponse

        template_id = request.query_params.get('template_id')
        fmt         = request.query_params.get('fmt', 'xlsx')
        date_from   = request.query_params.get('date_from', '')
        date_to     = request.query_params.get('date_to', '')

        if not template_id:
            return Response({'error': 'template_id 필요'}, status=400)
        try:
            template = ReportTemplate.objects.get(id=template_id)
        except ReportTemplate.DoesNotExist:
            return Response({'error': '템플릿 없음'}, status=404)

        qs = Report.objects.filter(
            template=template,
        ).select_related('school', 'school__support_center', 'created_by').order_by('-created_at')

        if date_from:
            qs = qs.filter(created_at__date__gte=date_from)
        if date_to:
            qs = qs.filter(created_at__date__lte=date_to)

        rtype = template.report_type
        if rtype == 'switch_install':
            rows = _export_rows_switch_install(qs)
        elif rtype == 'cable':
            rows = _export_rows_cable(qs)
        elif rtype == 'regular':
            rows = _export_rows_regular(qs)
        else:
            rows = _export_rows_generic(qs, template)

        if not rows:
            return Response({'error': '내보낼 데이터가 없습니다.'}, status=404)

        from urllib.parse import quote
        safe_name = template.name.replace('/', '_').replace('\\', '_')

        if fmt == 'csv':
            response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
            encoded = quote(f'{safe_name}.csv')
            response['Content-Disposition'] = f"attachment; filename*=UTF-8''{encoded}"
            writer = _csv.DictWriter(response, fieldnames=list(rows[0].keys()))
            writer.writeheader()
            writer.writerows(rows)
            return response

        # Excel
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            return Response({'error': 'openpyxl 미설치'}, status=500)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = template.name[:30]

        thin = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        hdr_fill = PatternFill(fill_type='solid', fgColor='1A5FA8')
        hdr_font = Font(bold=True, color='FFFFFF', size=10)
        alt_fill = PatternFill(fill_type='solid', fgColor='EEF4FF')

        headers = list(rows[0].keys())
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.border = thin
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[1].height = 22

        for row_idx, row_data in enumerate(rows, 2):
            fill = alt_fill if row_idx % 2 == 0 else None
            for col, key in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col, value=row_data.get(key, ''))
                cell.border = thin
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                if fill:
                    cell.fill = fill

        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 8), 40)

        ws.freeze_panes = 'A2'

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        encoded = quote(f'{safe_name}.xlsx')
        response['Content-Disposition'] = f"attachment; filename*=UTF-8''{encoded}"
        return response


def _common_cols(report):
    """모든 서류 공통 컬럼값 반환"""
    return {
        '보고서ID':   report.id,
        '학교명':     report.school.name,
        '교육지원청': getattr(report.school.support_center, 'name', '') if report.school.support_center else '',
        '작성자':     (report.created_by.name or report.created_by.username) if report.created_by else '',
        '작성일시':   report.created_at.strftime('%Y-%m-%d %H:%M') if report.created_at else '',
        '완료일시':   report.completed_at.strftime('%Y-%m-%d %H:%M') if report.completed_at else '',
        '상태':       dict(Report.STATUS_CHOICES).get(report.status, report.status),
    }


def _sig_cols(data, prefix_itl='설치자', prefix_sch='확인자'):
    sig_i = data.get('signature_itl') or {}
    sig_s = data.get('signature_school') or {}
    return {
        f'{prefix_itl}_소속':  sig_i.get('org', ''),
        f'{prefix_itl}_담당자': sig_i.get('name', ''),
        f'{prefix_itl}_연락처': sig_i.get('phone', ''),
        f'{prefix_sch}_소속':  sig_s.get('org', ''),
        f'{prefix_sch}_담당자': sig_s.get('name', ''),
        f'{prefix_sch}_연락처': sig_s.get('phone', ''),
    }


def _export_rows_switch_install(qs):
    rows = []
    for rpt in qs:
        data    = rpt.data or {}
        common  = _common_cols(rpt)
        common['설치일자'] = data.get('install_date', '')
        common['문서종류'] = 'AP 설치 확인서' if data.get('doc_type') == 'ap' else '스위치 설치 확인서'
        common['비고']    = data.get('notes', '')
        sigs = _sig_cols(data)
        devices = data.get('devices') or []
        if devices:
            for i, dv in enumerate(devices, 1):
                row = dict(common)
                row['장비번호']     = i
                row['제조번호(S/N)'] = dv.get('serial_number', '')
                row['모델명']       = dv.get('model_name', '')
                row['제조사']       = dv.get('manufacturer', '')
                row['분류']         = dv.get('category', '')
                row['장비ID']       = dv.get('asset_id', '')
                row['망구분']       = dv.get('network_type', '')
                row['건물']         = dv.get('building', '')
                row['층']           = dv.get('floor', '')
                row['설치위치']     = dv.get('location', '')
                row['교체전_모델명'] = dv.get('prev_model', '')
                row['교체전_제조사'] = dv.get('prev_manufacturer', '')
                row.update(sigs)
                rows.append(row)
        else:
            row = dict(common)
            for k in ('장비번호','제조번호(S/N)','모델명','제조사','분류','장비ID','망구분','건물','층','설치위치','교체전_모델명','교체전_제조사'):
                row[k] = ''
            row.update(sigs)
            rows.append(row)
    return rows


def _export_rows_cable(qs):
    rows = []
    for rpt in qs:
        data   = rpt.data or {}
        common = _common_cols(rpt)
        common['공사일자'] = data.get('work_date', '')
        common['비고']    = data.get('notes', '')
        sigs = _sig_cols(data)
        cables = data.get('cables') or []
        if cables:
            for i, c in enumerate(cables, 1):
                sp = c.get('start_point') or {}
                ep = c.get('end_point') or {}
                wt = c.get('work_types') or []
                row = dict(common)
                row['항목번호']   = i
                row['작업내용']   = ', '.join(wt) if wt else ''
                row['케이블종류'] = c.get('cable_type', '')
                row['길이(m)']    = c.get('cable_length', '')
                row['시점_모델명']   = sp.get('model_name', '')
                row['시점_건물']     = sp.get('building', '')
                row['시점_층']       = sp.get('floor', '')
                row['시점_설치장소'] = sp.get('location', '')
                row['시점_망']       = sp.get('network', '')
                row['시점_장비ID']   = sp.get('device_id', '')
                row['시점_Port']     = sp.get('port', '')
                row['종점_모델명']   = ep.get('model_name', '')
                row['종점_건물']     = ep.get('building', '')
                row['종점_층']       = ep.get('floor', '')
                row['종점_설치장소'] = ep.get('location', '')
                row['종점_망']       = ep.get('network', '')
                row['종점_장비ID']   = ep.get('device_id', '')
                row['종점_Port']     = ep.get('port', '')
                row.update(sigs)
                rows.append(row)
        else:
            row = dict(common)
            for k in ('항목번호','작업내용','케이블종류','길이(m)',
                      '시점_모델명','시점_건물','시점_층','시점_설치장소','시점_망','시점_장비ID','시점_Port',
                      '종점_모델명','종점_건물','종점_층','종점_설치장소','종점_망','종점_장비ID','종점_Port'):
                row[k] = ''
            row.update(sigs)
            rows.append(row)
    return rows


def _export_rows_regular(qs):
    """정기점검 보고서 전용 Excel export (신규 양식)"""
    rows = []
    for rpt in qs:
        data = rpt.data or {}
        row = _common_cols(rpt)
        row['차수'] = data.get('quarter', '')
        row['점검일자'] = data.get('inspect_date', '')
        row['선생님 의견'] = data.get('teacher_opinion', '')
        row['점검/조치 결과'] = data.get('action_result', '')
        # 장비
        row['스위치 수량'] = data.get('switch_count', '')
        row['PoE 수량'] = data.get('poe_count', '')
        row['AP 수량'] = data.get('ap_count', '')
        row['WEISS'] = data.get('ap_weiss_count', '')
        row['컨트롤러'] = data.get('ap_controller_count', '')
        row['자체구축'] = data.get('ap_self_count', '')
        row['스위치 결과'] = data.get('switch_result', '')
        row['PoE 결과'] = data.get('poe_result', '')
        row['AP 결과'] = data.get('ap_result', '')
        row['무선망 대역폭'] = data.get('wireless_band', '')
        # iperf 속도
        row['교사망 망id'] = f"K{data.get('teacher_vlan_id', '')}"
        row['교사망 측정장소'] = data.get('teacher_location', '교무실')
        row['교사망 Down'] = data.get('wired_teacher_down', '')
        row['교사망 Up'] = data.get('wired_teacher_up', '')
        row['학생망 망id'] = f"H{data.get('student_vlan_id', '')}"
        row['학생망 측정장소'] = data.get('student_location', '컴퓨터실')
        row['학생망 Down'] = data.get('wired_student_down', '')
        row['학생망 Up'] = data.get('wired_student_up', '')
        row['무선망 망id'] = f"PoE{data.get('wireless_vlan_id', '')}"
        row['무선망 측정장소'] = data.get('wireless_location', '교실')
        row['무선망 Down'] = data.get('wired_wireless_down', '')
        row['무선망 Up'] = data.get('wired_wireless_up', '')
        # Wifi 360
        row['Wifi360 APid'] = data.get('wifi360_ap_id', '')
        row['Wifi360 측정장소'] = data.get('wifi360_location', '')
        row['Wifi360 Down'] = data.get('wifi360_down', '')
        row['Wifi360 Up'] = data.get('wifi360_up', '')
        # 전산실환경
        row['전원콘센트 수량'] = data.get('outlet_count', '')
        row['패치판넬 수'] = data.get('patch_panel_count', '')
        row['케이블 타입'] = data.get('patch_cable_type', '')
        # 전화망
        row['전화 구분'] = data.get('phone_type', '')
        row['전화 구성'] = data.get('phone_config', '')
        # 케이블
        row['교무실 케이블'] = data.get('cable_office', '')
        row['컴퓨터실 케이블'] = data.get('cable_computer', '')
        # 기타
        row['전자칠판'] = data.get('smartboard_result', '')
        row['디벗'] = data.get('devit_result', '')
        row['기타사항'] = data.get('etc_request', '')
        # 차수별 점검
        q = str(data.get('quarter', ''))
        if q == '1':
            row['1차-구성도'] = data.get('chk_1_diagram', '')
            row['1차-랙실장도'] = data.get('chk_1_rack', '')
            row['1차-장비목록'] = data.get('chk_1_equipment', '')
        elif q == '2':
            row['2차-자산라벨'] = data.get('chk_2_asset_label', '')
            row['2차-스위치교체'] = data.get('chk_2_switch_replace', '')
        elif q == '3':
            row['3차-AP배치도'] = data.get('chk_3_ap_layout', '')
        # 서명
        row.update(_sig_cols(data))
        rows.append(row)
    return rows


def _export_rows_generic(qs, template):
    fields = (template.fields_schema or {}).get('fields', [])
    rows = []
    for rpt in qs:
        data = rpt.data or {}
        row  = _common_cols(rpt)
        for f in fields:
            row[f.get('label', f.get('name', ''))] = data.get(f['name'], '')
        rows.append(row)
    return rows


def _sync_school_equipment(report):
    """
    스위치 설치확인서 최종 확정 시 SchoolEquipment 갱신.
    - 교체전 장비(prev_model) → 해당 학교 SchoolEquipment 삭제
    - 신규 장비 → SchoolEquipment 신규 등록
    """
    from apps.schools.models import SchoolEquipment

    CATEGORY_MAP = {
        'switch':     '스위치',
        'poe_switch': 'PoE스위치',
        'ap':         'AP',
        'router':     '라우터',
        'server':     '서버',
    }

    data = report.data or {}
    devices = data.get('devices') or []

    for device in devices:
        prev_model = (device.get('prev_model') or '').strip()
        building   = (device.get('building') or '').strip()
        floor      = (device.get('floor') or '').strip()
        location   = (device.get('location') or '').strip()
        new_model  = (device.get('model_name') or '').strip()
        new_mfr    = (device.get('manufacturer') or '').strip()
        new_sn     = (device.get('serial_number') or '').strip()
        asset_id   = (device.get('asset_id') or '').strip()
        cat_code   = (device.get('category') or '').strip()
        category   = CATEGORY_MAP.get(cat_code, new_model or '스위치')

        # ── 교체전 장비 삭제 (신규설치 제외)
        if prev_model and prev_model not in ('신규설치', '-', ''):
            qs = SchoolEquipment.objects.filter(
                school=report.school,
                model_name=prev_model,
            )
            # 위치가 일치하면 정확히 삭제, 없으면 모델명만으로 첫 번째 삭제
            loc_qs = qs.filter(building=building, floor=floor) if building else qs
            if loc_qs.exists():
                loc_qs.first().delete()
            elif qs.exists():
                qs.first().delete()

        # ── 새 장비 등록 (S/N 기준 중복 방지)
        if new_model:
            defaults = {
                'category':         category,
                'model_name':       new_model,
                'manufacturer':     new_mfr,
                'building':         building,
                'floor':            floor,
                'install_location': location,
                'device_id':        asset_id,
            }
            if new_sn:
                # asset_tag(serial) 기준 upsert
                SchoolEquipment.objects.update_or_create(
                    school=report.school,
                    device_id=asset_id if asset_id else new_sn,
                    defaults=defaults,
                )
            else:
                SchoolEquipment.objects.create(school=report.school, **defaults)


def _record_attendance_work(report):
    """보고서 완료 시 작성자의 근태기록에 업무 내용 자동 기록"""
    from apps.workforce.models import AttendanceLog
    import logging
    logger = logging.getLogger(__name__)
    try:
        user = report.created_by
        if not user:
            return
        work_date = timezone.localdate()
        rtype_label = {
            'switch_install': '스위치 설치확인서',
            'cable': '소규모 네트워크 포설',
            'regular': '정기점검 보고서',
        }.get(report.template.report_type, report.template.name)
        work_note = f'{rtype_label} - {report.school.name}'

        log, created = AttendanceLog.objects.get_or_create(
            worker=user,
            work_date=work_date,
            defaults={'status': 'normal', 'note': work_note}
        )
        if not created:
            # 기존 기록에 업무 추가 (중복 방지)
            if work_note not in (log.note or ''):
                log.note = f'{log.note}\n{work_note}'.strip() if log.note else work_note
                log.save(update_fields=['note'])
        logger.info(f'근태 업무 기록: {user.name} / {work_date} / {work_note}')
    except Exception as e:
        logger.warning(f'근태 업무 기록 실패: {e}')


def _auto_register_school_contact(report):
    """스위치/AP 설치 확인서 완료 시 signature_school 정보를 SchoolContact에 자동 등록.
    이름+전화번호가 같은 연락처가 이미 존재하면 건너뜀."""
    from apps.schools.models import SchoolContact
    import logging
    logger = logging.getLogger(__name__)

    data = report.data or {}
    sig = data.get('signature_school') or {}
    name = (sig.get('name') or '').strip()
    phone = (sig.get('phone') or '').strip()
    position = (sig.get('position') or '').strip()

    if not name:
        return

    # 동일 학교·이름·전화번호 조합이 이미 있으면 등록 생략
    exists = SchoolContact.objects.filter(
        school=report.school,
        name=name,
        phone=phone,
    ).exists()
    if exists:
        return

    SchoolContact.objects.create(
        school=report.school,
        name=name,
        phone=phone,
        position=position or '담당자',
    )
    logger.info(
        f'SchoolContact 자동 등록: school={report.school.name}, name={name}, phone={phone}'
    )


def _sync_wbs_regular_inspect(report):
    """
    정기점검 보고서 완료 시:
    1) 진척관리(SchoolInspection) — 해당 차수 학교 점검 완료 처리
    2) WBS 2.3.1 / 2.3.2 / 2.3.3 진척률 자동 계산
    """
    from datetime import date
    from apps.wbs.models import WBSItem
    from apps.audit.models import AuditProject
    from apps.schools.models import School
    from apps.progress.models import InspectionPlan, SchoolInspection

    # 보고서 점검일 결정 (data.inspect_date → completed_at → today)
    inspect_date_str = (report.data or {}).get('inspect_date', '')
    try:
        inspect_date = date.fromisoformat(inspect_date_str) if inspect_date_str else None
    except ValueError:
        inspect_date = None
    if not inspect_date:
        inspect_date = (report.completed_at.date() if report.completed_at
                        else date.today())

    # ── 1) 진척관리 SchoolInspection 자동 완료 ──
    # 점검일이 포함되는 차수(InspectionPlan)를 찾아서 해당 학교 점검을 완료 처리
    if report.school_id:
        matching_plans = InspectionPlan.objects.filter(
            plan_type='regular',
            start_date__lte=inspect_date,
            end_date__gte=inspect_date,
            status__in=['active', 'draft'],
        )
        for plan in matching_plans:
            si = SchoolInspection.objects.filter(
                plan=plan, school_id=report.school_id
            ).first()
            if si and si.status != 'completed':
                si.status = 'completed'
                si.completed_date = inspect_date
                si.report = report
                si.save(update_fields=['status', 'completed_date', 'report', 'updated_at'])
                logger.info('진척관리 자동완료: plan=%s, school=%s, date=%s',
                            plan.name, report.school_id, inspect_date)

    # ── 2) WBS 진척률 자동 계산 ──
    project = AuditProject.objects.filter(is_active=True).first()
    if not project:
        return

    total_schools = School.objects.filter(is_active=True).count()
    if total_schools == 0:
        return

    WBS_CODES = ['2.3.1', '2.3.2', '2.3.3']

    for code in WBS_CODES:
        try:
            item = WBSItem.objects.get(project=project, code=code)
        except WBSItem.DoesNotExist:
            continue
        p_start = item.planned_start
        p_end = item.planned_end
        if not p_start or not p_end:
            continue
        if p_start <= inspect_date <= p_end:
            # 해당 기간 내 완료된 정기점검 보고서의 고유 학교 수
            completed_schools = Report.objects.filter(
                template__report_type='regular',
                status='completed',
                completed_at__date__gte=p_start,
                completed_at__date__lte=p_end,
            ).values('school_id').distinct().count()

            progress = min(100, round(completed_schools / total_schools * 100))

            item.progress = progress
            if item.actual_start is None:
                item.actual_start = inspect_date
            if progress >= 100:
                item.actual_end = inspect_date
            item.save(update_fields=['progress', 'actual_start', 'actual_end', 'updated_at'])
            try:
                from apps.wbs.signals import _bubble_up
                _bubble_up(item)
            except Exception as e:
                logger.warning('WBS 진척 버블업 실패 item=%s: %s', item.pk, e)
            break


# ═══════════════════════════════════════════════════════════
# 월간업무보고 데이터 API
# ═══════════════════════════════════════════════════════════
@login_required
def monthly_work_report_api(request):
    """월간업무보고 데이터 API — GET ?year=2026&month=4&center=dongbu"""
    import calendar as cal_mod
    from collections import defaultdict
    from datetime import date
    from django.http import JsonResponse
    from django.db.models import Count, Q
    from apps.schools.models import School, SupportCenter, SchoolType
    from apps.incidents.models import Incident, IncidentCategory
    from apps.progress.models import SchoolInspection
    from apps.reports.models import Report

    year  = int(request.GET.get('year', 2026))
    month = int(request.GET.get('month', 4))
    center_code = request.GET.get('center', '')
    is_all = (not center_code or center_code == 'all')

    # 지원청 필터 준비
    center_filter = {}       # school__support_center 용
    school_filter = {}       # support_center 용
    if is_all:
        sc = None
        center_name = '전체'
    else:
        sc = SupportCenter.objects.filter(code=center_code).first()
        if not sc:
            return JsonResponse({'error': '유효하지 않은 교육지원청입니다.'}, status=400)
        center_name = sc.name
        center_filter = {'school__support_center': sc}
        school_filter = {'support_center': sc}

    date_from = date(year, month, 1)
    date_to = date(year, month, cal_mod.monthrange(year, month)[1])

    # ── 1. 사업 개요: 학제별 학교 수 ──
    school_types = list(SchoolType.objects.order_by('order').values('code', 'name'))

    # 전체: 지원청별 크로스탭 생성
    centers_list = []
    if is_all:
        centers_qs = SupportCenter.objects.filter(is_active=True).order_by('id')
        for ctr in centers_qs:
            ctr_counts = dict(
                School.objects.filter(support_center=ctr, is_active=True)
                .values_list('school_type__code').annotate(cnt=Count('id'))
            )
            ctr_total = sum(ctr_counts.values())
            row = {'code': ctr.code, 'name': ctr.name, 'total': ctr_total}
            for st in school_types:
                row[st['code']] = ctr_counts.get(st['code'], 0)
            centers_list.append(row)

    school_counts = dict(
        School.objects.filter(is_active=True, **school_filter)
        .values_list('school_type__code').annotate(cnt=Count('id'))
    )
    total_schools = 0
    for st in school_types:
        cnt = school_counts.get(st['code'], 0)
        st['count'] = cnt
        total_schools += cnt
    overview = {
        'center_name': center_name,
        'project_period': '2025. 05. 06. ~ 2026. 12. 31.',
        'school_types': school_types,
        'total_schools': total_schools,
        'centers': centers_list,
    }

    # ── 2. 정기점검 일정: 해당월 점검 예정일 + 학교명 ──
    insp_qs = SchoolInspection.objects.filter(
        scheduled_date__gte=date_from,
        scheduled_date__lte=date_to,
        **center_filter,
    ).select_related('school')
    # 날짜별 학교명 매핑
    from collections import defaultdict as _dd
    insp_by_date = _dd(list)
    for si in insp_qs:
        if si.scheduled_date:
            insp_by_date[si.scheduled_date].append(si.school.name)
    inspection_dates = sorted(insp_by_date.keys())

    # 캘린더 데이터 (월~금 그리드)
    cal_weeks = []
    c = cal_mod.Calendar(firstweekday=0)  # 월요일 시작
    for week in c.monthdatescalendar(year, month):
        week_data = []
        for d in week:
            if d.month == month and d.weekday() < 5:  # 평일만
                schools = sorted(set(insp_by_date.get(d, [])))
                week_data.append({
                    'day': d.day,
                    'is_inspection': d in insp_by_date,
                    'schools': schools,
                })
            elif d.weekday() < 5:
                week_data.append(None)  # 다른 달
        # 주말 제외 — 5일만
        cal_weeks.append(week_data)
    inspection_schedule = {
        'weeks': cal_weeks,
        'inspection_dates': [d.isoformat() for d in inspection_dates],
    }

    # ── 3. 장애처리 현황 ──
    inc_qs = Incident.objects.filter(
        received_at__date__gte=date_from,
        received_at__date__lte=date_to,
        **center_filter,
    )

    # 학제별
    by_school_type = []
    st_counts = dict(
        inc_qs.values_list('school__school_type__code')
        .annotate(cnt=Count('id'))
    )
    inc_total = 0
    for st in school_types:
        cnt = st_counts.get(st['code'], 0)
        by_school_type.append({'code': st['code'], 'name': st['name'], 'count': cnt})
        inc_total += cnt

    # 전체: 학제별 × 지원청 크로스탭
    inc_centers_cross = []
    if is_all:
        centers_qs = SupportCenter.objects.filter(is_active=True).order_by('id')
        for ctr in centers_qs:
            ctr_inc = inc_qs.filter(school__support_center=ctr)
            ctr_st_counts = dict(
                ctr_inc.values_list('school__school_type__code').annotate(cnt=Count('id'))
            )
            row = {'code': ctr.code, 'name': ctr.name, 'total': sum(ctr_st_counts.values())}
            for st in school_types:
                row[st['code']] = ctr_st_counts.get(st['code'], 0)
            inc_centers_cross.append(row)

    # 장애 분류별
    categories = list(IncidentCategory.objects.filter(is_active=True).order_by('order').values('code', 'name'))
    cat_counts = dict(
        inc_qs.values_list('category__code')
        .annotate(cnt=Count('id'))
    )
    cat_total = 0
    for cat in categories:
        cnt = cat_counts.get(cat['code'], 0)
        cat['count'] = cnt
        cat_total += cnt

    incidents_summary = {
        'by_school_type': by_school_type,
        'total_by_school_type': inc_total,
        'by_category': categories,
        'total_by_category': cat_total,
        'centers_cross': inc_centers_cross,
    }

    # ── 4. 일자별 업무 현황 ──
    # 각 항목은 {'name': 학교명, 'id': 보고서/장애 ID, 'type': 'report'|'incident'} 형태
    daily_work = defaultdict(lambda: {
        'inspection': [], 'incident': [], 'switch': [], 'note': ''
    })

    # (a) 정기점검: Report(regular) → data['inspect_date'] 기준
    regular_reports = Report.objects.filter(
        template__report_type='regular',
        status='completed',
        **center_filter,
    ).select_related('school')
    for r in regular_reports:
        inspect_date_str = (r.data or {}).get('inspect_date', '')
        if inspect_date_str:
            try:
                d = date.fromisoformat(inspect_date_str)
                if date_from <= d <= date_to:
                    daily_work[d.isoformat()]['inspection'].append({
                        'name': r.school.name, 'id': r.id, 'type': 'report',
                    })
            except (ValueError, TypeError):
                pass

    # (b) 장애처리: Incident → received_at (장애 발생일) 기준
    # NAS에서 장애처리보고서 파일 매핑 (파일명: 장애처리보고서_{학교명}_{incident_number}.pdf)
    from apps.nas.models import File as NasFile
    inc_list = list(inc_qs.select_related('school'))
    nas_file_map = {}  # incident_number → nas_file_id
    if inc_list:
        inc_numbers = [inc.incident_number for inc in inc_list]
        expected_names = [
            f'장애처리보고서_{inc.school.name}_{inc.incident_number}.pdf'
            for inc in inc_list
        ]
        for nf_id, nf_name in NasFile.objects.filter(
            original_name__in=expected_names, is_deleted=False,
        ).values_list('id', 'original_name'):
            nas_file_map[nf_name] = nf_id

    for inc_obj in inc_list:
        d = inc_obj.received_at.date()
        if date_from <= d <= date_to:
            expected_name = f'장애처리보고서_{inc_obj.school.name}_{inc_obj.incident_number}.pdf'
            nas_id = nas_file_map.get(expected_name)
            daily_work[d.isoformat()]['incident'].append({
                'name': inc_obj.school.name,
                'id': nas_id or inc_obj.id,
                'type': 'nas' if nas_id else 'incident',
            })

    # (c) 스위치교체: Report(switch_install) → data['install_date'] 기준
    switch_reports = Report.objects.filter(
        template__report_type='switch_install',
        status='completed',
        **center_filter,
    ).select_related('school')
    for r in switch_reports:
        install_date_str = (r.data or {}).get('install_date', '')
        if install_date_str:
            try:
                d = date.fromisoformat(install_date_str)
                if date_from <= d <= date_to:
                    daily_work[d.isoformat()]['switch'].append({
                        'name': r.school.name, 'id': r.id, 'type': 'report',
                    })
            except (ValueError, TypeError):
                pass

    # 평일만 정렬하여 리스트로 변환 (학교명 중복 제거, 순서 유지)
    from datetime import timedelta
    WEEKDAY_NAMES = ['월', '화', '수', '목', '금', '토', '일']

    def _dedup_items(items):
        """학교명 기준 중복 제거 (순서 유지)"""
        seen = set()
        result = []
        for item in items:
            if item['name'] not in seen:
                seen.add(item['name'])
                result.append(item)
        return result

    daily_list = []
    current = date_from
    while current <= date_to:
        if current.weekday() < 5:  # 평일
            key = current.isoformat()
            entry = daily_work.get(key, {'inspection': [], 'incident': [], 'switch': [], 'note': ''})
            daily_list.append({
                'date': key,
                'day': current.day,
                'weekday': WEEKDAY_NAMES[current.weekday()],
                'inspection': _dedup_items(entry['inspection']),
                'incident': _dedup_items(entry['incident']),
                'switch': _dedup_items(entry['switch']),
                'note': entry['note'],
            })
        current += timedelta(days=1)

    return JsonResponse({
        'year': year,
        'month': month,
        'center_code': center_code or 'all',
        'center_name': center_name,
        'is_all': is_all,
        'overview': overview,
        'inspection_schedule': inspection_schedule,
        'incidents': incidents_summary,
        'daily_work': daily_list,
    })

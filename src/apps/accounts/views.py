import csv
import io
from django.shortcuts import render, redirect
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.utils.decorators import method_decorator
from django.views import View
from rest_framework import viewsets, status, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework_simplejwt.views import TokenObtainPairView
from core.throttling import LoginRateThrottle
from django.utils import timezone
from django.db.models import Q
from .models import User, UserSession, LoginHistory, UserActivityLog
from .serializers import (
    UserListSerializer, UserDetailSerializer, UserCreateSerializer,
    UserUpdateSerializer, PasswordChangeSerializer,
    CustomTokenObtainPairSerializer, UserSessionSerializer
)
from core.permissions.roles import IsAdmin, IsSuperAdmin


@login_required
def data_management_view(request):
    """CSV 데이터 관리 페이지 (관리자 전용)"""
    return render(request, 'admin/data_management.html')


# ─────────────────────────────────────────
# 템플릿 기반 뷰 (로그인/로그아웃)
# ─────────────────────────────────────────
class LoginView(View):
    template_name = 'accounts/login.html'

    def get(self, request):
        if request.user.is_authenticated:
            return redirect('dashboard:index')
        return render(request, self.template_name)

    def post(self, request):
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')

        from core.utils.network import get_client_ip
        ip = get_client_ip(request)
        ua = request.META.get('HTTP_USER_AGENT', '')

        # ── 계정 잠금 확인 (Redis 캐시) ─────────────────────────
        from django.core.cache import cache
        LOCK_KEY   = f'login_lock:{username}'
        FAIL_KEY   = f'login_fail:{username}'
        MAX_FAILS  = 5
        LOCK_SEC   = 30 * 60   # 잠금 30분
        WINDOW_SEC = 10 * 60   # 실패 카운트 창 10분

        if cache.get(LOCK_KEY):
            remaining = cache.ttl(LOCK_KEY) // 60 + 1
            return render(request, self.template_name, {
                'error': f'로그인 시도가 너무 많습니다. {remaining}분 후 다시 시도해 주세요.'
            })

        user = authenticate(request, username=username, password=password)

        if user is None:
            # 실패 횟수 증가
            fail_count = cache.get(FAIL_KEY, 0) + 1
            cache.set(FAIL_KEY, fail_count, WINDOW_SEC)
            if fail_count >= MAX_FAILS:
                cache.set(LOCK_KEY, True, LOCK_SEC)
                cache.delete(FAIL_KEY)
                fail_reason = f'로그인 {MAX_FAILS}회 실패 — 계정 잠금'
            else:
                fail_reason = f'인증 실패 ({fail_count}/{MAX_FAILS})'
            failed_user = User.objects.filter(username=username).first()
            LoginHistory.objects.create(
                user=failed_user, attempted_username=username,
                ip_address=ip, user_agent=ua, success=False, fail_reason=fail_reason
            )
            UserActivityLog.objects.create(
                user=failed_user, action='login',
                detail=f'로그인 실패: {fail_reason}',
                ip_address=ip,
            )
            if fail_count >= MAX_FAILS:
                return render(request, self.template_name, {
                    'error': f'로그인 {MAX_FAILS}회 연속 실패로 30분간 잠금됩니다.'
                })
            return render(request, self.template_name, {
                'error': f'아이디 또는 비밀번호가 올바르지 않습니다. ({fail_count}/{MAX_FAILS}회 실패)'
            })

        if not user.is_service_active():
            return render(request, self.template_name, {'error': '서비스 이용 기간이 만료되었습니다.'})

        # 로그인 성공 — 실패 카운터 초기화
        cache.delete(FAIL_KEY)
        cache.delete(LOCK_KEY)

        login(request, user)
        LoginHistory.objects.create(user=user, attempted_username=username, ip_address=ip, user_agent=ua, success=True)
        UserActivityLog.objects.create(
            user=user, action='login',
            detail='로그인 성공',
            ip_address=ip,
        )

        # 세션 등록
        UserSession.objects.update_or_create(
            session_key=request.session.session_key,
            defaults={'user': user, 'ip_address': ip, 'user_agent': ua, 'is_active': True}
        )
        next_url = request.POST.get('next') or request.GET.get('next') or '/npms/'
        return redirect(next_url)


class LogoutView(View):
    def post(self, request):
        if request.user.is_authenticated:
            UserSession.objects.filter(
                session_key=request.session.session_key
            ).update(is_active=False)
            from core.utils.network import get_client_ip
            UserActivityLog.objects.create(
                user=request.user, action='logout',
                detail='로그아웃',
                ip_address=get_client_ip(request),
            )
        logout(request)
        return redirect('accounts:login')


# ─────────────────────────────────────────
# REST API 뷰
# ─────────────────────────────────────────
class CustomTokenObtainPairView(TokenObtainPairView):
    serializer_class = CustomTokenObtainPairSerializer
    throttle_classes = [LoginRateThrottle]
    _ratelimit_config = True  # 레이트 리미팅 적용 표시

    def post(self, request, *args, **kwargs):
        # JWT API 5회 실패 잠금
        from django.core.cache import cache
        username = request.data.get('username', '')
        LOCK_KEY   = f'jwt_lock:{username}'
        FAIL_KEY   = f'jwt_fail:{username}'
        MAX_FAILS  = 5
        LOCK_SEC   = 30 * 60
        WINDOW_SEC = 10 * 60

        if cache.get(LOCK_KEY):
            from rest_framework.response import Response
            from rest_framework import status
            return Response(
                {'detail': '로그인 시도가 너무 많습니다. 잠시 후 다시 시도해 주세요.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            response = super().post(request, *args, **kwargs)
        except Exception:
            # 인증 실패 — 실패 카운터 증가
            fail_count = cache.get(FAIL_KEY, 0) + 1
            cache.set(FAIL_KEY, fail_count, WINDOW_SEC)
            if fail_count >= MAX_FAILS:
                cache.set(LOCK_KEY, True, LOCK_SEC)
                cache.delete(FAIL_KEY)
            raise  # DRF 예외 핸들러로 전달

        # 성공 — 카운터 초기화
        cache.delete(FAIL_KEY)
        cache.delete(LOCK_KEY)
        return response


class TwoFactorSetupView(View):
    """2FA TOTP 설정 (관리자·슈퍼관리자)"""

    def get(self, request):
        """TOTP QR URI 반환"""
        import pyotp
        user = request.user
        if not user.is_authenticated:
            return JsonResponse({'error': '로그인 필요'}, status=401)
        if not user.totp_secret:
            user.totp_secret = pyotp.random_base32()
            user.save(update_fields=['totp_secret'])
        totp = pyotp.TOTP(user.totp_secret)
        uri = totp.provisioning_uri(name=user.username, issuer_name='NPMS')
        return JsonResponse({'uri': uri, 'secret': user.totp_secret})

    def post(self, request):
        """TOTP 코드 검증 후 2FA 활성화"""
        import pyotp, json as _json
        user = request.user
        if not user.is_authenticated:
            return JsonResponse({'error': '로그인 필요'}, status=401)
        try:
            body = _json.loads(request.body)
        except Exception:
            return JsonResponse({'error': '잘못된 요청'}, status=400)
        code = body.get('code', '')
        if not user.totp_secret:
            return JsonResponse({'error': 'TOTP 설정이 없습니다.'}, status=400)
        totp = pyotp.TOTP(user.totp_secret)
        if totp.verify(code):
            user.is_2fa_enabled = True
            user.save(update_fields=['is_2fa_enabled'])
            return JsonResponse({'ok': True, 'message': '2FA가 활성화되었습니다.'})
        return JsonResponse({'error': '잘못된 코드입니다.'}, status=400)


class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.select_related('support_center').order_by('-created_at')
    permission_classes = [permissions.IsAuthenticated]

    def get_serializer_class(self):
        if self.action == 'create':
            return UserCreateSerializer
        if self.action in ('update', 'partial_update'):
            return UserUpdateSerializer
        if self.action == 'list':
            return UserListSerializer
        return UserDetailSerializer

    def get_permissions(self):
        if self.action in ('create', 'destroy', 'update'):
            return [IsAdmin()]
        return super().get_permissions()

    def get_queryset(self):
        qs = super().get_queryset()
        user = self.request.user
        # superadmin이 아니면 같은 지원청만
        if user.role not in ('superadmin', 'admin'):
            qs = qs.filter(id=user.id)
        q = self.request.query_params.get('q')
        role = self.request.query_params.get('role')
        center = self.request.query_params.get('center')
        if q:
            qs = qs.filter(Q(name__icontains=q) | Q(username__icontains=q) | Q(phone__icontains=q))
        if role:
            qs = qs.filter(role=role)
        if center:
            qs = qs.filter(support_center_id=center)
        return qs

    @action(detail=False, methods=['get', 'patch'], permission_classes=[permissions.IsAuthenticated])
    def me(self, request):
        """현재 로그인 사용자 정보"""
        if request.method == 'GET':
            return Response(UserDetailSerializer(request.user).data)
        serializer = UserUpdateSerializer(request.user, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(UserDetailSerializer(request.user).data)

    @action(detail=False, methods=['get'], permission_classes=[permissions.IsAuthenticated])
    def center_staff(self, request):
        """소속 지원청 인력 목록. 모달 인계자/인수자 선택용.
        - support_center 있는 경우: 해당 센터 worker 목록
        - support_center 없는 경우: center 쿼리파라미터(센터 ID)로 조회
        """
        center = request.user.support_center
        if not center:
            center_id = request.query_params.get('center')
            if center_id:
                from apps.schools.models import SupportCenter
                try:
                    center = SupportCenter.objects.get(id=center_id)
                except SupportCenter.DoesNotExist:
                    return Response([])
            else:
                return Response([])
        qs = User.objects.filter(
            support_center=center, is_active=True
        ).exclude(role__in=['superadmin', 'customer']).order_by('name')
        data = [{'id': u.id, 'name': u.name, 'role': u.role, 'phone': u.phone} for u in qs]
        return Response(data)

    @action(detail=False, methods=['get'], permission_classes=[permissions.IsAuthenticated])
    def warehouse_staff(self, request):
        """창고 인계자/인수자 후보 목록 (관리자·상주자)."""
        qs = User.objects.filter(
            is_active=True, role__in=['superadmin', 'admin', 'resident']
        ).order_by('name')
        data = [{'id': u.id, 'name': u.name, 'role': u.role, 'phone': u.phone} for u in qs]
        return Response(data)

    @action(detail=True, methods=['post'], permission_classes=[IsAdmin])
    def set_expiry(self, request, pk=None):
        """서비스 만료일 설정 (superadmin only)"""
        user = self.get_object()
        expiry = request.data.get('service_expiry')
        user.service_expiry = expiry
        user.save()
        return Response({'service_expiry': expiry})

    @action(detail=False, methods=['post'], permission_classes=[permissions.IsAuthenticated])
    def change_password(self, request):
        serializer = PasswordChangeSerializer(data=request.data, context={'request': request})
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response({'message': '비밀번호가 변경되었습니다.'})

    @action(detail=False, methods=['get'], permission_classes=[IsAdmin])
    def excel_download(self, request):
        """인력 정보 엑셀 다운로드"""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '인력목록'

        headers = ['username', '이름', '역할', '전화번호', '이메일',
                   '소속지원청', '자택주소', '자택위도', '자택경도',
                   '서비스만료일', '활성여부', '비밀번호(신규등록용)']

        header_fill = PatternFill('solid', fgColor='1F4E79')
        header_font = Font(bold=True, color='FFFFFF', size=10)
        thin = Side(style='thin', color='CCCCCC')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        ws.row_dimensions[1].height = 20

        role_map = dict(User.ROLE_CHOICES)
        qs = User.objects.select_related('support_center').order_by('role', 'name')
        for row_idx, u in enumerate(qs, 2):
            row_data = [
                u.username,
                u.name,
                role_map.get(u.role, u.role),
                u.phone or '',
                u.email or '',
                u.support_center.name if u.support_center else '',
                u.home_address or '',
                float(u.home_lat) if u.home_lat else '',
                float(u.home_lng) if u.home_lng else '',
                str(u.service_expiry) if u.service_expiry else '',
                '활성' if u.is_active else '비활성',
                '',
            ]
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = border
                cell.alignment = Alignment(vertical='center')
                if row_idx % 2 == 0:
                    cell.fill = PatternFill('solid', fgColor='F0F4FA')

        col_widths = [16, 12, 12, 16, 28, 20, 40, 12, 12, 14, 8, 20]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}1'

        import io as _io
        buf = _io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="npms_인력목록.xlsx"'
        return response

    @action(detail=False, methods=['post'], permission_classes=[IsAdmin])
    def excel_import(self, request):
        """엑셀 일괄 등록"""
        import openpyxl
        file = request.FILES.get('file')
        if not file:
            return Response({'error': '파일을 선택하세요.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
            ws = wb.active
        except Exception as e:
            return Response({'error': f'파일을 읽을 수 없습니다: {e}'}, status=status.HTTP_400_BAD_REQUEST)

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return Response({'error': '데이터가 없습니다.'}, status=status.HTTP_400_BAD_REQUEST)

        headers = [str(h).strip() if h else '' for h in rows[0]]

        def col_raw(row, name):
            """원본 값 그대로 반환"""
            try:
                idx = headers.index(name)
                return row[idx]
            except (ValueError, IndexError):
                return None

        def col(row, name):
            """문자열로 변환"""
            v = col_raw(row, name)
            if v is None:
                return ''
            return str(v).strip()

        from apps.schools.models import SupportCenter
        import datetime
        # 역할: 코드값(worker) 또는 표시명(현장기사) 모두 허용
        role_codes = {k for k, _ in User.ROLE_CHOICES}
        role_rev   = {v: k for k, v in dict(User.ROLE_CHOICES).items()}

        created, updated, errors = 0, 0, []
        for row_num, row in enumerate(rows[1:], 2):
            if not any(row):
                continue
            username = col(row, 'username')
            if not username:
                errors.append({'row': row_num, 'error': 'username 없음'})
                continue
            try:
                center = None
                center_name = col(row, '소속지원청')
                if center_name:
                    center = SupportCenter.objects.filter(
                        Q(name=center_name) | Q(name__icontains=center_name)
                    ).first()

                # 역할: 코드값이면 그대로, 표시명이면 변환
                role_val = col(row, '역할')
                if role_val in role_codes:
                    role = role_val
                else:
                    role = role_rev.get(role_val, 'worker')

                # 활성여부: 불리언 또는 문자열 모두 처리
                active_raw = col_raw(row, '활성여부')
                if isinstance(active_raw, bool):
                    is_active = active_raw
                else:
                    is_active = str(active_raw).strip() != '비활성'

                # 서비스 만료일: datetime 또는 문자열
                expiry_raw = col_raw(row, '서비스만료일')
                expiry = None
                if expiry_raw:
                    if isinstance(expiry_raw, (datetime.datetime, datetime.date)):
                        expiry = expiry_raw.date() if hasattr(expiry_raw, 'date') else expiry_raw
                    else:
                        expiry = str(expiry_raw).strip() or None

                defaults = {
                    'name':           col(row, '이름'),
                    'phone':          col(row, '전화번호'),
                    'email':          col(row, '이메일'),
                    'role':           role,
                    'support_center': center,
                    'home_address':   col(row, '자택주소'),
                    'is_active':      is_active,
                    'service_expiry': expiry,
                }
                lat = col(row, '자택위도')
                lng = col(row, '자택경도')
                if lat: defaults['home_lat'] = lat
                if lng: defaults['home_lng'] = lng

                user, is_new = User.objects.get_or_create(username=username, defaults=defaults)
                if is_new:
                    pw = col(row, '비밀번호(신규등록용)') or 'Change1234!'
                    user.set_password(pw)
                    user.save()
                    created += 1
                else:
                    for k, v in defaults.items():
                        setattr(user, k, v)
                    user.save()
                    updated += 1
            except Exception as e:
                errors.append({'row': row_num, 'error': str(e)})

        return Response({'created': created, 'updated': updated, 'errors': errors})

    @action(detail=False, methods=['get'], permission_classes=[IsAdmin])
    def csv_template(self, request):
        """수행인력 CSV 양식 다운로드"""
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = 'attachment; filename="worker_template.csv"'
        writer = csv.writer(response)
        writer.writerow(['username', '이름', '역할', '전화번호', '이메일', '소속지원청', '비밀번호', '활성여부'])
        writer.writerow(['hong1', '홍길동', 'worker', '010-1234-5678', 'hong@example.com', '동부교육지원청', 'Pass1234!', '활성'])
        return response

    @action(detail=False, methods=['post'], permission_classes=[IsAdmin])
    def csv_import(self, request):
        """수행인력 CSV 일괄 등록 (추가/덮어쓰기)"""
        import io
        f = request.FILES.get('file')
        mode = request.data.get('mode', 'add_update')
        if not f:
            return Response({'error': 'CSV 파일을 첨부해 주세요.'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            text = f.read().decode('utf-8-sig')
        except Exception:
            return Response({'error': 'UTF-8 인코딩 파일만 지원합니다.'}, status=status.HTTP_400_BAD_REQUEST)

        reader = csv.DictReader(io.StringIO(text))
        rows = list(reader)

        if mode == 'reset':
            User.objects.filter(role__in=['worker', 'resident']).delete()

        from apps.schools.models import SupportCenter
        role_codes = {k for k, _ in User.ROLE_CHOICES}
        role_rev = {v: k for k, v in dict(User.ROLE_CHOICES).items()}

        created, updated, errors = 0, 0, []
        for row_num, row in enumerate(rows, 2):
            username = (row.get('username') or '').strip()
            if not username:
                errors.append({'row': row_num, 'error': 'username 없음'})
                continue
            try:
                center = None
                center_name = (row.get('소속지원청') or '').strip()
                if center_name:
                    center = SupportCenter.objects.filter(
                        Q(name=center_name) | Q(name__icontains=center_name)
                    ).first()
                role_val = (row.get('역할') or '').strip()
                role = role_val if role_val in role_codes else role_rev.get(role_val, 'worker')
                is_active = (row.get('활성여부') or '활성').strip() != '비활성'
                defaults = {
                    'name': (row.get('이름') or '').strip(),
                    'phone': (row.get('전화번호') or '').strip(),
                    'email': (row.get('이메일') or '').strip(),
                    'role': role,
                    'support_center': center,
                    'is_active': is_active,
                }
                user, is_new = User.objects.get_or_create(username=username, defaults=defaults)
                if is_new:
                    pwd = (row.get('비밀번호') or '').strip()
                    user.set_password(pwd if pwd else User.objects.make_random_password())
                    user.save()
                    created += 1
                else:
                    for k, v in defaults.items():
                        setattr(user, k, v)
                    user.save()
                    updated += 1
            except Exception as e:
                errors.append({'row': row_num, 'error': str(e)})
        return Response({'created': created, 'updated': updated, 'errors': errors})


class ActiveSessionViewSet(viewsets.ReadOnlyModelViewSet):
    """현재 접속 중인 사용자 목록"""
    serializer_class = UserSessionSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        from datetime import timedelta
        cutoff = timezone.now() - timedelta(minutes=30)
        return UserSession.objects.filter(
            is_active=True, last_active__gte=cutoff
        ).select_related('user').order_by('-last_active')

    @action(detail=False, methods=['post'])
    def update_page(self, request):
        """현재 접속 페이지 갱신"""
        page = request.data.get('page', '')
        UserSession.objects.filter(
            session_key=request.session.session_key
        ).update(current_page=page, last_active=timezone.now())
        return Response({'status': 'ok'})

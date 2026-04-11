"""
progress 앱 비즈니스 로직

- get_business_days()         : 기간 내 영업일 목록 (토/일/공휴일 제외)
- auto_assign()               : 학교-인력-날짜 자동 배정 (근접거리 최적화)
- _nearest_neighbor_sort()    : 근접거리 순 정렬 (greedy nearest-neighbor)
- _dist()                     : 유클리드 거리
- replace_worker()            : 인력 교체 및 업무 인수인계
- process_upload()            : Excel/CSV 업로드 → 학교 매칭 결과 반환
- confirm_upload()            : 매칭 결과 확정 → SchoolInspection 등록
"""
import csv
import io
import math
import mimetypes
import os
from datetime import date, timedelta
from difflib import SequenceMatcher

from django.conf import settings
from django.db import transaction
from django.utils import timezone


# ──────────────────────────────────────────────────
# 1. 영업일 계산
# ──────────────────────────────────────────────────
def get_business_days(start: date, end: date) -> list[date]:
    """
    start~end 사이 영업일 목록 반환
    제외 기준: 토요일(5), 일요일(6) + DB 등록 공휴일
    """
    from .models import Holiday

    # DB 공휴일 목록 수집
    holiday_dates: set[date] = set()
    year_set = {y for y in range(start.year, end.year + 1)}

    for h in Holiday.objects.filter(is_active=True):
        if h.is_recurring and h.month and h.day:
            for y in year_set:
                try:
                    holiday_dates.add(date(y, h.month, h.day))
                except ValueError:
                    pass
        elif not h.is_recurring and h.specific_date:
            holiday_dates.add(h.specific_date)

    days = []
    cur = start
    while cur <= end:
        if cur.weekday() < 5 and cur not in holiday_dates:  # 월~금
            days.append(cur)
        cur += timedelta(days=1)
    return days


# ──────────────────────────────────────────────────
# 2. 자동 배정
# ──────────────────────────────────────────────────
def create_worker_areas_from_users() -> dict:
    """
    User.support_center 기반으로 WorkerArea 자동 생성
    role='worker' 인 인력을 대상으로 함
    """
    from .models import WorkerArea
    from apps.accounts.models import User

    created = already = skipped = 0
    with transaction.atomic():
        for user in User.objects.filter(role='worker', support_center__isnull=False).select_related('support_center'):
            _, is_new = WorkerArea.objects.get_or_create(
                worker=user,
                support_center=user.support_center,
                defaults={'is_primary': True},
            )
            if is_new:
                created += 1
            else:
                already += 1
        for user in User.objects.filter(role='worker', support_center__isnull=True):
            skipped += 1

    return {'created': created, 'already_existed': already, 'skipped_no_center': skipped}


def reset_assignments(plan_id: int) -> dict:
    """scheduled 상태의 점검 항목을 pending으로 초기화 (재배정 준비)"""
    from .models import SchoolInspection
    count = SchoolInspection.objects.filter(
        plan_id=plan_id, status='scheduled'
    ).update(status='pending', scheduled_date=None, assigned_worker=None,
             work_schedule=None)
    return {'reset': count}


def _dist(lat1, lng1, lat2, lng2) -> float:
    """유클리드 거리 (서울 내 소규모 영역 적용)"""
    return ((float(lat1) - float(lat2)) ** 2 + (float(lng1) - float(lng2)) ** 2) ** 0.5


def _nearest_neighbor_sort(sis: list) -> list:
    """
    근접 거리 순 정렬 (greedy nearest-neighbor)
    학교가 몰려있는 구역을 같은 날짜·기사에 할당하기 위해 사용
    시작점: 북서쪽(위도 높고 경도 낮은) 학교
    """
    if not sis:
        return []
    remaining = list(sis)
    current = min(remaining, key=lambda s: (-float(s.school.lat), float(s.school.lng)))
    result = [current]
    remaining.remove(current)
    while remaining:
        clat = float(current.school.lat)
        clng = float(current.school.lng)
        nearest = min(remaining,
                      key=lambda s: _dist(clat, clng, s.school.lat, s.school.lng))
        result.append(nearest)
        remaining.remove(nearest)
        current = nearest
    return result


def auto_assign(plan_id: int, force_reassign: bool = False) -> dict:
    """
    점검계획 미배정 학교에 대해 인력-날짜 자동 배정 (근접거리 최적화)

    배정 기준:
      1. pending + scheduled-but-unassigned 모두 처리
      2. WorkerArea 기반 지원청별 담당 인력 결정
         (WorkerArea 미등록 시 User.support_center 폴백)
      3. 기사별 지리적 구역 분할 (위도+경도 합산 기준 등분)
      4. 각 구역 내 근접거리 순 정렬 (greedy nearest-neighbor)
      5. 계획 기간 영업일 / 기사별 학교 수 기준 일당 건수 자동 계산
         → 인근 학교가 같은 날 묶여 배정됨
      6. task_type 미입력 시 계획명으로 자동 설정

    force_reassign=True: 기존 scheduled 항목도 초기화 후 재배정
    """
    from .models import InspectionPlan, SchoolInspection, WorkerArea
    from apps.accounts.models import User as UserModel
    from apps.workforce.models import WorkSchedule, TaskAssignment, WorkScheduleType

    plan = InspectionPlan.objects.select_related('created_by').get(pk=plan_id)
    biz_days = get_business_days(plan.start_date, plan.end_date)
    if not biz_days:
        return {'assigned': 0, 'skipped': 0, 'biz_days': 0, 'error': '영업일이 없습니다.'}

    if force_reassign:
        # 기존 WorkSchedule 연결 ID 수집 후 먼저 FK 해제, 그 다음 삭제
        old_ws_ids = list(
            SchoolInspection.objects.filter(
                plan_id=plan_id, status='scheduled', work_schedule__isnull=False
            ).values_list('work_schedule_id', flat=True)
        )
        SchoolInspection.objects.filter(
            plan_id=plan_id, status='scheduled'
        ).update(status='pending', scheduled_date=None, assigned_worker=None, work_schedule=None)
        if old_ws_ids:
            WorkSchedule.objects.filter(id__in=old_ws_ids).delete()

    # 미배정 항목 조회 (pending + scheduled-but-unassigned 모두 처리)
    target_qs = (
        plan.school_inspections
        .filter(assigned_worker__isnull=True, status__in=['pending', 'scheduled'])
        .select_related('school__support_center', 'school')
    )

    # 지원청별 그룹화
    by_center: dict[int, list] = {}
    for si in target_qs:
        by_center.setdefault(si.school.support_center_id, []).append(si)

    # WorkerArea 기반 지원청별 담당 인력 맵
    area_map: dict[int, list] = {}
    for wa in WorkerArea.objects.select_related('worker', 'support_center').order_by('-is_primary', 'worker__name'):
        area_map.setdefault(wa.support_center_id, []).append(wa.worker)

    # plan_type → WorkScheduleType.code 매핑
    PLAN_TO_STYPE = {
        'regular':   'regular_check',
        'quarterly': 'regular_check',
        'special':   'special_check',
        'project':   'other',
        'survey':    'other',
        'followup':  'other',
    }
    stype_code = PLAN_TO_STYPE.get(plan.plan_type, 'other')
    schedule_type = (
        WorkScheduleType.objects.filter(code=stype_code, is_active=True).first()
        or WorkScheduleType.objects.filter(is_active=True).first()
    )
    n_days = len(biz_days)

    assigned_count = 0
    skipped_count = 0
    by_worker: dict[str, int] = {}
    no_worker_schools: list[str] = []

    with transaction.atomic():
        for center_id, sis in by_center.items():
            workers = area_map.get(center_id, [])

            # WorkerArea 미등록 시 User.support_center 기반 폴백
            if not workers and center_id is not None:
                workers = list(
                    UserModel.objects.filter(
                        support_center_id=center_id,
                        role__in=['worker', 'resident'],
                        is_active=True,
                    ).order_by('name')
                )

            if not workers:
                for si in sis:
                    skipped_count += 1
                    no_worker_schools.append(si.school.name)
                continue

            n_workers = len(workers)

            # 좌표 있는 학교 / 없는 학교 분리
            with_coords = [
                si for si in sis
                if si.school.lat and si.school.lng
                and float(si.school.lat) != 0 and float(si.school.lng) != 0
            ]
            with_coords_ids = {id(si) for si in with_coords}
            without_coords = [si for si in sis if id(si) not in with_coords_ids]

            # 위도+경도 합산으로 정렬 → 지리적 구역 분할
            with_coords.sort(key=lambda s: float(s.school.lat) + float(s.school.lng))
            zone_size = math.ceil(len(with_coords) / n_workers) if with_coords else 0

            # 기사별 담당 학교 구역 할당
            worker_school_map: dict[int, list] = {w.id: [] for w in workers}
            for w_idx, worker in enumerate(workers):
                start = w_idx * zone_size
                end = min((w_idx + 1) * zone_size, len(with_coords))
                worker_school_map[worker.id].extend(with_coords[start:end])

            # 좌표 없는 학교는 순서대로 기사에 배분
            for i, si in enumerate(without_coords):
                worker_school_map[workers[i % n_workers].id].append(si)

            # 각 기사: 근접거리 순 정렬 → max_per_day 기준 날짜 배분
            for worker in workers:
                wsis = worker_school_map[worker.id]
                if not wsis:
                    continue

                # 근접거리 순 정렬 (lat·lng 모두 유효한 것만, 없는 것은 뒤에 붙임)
                wc = [si for si in wsis
                      if si.school.lat and si.school.lng
                      and float(si.school.lat) != 0 and float(si.school.lng) != 0]
                nc = [si for si in wsis if si not in set(wc)]
                sorted_wsis = (_nearest_neighbor_sort(wc) if wc else []) + nc

                n_w = len(sorted_wsis)
                # 계획 기간 영업일 기준 일당 건수 자동 계산
                spd = max(1, math.ceil(n_w / n_days))

                for i, si in enumerate(sorted_wsis):
                    day_idx = min(i // spd, n_days - 1)
                    chosen_date = biz_days[day_idx]

                    si.assigned_worker = worker
                    si.scheduled_date = chosen_date
                    si.status = 'scheduled'
                    if not si.task_type:
                        si.task_type = plan.name

                    update_fields = [
                        'assigned_worker', 'scheduled_date', 'status',
                        'task_type', 'updated_at',
                    ]

                    # WorkSchedule 연동 (실패해도 배정 자체는 유지)
                    if schedule_type:
                        try:
                            plan_type_label = dict(plan.TYPE_CHOICES).get(plan.plan_type, plan.plan_type)
                            ws_title = f'[{plan_type_label}] {si.school.name}'
                            ws, _ = WorkSchedule.objects.get_or_create(
                                worker=worker,
                                start_dt__date=chosen_date,
                                schedule_type=schedule_type,
                                defaults={
                                    'title': ws_title,
                                    'start_dt': timezone.make_aware(
                                        timezone.datetime.combine(
                                            chosen_date,
                                            timezone.datetime.strptime('09:00', '%H:%M').time()
                                        )
                                    ),
                                    'end_dt': timezone.make_aware(
                                        timezone.datetime.combine(
                                            chosen_date,
                                            timezone.datetime.max.time().replace(microsecond=0)
                                        )
                                    ),
                                    'school': si.school,
                                    'status': 'planned',
                                    'created_by': plan.created_by,
                                }
                            )
                            order = TaskAssignment.objects.filter(schedule=ws).count() + 1
                            TaskAssignment.objects.get_or_create(
                                schedule=ws,
                                school=si.school,
                                defaults={
                                    'description': ws_title,
                                    'status': 'pending',
                                    'order': order,
                                }
                            )
                            si.work_schedule = ws
                            update_fields.append('work_schedule')
                        except Exception:
                            pass  # WorkSchedule 연동 실패 무시

                    si.save(update_fields=update_fields)
                    by_worker[worker.name] = by_worker.get(worker.name, 0) + 1
                    assigned_count += 1

    return {
        'assigned':          assigned_count,
        'skipped':           skipped_count,
        'biz_days':          len(biz_days),
        'by_worker':         by_worker,
        'no_worker_schools': no_worker_schools,
    }


# ──────────────────────────────────────────────────
# 3. 인력 교체 (업무 인수인계)
# ──────────────────────────────────────────────────
def replace_worker(plan_id: int, old_worker_id: int, new_worker_id: int,
                   from_date: date | None = None) -> dict:
    """
    점검계획 내 old_worker → new_worker 인수인계

    - from_date 지정: 해당 날짜 이후 미완료 건만 이관
    - from_date 없음: 미완료 전체 이관
    - WorkSchedule도 함께 new_worker로 재배정
    - 교체 이력(replaced_from, replaced_at) 기록
    """
    from .models import SchoolInspection
    from apps.workforce.models import WorkSchedule
    from apps.accounts.models import User

    try:
        new_worker = User.objects.get(pk=new_worker_id)
    except User.DoesNotExist:
        return {'error': '신규 인력을 찾을 수 없습니다.'}

    qs = SchoolInspection.objects.filter(
        plan_id=plan_id,
        assigned_worker_id=old_worker_id,
        status__in=['pending', 'scheduled'],
    ).select_related('work_schedule')

    if from_date:
        qs = qs.filter(
            scheduled_date__gte=from_date
        ) | SchoolInspection.objects.filter(
            plan_id=plan_id,
            assigned_worker_id=old_worker_id,
            status='pending',
            scheduled_date__isnull=True,
        )

    now = timezone.now()
    transferred = 0

    with transaction.atomic():
        for si in qs:
            si.replaced_from = si.assigned_worker
            si.replaced_at   = now
            si.assigned_worker = new_worker

            # WorkSchedule worker 교체
            if si.work_schedule:
                si.work_schedule.worker = new_worker
                si.work_schedule.save(update_fields=['worker'])

            si.save(update_fields=['assigned_worker', 'replaced_from', 'replaced_at', 'updated_at'])
            transferred += 1

    return {
        'transferred': transferred,
        'new_worker':  new_worker.name,
        'from_date':   str(from_date) if from_date else '전체',
    }


# ──────────────────────────────────────────────────
# 4. 파일 업로드 처리 (Excel / CSV)
# ──────────────────────────────────────────────────
def process_upload(file_obj) -> dict:
    """
    업로드 파일(Excel/CSV) 파싱 → 학교 매칭 결과 반환
    실제 DB 변경 없음 (confirm_upload에서 확정)

    반환:
        {
          rows: [{school_id, school_name, center_name,
                  assigned_worker_id, scheduled_date,
                  priority, task_type, notes,
                  match_type, raw_name}],
          failed: [{row_num, raw_name, reason}],
          total: int, matched: int, failed_count: int
        }
    """
    from apps.schools.models import School

    # ── 파일 파싱 ──────────────────────────────────
    fname = getattr(file_obj, 'name', '')
    ext   = os.path.splitext(fname)[1].lower()

    if ext in ('.xlsx', '.xlsm', '.xls'):
        rows_raw = _parse_excel(file_obj)
    else:
        rows_raw = _parse_csv(file_obj)

    # 학교 DB 전체 캐시 (name → [School])
    school_map: dict[str, list] = {}
    for s in School.objects.select_related('support_center').filter(is_active=True):
        school_map.setdefault(s.name, []).append(s)

    matched = []
    failed  = []

    for row_num, row in enumerate(rows_raw, start=2):  # 헤더 제외, 2행부터
        raw_name = str(row.get('school_name') or row.get('학교명') or '').strip()
        center_hint = str(row.get('support_center') or row.get('교육지원청') or '').strip()

        if not raw_name:
            failed.append({'row_num': row_num, 'raw_name': '', 'reason': '학교명 없음'})
            continue

        school, match_type = _match_school(raw_name, center_hint, school_map)

        if school is None:
            failed.append({'row_num': row_num, 'raw_name': raw_name, 'reason': '매칭 실패'})
            continue

        # 담당기사 매칭
        worker_id = None
        worker_name_raw = str(row.get('assigned_worker') or row.get('담당기사') or '').strip()
        if worker_name_raw:
            from apps.accounts.models import User
            u = User.objects.filter(name__icontains=worker_name_raw).first()
            if u:
                worker_id = u.id

        matched.append({
            'school_id':          school.id,
            'school_name':        school.name,
            'center_name':        school.support_center.name if school.support_center else '',
            'match_type':         match_type,   # exact / center / fuzzy
            'raw_name':           raw_name,
            'assigned_worker_id': worker_id,
            'scheduled_date':     str(row.get('scheduled_date') or row.get('예정일') or '') or None,
            'priority':           str(row.get('priority') or row.get('우선순위') or 'normal').strip() or 'normal',
            'task_type':          str(row.get('task_type') or row.get('작업유형') or '').strip(),
            'notes':              str(row.get('notes') or row.get('비고') or '').strip(),
        })

    return {
        'rows':         matched,
        'failed':       failed,
        'total':        len(rows_raw),
        'matched':      len(matched),
        'failed_count': len(failed),
    }


def _match_school(name: str, center_hint: str, school_map: dict):
    """학교명으로 DB 매칭. (school, match_type) 반환"""
    # 1순위: 완전 일치
    candidates = school_map.get(name)
    if candidates:
        if len(candidates) == 1:
            return candidates[0], 'exact'
        # 동명 학교 여러 개 → 지원청으로 구분
        if center_hint:
            for s in candidates:
                if s.support_center and center_hint in s.support_center.name:
                    return s, 'exact'
        return candidates[0], 'exact'

    # 2순위: 유사 매칭 (80% 이상)
    best_score = 0.0
    best_school = None
    for sname, schools in school_map.items():
        score = SequenceMatcher(None, name, sname).ratio()
        if score > best_score:
            best_score = score
            best_school = schools[0]

    if best_score >= 0.80:
        return best_school, f'fuzzy({best_score:.0%})'

    return None, None


def _parse_csv(file_obj) -> list[dict]:
    """CSV 파싱"""
    content = file_obj.read()
    for enc in ('utf-8-sig', 'cp949', 'utf-8'):
        try:
            text = content.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = content.decode('utf-8', errors='replace')

    reader = csv.DictReader(io.StringIO(text))
    return [dict(row) for row in reader]


def _parse_excel(file_obj) -> list[dict]:
    """Excel 파싱 (openpyxl)"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h).strip().rstrip(' *').strip() if h else f'col{i}' for i, h in enumerate(rows[0])]
        result = []
        for row in rows[1:]:
            if all(v is None for v in row):
                continue
            result.append(dict(zip(headers, row)))
        return result
    except ImportError:
        raise ValueError('Excel 파싱을 위해 openpyxl이 필요합니다.')


# ──────────────────────────────────────────────────
# 5. 업로드 확정 (SchoolInspection 생성)
# ──────────────────────────────────────────────────
@transaction.atomic
def confirm_upload(plan_id: int, rows: list[dict], uploaded_by,
                   file_name: str = '') -> dict:
    """
    process_upload() 결과를 확정 → SchoolInspection 일괄 생성
    이미 존재하는 항목은 건너뜀 (중복 방지)
    """
    from .models import InspectionPlan, SchoolInspection, InspectionUploadLog

    plan = InspectionPlan.objects.get(pk=plan_id)
    added = 0
    skipped = 0

    for row in rows:
        school_id = row.get('school_id')
        if not school_id:
            continue
        _, is_new = SchoolInspection.objects.get_or_create(
            plan=plan,
            school_id=school_id,
            defaults={
                'status':           'scheduled' if row.get('scheduled_date') else 'pending',
                'scheduled_date':   row.get('scheduled_date') or None,
                'assigned_worker_id': row.get('assigned_worker_id') or None,
                'priority':         row.get('priority') or 'normal',
                'task_type':        row.get('task_type') or '',
                'notes':            row.get('notes') or '',
            }
        )
        if is_new:
            added += 1
        else:
            skipped += 1

    # 업로드 이력 저장
    InspectionUploadLog.objects.create(
        plan=plan,
        uploaded_by=uploaded_by,
        file_name=file_name,
        total_rows=len(rows),
        matched_count=len(rows),
        failed_count=0,
        added_count=added,
        result_json={'added': added, 'skipped': skipped},
    )

    return {'added': added, 'skipped': skipped}


# ──────────────────────────────────────────────────
# 6. Excel 템플릿 생성 (다운로드용)
# ──────────────────────────────────────────────────
def generate_template_excel(plan_type: str = 'regular') -> bytes:
    """
    업로드용 Excel 템플릿 생성
    정기점검: 전체 학교 pre-filled
    기타: 빈 양식 + 예시 1행
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from apps.schools.models import School

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '점검대상학교'

    # 헤더 정의
    headers = [
        ('학교명',      'school_name',      True,  25),
        ('교육지원청',  'support_center',   False, 18),
        ('예정일',      'scheduled_date',   False, 14),
        ('담당기사',    'assigned_worker',  False, 14),
        ('우선순위',    'priority',         False, 12),
        ('작업유형',    'task_type',        False, 20),
        ('비고',        'notes',            False, 30),
    ]

    # 헤더 스타일
    header_fill  = PatternFill('solid', fgColor='1F4E79')
    req_fill     = PatternFill('solid', fgColor='C00000')
    header_font  = Font(color='FFFFFF', bold=True, size=10)
    thin         = Side(style='thin', color='CCCCCC')
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, (label, key, required, width) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label + (' *' if required else ''))
        cell.fill    = req_fill if required else header_fill
        cell.font    = header_font
        cell.border  = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    ws.row_dimensions[1].height = 20

    # 데이터 행
    if plan_type == 'regular':
        schools = School.objects.select_related('support_center').filter(is_active=True).order_by(
            'support_center__name', 'name'
        )
        for r, s in enumerate(schools, start=2):
            ws.cell(r, 1, s.name)
            ws.cell(r, 2, s.support_center.name if s.support_center else '')
            ws.cell(r, 5, 'normal')
            for c in range(1, 8):
                ws.cell(r, c).border = border
    else:
        # 예시 행
        ws.cell(2, 1, '예) 가락고등학교')
        ws.cell(2, 2, '예) 강동송파')
        ws.cell(2, 3, '예) 2026-05-15')
        ws.cell(2, 4, '예) 홍길동')
        ws.cell(2, 5, 'normal')
        ws.cell(2, 6, '예) 스위치교체')
        ws.cell(2, 7, '예) 방과후 방문 요망')
        example_fill = PatternFill('solid', fgColor='FFF2CC')
        for c in range(1, 8):
            ws.cell(2, c).fill   = example_fill
            ws.cell(2, c).border = border

    # 안내 시트
    ws2 = wb.create_sheet('작성안내')
    guide = [
        ['필드명',      '영문키',           '필수', '설명'],
        ['학교명',      'school_name',      'O',    'DB에 등록된 학교명과 유사하면 자동 매칭'],
        ['교육지원청',  'support_center',   '-',    '동명 학교 구분에 사용 (강동송파, 강남서초 등)'],
        ['예정일',      'scheduled_date',   '-',    'YYYY-MM-DD 형식 (예: 2026-05-15)'],
        ['담당기사',    'assigned_worker',  '-',    '인력 이름 (DB에 없으면 미배정)'],
        ['우선순위',    'priority',         '-',    'high / normal / low'],
        ['작업유형',    'task_type',        '-',    '스위치교체, AP설치, 망구성도 등 자유 입력'],
        ['비고',        'notes',            '-',    '특이사항 자유 입력'],
    ]
    for r, row_data in enumerate(guide, start=1):
        for c, val in enumerate(row_data, start=1):
            cell = ws2.cell(r, c, val)
            if r == 1:
                cell.font = Font(bold=True)
                cell.fill = PatternFill('solid', fgColor='DDEBF7')
            cell.border = border
    for c in [1, 2, 3, 4]:
        ws2.column_dimensions[ws2.cell(1, c).column_letter].width = [14, 18, 6, 50][c-1]

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ──────────────────────────────────────────────────
# 7. CSV 다운로드 생성
# ──────────────────────────────────────────────────
def generate_csv_download(queryset) -> bytes:
    """
    SchoolInspection queryset → CSV bytes (UTF-8 BOM, Excel 호환)
    """
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['지원청', '학교', '학제', '담당기사', '예정일', '완료일',
                     '작업유형', '우선순위', '상태', '비고'])
    STATUS_KO = {'pending': '미정', 'scheduled': '예정', 'completed': '완료', 'skipped': '제외'}
    PRIORITY_KO = {'high': '높음', 'normal': '보통', 'low': '낮음'}
    for si in queryset.select_related(
        'school__support_center', 'school__school_type', 'assigned_worker'
    ):
        writer.writerow([
            si.school.support_center.name if si.school.support_center else '',
            si.school.name,
            si.school.school_type.name if hasattr(si.school, 'school_type') and si.school.school_type else '',
            si.assigned_worker.name if si.assigned_worker else '',
            str(si.scheduled_date) if si.scheduled_date else '',
            str(si.completed_date) if si.completed_date else '',
            si.task_type or '',
            PRIORITY_KO.get(si.priority, si.priority),
            STATUS_KO.get(si.status, si.status),
            si.notes or '',
        ])
    return ('\ufeff' + output.getvalue()).encode('utf-8')

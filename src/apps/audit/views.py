"""
audit 앱 뷰
감리 대응 시스템 - 요구사항추적/산출물템플릿/산출물/체크리스트/시정조치
"""
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import FileResponse, Http404
from django.conf import settings
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.utils import timezone
import os

# audit 앱은 전체 목록을 한번에 반환해야 하므로 페이지네이션 비활성화
_NO_PAGINATION = None

from .models import AuditProject, Requirement, ArtifactTemplate, Artifact, AuditPlan, ChecklistItem, CorrectiveAction, ArtifactFile
from .serializers import (
    AuditProjectSerializer,
    RequirementSerializer,
    ArtifactTemplateSerializer,
    ArtifactSerializer, ArtifactCreateSerializer,
    AuditPlanSerializer,
    ChecklistItemSerializer,
    CorrectiveActionSerializer,
    ArtifactFileSerializer,
)


# ──────────────────────────────────────────────────
# 템플릿 뷰
# ──────────────────────────────────────────────────
@login_required
def audit_view(request):
    project = AuditProject.objects.filter(is_active=True).first()
    return render(request, 'audit/index.html', {'project': project})


# ──────────────────────────────────────────────────
# REST API ViewSets
# ──────────────────────────────────────────────────
class AuditProjectViewSet(viewsets.ModelViewSet):
    queryset           = AuditProject.objects.all()
    serializer_class   = AuditProjectSerializer
    permission_classes = [IsAuthenticated]
    pagination_class   = _NO_PAGINATION

    @action(detail=True, methods=['get'])
    def summary(self, request, pk=None):
        project = self.get_object()
        req_qs  = project.requirements.all()
        art_qs  = project.artifacts.all()
        tmpl_qs = project.artifact_templates.all()
        plan_qs = project.audit_plans.all()

        # 산출물 제출율 (템플릿 대비)
        tmpl_required = tmpl_qs.filter(is_required=True).count()
        art_submitted  = art_qs.filter(status__in=['submitted', 'approved']).count()

        ca_qs = CorrectiveAction.objects.filter(checklist_item__audit_plan__project=project)

        data = {
            'requirements': {
                'total':       req_qs.count(),
                'not_started': req_qs.filter(status='not_started').count(),
                'in_progress': req_qs.filter(status='in_progress').count(),
                'completed':   req_qs.filter(status='completed').count(),
                'excluded':    req_qs.filter(status='excluded').count(),
                'additional':  req_qs.filter(is_additional=True).count(),
            },
            'artifacts': {
                'template_total':    tmpl_qs.count(),
                'template_required': tmpl_required,
                'submitted':         art_submitted,
                'pending':           art_qs.filter(status='pending').count(),
                'draft':             art_qs.filter(status='draft').count(),
            },
            'corrective_actions': {
                'total':    ca_qs.count(),
                'open':     ca_qs.filter(status='open').count(),
                'verified': ca_qs.filter(status='verified').count(),
            },
            'audit_plans': AuditPlanSerializer(plan_qs.order_by('phase'), many=True).data,
            'phase_artifact_counts': {
                phase: {
                    'total':     tmpl_qs.filter(audit_phase=phase).count(),
                    'submitted': art_qs.filter(audit_phase=phase, status__in=['submitted','approved']).count(),
                }
                for phase in ['initiation', 'midterm', 'closing', 'all']
            },
        }
        return Response(data)

    @action(detail=True, methods=['post'], url_path='scan_files')
    def scan_files(self, request, pk=None):
        """NAS 폴더 스캔 → ArtifactFile 자동 등록"""
        from .management.commands.scan_audit_files import parse_filename
        project = self.get_object()
        base_dir = os.path.join(settings.MEDIA_ROOT, '2026감리산출물')
        if not os.path.isdir(base_dir):
            return Response({'error': f'폴더 없음: {base_dir}'}, status=status.HTTP_404_NOT_FOUND)

        tmpl_map = {
            t.code.upper(): t
            for t in ArtifactTemplate.objects.filter(project=project)
        }
        new_count = skip_count = err_count = 0
        results = []

        for entry in sorted(os.scandir(base_dir), key=lambda e: e.name):
            if not entry.is_dir():
                continue
            tmpl = tmpl_map.get(entry.name.upper())
            if not tmpl:
                continue
            for fentry in sorted(os.scandir(entry.path), key=lambda e: e.name):
                if not fentry.is_file() or fentry.name.startswith('.'):
                    continue
                fname = fentry.name
                exists = ArtifactFile.objects.filter(project=project, template=tmpl, file_name=fname).exists()
                if exists:
                    skip_count += 1
                    continue
                try:
                    display_name, occ_date, loc_note = parse_filename(fname, tmpl.code)
                    rel_path = os.path.relpath(fentry.path, settings.MEDIA_ROOT)
                    ArtifactFile.objects.create(
                        project=project, template=tmpl,
                        file=rel_path, file_name=fname,
                        display_name=display_name,
                        file_size=fentry.stat().st_size,
                        occurrence_date=occ_date,
                        location_note=loc_note,
                        is_scanned=True,
                    )
                    new_count += 1
                    results.append({'code': tmpl.code, 'file': fname, 'action': 'new'})
                except Exception as e:
                    err_count += 1
                    results.append({'code': tmpl.code, 'file': fname, 'action': 'error', 'msg': str(e)})

        return Response({
            'new': new_count, 'skipped': skip_count, 'errors': err_count,
            'results': results,
        })

    @action(detail=True, methods=['get'], url_path='file_tree')
    def file_tree(self, request, pk=None):
        """산출물 보관함 트리 데이터 — 단계 > 카테고리 > 템플릿 > 파일수"""
        project  = self.get_object()
        tmpls    = ArtifactTemplate.objects.filter(project=project).select_related('requirement')
        files_qs = ArtifactFile.objects.filter(project=project)

        # 템플릿별 파일 수
        from django.db.models import Count
        file_counts = {
            r['template_id']: r['cnt']
            for r in files_qs.values('template_id').annotate(cnt=Count('id'))
        }

        PHASE_ORDER  = {'initiation': 1, 'midterm': 2, 'closing': 3, 'all': 4}
        PHASE_LABEL  = {'initiation': '착수감리', 'midterm': '중간감리', 'closing': '종료감리', 'all': '전 단계'}
        CAT_LABEL    = {'PM':'사업관리','IM':'통합관리','HR':'운영인력','SEC':'보안',
                        'SVC':'서비스수행','ADD':'추가제안','FIN':'완료'}

        tree = {}
        for t in sorted(tmpls, key=lambda x: (PHASE_ORDER.get(x.audit_phase,9), x.seq)):
            phase = t.audit_phase
            cat   = t.category
            if phase not in tree:
                tree[phase] = {'label': PHASE_LABEL.get(phase, phase), 'order': PHASE_ORDER.get(phase, 9), 'categories': {}}
            if cat not in tree[phase]['categories']:
                tree[phase]['categories'][cat] = {'label': CAT_LABEL.get(cat, cat), 'templates': []}
            tree[phase]['categories'][cat]['templates'].append({
                'id':           t.id,
                'code':         t.code,
                'name':         t.name,
                'is_required':  t.is_required,
                'submit_timing': t.submit_timing,
                'file_count':   file_counts.get(t.id, 0),
            })

        # 정렬 후 리스트로
        result = []
        for phase_key in sorted(tree.keys(), key=lambda k: PHASE_ORDER.get(k, 9)):
            ph = tree[phase_key]
            cats = []
            for cat_key, cat_val in ph['categories'].items():
                cats.append({'key': cat_key, 'label': cat_val['label'], 'templates': cat_val['templates']})
            result.append({'key': phase_key, 'label': ph['label'], 'categories': cats})

        return Response(result)


class RequirementViewSet(viewsets.ModelViewSet):
    queryset           = Requirement.objects.select_related('project').all()
    serializer_class   = RequirementSerializer
    permission_classes = [IsAuthenticated]
    pagination_class   = _NO_PAGINATION

    def get_queryset(self):
        qs = super().get_queryset()
        project_id  = self.request.query_params.get('project')
        category    = self.request.query_params.get('category')
        status_     = self.request.query_params.get('status')
        is_add      = self.request.query_params.get('is_additional')
        if project_id: qs = qs.filter(project_id=project_id)
        if category:   qs = qs.filter(category=category)
        if status_:    qs = qs.filter(status=status_)
        if is_add is not None:
            qs = qs.filter(is_additional=(is_add.lower() == 'true'))
        return qs

    @action(detail=False, methods=['get'])
    def by_category(self, request):
        project_id = request.query_params.get('project')
        qs = self.get_queryset()
        if project_id:
            qs = qs.filter(project_id=project_id)
        result = {}
        for cat_code, cat_name in Requirement.CATEGORY_CHOICES:
            cat_qs = qs.filter(category=cat_code)
            result[cat_code] = {
                'name':        cat_name,
                'total':       cat_qs.count(),
                'completed':   cat_qs.filter(status='completed').count(),
                'in_progress': cat_qs.filter(status='in_progress').count(),
                'not_started': cat_qs.filter(status='not_started').count(),
            }
        return Response(result)

    @action(detail=True, methods=['post'])
    def update_status(self, request, pk=None):
        req        = self.get_object()
        new_status = request.data.get('status')
        evidence   = request.data.get('evidence', '')
        if new_status not in dict(Requirement.STATUS_CHOICES):
            return Response({'error': '유효하지 않은 상태값'}, status=status.HTTP_400_BAD_REQUEST)
        req.status   = new_status
        req.evidence = evidence
        req.save()
        return Response(RequirementSerializer(req).data)


class ArtifactTemplateViewSet(viewsets.ModelViewSet):
    queryset           = ArtifactTemplate.objects.select_related('project', 'requirement').prefetch_related('artifacts').all()
    serializer_class   = ArtifactTemplateSerializer
    permission_classes = [IsAuthenticated]
    pagination_class   = _NO_PAGINATION

    def get_queryset(self):
        qs = super().get_queryset()
        project_id  = self.request.query_params.get('project')
        phase       = self.request.query_params.get('phase')
        category    = self.request.query_params.get('category')
        is_add      = self.request.query_params.get('is_additional')
        is_req      = self.request.query_params.get('is_required')
        if project_id: qs = qs.filter(project_id=project_id)
        if phase:      qs = qs.filter(audit_phase=phase)
        if category:   qs = qs.filter(category=category)
        if is_add is not None:
            qs = qs.filter(is_additional=(is_add.lower() == 'true'))
        if is_req is not None:
            qs = qs.filter(is_required=(is_req.lower() == 'true'))
        return qs

    @action(detail=True, methods=['post'])
    def create_artifact(self, request, pk=None):
        """템플릿에서 실제 산출물 생성"""
        tmpl = self.get_object()
        art  = Artifact.objects.create(
            project=tmpl.project,
            template=tmpl,
            code=tmpl.code,
            name=tmpl.name,
            audit_phase=tmpl.audit_phase,
            status='draft',
        )
        if tmpl.requirement:
            art.requirements.add(tmpl.requirement)
        return Response(ArtifactSerializer(art).data, status=status.HTTP_201_CREATED)


class ArtifactViewSet(viewsets.ModelViewSet):
    queryset           = Artifact.objects.select_related('project', 'template', 'submitted_by').prefetch_related('requirements').all()
    permission_classes = [IsAuthenticated]
    pagination_class   = _NO_PAGINATION

    def get_serializer_class(self):
        if self.action in ('create', 'update', 'partial_update'):
            return ArtifactCreateSerializer
        return ArtifactSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        project_id = self.request.query_params.get('project')
        phase      = self.request.query_params.get('phase')
        status_    = self.request.query_params.get('status')
        if project_id: qs = qs.filter(project_id=project_id)
        if phase:      qs = qs.filter(audit_phase=phase)
        if status_:    qs = qs.filter(status=status_)
        return qs

    @action(detail=True, methods=['post'])
    def submit(self, request, pk=None):
        artifact = self.get_object()
        artifact.status       = 'submitted'
        artifact.submitted_by = request.user
        artifact.submitted_at = timezone.now().date()
        artifact.save()
        return Response(ArtifactSerializer(artifact).data)

    @action(detail=True, methods=['post'], url_path='upload')
    def upload_file(self, request, pk=None):
        """산출물 파일 업로드"""
        artifact = self.get_object()
        f = request.FILES.get('file')
        if not f:
            return Response({'error': '파일이 없습니다'}, status=status.HTTP_400_BAD_REQUEST)
        if artifact.file:
            try:
                old_path = artifact.file.path
                if os.path.exists(old_path):
                    os.remove(old_path)
            except Exception:
                pass
        artifact.file      = f
        artifact.file_name = f.name
        artifact.status    = 'draft' if artifact.status == 'pending' else artifact.status
        occ = request.data.get('occurrence_date') or request.POST.get('occurrence_date')
        loc = request.data.get('location_note')   or request.POST.get('location_note')
        if occ: artifact.occurrence_date = occ
        if loc: artifact.location_note   = loc
        artifact.save()
        return Response(ArtifactSerializer(artifact).data)

    @action(detail=True, methods=['get'], url_path='download')
    def download_file(self, request, pk=None):
        """산출물 파일 다운로드"""
        artifact = self.get_object()
        if not artifact.file:
            raise Http404('파일이 없습니다')
        try:
            file_path = artifact.file.path
            response = FileResponse(
                open(file_path, 'rb'),
                as_attachment=True,
                filename=artifact.file_name or os.path.basename(file_path)
            )
            return response
        except FileNotFoundError:
            raise Http404('파일을 찾을 수 없습니다')


    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """B1: 산출물 승인"""
        artifact = self.get_object()
        artifact.status = 'approved'
        artifact.save()
        # B3: 연결된 요구사항 자동 갱신
        for req in artifact.requirements.all():
            if req.status != 'completed':
                req.status = 'completed'
                req.save(update_fields=['status'])
        return Response(ArtifactSerializer(artifact).data)

    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        """B1: 산출물 반려"""
        artifact = self.get_object()
        artifact.status = 'rejected'
        artifact.save()
        return Response(ArtifactSerializer(artifact).data)


class AuditPlanViewSet(viewsets.ModelViewSet):
    queryset           = AuditPlan.objects.select_related('project').prefetch_related('checklist_items').all()
    serializer_class   = AuditPlanSerializer
    permission_classes = [IsAuthenticated]
    pagination_class   = _NO_PAGINATION

    def get_queryset(self):
        qs = super().get_queryset()
        project_id = self.request.query_params.get('project')
        if project_id:
            qs = qs.filter(project_id=project_id)
        return qs


class ChecklistItemViewSet(viewsets.ModelViewSet):
    queryset           = ChecklistItem.objects.select_related('audit_plan', 'requirement').all()
    serializer_class   = ChecklistItemSerializer
    permission_classes = [IsAuthenticated]
    pagination_class   = _NO_PAGINATION

    def get_queryset(self):
        qs = super().get_queryset()
        plan_id = self.request.query_params.get('plan')
        area    = self.request.query_params.get('area')
        result  = self.request.query_params.get('result')
        if plan_id: qs = qs.filter(audit_plan_id=plan_id)
        if area:    qs = qs.filter(area=area)
        if result:  qs = qs.filter(result=result)
        return qs

    @action(detail=True, methods=['post'])
    def set_result(self, request, pk=None):
        item     = self.get_object()
        result   = request.data.get('result')
        evidence = request.data.get('evidence', '')
        finding  = request.data.get('finding', '')
        if result not in dict(ChecklistItem.RESULT_CHOICES):
            return Response({'error': '유효하지 않은 결과값'}, status=status.HTTP_400_BAD_REQUEST)
        item.result   = result
        item.evidence = evidence
        item.finding  = finding
        item.save()

        # B2: 부적합 시 시정조치 자동 생성
        if result == 'fail' and finding:
            existing = CorrectiveAction.objects.filter(checklist_item=item).first()
            if not existing:
                CorrectiveAction.objects.create(
                    checklist_item=item,
                    action_type='mandatory',
                    issue_description=finding,
                    status='open',
                )

        return Response(ChecklistItemSerializer(item).data)


class CorrectiveActionViewSet(viewsets.ModelViewSet):
    queryset           = CorrectiveAction.objects.select_related(
        'checklist_item__audit_plan', 'completed_by', 'verified_by'
    ).all()
    serializer_class   = CorrectiveActionSerializer
    permission_classes = [IsAuthenticated]
    pagination_class   = _NO_PAGINATION

    def get_queryset(self):
        qs = super().get_queryset()
        plan_id     = self.request.query_params.get('plan')
        project_id  = self.request.query_params.get('project')
        status_     = self.request.query_params.get('status')
        action_type = self.request.query_params.get('action_type')
        if plan_id:     qs = qs.filter(checklist_item__audit_plan_id=plan_id)
        if project_id:  qs = qs.filter(checklist_item__audit_plan__project_id=project_id)
        if status_:     qs = qs.filter(status=status_)
        if action_type: qs = qs.filter(action_type=action_type)
        return qs

    @action(detail=True, methods=['post'])
    def start(self, request, pk=None):
        """미조치 → 조치중"""
        obj = self.get_object()
        if obj.status != 'open':
            return Response({'error': '미조치 상태에서만 시작할 수 있습니다'}, status=400)
        obj.status = 'in_progress'
        obj.save(update_fields=['status', 'updated_at'])
        return Response(CorrectiveActionSerializer(obj).data)

    @action(detail=True, methods=['post'])
    def complete(self, request, pk=None):
        """조치중 → 조치완료"""
        obj = self.get_object()
        obj.status               = 'completed'
        obj.completed_by         = request.user
        obj.completed_at         = timezone.now().date()
        obj.action_description   = request.data.get('action_description', obj.action_description)
        obj.evidence_note        = request.data.get('evidence_note', obj.evidence_note)
        obj.save()
        return Response(CorrectiveActionSerializer(obj).data)

    @action(detail=True, methods=['post'])
    def verify(self, request, pk=None):
        """조치완료 → 확인완료"""
        obj = self.get_object()
        if obj.status != 'completed':
            return Response({'error': '조치완료 상태에서만 확인할 수 있습니다'}, status=400)
        obj.status      = 'verified'
        obj.verified_by = request.user
        obj.verified_at = timezone.now().date()
        obj.save()
        return Response(CorrectiveActionSerializer(obj).data)


class ArtifactFileViewSet(viewsets.ModelViewSet):
    """산출물 보관함 개별 파일 CRUD + 업로드/다운로드"""
    queryset           = ArtifactFile.objects.select_related('project', 'template', 'uploaded_by').all()
    serializer_class   = ArtifactFileSerializer
    permission_classes = [IsAuthenticated]
    pagination_class   = _NO_PAGINATION

    def get_queryset(self):
        qs         = super().get_queryset()
        project_id = self.request.query_params.get('project')
        tmpl_id    = self.request.query_params.get('template')
        phase      = self.request.query_params.get('phase')
        keyword    = self.request.query_params.get('q')
        if project_id: qs = qs.filter(project_id=project_id)
        if tmpl_id:    qs = qs.filter(template_id=tmpl_id)
        if phase:      qs = qs.filter(template__audit_phase=phase)
        if keyword:    qs = qs.filter(file_name__icontains=keyword) | qs.filter(location_note__icontains=keyword)
        return qs

    def perform_create(self, serializer):
        serializer.save(uploaded_by=self.request.user)

    @action(detail=False, methods=['post'], url_path='upload')
    def upload(self, request):
        """파일 업로드 — multipart: file, template(id), project(id), occurrence_date, location_note"""
        f          = request.FILES.get('file')
        tmpl_id    = request.data.get('template')
        project_id = request.data.get('project')
        if not f:
            return Response({'error': '파일이 없습니다'}, status=status.HTTP_400_BAD_REQUEST)
        if not tmpl_id or not project_id:
            return Response({'error': 'template, project 필수'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            tmpl    = ArtifactTemplate.objects.get(pk=tmpl_id)
            project = AuditProject.objects.get(pk=project_id)
        except (ArtifactTemplate.DoesNotExist, AuditProject.DoesNotExist):
            return Response({'error': '프로젝트 또는 템플릿 없음'}, status=status.HTTP_404_NOT_FOUND)

        from .management.commands.scan_audit_files import parse_filename
        display_name, occ_date, loc_note = parse_filename(f.name, tmpl.code)

        occ_override = request.data.get('occurrence_date')
        loc_override = request.data.get('location_note', '').strip()
        if occ_override:
            try:
                from datetime import date
                parts = occ_override.split('-')
                occ_date = date(int(parts[0]), int(parts[1]), int(parts[2]))
            except Exception:
                pass
        if loc_override:
            loc_note = loc_override

        af = ArtifactFile.objects.create(
            project=project, template=tmpl,
            file=f, file_name=f.name,
            display_name=display_name,
            file_size=f.size,
            occurrence_date=occ_date,
            location_note=loc_note,
            uploaded_by=request.user,
            is_scanned=False,
        )
        return Response(ArtifactFileSerializer(af, context={'request': request}).data,
                        status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['get'], url_path='download')
    def download(self, request, pk=None):
        af = self.get_object()
        if not af.file:
            raise Http404('파일이 없습니다')
        try:
            response = FileResponse(
                open(af.file.path, 'rb'),
                as_attachment=True,
                filename=af.file_name or os.path.basename(af.file.path)
            )
            return response
        except FileNotFoundError:
            raise Http404('파일을 찾을 수 없습니다')


# ──────────────────────────────────────────────────────────────────
# RTM Excel 내보내기
# ──────────────────────────────────────────────────────────────────
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter


@login_required
def export_rtm_excel(request):
    """요구사항 추적 매트릭스(RTM) Excel 내보내기"""
    project_id = request.GET.get('project')
    project = AuditProject.objects.filter(
        id=project_id if project_id else None,
        is_active=True,
    ).first() or AuditProject.objects.filter(is_active=True).first()

    if not project:
        return HttpResponse('감리 프로젝트가 없습니다', status=404)

    reqs   = Requirement.objects.filter(project=project).order_by('category', 'code')
    tmpls  = {t.requirement_id: t for t in ArtifactTemplate.objects.filter(project=project).select_related('requirement')}
    arts   = {}
    for a in ArtifactViewSet().get_queryset().filter(project=project):
        if a.template_id:
            arts[a.template_id] = a

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'RTM'

    # ── 스타일 ───────────────────────────────────
    HEADER_FILL = PatternFill('solid', fgColor='1E2D40')
    CAT_FILL    = PatternFill('solid', fgColor='3D5A7A')
    ALT_FILL    = PatternFill('solid', fgColor='F5F8FB')
    WHITE_FILL  = PatternFill('solid', fgColor='FFFFFF')
    OK_FILL     = PatternFill('solid', fgColor='D4EDDA')
    NG_FILL     = PatternFill('solid', fgColor='F8D7DA')
    IP_FILL     = PatternFill('solid', fgColor='FFF3CD')

    HDR_FONT    = Font(bold=True, color='FFFFFF', size=9)
    CAT_FONT    = Font(bold=True, color='FFFFFF', size=9)
    BODY_FONT   = Font(size=9)
    TITLE_FONT  = Font(bold=True, size=14, color='1E2D40')

    thin = Side(style='thin', color='CBD5E1')
    bd   = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left   = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    STATUS_KO = {
        'not_started': '미착수',
        'in_progress': '진행중',
        'completed':   '완료',
        'excluded':    '점검제외',
    }
    ART_STATUS_KO = {
        'pending':   '미작성',
        'draft':     '작성중',
        'submitted': '제출완료',
        'approved':  '승인완료',
        'rejected':  '반려',
    }

    # ── 타이틀 행 ─────────────────────────────────
    ws.merge_cells('A1:J1')
    ws['A1'] = f'요구사항 추적 매트릭스 (RTM) — {project.name}'
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:J2')
    ws['A2'] = f'사업연도: {project.year}년  /  감리법인: {project.audit_firm}  /  출력일: {timezone.now().strftime("%Y-%m-%d %H:%M")}'
    ws['A2'].font = Font(size=9, color='5A6A7A')
    ws['A2'].alignment = center
    ws.row_dimensions[2].height = 18

    # ── 헤더 ──────────────────────────────────────
    headers = ['No.', '분류', '요구사항 코드', '요구사항명', '이행 상태',
               '산출물명', '산출물 상태', '제출일', '감리단계', '비고']
    widths  = [5,      10,    14,             40,         12,
               30,     12,    12,             12,         25]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font      = HDR_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = center
        cell.border    = bd
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[3].height = 22

    # ── 데이터 행 ─────────────────────────────────
    CAT_LABELS = dict(Requirement.CATEGORY_CHOICES)
    row_num = 4
    seq = 0

    cat_groups = {}
    for r in reqs:
        cat_groups.setdefault(r.category, []).append(r)

    for cat_code, cat_reqs in cat_groups.items():
        # 카테고리 헤더 행
        ws.merge_cells(f'A{row_num}:J{row_num}')
        cat_cell = ws.cell(row=row_num, column=1, value=f'■  {CAT_LABELS.get(cat_code, cat_code)}')
        cat_cell.font      = CAT_FONT
        cat_cell.fill      = CAT_FILL
        cat_cell.alignment = left
        cat_cell.border    = bd
        ws.row_dimensions[row_num].height = 18
        row_num += 1

        for req in cat_reqs:
            seq += 1
            # 연결된 산출물 템플릿/실체
            tmpl = tmpls.get(req.id)
            art  = arts.get(tmpl.id) if tmpl else None

            art_name   = tmpl.name      if tmpl else ''
            art_status = ART_STATUS_KO.get(art.status, '') if art else ('미제출' if tmpl else '')
            art_date   = art.submitted_at.strftime('%Y-%m-%d') if art and art.submitted_at else ''
            art_phase  = tmpl.get_audit_phase_display() if tmpl else ''

            st = req.status
            if   st == 'completed':   row_fill = OK_FILL
            elif st == 'in_progress': row_fill = IP_FILL
            elif st == 'excluded':    row_fill = ALT_FILL
            else:                      row_fill = WHITE_FILL

            values = [
                seq, CAT_LABELS.get(cat_code, cat_code), req.code, req.name,
                STATUS_KO.get(st, st), art_name, art_status, art_date, art_phase,
                req.notes or '',
            ]
            aligns = [center, center, center, left, center, left, center, center, center, left]

            for col, (val, aln) in enumerate(zip(values, aligns), 1):
                cell = ws.cell(row=row_num, column=col, value=val)
                cell.font      = BODY_FONT
                cell.fill      = row_fill
                cell.alignment = aln
                cell.border    = bd
            ws.row_dimensions[row_num].height = 18
            row_num += 1

    # ── 요약 통계 (하단) ─────────────────────────────
    ws.row_dimensions[row_num].height = 8
    row_num += 1

    ws.merge_cells(f'A{row_num}:B{row_num}')
    ws.cell(row=row_num, column=1, value='구분').font = HDR_FONT
    ws.cell(row=row_num, column=1).fill = HEADER_FILL
    ws.cell(row=row_num, column=1).alignment = center
    ws.merge_cells(f'C{row_num}:D{row_num}')
    ws.cell(row=row_num, column=3, value='건수').font = HDR_FONT
    ws.cell(row=row_num, column=3).fill = HEADER_FILL
    ws.cell(row=row_num, column=3).alignment = center
    ws.merge_cells(f'E{row_num}:F{row_num}')
    ws.cell(row=row_num, column=5, value='비율(%)').font = HDR_FONT
    ws.cell(row=row_num, column=5).fill = HEADER_FILL
    ws.cell(row=row_num, column=5).alignment = center
    row_num += 1

    total = reqs.count()
    for label, qs in [
        ('전체',    reqs),
        ('완료',    reqs.filter(status='completed')),
        ('진행중',  reqs.filter(status='in_progress')),
        ('미착수',  reqs.filter(status='not_started')),
        ('점검제외', reqs.filter(status='excluded')),
    ]:
        cnt  = qs.count()
        pct  = f'{cnt/total*100:.1f}%' if total else '0%'
        ws.merge_cells(f'A{row_num}:B{row_num}')
        ws.cell(row=row_num, column=1, value=label).font      = BODY_FONT
        ws.cell(row=row_num, column=1).alignment = center
        ws.merge_cells(f'C{row_num}:D{row_num}')
        ws.cell(row=row_num, column=3, value=cnt).font        = BODY_FONT
        ws.cell(row=row_num, column=3).alignment = center
        ws.merge_cells(f'E{row_num}:F{row_num}')
        ws.cell(row=row_num, column=5, value=pct).font        = BODY_FONT
        ws.cell(row=row_num, column=5).alignment = center
        row_num += 1

    # ── 헤더 고정 ────────────────────────────────
    ws.freeze_panes = 'A4'

    # ── 응답 ────────────────────────────────────
    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f'RTM_{project.year}_{project.name[:20]}.xlsx'
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def export_audit_data(request, data_type):
    """감리 데이터 Excel 내보내기 (산출물/체크리스트/시정조치)"""
    import openpyxl, urllib.parse
    from io import BytesIO
    from openpyxl.styles import Font as F2, PatternFill as P2

    project_id = request.GET.get('project')
    project = AuditProject.objects.filter(id=project_id).first() if project_id else AuditProject.objects.filter(is_active=True).first()
    if not project:
        return HttpResponse('프로젝트 없음', status=404)

    wb = openpyxl.Workbook()
    ws = wb.active
    hf = P2('solid', fgColor='1F497D')
    hfont = F2(bold=True, color='FFFFFF', size=10)

    SK = {'pending':'대기','draft':'작성중','submitted':'제출','approved':'승인','rejected':'반려',
          'not_started':'미착수','in_progress':'진행중','completed':'완료','excluded':'점검제외',
          'not_checked':'미점검','pass':'적합','fail':'부적합',
          'open':'미조치','verified':'검증완료'}

    if data_type == 'artifacts':
        ws.title = '산출물 현황'
        headers = ['코드','산출물명','감리단계','제출시점','연관요구사항','상태','제출일','제출자']
        ws.append(headers)
        for t in ArtifactTemplate.objects.filter(project=project).order_by('seq'):
            art = Artifact.objects.filter(template=t).first()
            req_codes = ', '.join(t.requirement.values_list('code', flat=True)) if hasattr(t,'requirement') else ''
            ws.append([
                t.code, t.name, t.get_audit_phase_display(), t.get_submit_timing_display(),
                req_codes, SK.get(art.status, art.status) if art else '대기',
                str(art.submitted_at) if art and art.submitted_at else '',
                art.submitted_by.name if art and art.submitted_by else '',
            ])
        fname = '산출물현황.xlsx'

    elif data_type == 'checklist':
        ws.title = '감리 체크리스트'
        headers = ['No.','감리영역','감리단계','요구사항','점검항목','확인포인트','결과','확인내용','지적사항']
        ws.append(headers)
        for i, c in enumerate(ChecklistItem.objects.filter(audit_plan__project=project).select_related('audit_plan','requirement').order_by('audit_plan__phase','area','seq'), 1):
            ws.append([
                i, c.get_area_display(), c.get_phase_display(),
                c.requirement.code if c.requirement else '',
                c.description, c.check_point,
                SK.get(c.result, c.result), c.evidence or '', c.finding or '',
            ])
        fname = '감리체크리스트.xlsx'

    elif data_type == 'corrective':
        ws.title = '시정조치'
        headers = ['No.','유형','감리단계','지적사항','조치내용','기한','상태','완료자','검증자']
        ws.append(headers)
        for i, c in enumerate(CorrectiveAction.objects.filter(checklist_item__audit_plan__project=project).select_related('checklist_item__audit_plan','completed_by','verified_by').order_by('-id'), 1):
            ws.append([
                i, c.get_action_type_display(),
                c.checklist_item.audit_plan.get_phase_display() if c.checklist_item and c.checklist_item.audit_plan else '',
                c.issue_description[:80], c.action_description[:80],
                str(c.due_date) if c.due_date else '',
                SK.get(c.status, c.status),
                c.completed_by.name if c.completed_by else '',
                c.verified_by.name if c.verified_by else '',
            ])
        fname = '시정조치.xlsx'
    else:
        return HttpResponse('알 수 없는 유형', status=400)

    for cell in ws[1]:
        cell.font = hfont; cell.fill = hf

    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    encoded = urllib.parse.quote(fname)
    resp = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f"attachment; filename*=UTF-8''{encoded}"
    return resp

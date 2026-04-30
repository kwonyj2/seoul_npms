import csv, io
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.views import View
from django.http import JsonResponse, HttpResponse
from rest_framework import viewsets, status, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from django.db.models import Q, Count
from .models import (SupportCenter, SchoolType, School, SchoolBuilding,
                     SchoolFloor, SchoolRoom, SchoolContact, VsdxImportLog)
from .serializers import (
    SupportCenterSerializer, SchoolTypeSerializer,
    SchoolListSerializer, SchoolDetailSerializer, SchoolGISSerializer,
    SchoolBuildingSerializer, SchoolContactSerializer
)
from core.permissions.roles import IsAdmin


@login_required
def school_list_view(request):
    return render(request, 'schools/index.html')


@login_required
def school_map_view(request):
    return render(request, 'schools/map.html')


@login_required
def school_detail_view(request, pk):
    from django.shortcuts import redirect
    tab = request.GET.get('tab', 'info')
    return redirect(f'/npms/schools/list/?school={pk}&tab={tab}')


@login_required
def building_docs_api(request, pk):
    """건물 정보 PDF 파일 목록 조회 / 업로드"""
    import os
    from django.conf import settings
    from django.shortcuts import get_object_or_404
    school = get_object_or_404(School, pk=pk)

    doc_dir = os.path.join(settings.MEDIA_ROOT, 'data', '건물 정보', str(pk))
    doc_url_base = f"{settings.MEDIA_URL}data/건물 정보/{pk}/"

    if request.method == 'GET':
        files = []
        # pk 기반 서브폴더
        if os.path.isdir(doc_dir):
            for fname in sorted(os.listdir(doc_dir)):
                if fname.lower().endswith('.pdf'):
                    files.append({'name': fname, 'url': doc_url_base + fname})
        # 기존 파일: media/data/건물 정보/건물정보_{학교명}.pdf
        legacy_name = f'건물정보_{school.name}.pdf'
        legacy_abs = os.path.join(settings.MEDIA_ROOT, 'data', '건물 정보', legacy_name)
        if os.path.exists(legacy_abs):
            legacy_url = f"{settings.MEDIA_URL}data/건물 정보/{legacy_name}"
            if not any(f['name'] == legacy_name for f in files):
                files.insert(0, {'name': legacy_name, 'url': legacy_url})
        return JsonResponse({'files': files})

    if request.method == 'POST':
        upload = request.FILES.get('file')
        if not upload:
            return JsonResponse({'error': '파일을 선택하세요.'}, status=400)
        if not upload.name.lower().endswith('.pdf'):
            return JsonResponse({'error': 'PDF 파일만 업로드할 수 있습니다.'}, status=400)
        os.makedirs(doc_dir, exist_ok=True)
        dest = os.path.join(doc_dir, upload.name)
        with open(dest, 'wb') as f:
            for chunk in upload.chunks():
                f.write(chunk)
        return JsonResponse({'name': upload.name, 'url': doc_url_base + upload.name})

    return JsonResponse({'error': 'Method not allowed'}, status=405)


# 네트워크 문서 카테고리 정의
# keyword: 산출물/2025년 테크센터/ 하위에서 이 글자가 포함된 폴더를 찾음
NETDOC_CATEGORIES = [
    {
        'key':      '구성도',
        'keyword':  '구성도',
        'exts':     ['.pptx', '.ppt'],
        'accept':   '.pptx,.ppt',
        'icon':     'ppt',
    },
    {
        'key':      '선번장',
        'keyword':  '선번장',
        'exts':     ['.xlsx', '.xlsm'],
        'accept':   '.xlsx,.xlsm',
        'icon':     'xlsx',
    },
    {
        'key':      '랙실장도',
        'keyword':  '랙실장도',
        'exts':     ['.xlsx', '.xlsm'],
        'accept':   '.xlsx,.xlsm',
        'icon':     'xlsx',
    },
    {
        'key':      '건물정보',
        'keyword':  '건물 정보',
        'exts':     ['.pdf'],
        'accept':   '.pdf',
        'icon':     'pdf',
    },
    {
        'key':      '전산실랙',
        'keyword':  '전산실랙',
        'exts':     ['.jpg', '.jpeg', '.png'],
        'accept':   '.jpg,.jpeg,.png',
        'icon':     'image',
    },
]

# 산출물 기준 폴더 (2025년/2026년 테크센터)
_TECHCENTER_PATTERN = '테크센터'


def _find_nas_doc_folders(nas_root, keyword):
    """산출물/ 하위에서 keyword가 포함된 폴더 경로 목록 반환"""
    import os
    result = []
    output_root = os.path.join(nas_root, '산출물')
    if not os.path.isdir(output_root):
        return result
    for year_dir in sorted(os.listdir(output_root)):
        if _TECHCENTER_PATTERN not in year_dir:
            continue
        year_path = os.path.join(output_root, year_dir)
        if not os.path.isdir(year_path):
            continue
        for sub_dir in sorted(os.listdir(year_path)):
            if keyword in sub_dir and os.path.isdir(os.path.join(year_path, sub_dir)):
                result.append(os.path.join(year_path, sub_dir))
    return result


@login_required
def network_docs_api(request, pk):
    """네트워크 문서 목록 조회 / 업로드 (구성도·선번장·랙실장도·건물정보·전산실랙)"""
    import os
    from django.conf import settings
    from django.shortcuts import get_object_or_404
    school = get_object_or_404(School, pk=pk)
    nas_root = getattr(settings, 'NAS_MEDIA_ROOT', settings.MEDIA_ROOT)

    if request.method == 'GET':
        from urllib.parse import quote
        result = {}
        for cat in NETDOC_CATEGORIES:
            files = []

            # 1) 산출물/테크센터/ 하위에서 keyword 폴더 찾기
            for folder_abs in _find_nas_doc_folders(nas_root, cat['keyword']):
                rel_path = os.path.relpath(folder_abs, nas_root)
                url_base = f"{settings.MEDIA_URL}{quote(rel_path, safe='/')}/"
                for fname in sorted(os.listdir(folder_abs)):
                    if os.path.isdir(os.path.join(folder_abs, fname)):
                        continue
                    ext = os.path.splitext(fname)[1].lower()
                    if ext not in cat['exts']:
                        continue
                    # 파일명에 학교명이 포함되어 있는지 확인
                    if school.name in fname:
                        files.append({'name': fname, 'url': url_base + quote(fname)})

            # 2) 기존 data/ pk 서브폴더 업로드 파일 (하위 호환)
            pk_dir = os.path.join(nas_root, 'data', cat['key'], str(pk))
            pk_url_base = f"{settings.MEDIA_URL}data/{quote(cat['key'])}/{pk}/"
            if os.path.isdir(pk_dir):
                for fname in sorted(os.listdir(pk_dir)):
                    ext = os.path.splitext(fname)[1].lower()
                    if ext in cat['exts']:
                        files.append({'name': fname, 'url': pk_url_base + quote(fname)})

            result[cat['key']] = files
        return JsonResponse({'docs': result})

    if request.method == 'POST':
        doc_type = request.POST.get('doc_type', '')
        upload = request.FILES.get('file')
        if not upload:
            return JsonResponse({'error': '파일을 선택하세요.'}, status=400)

        cat = next((c for c in NETDOC_CATEGORIES if c['key'] == doc_type), None)
        if not cat:
            return JsonResponse({'error': '문서 유형이 올바르지 않습니다.'}, status=400)

        ext = os.path.splitext(upload.name)[1].lower()
        if ext not in cat['exts']:
            return JsonResponse({'error': f"{cat['key']} 파일은 {', '.join(cat['exts'])} 형식만 허용됩니다."}, status=400)

        from urllib.parse import quote
        dest_dir = os.path.join(nas_root, 'data', cat['key'], str(pk))
        os.makedirs(dest_dir, exist_ok=True)
        dest = os.path.join(dest_dir, upload.name)
        with open(dest, 'wb') as f:
            for chunk in upload.chunks():
                f.write(chunk)
        url = f"{settings.MEDIA_URL}data/{quote(cat['key'])}/{pk}/{quote(upload.name)}"
        return JsonResponse({'name': upload.name, 'url': url, 'doc_type': doc_type})

    return JsonResponse({'error': 'Method not allowed'}, status=405)


class SupportCenterViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = SupportCenter.objects.filter(is_active=True)
    serializer_class = SupportCenterSerializer
    permission_classes = [permissions.IsAuthenticated]


class SchoolTypeViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = SchoolType.objects.all()
    serializer_class = SchoolTypeSerializer
    permission_classes = [permissions.IsAuthenticated]


class SchoolViewSet(viewsets.ModelViewSet):
    queryset = School.objects.select_related('support_center', 'school_type').order_by('support_center', 'name')
    permission_classes = [permissions.IsAuthenticated]

    def get_serializer_class(self):
        if self.action == 'list':
            return SchoolListSerializer
        if self.action == 'gis':
            return SchoolGISSerializer
        return SchoolDetailSerializer

    def get_permissions(self):
        if self.action in ('create', 'update', 'partial_update', 'destroy'):
            return [IsAdmin()]
        return super().get_permissions()

    def get_queryset(self):
        qs = super().get_queryset()
        params  = self.request.query_params
        center  = params.get('center')
        stype   = params.get('type')
        q       = params.get('q')
        school  = params.get('school')
        if school:
            qs = qs.filter(id=school)
        elif center:
            qs = qs.filter(support_center__code=center)
        if stype and not school:
            qs = qs.filter(school_type__code=stype)
        if q:
            qs = qs.filter(Q(name__icontains=q) | Q(address__icontains=q))
        return qs

    @action(detail=False, methods=['get'], pagination_class=None)
    def filter(self, request):
        """일정 등록용 학교 필터 (페이지네이션 없음 — center 필수)"""
        center = request.query_params.get('center', '')
        stype  = request.query_params.get('type', '')
        if not center:
            return Response({'error': '지원청(center)은 필수입니다.'}, status=400)
        qs = School.objects.filter(
            is_active=True,
            support_center__code=center,
        ).order_by('school_type__order', 'name')
        if stype:
            qs = qs.filter(school_type__code=stype)
        data = [{'id': s.id, 'name': s.name,
                  'type_code': s.school_type.code if s.school_type else ''} for s in qs]
        return Response(data)

    @action(detail=False, methods=['get'])
    def search(self, request):
        """학교명 자동완성 검색 (이름 부분 일치, 최대 15건)"""
        q = request.query_params.get('q', '').strip()
        if len(q) < 1:
            return Response([])
        qs = School.objects.filter(
            is_active=True, name__icontains=q
        ).select_related('support_center').order_by('name')[:15]
        data = [{'id': s.id, 'name': s.name,
                 'support_center_name': s.support_center.name if s.support_center else ''}
                for s in qs]
        return Response(data)

    @action(detail=False, methods=['get'])
    def labeling_summary(self, request):
        """전체 학교 라벨링 현황 요약"""
        from django.db.models import Count, Q
        from .models import LabelingCompletion
        qs = School.objects.filter(is_active=True).order_by('support_center__id', 'name')
        center = request.query_params.get('center')
        school_type = request.query_params.get('school_type')
        q = request.query_params.get('q')
        if center:
            qs = qs.filter(support_center__code=center)
        if school_type:
            qs = qs.filter(school_type_id=school_type)
        if q:
            qs = qs.filter(name__icontains=q)
        qs = qs.annotate(
            equip_total=Count('equipment_list'),
            equip_tagged=Count('equipment_list', filter=Q(equipment_list__asset_tag__gt='')),
        ).filter(equip_total__gt=0)
        completed_ids = set(LabelingCompletion.objects.values_list('school_id', flat=True))
        data = [{
            'id': s.id, 'name': s.name,
            'center_name': s.support_center.name if s.support_center else '',
            'school_type_name': s.school_type.name if s.school_type else '',
            'total': s.equip_total, 'tagged': s.equip_tagged,
            'completed': s.id in completed_ids,
        } for s in qs.select_related('support_center', 'school_type')]
        return Response(data)

    @action(detail=False, methods=['get'])
    def tree(self, request):
        """교육청 → 학제 → 학교 3단계 트리 데이터"""
        from collections import defaultdict
        centers = list(SupportCenter.objects.filter(is_active=True).order_by('id').values('id', 'code', 'name'))
        schools = School.objects.filter(is_active=True).order_by('name').values(
            'id', 'name', 'support_center_id',
            'school_type__id', 'school_type__name', 'school_type__order'
        )

        # {center_id: {type_id: [school, ...]}}
        grouped = defaultdict(lambda: defaultdict(list))
        type_meta = {}  # type_id → {name, order}
        for s in schools:
            tid = s['school_type__id']
            type_meta[tid] = {'name': s['school_type__name'], 'order': s['school_type__order']}
            grouped[s['support_center_id']][tid].append({'id': s['id'], 'name': s['name']})

        result = []
        for c in centers:
            by_type = grouped.get(c['id'], {})
            types = []
            for tid, tinfo in sorted(type_meta.items(), key=lambda x: x[1]['order']):
                schools_in = by_type.get(tid, [])
                if not schools_in:
                    continue
                types.append({
                    'id':      tid,
                    'name':    tinfo['name'],
                    'schools': sorted(schools_in, key=lambda s: s['name']),
                    'count':   len(schools_in),
                })
            result.append({
                'id':    c['id'],
                'code':  c['code'],
                'name':  c['name'],
                'count': sum(t['count'] for t in types),
                'types': types,
            })
        return Response(result)

    @action(detail=False, methods=['get'])
    def gis(self, request):
        """카카오맵 마커 데이터"""
        from django.db.models import Count, Q
        qs = self.get_queryset().filter(lat__isnull=False, lng__isnull=False).annotate(
            _active_incidents=Count('incidents', filter=~Q(incidents__status='completed'))
        )
        return Response(SchoolGISSerializer(qs, many=True).data)

    @action(detail=False, methods=['get'])
    def stats(self, request):
        """지원청×학제 통계 (대시보드용)"""
        data = School.objects.filter(is_active=True).values(
            'support_center__name', 'school_type__name'
        ).annotate(count=Count('id')).order_by('support_center__id', 'school_type__order')
        return Response(list(data))

    @action(detail=True, methods=['get'])
    def equipment(self, request, pk=None):
        """학교 장비 목록"""
        from .models import SchoolEquipment
        school = self.get_object()
        qs = school.equipment_list.order_by('category', 'device_id')
        data = list(qs.values(
            'id', 'category', 'model_name', 'manufacturer',
            'building', 'floor', 'install_location', 'device_id', 'network_type',
            'speed', 'tier', 'origin', 'mgmt', 'install_year',
            'asset_tag', 'tagged_at', 'tagged_by__name', 'tag_photo',
        ))
        return Response(data)

    @action(detail=True, methods=['post'], url_path='equipment_tag')
    def equipment_tag(self, request, pk=None):
        """장비 관리번호(라벨링) 부여"""
        from .models import SchoolEquipment
        from django.utils import timezone
        school = self.get_object()
        equip_id = request.data.get('equipment_id')
        asset_tag = request.data.get('asset_tag', '').strip()
        if not equip_id:
            return Response({'error': '장비 ID를 입력하세요.'}, status=400)
        if not asset_tag and asset_tag != '':
            return Response({'error': '관리번호를 입력하세요.'}, status=400)
        try:
            equip = school.equipment_list.get(pk=equip_id)
        except SchoolEquipment.DoesNotExist:
            return Response({'error': '해당 장비를 찾을 수 없습니다.'}, status=404)
        equip.asset_tag = asset_tag
        equip.tagged_at = timezone.now() if asset_tag else None
        equip.tagged_by = request.user if asset_tag else None
        equip.save(update_fields=['asset_tag', 'tagged_at', 'tagged_by', 'updated_at']
                   if hasattr(equip, 'updated_at') else ['asset_tag', 'tagged_at', 'tagged_by'])
        return Response({'success': True, 'asset_tag': asset_tag, 'equipment_id': equip.id})

    @action(detail=True, methods=['post'], url_path='equipment_add')
    def equipment_add(self, request, pk=None):
        """장비 추가 (라벨링 시 리스트에 없는 장비)"""
        from .models import SchoolEquipment
        from django.utils import timezone
        school = self.get_object()
        equip = SchoolEquipment.objects.create(
            school=school,
            category=request.data.get('category', ''),
            model_name=request.data.get('model_name', ''),
            manufacturer=request.data.get('manufacturer', ''),
            building=request.data.get('building', ''),
            floor=request.data.get('floor', ''),
            install_location=request.data.get('install_location', ''),
            device_id=request.data.get('device_id', ''),
            network_type=request.data.get('network_type', ''),
            speed=request.data.get('speed', ''),
            tier=request.data.get('tier', ''),
            mgmt=request.data.get('mgmt', ''),
            asset_tag=request.data.get('asset_tag', ''),
            tagged_at=timezone.now() if request.data.get('asset_tag') else None,
            tagged_by=request.user if request.data.get('asset_tag') else None,
        )
        return Response({'success': True, 'equipment_id': equip.id}, status=201)

    @action(detail=True, methods=['post'], url_path='equipment_photo')
    def equipment_photo(self, request, pk=None):
        """라벨링 스티커 사진 NAS 저장"""
        from .models import SchoolEquipment
        import os, base64
        school = self.get_object()
        equip_id = request.data.get('equipment_id')
        photo_data = request.data.get('photo')  # base64
        asset_tag = request.data.get('asset_tag', '')
        if not equip_id or not photo_data:
            return Response({'error': '장비 ID와 사진 데이터가 필요합니다.'}, status=400)
        try:
            equip = school.equipment_list.get(pk=equip_id)
        except SchoolEquipment.DoesNotExist:
            return Response({'error': '해당 장비를 찾을 수 없습니다.'}, status=404)

        # NAS 저장
        nas_dir = '/app/nas/media/npms/산출물/라벨링'
        os.makedirs(nas_dir, exist_ok=True)
        safe_school = school.name.replace('/', '_')
        safe_tag = (asset_tag or equip.asset_tag or str(equip.id)).replace('/', '_')
        fname = f'라벨링_{safe_school}_{safe_tag}.jpg'
        fpath = os.path.join(nas_dir, fname)

        # base64 디코딩 저장
        if ',' in photo_data:
            photo_data = photo_data.split(',', 1)[1]
        with open(fpath, 'wb') as f:
            f.write(base64.b64decode(photo_data))

        equip.tag_photo = fpath
        equip.save(update_fields=['tag_photo'])
        return Response({'success': True, 'path': fpath})

    @action(detail=True, methods=['post'], url_path='equipment_update')
    def equipment_update(self, request, pk=None):
        """장비 정보 수정 (변경 전 원본 자동 보존)"""
        from .models import SchoolEquipment
        school = self.get_object()
        equip_id = request.data.get('equipment_id')
        try:
            equip = school.equipment_list.get(pk=equip_id)
        except SchoolEquipment.DoesNotExist:
            return Response({'error': '해당 장비를 찾을 수 없습니다.'}, status=404)
        # 원본 보존
        equip.save_original()
        # 수정 가능 필드
        for field in ['category','model_name','manufacturer','building','floor',
                       'install_location','device_id','network_type','speed','tier','mgmt']:
            if field in request.data:
                setattr(equip, field, request.data[field])
        equip.save()
        return Response({'success': True})

    @action(detail=True, methods=['post'], url_path='labeling_complete')
    def labeling_complete(self, request, pk=None):
        """학교 라벨링 완료 처리 + NAS 엑셀 저장"""
        import io, os
        from urllib.parse import quote
        from django.utils import timezone
        from .models import LabelingCompletion
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            return Response({'error': 'openpyxl 필요'}, status=500)

        school = self.get_object()
        qs = school.equipment_list.order_by('category', 'id')

        wb = openpyxl.Workbook()
        # ── 시트1: 현재(변경 후) ──
        ws = wb.active
        ws.title = '라벨링 결과'
        hdr_fill = PatternFill('solid', fgColor='1F497D')
        hdr_font = Font(bold=True, color='FFFFFF', size=10)
        ctr = Alignment(horizontal='center', vertical='center')
        thin = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
        green = PatternFill('solid', fgColor='C6EFCE')

        ws.merge_cells('A1:J1')
        ws.cell(1, 1, f'{school.name} — 라벨링 결과').font = Font(bold=True, size=13)
        headers = ['#','구분','모델명','제조사','건물/층','설치장소','망구분','장비ID','관리번호','유무']
        for ci, h in enumerate(headers, 1):
            c = ws.cell(3, ci, h)
            c.font, c.fill, c.alignment, c.border = hdr_font, hdr_fill, ctr, thin
        for ri, eq in enumerate(qs, 1):
            exist = '유' if (eq.asset_tag and eq.asset_tag != '장비없음') else ('무' if eq.asset_tag == '장비없음' else '-')
            vals = [ri, eq.category, eq.model_name, eq.manufacturer,
                    f'{eq.building}/{eq.floor}', eq.install_location, eq.network_type,
                    eq.device_id, eq.asset_tag or '미부여', exist]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(ri+3, ci, v)
                c.border = thin
                if eq.asset_tag and eq.asset_tag != '장비없음' and ci == 9:
                    c.fill = green

        # ── 시트2: 변경 전 (original_data) ──
        ws2 = wb.create_sheet('변경 전')
        ws2.merge_cells('A1:J1')
        ws2.cell(1, 1, f'{school.name} — 변경 전 원본').font = Font(bold=True, size=13)
        for ci, h in enumerate(headers, 1):
            c = ws2.cell(3, ci, h)
            c.font, c.fill, c.alignment, c.border = hdr_font, hdr_fill, ctr, thin
        changed_count = 0
        for ri, eq in enumerate(qs, 1):
            orig = eq.original_data or {}
            if orig:
                changed_count += 1
                vals = [ri, orig.get('category',''), orig.get('model_name',''),
                        orig.get('manufacturer',''), f"{orig.get('building','')}/{orig.get('floor','')}",
                        orig.get('install_location',''), orig.get('network_type',''),
                        orig.get('device_id',''), '', '-']
            else:
                vals = [ri, eq.category, eq.model_name, eq.manufacturer,
                        f'{eq.building}/{eq.floor}', eq.install_location, eq.network_type,
                        eq.device_id, '', '-']
            for ci, v in enumerate(vals, 1):
                c = ws2.cell(ri+3, ci, v)
                c.border = thin
                if orig:
                    c.fill = PatternFill('solid', fgColor='FFF3CD')

        col_widths = [5, 8, 18, 12, 12, 20, 10, 14, 20, 6]
        for ws_ in [ws, ws2]:
            for ci, w in enumerate(col_widths, 1):
                ws_.column_dimensions[get_column_letter(ci)].width = w

        # NAS 저장
        nas_dir = '/app/nas/media/npms/산출물/라벨링 엑셀'
        os.makedirs(nas_dir, exist_ok=True)
        safe_name = school.name.replace('/', '_')
        fpath = os.path.join(nas_dir, f'라벨링목록_{safe_name}.xlsx')
        wb.save(fpath)

        # 완료 기록
        now = timezone.now()
        comp, _ = LabelingCompletion.objects.update_or_create(
            school=school,
            defaults={'completed_at': now, 'completed_by': request.user, 'excel_path': fpath}
        )
        return Response({'success': True, 'path': fpath, 'changed': changed_count})

    @action(detail=False, methods=['get'], url_path='labeling_stats')
    def labeling_stats(self, request):
        """전체 지원청별 라벨링 통계"""
        from django.db.models import Count, Q
        from .models import LabelingCompletion
        centers = SupportCenter.objects.filter(is_active=True).order_by('id')
        result = []
        for ctr in centers:
            schools = School.objects.filter(is_active=True, support_center=ctr)
            school_ids = list(schools.values_list('id', flat=True))
            from .models import SchoolEquipment
            eq_qs = SchoolEquipment.objects.filter(school_id__in=school_ids)
            total = eq_qs.count()
            tagged = eq_qs.filter(asset_tag__gt='').count()
            # 카테고리별
            cats = {}
            for cat_key, cat_filter in [('스위치', ['스위치']), ('PoE', ['PoE', 'PoE스위치']), ('AP', ['AP'])]:
                cat_qs = eq_qs.filter(category__in=cat_filter)
                cats[cat_key] = {'total': cat_qs.count(), 'tagged': cat_qs.filter(asset_tag__gt='').count()}
            # 완료 학교 수
            completed = LabelingCompletion.objects.filter(school_id__in=school_ids).count()
            # 추가/삭감 (original_data가 있는 건 = 변경됨)
            changed = eq_qs.filter(original_data__isnull=False).count()
            # 장비 추가된 건 (install_year가 없고 asset_tag가 있는 건)
            added = eq_qs.filter(install_year__isnull=True, asset_tag__gt='').count()
            result.append({
                'center_name': ctr.name, 'center_code': ctr.code,
                'school_count': schools.count(), 'completed_count': completed,
                'total': total, 'tagged': tagged,
                'categories': cats, 'changed': changed, 'added': added,
            })
        return Response(result)

    @action(detail=True, methods=['post'], permission_classes=[IsAdmin])
    def equipment_to_assets(self, request, pk=None):
        """SchoolEquipment → Asset 일괄 변환 (슈퍼어드민/관리자 전용)"""
        from datetime import date
        from apps.assets.models import AssetCategory, AssetModel, Asset, AssetHistory

        school = self.get_object()
        equips = school.equipment_list.all()
        if not equips.exists():
            return Response({'message': '변환할 장비 목록이 없습니다.'})

        # SchoolEquipment.category → AssetCategory.code 매핑
        cat_map = {
            '스위치':    'switch',
            'switch':    'switch',
            'PoE스위치': 'poe_switch',
            'PoE':       'poe_switch',
            'poe':       'poe_switch',
            'AP':        'ap',
            'ap':        'ap',
            '무선AP':    'ap',
            '라우터':    'router',
            'router':    'router',
            '서버':      'server',
            'server':    'server',
        }

        created_count = 0
        skipped = []

        for eq in equips:
            cat_code = cat_map.get(eq.category.strip(), 'switch')
            try:
                cat = AssetCategory.objects.get(code=cat_code)
            except AssetCategory.DoesNotExist:
                cat = AssetCategory.objects.filter(code='switch').first()

            mfr   = eq.manufacturer or '미상'
            model = eq.model_name   or '미상'
            asset_model, _ = AssetModel.objects.get_or_create(
                manufacturer=mfr,
                model_name=model,
                defaults={'category': cat}
            )

            # 고유 식별자: device_id 우선, 없으면 학교ID+순번으로 생성
            sn = eq.device_id.strip() if eq.device_id and eq.device_id.strip() else f'EQ-{school.id}-{eq.id}'
            installed_date = date(eq.install_year, 1, 1) if eq.install_year else None

            asset, created = Asset.objects.get_or_create(
                serial_number=sn,
                defaults={
                    'asset_model':      asset_model,
                    'status':           'installed',
                    'current_school':   school,
                    'install_location': eq.install_location or '',
                    'installed_at':     installed_date,
                    'asset_tag':        sn[:50],
                }
            )
            if created:
                AssetHistory.objects.create(
                    asset=asset,
                    action='install',
                    from_location='학교장비목록 가져오기',
                    to_location=f'{school.name} {eq.install_location or ""}',
                    worker=request.user,
                    note=f'SchoolEquipment 일괄 변환 (망:{eq.network_type} 속도:{eq.speed})',
                )
                created_count += 1
            else:
                skipped.append({'sn': sn, 'reason': '이미 등록된 S/N'})

        return Response({
            'created': created_count,
            'skipped_count': len(skipped),
            'skipped': skipped[:20],  # 최대 20건만 반환
        })

    @action(detail=True, methods=['get'])
    def buildings(self, request, pk=None):
        school = self.get_object()
        return Response(SchoolBuildingSerializer(school.buildings.all(), many=True).data)

    @action(detail=True, methods=['get'])
    def contacts(self, request, pk=None):
        school = self.get_object()
        return Response(SchoolContactSerializer(school.contacts.all(), many=True).data)

    @action(detail=True, methods=['get'])
    def all_contacts(self, request, pk=None):
        """등록 담당자 + 장애 이력 신고자 통합 목록"""
        from apps.incidents.models import Incident

        school = self.get_object()

        # 1) 등록된 담당자
        result = []
        registered_phones = set()
        for c in school.contacts.all():
            result.append({
                'name':       c.name,
                'position':   c.position or '',
                'phone':      c.phone,
                'email':      c.email or '',
                'is_primary': c.is_primary,
                'source':     'registered',
            })
            if c.phone:
                registered_phones.add(c.phone.replace('-', '').strip())

        # 2) 장애 이력 신고자 (등록 담당자와 전화번호 중복 제외)
        reporters = (
            Incident.objects
            .filter(school=school)
            .exclude(requester_name='')
            .values('requester_name', 'requester_phone', 'requester_position')
            .distinct()
            .order_by('requester_name')
        )
        seen = set()
        for r in reporters:
            phone_norm = (r['requester_phone'] or '').replace('-', '').strip()
            key = (r['requester_name'], phone_norm)
            if key in seen:
                continue
            seen.add(key)
            if phone_norm and phone_norm in registered_phones:
                continue
            result.append({
                'name':       r['requester_name'],
                'position':   r['requester_position'] or '',
                'phone':      r['requester_phone'] or '',
                'email':      '',
                'is_primary': False,
                'source':     'incident',
            })

        return Response(result)

    @action(detail=True, methods=['get'])
    def inspection_history(self, request, pk=None):
        """학교별 점검 이력 (SchoolInspection + 완료 Report)"""
        school = self.get_object()

        from apps.progress.models import SchoolInspection
        from apps.reports.models import Report

        PLAN_TYPE_LABELS = {
            'regular': '정기점검', 'special': '특별점검',
            'quarterly': '분기점검', 'project': '사업점검',
            'survey': '실태조사', 'followup': '사후점검',
        }
        STATUS_LABELS = {
            'scheduled': '예정', 'in_progress': '진행중',
            'completed': '완료', 'cancelled': '취소',
        }

        # SchoolInspection 이력 (최근 20건)
        insp_qs = SchoolInspection.objects.filter(
            school=school
        ).select_related('plan', 'assigned_worker').order_by('-plan__start_date')[:20]

        inspections = []
        for si in insp_qs:
            inspections.append({
                'plan_name':   si.plan.name,
                'plan_type':   PLAN_TYPE_LABELS.get(si.plan.plan_type, si.plan.plan_type),
                'start_date':  si.plan.start_date.isoformat(),
                'end_date':    si.plan.end_date.isoformat(),
                'status':      STATUS_LABELS.get(si.status, si.status),
                'status_code': si.status,
                'worker':      si.assigned_worker.name if si.assigned_worker else '-',
                'inspect_date': si.inspect_date.isoformat() if si.inspect_date else None,
                'note':        si.note or '',
            })

        # 완료 정기점검 보고서 (최근 10건)
        rpt_qs = Report.objects.filter(
            school=school,
            template__report_type='regular',
            status='completed',
        ).select_related('created_by').order_by('-completed_at')[:10]

        reports = [{
            'title':        r.title,
            'completed_at': r.completed_at.strftime('%Y-%m-%d') if r.completed_at else '-',
            'created_by':   r.created_by.name if r.created_by else '-',
        } for r in rpt_qs]

        return Response({'inspections': inspections, 'reports': reports})

    @action(detail=True, methods=['get'])
    def floor_plan(self, request, pk=None):
        """건물·층·호실 트리 + 평면도 좌표 (건물정보 탭용)"""
        school = self.get_object()
        result = []
        for bld in school.buildings.prefetch_related(
            'floor_list__rooms'
        ).order_by('order'):
            floors = []
            for fl in bld.floor_list.order_by('-floor_num'):
                rooms = []
                for rm in fl.rooms.order_by('room_number', 'name'):
                    rooms.append({
                        'id':          rm.id,
                        'name':        rm.name,
                        'room_number': rm.room_number,
                        'room_type':   rm.room_type,
                        'area_m2':     float(rm.area_m2) if rm.area_m2 else None,
                        'pos_x':       float(rm.pos_x)   if rm.pos_x   else None,
                        'pos_y':       float(rm.pos_y)   if rm.pos_y   else None,
                        'pos_w':       float(rm.pos_w)   if rm.pos_w   else None,
                        'pos_h':       float(rm.pos_h)   if rm.pos_h   else None,
                    })
                floors.append({
                    'id':         fl.id,
                    'floor_num':  fl.floor_num,
                    'floor_name': fl.floor_name,
                    'rooms':      rooms,
                })
            result.append({
                'id':       bld.id,
                'name':     bld.name,
                'floors':   bld.floors,
                'basement': bld.basement,
                'floor_list': floors,
            })
        # VSDX 임포트 로그 (최근 5건)
        from django.utils import timezone as tz
        logs = VsdxImportLog.objects.filter(school=school).order_by('-imported_at')[:5]
        log_data = [{'file_name': l.file_name, 'status': l.status,
                     'room_count': l.room_count,
                     'imported_at': tz.localtime(l.imported_at).strftime('%Y-%m-%d %H:%M'),
                     'error_msg': l.error_msg} for l in logs]
        return Response({'buildings': result, 'vsdx_logs': log_data})

    @action(detail=True, methods=['post'], permission_classes=[IsAdmin])
    def trigger_vsdx(self, request, pk=None):
        """VSDX 수동 재파싱 트리거 (관리자 전용)"""
        import os
        from django.conf import settings
        from .tasks import import_vsdx_file

        school = self.get_object()
        vsdx_folder = os.path.join(
            getattr(settings, 'NAS_MEDIA_ROOT', '/app/nas/media/npms'),
            'data', '건물정보_비지오'
        )
        # 학교명과 일치하는 파일 검색
        matched = []
        if os.path.isdir(vsdx_folder):
            for fname in os.listdir(vsdx_folder):
                if fname.lower().endswith('.vsdx') and school.name in fname:
                    import_vsdx_file.delay(os.path.join(vsdx_folder, fname))
                    matched.append(fname)
        if not matched:
            return Response({'error': f'{school.name} VSDX 파일 없음'}, status=404)
        return Response({'queued': matched})

    @action(detail=True, methods=['get'])
    def rooms_for_select(self, request, pk=None):
        """사진·장애접수 등 타 앱용 cascade 셀렉트 데이터
        ?building=건물ID&floor=층ID"""
        school    = self.get_object()
        bld_id    = request.query_params.get('building')
        floor_id  = request.query_params.get('floor')

        if floor_id:
            rooms = SchoolRoom.objects.filter(
                floor_id=floor_id
            ).exclude(room_type='support').order_by('room_number', 'name')
            return Response([{'id': r.id, 'name': r.name,
                               'room_number': r.room_number} for r in rooms])

        if bld_id:
            floors = SchoolFloor.objects.filter(
                building_id=bld_id
            ).order_by('-floor_num')
            return Response([{'id': f.id, 'floor_name': f.floor_name,
                               'floor_num': f.floor_num} for f in floors])

        # 건물 목록
        buildings = school.buildings.order_by('order')
        return Response([{'id': b.id, 'name': b.name} for b in buildings])

    @action(detail=True, methods=['get'], url_path='equipment_excel')
    def equipment_excel(self, request, pk=None):
        """장비 라벨링 현황 엑셀 다운로드"""
        import io
        from urllib.parse import quote
        from django.http import HttpResponse
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            return HttpResponse('openpyxl 필요', status=500)

        school = self.get_object()
        qs = school.equipment_list.order_by('category', 'id')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '라벨링 현황'

        hdr_fill = PatternFill('solid', fgColor='1F497D')
        hdr_font = Font(bold=True, color='FFFFFF', size=10)
        ctr = Alignment(horizontal='center', vertical='center')
        thin = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
        green = PatternFill('solid', fgColor='C6EFCE')

        ws.merge_cells('A1:H1')
        c = ws.cell(1, 1, f'{school.name} — 장비 라벨링 현황')
        c.font = Font(bold=True, size=13)
        c.alignment = ctr

        headers = ['#', '구분', '모델명', '건물/층', '설치장소', '망구분', '관리번호', '부여일시']
        for ci, h in enumerate(headers, 1):
            c = ws.cell(3, ci, h)
            c.font, c.fill, c.alignment, c.border = hdr_font, hdr_fill, ctr, thin

        for ri, eq in enumerate(qs, 1):
            vals = [
                ri, eq.category, eq.model_name or '',
                f'{eq.building or ""}/{eq.floor or ""}',
                eq.install_location or '', eq.network_type or '',
                eq.asset_tag or '미부여',
                eq.tagged_at.strftime('%Y-%m-%d %H:%M') if eq.tagged_at else '',
            ]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(ri + 3, ci, v)
                c.border = thin
                c.alignment = ctr if ci in (1, 2, 6) else Alignment(vertical='center')
                if eq.asset_tag and ci == 7:
                    c.fill = green

        col_widths = [5, 8, 18, 12, 20, 10, 20, 16]
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f'라벨링현황_{school.name}.xlsx'
        resp = HttpResponse(buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(fname)}"
        return resp

    @action(detail=True, methods=['get'], url_path='equipment_locations')
    def equipment_locations(self, request, pk=None):
        """SchoolEquipment 기반 건물/층/위치 cascade (건물 데이터 없는 학교용)
        파라미터 없음 → 건물 목록, ?building=X → 층 목록, ?building=X&floor=Y → 위치 목록"""
        school = self.get_object()
        qs = school.equipment_list
        building = request.query_params.get('building')
        floor = request.query_params.get('floor')

        if building and floor:
            # 위치(교실) 목록
            locations = (
                qs.filter(building=building, floor=floor)
                .exclude(install_location='')
                .values_list('install_location', flat=True)
                .distinct()
                .order_by('install_location')
            )
            return Response(list(locations))

        if building:
            # 층 목록
            floors = (
                qs.filter(building=building)
                .exclude(floor='')
                .values_list('floor', flat=True)
                .distinct()
                .order_by('floor')
            )
            return Response(list(floors))

        # 건물 목록
        buildings = (
            qs.exclude(building='')
            .values_list('building', flat=True)
            .distinct()
            .order_by('building')
        )
        return Response(list(buildings))

    @action(detail=True, methods=['post'], url_path='add_building')
    def add_building(self, request, pk=None):
        """건물 신규 추가 (사진/장애 등록 모달에서 직접 입력 시)"""
        school = self.get_object()
        name = (request.data.get('name') or '').strip()
        if not name:
            return Response({'error': '건물명을 입력하세요.'}, status=status.HTTP_400_BAD_REQUEST)
        bld, _ = SchoolBuilding.objects.get_or_create(school=school, name=name, defaults={'order': 99})
        return Response({'id': bld.id, 'name': bld.name})

    @action(detail=True, methods=['post'], url_path='add_floor')
    def add_floor(self, request, pk=None):
        """층 신규 추가"""
        building_id = request.data.get('building_id')
        floor_name  = (request.data.get('floor_name') or '').strip()
        if not building_id or not floor_name:
            return Response({'error': '건물ID와 층이름을 입력하세요.'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            building = SchoolBuilding.objects.get(pk=building_id)
        except SchoolBuilding.DoesNotExist:
            return Response({'error': '건물을 찾을 수 없습니다.'}, status=status.HTTP_404_NOT_FOUND)
        floor, _ = SchoolFloor.objects.get_or_create(
            building=building, floor_name=floor_name,
            defaults={'floor_num': SchoolFloor.objects.filter(building=building).count() + 1}
        )
        return Response({'id': floor.id, 'floor_name': floor.floor_name})

    @action(detail=True, methods=['post'], url_path='add_room')
    def add_room(self, request, pk=None):
        """교실 신규 추가"""
        floor_id  = request.data.get('floor_id')
        room_name = (request.data.get('room_name') or '').strip()
        if not floor_id or not room_name:
            return Response({'error': '층ID와 교실명을 입력하세요.'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            floor = SchoolFloor.objects.get(pk=floor_id)
        except SchoolFloor.DoesNotExist:
            return Response({'error': '층을 찾을 수 없습니다.'}, status=status.HTTP_404_NOT_FOUND)
        room, _ = SchoolRoom.objects.get_or_create(
            floor=floor, name=room_name,
            defaults={'room_type': 'other'}
        )
        return Response({'id': room.id, 'name': room.name})

    @action(detail=False, methods=['post'], permission_classes=[IsAdmin])
    def csv_upload(self, request):
        """학교 CSV 업로드 (학교명,교육지원청,학제,주소,위도,경도)
        mode: add_update(기본) | reset(초기화 후 입력)
        """
        file = request.FILES.get('file')
        if not file:
            return Response({'error': '파일이 없습니다.'}, status=status.HTTP_400_BAD_REQUEST)
        mode = request.data.get('mode', 'add_update')
        if mode == 'reset':
            School.objects.all().update(is_active=False)
        reader = csv.DictReader(io.StringIO(file.read().decode('utf-8-sig')))
        created, updated, errors = 0, 0, []
        for row in reader:
            try:
                center = SupportCenter.objects.get(name=row['교육지원청'])
                stype  = SchoolType.objects.get(name=row['학제'])
                school, is_new = School.objects.update_or_create(
                    name=row['학교명'], support_center=center,
                    defaults={
                        'school_type': stype,
                        'address':     row.get('주소', ''),
                        'lat':         row.get('위도') or None,
                        'lng':         row.get('경도') or None,
                        'phone':       row.get('전화번호', ''),
                        'is_active':   True,
                    }
                )
                if is_new:
                    created += 1
                else:
                    updated += 1
            except Exception as e:
                errors.append({'row': row, 'error': str(e)})
        return Response({'created': created, 'updated': updated, 'errors': errors})

    @action(detail=False, methods=['post'], permission_classes=[IsAdmin])
    def contacts_csv_upload(self, request):
        """교직원(담당자) CSV 업로드 (학교명,성명,직위,전화번호,이메일)
        mode: add_update(기본) | reset(학교별 담당자 초기화 후 입력)
        """
        file = request.FILES.get('file')
        if not file:
            return Response({'error': '파일이 없습니다.'}, status=status.HTTP_400_BAD_REQUEST)
        mode = request.data.get('mode', 'add_update')
        try:
            content = file.read().decode('utf-8-sig')
        except UnicodeDecodeError:
            content = file.read().decode('cp949', errors='replace')
        reader = csv.DictReader(io.StringIO(content))
        rows = list(reader)
        if mode == 'reset':
            school_names = {r.get('학교명', '').strip() for r in rows if r.get('학교명')}
            School.objects.filter(name__in=school_names).prefetch_related('contacts')
            from apps.schools.models import SchoolContact
            SchoolContact.objects.filter(school__name__in=school_names).delete()
        from apps.schools.models import SchoolContact
        created, updated, errors = 0, 0, []
        for row in rows:
            try:
                school = School.objects.get(name=row['학교명'].strip())
                phone  = row.get('전화번호', '').strip()
                name   = row.get('성명', '').strip()
                if not name:
                    errors.append({'row': row, 'error': '성명 필수'})
                    continue
                contact, is_new = SchoolContact.objects.update_or_create(
                    school=school, name=name,
                    defaults={
                        'position': row.get('직위', '').strip(),
                        'phone':    phone,
                        'email':    row.get('이메일', '').strip(),
                    }
                )
                if is_new:
                    created += 1
                else:
                    updated += 1
            except School.DoesNotExist:
                errors.append({'row': row, 'error': f"학교 없음: {row.get('학교명')}"})
            except Exception as e:
                errors.append({'row': row, 'error': str(e)})
        return Response({'created': created, 'updated': updated, 'errors': errors})

    @action(detail=False, methods=['get'], permission_classes=[IsAdmin])
    def contacts_csv_download(self, request):
        """교직원(담당자) CSV 다운로드"""
        from apps.schools.models import SchoolContact
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = 'attachment; filename="contacts.csv"'
        writer = csv.writer(response)
        writer.writerow(['학교명', '교육지원청', '성명', '직위', '전화번호', '이메일'])
        for c in SchoolContact.objects.select_related('school', 'school__support_center').order_by('school__name', 'name'):
            writer.writerow([
                c.school.name,
                c.school.support_center.name if c.school.support_center else '',
                c.name, c.position, c.phone, c.email,
            ])
        return response

    @action(detail=False, methods=['get'])
    def csv_download(self, request):
        """학교 목록 CSV 다운로드"""
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = 'attachment; filename="schools.csv"'
        writer = csv.writer(response)
        writer.writerow(['학교명', '교육지원청', '학제', '주소', '위도', '경도', '전화번호'])
        for s in School.objects.select_related('support_center', 'school_type').filter(is_active=True):
            writer.writerow([s.name, s.support_center.name, s.school_type.name,
                             s.address, s.lat, s.lng, s.phone])
        return response


class SchoolBuildingViewSet(viewsets.ModelViewSet):
    queryset = SchoolBuilding.objects.select_related('school')
    serializer_class = SchoolBuildingSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        school_id = self.request.query_params.get('school')
        qs = super().get_queryset()
        if school_id:
            qs = qs.filter(school_id=school_id)
        return qs

    @action(detail=False, methods=['get'])
    def csv_download(self, request):
        """건물명 CSV 다운로드"""
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = 'attachment; filename="buildings.csv"'
        writer = csv.writer(response)
        writer.writerow(['학교명', '건물명', '교육지원청', '학제', '층수'])
        for b in SchoolBuilding.objects.select_related('school', 'school__support_center', 'school__school_type'):
            writer.writerow([
                b.school.name,
                b.name,
                b.school.support_center.name if b.school.support_center else '',
                b.school.school_type.name    if b.school.school_type    else '',
                b.floor_count if hasattr(b, 'floor_count') else '',
            ])
        return response

    @action(detail=False, methods=['get'])
    def rooms_csv_download(self, request):
        """건물·층·호실 전체 CSV 다운로드
        ?school=학교ID  (생략 시 전체 — 슈퍼관리자 전용)
        ?center=지원청코드
        """
        from urllib.parse import quote
        school_id = request.query_params.get('school')
        center    = request.query_params.get('center')

        # 전체 다운로드는 슈퍼관리자 전용
        if not school_id and getattr(request.user, 'role', '') != 'superadmin':
            return Response({'error': '전체 다운로드는 슈퍼관리자만 가능합니다.'}, status=403)

        qs = SchoolRoom.objects.select_related(
            'floor__building__school',
            'floor__building__school__support_center',
            'floor__building__school__school_type',
        ).order_by(
            'floor__building__school__support_center__name',
            'floor__building__school__name',
            'floor__building__name',
            '-floor__floor_num',
            'room_number',
            'name',
        )
        if school_id:
            qs = qs.filter(floor__building__school_id=school_id)
        if center:
            qs = qs.filter(floor__building__school__support_center__code=center)

        # 파일명 결정: 건물정보_학교명.csv 또는 건물정보_전체.csv
        if school_id:
            school_obj = School.objects.filter(pk=school_id).first()
            label = school_obj.name if school_obj else school_id
        else:
            label = '전체'
        fname_ascii = quote(f'건물정보_{label}.csv', safe='')
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = f"attachment; filename*=UTF-8''{fname_ascii}"
        import re as _re
        _date_pat = _re.compile(r'^\d{1,2}[-/]\d{1,2}$')

        def _safe(v):
            """Excel 날짜 자동변환 방지: '1-4' → ="1-4" """
            s = str(v) if v is not None else ''
            if _date_pat.match(s):
                return f'="{s}"'
            return s

        writer = csv.writer(response)
        writer.writerow([
            '교육지원청', '학제', '학교명',
            '건물명', '지상층수', '지하층수',
            '층명', '층번호',
            '호실번호', '호실명', '유형',
            '면적(㎡)',
        ])
        ROOM_TYPE_KO = {
            'class': '학급교실', 'special': '특별교실', 'office': '교무/행정',
            'toilet': '화장실', 'support': '기타',
        }
        for rm in qs:
            fl  = rm.floor
            bld = fl.building
            sch = bld.school
            writer.writerow([
                sch.support_center.name if sch.support_center else '',
                sch.school_type.name    if sch.school_type    else '',
                sch.name,
                bld.name, bld.floors, bld.basement,
                fl.floor_name, fl.floor_num,
                _safe(rm.room_number), _safe(rm.name),
                ROOM_TYPE_KO.get(rm.room_type, rm.room_type),
                float(rm.area_m2) if rm.area_m2 else '',
            ])
        return response

    @action(detail=False, methods=['post'], permission_classes=[IsAdmin])
    def csv_upload(self, request):
        """건물명 CSV 업로드 (학교명,건물명)
        mode: add_update(기본) | reset(학교별 건물 초기화 후 입력)
        """
        file = request.FILES.get('file')
        if not file:
            return Response({'error': '파일 없음'}, status=status.HTTP_400_BAD_REQUEST)
        mode = request.data.get('mode', 'add_update')
        content = file.read().decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(content))
        rows = list(reader)
        if mode == 'reset':
            school_names = {r.get('학교명', '').strip() for r in rows if r.get('학교명')}
            SchoolBuilding.objects.filter(school__name__in=school_names).delete()
        created, updated, errors = 0, 0, []
        for row in rows:
            try:
                school = School.objects.get(name=row['학교명'].strip())
                _, is_new = SchoolBuilding.objects.update_or_create(
                    school=school, name=row['건물명'].strip(),
                    defaults={'floors': int(row['층수']) if row.get('층수', '').strip().isdigit() else 1}
                )
                if is_new:
                    created += 1
                else:
                    updated += 1
            except Exception as e:
                errors.append({'row': row, 'error': str(e)})
        return Response({'created': created, 'updated': updated, 'errors': errors})

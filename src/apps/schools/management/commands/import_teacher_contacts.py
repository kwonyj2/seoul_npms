"""
선생님정보.xlsx → school_contacts 테이블 임포트
사용법:
  python manage.py import_teacher_contacts
  python manage.py import_teacher_contacts --file /path/to/other.xlsx
  python manage.py import_teacher_contacts --dry-run
"""
import os
from django.core.management.base import BaseCommand
from django.conf import settings
from django.db import transaction
from apps.schools.models import School, SchoolContact


class Command(BaseCommand):
    help = '선생님정보.xlsx를 읽어 school_contacts 테이블에 저장'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            default=os.path.join(
                getattr(settings, 'NAS_MEDIA_ROOT', '/app/nas/media/npms'),
                'data', '선생님정보.xlsx'
            ),
            help='엑셀 파일 경로 (기본: NAS_MEDIA_ROOT/data/선생님정보.xlsx)',
        )
        parser.add_argument('--dry-run', action='store_true', help='DB 저장 없이 결과만 출력')

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError:
            self.stderr.write('openpyxl 없음: pip install openpyxl')
            return

        filepath = options['file']
        dry_run = options['dry_run']

        if not os.path.exists(filepath):
            self.stderr.write(f'파일 없음: {filepath}')
            return

        wb = openpyxl.load_workbook(filepath)
        ws = wb.active

        # 헤더 행 확인 (1행)
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        self.stdout.write(f'헤더: {headers}')
        self.stdout.write(f'총 {ws.max_row - 1}개 행 처리 시작...\n')

        # 컬럼 인덱스 자동 감지
        def col(name_candidates):
            for name in name_candidates:
                for i, h in enumerate(headers, 1):
                    if h and name in str(h):
                        return i
            return None

        col_school = col(['학교명', '학교'])
        col_name   = col(['선생님', '담당자', '성명', '이름'])
        col_phone  = col(['전화번호', '연락처', '전화'])
        col_pos    = col(['직책', '직위'])  # 없을 수 있음

        if not all([col_school, col_name, col_phone]):
            self.stderr.write(f'필수 컬럼을 찾지 못했습니다. 헤더: {headers}')
            return

        saved = skipped = updated = 0
        not_found = []

        with transaction.atomic():
            for row in range(2, ws.max_row + 1):
                school_name = ws.cell(row, col_school).value
                name        = ws.cell(row, col_name).value
                phone       = ws.cell(row, col_phone).value
                position    = ws.cell(row, col_pos).value if col_pos else '담당자'

                if not school_name or not name:
                    skipped += 1
                    continue

                school_name = str(school_name).strip()
                name        = str(name).strip()[:50]
                phone       = str(phone).strip()[:20] if phone else ''
                position    = str(position).strip()[:50] if position else '담당자'

                school = School.objects.filter(name=school_name).first()
                if not school:
                    not_found.append(school_name)
                    skipped += 1
                    continue

                if dry_run:
                    self.stdout.write(f'  [DRY] {school_name} | {name} | {phone}')
                    saved += 1
                    continue

                obj, created = SchoolContact.objects.update_or_create(
                    school=school,
                    position=position,
                    defaults={
                        'name':       name,
                        'phone':      phone,
                        'is_primary': True,
                    },
                )
                if created:
                    saved += 1
                else:
                    updated += 1

            if dry_run:
                transaction.set_rollback(True)

        self.stdout.write(self.style.SUCCESS(
            f'\n완료: 신규 {saved}건 / 업데이트 {updated}건 / 건너뜀 {skipped}건'
        ))
        if not_found:
            unique_not_found = sorted(set(not_found))
            self.stdout.write(f'미매칭 학교 {len(unique_not_found)}개:')
            for s in unique_not_found[:20]:
                self.stdout.write(f'  - {s}')
            if len(unique_not_found) > 20:
                self.stdout.write(f'  ... 외 {len(unique_not_found)-20}개')

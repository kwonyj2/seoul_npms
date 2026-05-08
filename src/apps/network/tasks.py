from config.celery import app as celery_app
import logging
import os
import json
import gc

logger = logging.getLogger(__name__)

# 일괄 분석 진행 상태 캐시 키
BULK_DIAGRAM_KEY = 'bulk_diagram_progress'


@celery_app.task(bind=True, max_retries=3)
def execute_network_command(self, command_id):
    """원격 명령 비동기 실행"""
    from .models import NetworkCommand
    from django.utils import timezone
    try:
        cmd = NetworkCommand.objects.select_related('device').get(id=command_id)
        cmd.status = 'running'
        cmd.save(update_fields=['status'])

        device = cmd.device
        result = ''

        if device.ssh_enabled and device.ip_address:
            try:
                import paramiko
                client = paramiko.SSHClient()
                client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                client.connect(device.ip_address, timeout=10, username='admin', password='')
                stdin, stdout, stderr = client.exec_command(cmd.command, timeout=30)
                result = stdout.read().decode() + stderr.read().decode()
                client.close()
                cmd.status = 'success'
            except Exception as e:
                result = str(e)
                cmd.status = 'failed'
        else:
            result = 'SSH가 비활성화된 장비입니다.'
            cmd.status = 'failed'

        cmd.result = result
        cmd.save(update_fields=['result', 'status'])
    except NetworkCommand.DoesNotExist:
        logger.error(f'NetworkCommand {command_id} not found')
    except Exception as exc:
        logger.error(f'Command execution error: {exc}')
        raise self.retry(exc=exc, countdown=30)


@celery_app.task
def poll_snmp_devices():
    """SNMP 장비 주기적 폴링"""
    from .models import SnmpDevice, SnmpMetric, NetworkDevice
    from django.utils import timezone

    try:
        from pysnmp.hlapi import (
            getCmd, SnmpEngine, CommunityData, UdpTransportTarget,
            ContextData, ObjectType, ObjectIdentity
        )
    except ImportError:
        logger.warning('pysnmp not installed, skipping SNMP polling')
        return

    active_devices = SnmpDevice.objects.filter(is_active=True).select_related('device')
    polled = 0
    for snmp_dev in active_devices:
        device = snmp_dev.device
        if not device.ip_address:
            continue
        try:
            for error_indication, error_status, error_index, var_binds in getCmd(
                SnmpEngine(),
                CommunityData(snmp_dev.community),
                UdpTransportTarget((device.ip_address, snmp_dev.port), timeout=3, retries=1),
                ContextData(),
                ObjectType(ObjectIdentity('1.3.6.1.2.1.1.1.0')),
                ObjectType(ObjectIdentity('1.3.6.1.2.1.1.3.0')),
            ):
                if not error_indication and not error_status:
                    now = timezone.now()
                    metrics = []
                    for varBind in var_binds:
                        metrics.append(SnmpMetric(
                            device=device,
                            metric_name=str(varBind[0]),
                            oid=str(varBind[0]),
                            value=str(varBind[1]),
                            collected_at=now,
                        ))
                    if metrics:
                        SnmpMetric.objects.bulk_create(metrics)
                    device.status = 'up'
                    device.last_seen = now
                    device.save(update_fields=['status', 'last_seen'])
                else:
                    device.status = 'down'
                    device.save(update_fields=['status'])
                polled += 1
                snmp_dev.last_poll_at = timezone.now()
                snmp_dev.save(update_fields=['last_poll_at'])
                break
        except Exception as e:
            logger.warning(f'SNMP poll error {device.ip_address}: {e}')
            device.status = 'unknown'
            device.save(update_fields=['status'])
    logger.info(f'SNMP polling complete: {polled} devices')


# ── PPTX 구성도 파싱 ─────────────────────────────────────
# 기존 이미지 AI 분석(Claude Vision) → PPTX XML 파싱으로 전환
# 정확도 70~80% → 95%+, 속도 60배, 비용 0원

def _analyze_image_with_claude(image_path: str) -> dict:
    """(Deprecated) 이미지 분석 — PPTX 파서로 대체됨. 호환성을 위해 남겨둠"""
    raise NotImplementedError('이미지 AI 분석은 제거되었습니다. PPTX 사용')

    api_key = os.environ.get('ANTHROPIC_API_KEY', '')
    if not api_key:
        raise ValueError('ANTHROPIC_API_KEY가 설정되지 않았습니다.')

    import base64
    with open(image_path, 'rb') as f:
        image_data = base64.standard_b64encode(f.read()).decode('utf-8')

    ext = os.path.splitext(image_path)[1].lower()
    media_type = {'jpg': 'image/jpeg', 'jpeg': 'image/jpeg', 'png': 'image/png'}.get(ext.lstrip('.'), 'image/jpeg')

    client = anthropic.Anthropic(api_key=api_key)
    message = client.messages.create(
        model='claude-sonnet-4-6',
        max_tokens=4096,
        messages=[{
            'role': 'user',
            'content': [
                {'type': 'image', 'source': {'type': 'base64', 'media_type': media_type, 'data': image_data}},
                {'type': 'text', 'text': DIAGRAM_PROMPT},
            ],
        }],
    )
    text = message.content[0].text.strip()
    # ```json ... ``` 블록 제거
    if text.startswith('```'):
        text = text.split('```')[1]
        if text.startswith('json'):
            text = text[4:]
    return json.loads(text.strip())


def _import_topology_data(school, data: dict) -> tuple:
    """topology dict → NetworkDevice + NetworkLink 저장, (devices, links) 반환"""
    from .models import NetworkDevice, NetworkLink

    nodes = data.get('nodes', [])
    edges = data.get('edges', [])

    NetworkLink.objects.filter(from_device__school=school, link_type='manual').delete()
    NetworkDevice.objects.filter(school=school, ip_address__isnull=True, snmp_enabled=False).delete()

    CABLE_MAP = {'광': 'fiber', 'Cat6': 'cat6', 'Cat5e': 'cat5e', 'Cat5': 'cat5'}
    name_to_device = {}
    created_devices = 0
    for node in nodes:
        name = node.get('name', '').strip()
        if not name:
            continue
        dev, created = NetworkDevice.objects.get_or_create(
            school=school, name=name,
            defaults={
                'device_type':  node.get('device_type', 'switch'),
                'model':        node.get('model', ''),
                'location':     node.get('location', ''),
                'network_type': node.get('network_type', ''),
                'status':       'unknown',
            },
        )
        if created:
            created_devices += 1
        name_to_device[name] = dev

    created_links = 0
    for edge in edges:
        fd = name_to_device.get(edge.get('from', ''))
        td = name_to_device.get(edge.get('to', ''))
        if fd and td and fd != td:
            NetworkLink.objects.create(
                from_device=fd, to_device=td,
                link_type='manual', is_active=True,
                cable_type=CABLE_MAP.get(edge.get('cable_type', ''), 'unknown'),
                network_type=edge.get('network_type', ''),
            )
            created_links += 1

    return created_devices, created_links


@celery_app.task
def analyze_single_diagram(school_id: int, image_path: str):
    """단일 학교 구성도 이미지 분석 → 토폴로지 저장"""
    from apps.schools.models import School
    try:
        school = School.objects.get(id=school_id)
    except School.DoesNotExist:
        logger.error(f'School {school_id} not found')
        return {'status': 'error', 'message': '학교 없음'}

    try:
        data = _analyze_image_with_claude(image_path)
        devices, links = _import_topology_data(school, data)
        logger.info(f'[{school.name}] 장비 {devices}개, 링크 {links}개 등록')
        return {'status': 'ok', 'devices': devices, 'links': links}
    except Exception as e:
        logger.error(f'[{school.name}] 분석 오류: {e}')
        return {'status': 'error', 'message': str(e)}


def _import_pptx_topology(school, data: dict, pptx_path: str = '') -> tuple:
    """PPTX 파서 결과를 DB에 저장
    - 수동 입력 장비(ip_address 또는 snmp_enabled)는 보존
    - 자동 등록 장비/링크만 삭제 후 재생성
    """
    from .models import NetworkDevice, NetworkLink, NetworkTopology
    from django.utils import timezone
    import os

    nodes = data.get('nodes', [])
    edges = data.get('edges', [])
    slides = data.get('slides', [])

    # 기존 자동 등록 장비·링크만 삭제 (수동 데이터 보호)
    NetworkLink.objects.filter(from_device__school=school, link_type='manual').delete()
    auto_devices = NetworkDevice.objects.filter(
        school=school, ip_address__isnull=True, snmp_enabled=False
    )
    auto_devices.delete()

    # 장비 생성
    name_to_device = {}
    created_devices = 0
    for node in nodes:
        name = node.get('name', '').strip()
        if not name:
            continue
        dev, created = NetworkDevice.objects.get_or_create(
            school=school, name=name,
            defaults={
                'device_type': node.get('device_type') or 'switch',
                'model': node.get('model', ''),
                'location': node.get('location', ''),
                'network_type': node.get('network_type', ''),
                'slide_source': node.get('slide_title', ''),
                'status': 'unknown',
            },
        )
        if created:
            created_devices += 1
        name_to_device[name] = dev

    # 링크 생성
    created_links = 0
    for edge in edges:
        fd = name_to_device.get(edge.get('from_name', ''))
        td = name_to_device.get(edge.get('to_name', ''))
        if fd and td and fd != td:
            NetworkLink.objects.create(
                from_device=fd, to_device=td,
                link_type='manual', is_active=True,
                cable_type=edge.get('cable_type', 'unknown'),
                network_type=edge.get('network_type', ''),
            )
            created_links += 1

    # NetworkTopology 저장 (슬라이드/통합 데이터 + PPTX 메타)
    mtime = None
    if pptx_path and os.path.exists(pptx_path):
        mtime = timezone.datetime.fromtimestamp(
            os.path.getmtime(pptx_path), tz=timezone.get_current_timezone()
        )

    NetworkTopology.objects.update_or_create(
        school=school,
        defaults={
            'topology_data': {
                'nodes': nodes, 'edges': edges, 'slides': slides,
                'stats': data.get('stats', {}),
            },
            'pptx_path': pptx_path,
            'pptx_mtime': mtime,
            'slide_titles': [s.get('title', '') for s in slides],
        }
    )

    # NAS에 토폴로지 CSV + SNMP 가이드 DOCX 자동 생성
    try:
        from .services import write_topology_files_to_nas
        write_topology_files_to_nas(school)
    except Exception as e:
        logger.warning(f'[{school.name}] NAS 파일 자동 생성 실패: {e}')

    return created_devices, created_links


def _extract_school_name_from_filename(filename: str) -> str:
    """파일명에서 학교명 추출
    예: '2025년 테크센터-네트워크 구성도_가락고등학교.pptx' → '가락고등학교'
    예: '2025년 테크센터-유치원 네트워크 구성도_새솔유치원.pptx' → '새솔유치원'
    예: '서울새솔유치원.pptx' → '서울새솔유치원'
    """
    import re
    name = os.path.splitext(filename)[0]
    # "_학교명" 패턴이 있으면 그 뒤만 추출
    if '_' in name:
        name = name.rsplit('_', 1)[-1]
    # 공백 제거·정리
    return name.strip()


def _find_school_by_name(name: str, school_map: dict):
    """학교명에서 School 객체 찾기 (유연한 매칭)"""
    n = name.strip()
    # 정확 일치
    if n in school_map:
        return school_map[n]
    # 접두어 "서울" 추가/제거 시도
    for variant in (n, '서울' + n, n.replace('서울', '', 1)):
        if variant in school_map:
            return school_map[variant]
    # 부분 일치 (가락고 → 가락고등학교)
    for sname, school in school_map.items():
        if n in sname or sname in n:
            if len(n) >= 3:  # 너무 짧은 매칭 방지
                return school
    return None


@celery_app.task
def scan_network_pptx(school_id: int = None):
    """NAS PPTX 파일을 스캔하여 토폴로지 생성

    지원 경로:
    1. /app/nas/media/npms/산출물/{school_id}/구성도/*.pptx (학교ID 기반)
    2. /app/nas/media/npms/산출물/2025년 테크센터/2025년 테크센터-네트워크 구성도/*.pptx
       (파일명 뒤 '_학교명.pptx' 기반 자동 매칭)
    """
    from apps.schools.models import School
    from .pptx_parser import parse_pptx_topology
    import os

    base_dir = os.environ.get('NAS_ARTIFACT_ROOT', '/app/nas/media/npms/산출물')
    stats = {'total': 0, 'ok': 0, 'skip': 0, 'fail': 0, 'results': []}

    # 학교명 → School 매핑
    school_map = {s.name: s for s in School.objects.all()}

    if school_id:
        schools = School.objects.filter(id=school_id)
    else:
        schools = School.objects.all()

    # ── 방식 1: 학교ID 기반 폴더 ──
    handled_schools = set()
    for school in schools:
        pptx_dir = os.path.join(base_dir, str(school.id), '구성도')
        if os.path.isdir(pptx_dir):
            pptx_files = sorted([
                f for f in os.listdir(pptx_dir)
                if f.lower().endswith('.pptx') and not f.startswith('.') and not f.startswith('~')
            ], key=lambda f: os.path.getmtime(os.path.join(pptx_dir, f)), reverse=True)
            if pptx_files:
                pptx_path = os.path.join(pptx_dir, pptx_files[0])
                _process_one(school, pptx_path, pptx_files[0], stats)
                handled_schools.add(school.id)

    # ── 방식 2: 테크센터 구성도 폴더 스캔 (파일명 매칭) ──
    # school_id 지정된 경우에도 해당 학교 PPTX가 있는지 체크
    tech_dirs = []
    for root, dirs, files in os.walk(base_dir):
        if '네트워크 구성도' in os.path.basename(root) or '네트워크구성도' in os.path.basename(root):
            tech_dirs.append(root)

    target_school_ids = {s.id for s in schools}

    for tech_dir in tech_dirs:
        for fname in sorted(os.listdir(tech_dir)):
            if not fname.lower().endswith('.pptx') or fname.startswith('.') or fname.startswith('~'):
                continue
            school_name = _extract_school_name_from_filename(fname)
            school = _find_school_by_name(school_name, school_map)
            if not school:
                if not school_id:  # 전체 스캔일 때만 skip 기록
                    stats['total'] += 1
                    stats['skip'] += 1
                    stats['results'].append({
                        'school_id': None, 'school_name': school_name,
                        'status': 'skip', 'note': f'DB 학교 매칭 실패: {fname}',
                        'pptx': fname,
                    })
                continue
            if school.id in handled_schools:
                continue
            # 개별 스캔 시 대상 학교만 처리
            if school_id and school.id not in target_school_ids:
                continue
            pptx_path = os.path.join(tech_dir, fname)
            _process_one(school, pptx_path, fname, stats)
            handled_schools.add(school.id)

    # ── 3. 처리 안된 학교들은 skip ──
    if not school_id:
        for school in schools:
            if school.id not in handled_schools:
                stats['total'] += 1
                stats['skip'] += 1
                stats['results'].append({
                    'school_id': school.id, 'school_name': school.name,
                    'status': 'skip', 'note': 'PPTX 파일 없음',
                })

    # 결과 리스트 너무 크면 잘라냄
    if len(stats['results']) > 500:
        stats['results'] = stats['results'][:500]

    return stats


def _process_one(school, pptx_path, fname, stats):
    """한 학교의 PPTX 파싱 + 저장 + stats 업데이트"""
    from .pptx_parser import parse_pptx_topology
    stats['total'] += 1
    result = {'school_id': school.id, 'school_name': school.name, 'pptx': fname}
    try:
        data = parse_pptx_topology(pptx_path)
        devices, links = _import_pptx_topology(school, data, pptx_path)
        result.update({
            'status': 'ok',
            'devices': devices,
            'links': links,
            'slides': data.get('stats', {}).get('slides', 0),
        })
        stats['ok'] += 1
        logger.info(f'[{school.name}] PPTX 파싱 완료: 장비 {devices}, 링크 {links}')
    except Exception as e:
        result['status'] = 'fail'
        result['note'] = str(e)[:200]
        stats['fail'] += 1
        logger.error(f'[{school.name}] PPTX 파싱 실패: {e}')
    stats['results'].append(result)


@celery_app.task
def bulk_analyze_diagrams(image_dir: str = None):
    """(Deprecated) 이미지 AI 분석 → scan_network_pptx 로 대체됨"""
    return scan_network_pptx.apply()


def _old_bulk_analyze_deprecated(image_dir: str = None):
    """[DEPRECATED] 구 이미지 분석 코드"""
    import django.core.cache as cache_module
    from apps.schools.models import School
    from django.core.cache import cache

    if image_dir is None:
        image_dir = os.path.join(os.environ.get('MEDIA_ROOT', '/app/media'), 'data', '구성도이미지')

    if not os.path.isdir(image_dir):
        logger.error(f'이미지 폴더 없음: {image_dir}')
        return

    # 학교명 → School 매핑
    school_map = {s.name: s for s in School.objects.all()}

    # 이미지 파일 목록 (구성도_학교명.jpg)
    files = sorted([
        f for f in os.listdir(image_dir)
        if f.lower().endswith(('.jpg', '.jpeg', '.png'))
    ])
    total = len(files)
    logger.info(f'일괄 분석 시작: {total}개 파일')

    progress = {
        'started': True, 'finished': False,
        'total': total, 'done': 0,
        'ok': 0, 'skip': 0, 'fail': 0,
        'results': [],
    }
    cache.set(BULK_DIAGRAM_KEY, progress, timeout=7200)

    for idx, filename in enumerate(files):
        # 파일명에서 학교명 추출: 구성도_가락고등학교.jpg → 가락고등학교
        school_name = filename
        for prefix in ('구성도_', '네트워크구성도_', '망구성도_'):
            if school_name.startswith(prefix):
                school_name = school_name[len(prefix):]
        school_name = os.path.splitext(school_name)[0]

        school = school_map.get(school_name)
        result_entry = {'school': school_name, 'file': filename}

        if not school:
            result_entry.update({'status': 'skip', 'note': 'DB 미등록 학교'})
            progress['skip'] += 1
            logger.warning(f'[{school_name}] DB에 없는 학교, 스킵')
        else:
            image_path = os.path.join(image_dir, filename)
            try:
                data = _analyze_image_with_claude(image_path)
                devices, links = _import_topology_data(school, data)
                result_entry.update({'status': 'ok', 'devices': devices, 'links': links})
                progress['ok'] += 1
                logger.info(f'[{school_name}] 완료: 장비 {devices}개, 링크 {links}개')
            except Exception as e:
                result_entry.update({'status': 'fail', 'note': str(e)[:80]})
                progress['fail'] += 1
                logger.error(f'[{school_name}] 실패: {e}')

        progress['done'] = idx + 1
        progress['results'].append(result_entry)
        # 최근 200건만 유지
        if len(progress['results']) > 200:
            progress['results'] = progress['results'][-200:]
        cache.set(BULK_DIAGRAM_KEY, progress, timeout=7200)

    progress['finished'] = True
    cache.set(BULK_DIAGRAM_KEY, progress, timeout=7200)
    logger.info(f'일괄 분석 완료: 성공 {progress["ok"]}, 스킵 {progress["skip"]}, 실패 {progress["fail"]}')


# ── 선번장 NAS 사전 파싱 → DB 저장 ─────────────────────────

def _detect_cable_by_color(ws, row, col, cable_labels=None):
    """셀 배경색으로 실제 연결 케이블 감지
    theme=0(흰색) 배경은 제외, RGB 명시 색상만 인정.
    """
    if cable_labels is None:
        cable_labels = ['C6', 'C5.e', 'C5']
    for offset, label in enumerate(cable_labels):
        c = ws.cell(row, col + offset)
        if c.fill and c.fill.patternType == 'solid':
            fg = c.fill.fgColor
            if not fg:
                continue
            try:
                theme = fg.theme
                if theme is not None and isinstance(theme, int):
                    continue
            except (TypeError, AttributeError):
                pass
            try:
                rgb = fg.rgb
                if rgb and isinstance(rgb, str) and rgb.startswith('FF') and rgb != 'FF000000':
                    return label
            except (TypeError, AttributeError):
                continue
    return ''


def _parse_portmap_type_b(ws, sname):
    """양식 B (xlsm): 67열, 19행 간격, C열 시작"""
    PORT_COLS = [3, 7, 11, 15, 19, 23, 27, 31, 35, 39, 43, 47, 53, 57]
    UPLINK_COLS = [53, 57]
    switches = []
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 3).value
        if not (v and '허' in str(v) and 'I' in str(v)):
            continue
        base = r
        hub_id = str(ws.cell(base, 7).value or '').strip()
        model = str(ws.cell(base, 19).value or '').strip()
        location = str(ws.cell(base, 31).value or '').strip()
        net_type = str(ws.cell(base + 2, 7).value or '').strip()
        mfr = str(ws.cell(base + 2, 31).value or '').strip()
        if not hub_id:
            continue
        ports = []
        for col in PORT_COLS:
            labels = ['UTP', 'SM', 'MM'] if col in UPLINK_COLS else ['C6', 'C5.e', 'C5']
            pnum = ws.cell(base + 6, col).value
            conn = str(ws.cell(base + 8, col).value or '').strip()
            cable = _detect_cable_by_color(ws, base + 7, col, labels)
            if pnum and str(pnum).strip().isdigit() and int(pnum) > 0:
                ports.append({'port': int(pnum), 'connected_to': conn if conn != '/' else '', 'cable': cable, 'vlan': '', 'note': ''})
            pnum2 = ws.cell(base + 11, col).value
            conn2 = str(ws.cell(base + 13, col).value or '').strip()
            cable2 = _detect_cable_by_color(ws, base + 12, col, labels)
            if pnum2 and str(pnum2).strip().isdigit() and int(pnum2) > 0:
                ports.append({'port': int(pnum2), 'connected_to': conn2 if conn2 != '/' else '', 'cable': cable2, 'vlan': '', 'note': ''})
        ports.sort(key=lambda p: p['port'])
        switches.append({
            'device_id': hub_id, 'model_name': model, 'network_type': net_type,
            'manufacturer': mfr, 'install_location': location,
            'category': '스위치',
            'port_count': len(ports), 'ports': ports, 'sheet': sname,
        })
    return switches


def _parse_portmap_type_a(ws, sname):
    """양식 A (xlsx): 66열, 10행 간격, A열 또는 C열 시작"""
    switches = []
    for r in range(1, ws.max_row + 1):
        # col 1 또는 col 3에서 '망구분' 찾기
        start_col = 0
        for sc in [1, 3]:
            v = ws.cell(r, sc).value
            if v and str(v).strip() == '망구분':
                start_col = sc
                break
        if not start_col:
            continue
        offset = start_col - 1  # 0 또는 2
        PORT_COLS = [1 + offset + i * 3 for i in range(20)]
        UPLINK_COLS = PORT_COLS[-2:]
        base = r
        net_type = str(ws.cell(base, 4 + offset).value or '').strip()
        model = str(ws.cell(base, 13 + offset).value or '').strip()
        location = str(ws.cell(base, 22 + offset).value or '').strip()
        poe = str(ws.cell(base, 31 + offset).value or '').strip()
        hub_id = str(ws.cell(base + 1, 4 + offset).value or '').strip()
        mfr = str(ws.cell(base + 1, 13 + offset).value or '').strip()
        if not hub_id and not model:
            continue
        ports = []
        for col in PORT_COLS:
            labels = ['UTP', 'SM', 'MM'] if col in UPLINK_COLS else ['C6', 'C5.e', 'C5']
            pnum = ws.cell(base + 2, col).value
            cable = _detect_cable_by_color(ws, base + 3, col, labels)
            conn = str(ws.cell(base + 4, col).value or '').strip()
            if pnum and str(pnum).strip().isdigit():
                ports.append({'port': int(pnum), 'connected_to': conn if conn and conn != '/' else '', 'cable': cable, 'vlan': '', 'note': ''})
            pnum2 = ws.cell(base + 5, col).value
            cable2 = _detect_cable_by_color(ws, base + 6, col, labels)
            conn2 = str(ws.cell(base + 7, col).value or '').strip()
            if pnum2 and str(pnum2).strip().isdigit():
                ports.append({'port': int(pnum2), 'connected_to': conn2 if conn2 and conn2 != '/' else '', 'cable': cable2, 'vlan': '', 'note': ''})
        ports.sort(key=lambda p: p['port'])
        switches.append({
            'device_id': hub_id or model, 'model_name': model, 'network_type': net_type,
            'manufacturer': mfr, 'install_location': location,
            'category': 'PoE' if poe and poe not in ('N', '없음', '') else '스위치',
            'port_count': len(ports), 'ports': ports, 'sheet': sname,
        })
    return switches


def _parse_nas_portmap_file(filepath):
    """NAS 선번장 파일 1개 파싱 → 스위치 목록 반환"""
    import openpyxl
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        logger.warning(f'선번장 파일 로드 실패: {filepath} — {e}')
        return None

    ext = filepath.rsplit('.', 1)[-1].lower()
    result = []
    try:
        for sname in wb.sheetnames:
            if sname in ('장비스펙', 'FDF', 'FDF현황'):
                continue
            try:
                ws = wb[sname]
                if not ws.max_row:
                    continue
                is_type_b = ext == 'xlsm'
                if not is_type_b:
                    for r in range(1, min(10, ws.max_row + 1)):
                        v = ws.cell(r, 3).value
                        if v and '허' in str(v):
                            is_type_b = True
                            break
                if is_type_b:
                    result.extend(_parse_portmap_type_b(ws, sname))
                else:
                    result.extend(_parse_portmap_type_a(ws, sname))
            except Exception as e:
                logger.warning(f'선번장 시트 파싱 오류: {os.path.basename(filepath)}/{sname} — {e}')
                continue
    finally:
        wb.close()
    return result if result else None


@celery_app.task(soft_time_limit=3600, time_limit=3900)
def sync_nas_portmap():
    """NAS 선번장 파일 사전 파싱 → SchoolEquipment.port_map에 저장

    - 100개마다 gc.collect()로 메모리 해제
    - 매일 새벽 자동 실행 (CELERY_BEAT_SCHEDULE)
    """
    from apps.schools.models import School, SchoolEquipment

    NAS_BASE = os.environ.get('NAS_ARTIFACT_ROOT', '/app/nas/media/npms/산출물')
    portmap_folder = os.path.join(NAS_BASE, '2025년 테크센터', '2025년 테크센터-네트워크 선번장')
    if not os.path.isdir(portmap_folder):
        logger.warning(f'선번장 폴더 없음: {portmap_folder}')
        return {'error': '폴더 없음'}

    school_map = {s.name: s for s in School.objects.all()}
    files = [f for f in os.listdir(portmap_folder)
             if (f.endswith('.xlsx') or f.endswith('.xlsm')) and not f.startswith('~')]

    stats = {'total': len(files), 'parsed': 0, 'skipped': 0, 'failed': 0}
    logger.info(f'선번장 파싱 시작: {len(files)}개 파일')

    for i, fname in enumerate(files):
        filepath = os.path.join(portmap_folder, fname)
        if not os.path.isfile(filepath):
            stats['skipped'] += 1
            continue

        school_name = fname.rsplit('.', 1)[0]
        if '_' in school_name:
            school_name = school_name.rsplit('_', 1)[-1]
        school_name = school_name.strip()

        school = school_map.get(school_name)
        if not school:
            for variant in (school_name, '서울' + school_name, school_name.replace('서울', '', 1)):
                if variant in school_map:
                    school = school_map[variant]; break
        if not school:
            stats['skipped'] += 1
            continue

        # mtime 비교
        file_mtime = os.path.getmtime(filepath)
        existing = SchoolEquipment.objects.filter(
            school=school, category__in=['스위치', 'PoE', 'PoE스위치'],
            port_map__isnull=False,
        ).first()
        if existing and existing.port_map:
            if isinstance(existing.port_map, list) and len(existing.port_map) > 0:
                meta = existing.port_map[0] if isinstance(existing.port_map[0], dict) else {}
                saved_mtime = meta.get('_file_mtime', 0)
                if abs(saved_mtime - file_mtime) < 1:
                    stats['skipped'] += 1
                    continue

        # 파싱 + DB 저장 (전체를 try/except로 감싸서 한 파일 오류가 전체를 중단하지 않도록)
        try:
            switches = _parse_nas_portmap_file(filepath)
            gc.collect()

            if not switches:
                stats['skipped'] += 1
                continue

            # DB 저장
            for sw in switches:
                device_id = sw['device_id']
                eq, created = SchoolEquipment.objects.get_or_create(
                    school=school,
                    device_id=device_id,
                    defaults={
                        'category': sw['category'],
                        'model_name': sw['model_name'],
                        'network_type': sw['network_type'],
                        'manufacturer': sw['manufacturer'],
                        'install_location': sw['install_location'],
                    }
                )
                ports = sw['ports']
                if ports:
                    ports[0]['_file_mtime'] = file_mtime
                eq.port_map = ports
                if not eq.model_name and sw['model_name']:
                    eq.model_name = sw['model_name']
                if not eq.network_type and sw['network_type']:
                    eq.network_type = sw['network_type']
                if not eq.manufacturer and sw['manufacturer']:
                    eq.manufacturer = sw['manufacturer']
                if not eq.install_location and sw['install_location']:
                    eq.install_location = sw['install_location']
                eq.save()

            stats['parsed'] += 1
        except Exception as e:
            logger.error(f'[{school_name}] 선번장 처리 오류: {e}')
            stats['failed'] += 1
            gc.collect()
            continue

        # 100개마다 진행률 로그 + 강제 메모리 해제
        done = stats['parsed'] + stats['skipped'] + stats['failed']
        if done % 100 == 0:
            gc.collect()
            logger.info(f'선번장 파싱 진행: {done}/{stats["total"]} (파싱 {stats["parsed"]}, 스킵 {stats["skipped"]})')

    logger.info(f'선번장 사전 파싱 완료: 총 {stats["total"]}, 파싱 {stats["parsed"]}, 스킵 {stats["skipped"]}, 실패 {stats["failed"]}')
    return stats


# ── 랙실장도 NAS 사전 파싱 → DB 저장 ────────────────────────

def _parse_rack_file(filepath):
    """NAS 랙실장도 파일 1개 파싱 → 랙 목록 반환"""
    import openpyxl, re

    def detect_type(name):
        nl = (name or '').upper()
        if '패치' in name: return 'patch'
        if 'FDF' in nl or 'OFD' in nl: return 'fdf'
        if '방화벽' in name or 'FW/' in nl or 'FW#' in nl or 'NGF' in nl or '200E' in nl: return 'firewall'
        if 'UPS' in nl: return 'ups'
        if '서버' in name or 'SERVER' in nl: return 'server'
        if 'POE' in nl: return 'poe'
        return 'switch'

    def split_id(name):
        if not name: return '', name or ''
        m = re.match(r'^([KHMPGFB][#B]\S*)\s+(.+)$', name.strip())
        if m: return m.group(1), m.group(2).strip('() ')
        m2 = re.match(r'^(IP#\S*)\s+(.+)$', name.strip())
        if m2: return m2.group(1), m2.group(2).strip('() ')
        return '', name

    def is_location(text):
        if not text: return False
        return any(k in text for k in ['서버', '전산', '층', 'EPS', '교무', '복도', '실장도', '통신랙', '통신렉'])

    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        logger.warning(f'랙 파일 로드 실패: {filepath} — {e}')
        return None

    # 시트 선택 우선순위: 랙/렉 키워드 → 학교명(마지막 시트) → 포트맵/층/계위도 제외
    ws = None
    for sname in wb.sheetnames:
        if '랙' in sname or '렉' in sname:
            ws = wb[sname]; break
    if not ws:
        # 마지막 시트가 학교명인 경우가 많음
        last = wb.sheetnames[-1]
        if '포트맵' not in last and '층' not in last and '계위' not in last:
            ws = wb[last]
    if not ws:
        for sname in wb.sheetnames:
            if '포트맵' not in sname and '층' not in sname and '계위' not in sname:
                ws = wb[sname]; break
    if not ws:
        ws = wb[wb.sheetnames[0]]

    # 통신랙 이름 열 수집
    rack_name_cols = {}
    for r in range(1, 8):
        for c in range(1, min((ws.max_column or 30) + 1, 40)):
            try:
                v = ws.cell(r, c).value
            except Exception:
                continue
            if v and '통신' in str(v) and ('랙' in str(v) or '렉' in str(v)):
                vs = str(v).strip()
                if '실장도' in vs:
                    continue
                rack_name_cols[c] = vs

    rack_positions = []
    if rack_name_cols:
        for eq_col, rname in sorted(rack_name_cols.items()):
            u_col = eq_col - 1
            has_u = False
            # U#1 찾기
            for r in range(3, 15):
                try:
                    v = ws.cell(r, u_col).value
                except Exception:
                    continue
                if v and str(v).strip() == '1':
                    has_u = True; break
            if not has_u:
                for alt_c in [c2 for c2 in [eq_col - 2, eq_col + 1] if c2 >= 1]:
                    for r in range(3, 15):
                        try:
                            v = ws.cell(r, alt_c).value
                        except Exception:
                            continue
                        if v and str(v).strip() == '1':
                            u_col = alt_c; has_u = True; break
                    if has_u: break
            # U#1이 없으면 숫자가 있는 U열 확인 (짝수만 있는 랙 대응)
            if not has_u:
                for r in range(3, min((ws.max_row or 40) + 1, 50)):
                    try:
                        v = ws.cell(r, u_col).value
                    except Exception:
                        continue
                    if v and str(v).strip().isdigit():
                        has_u = True; break
                if not has_u:
                    for alt_c in [c2 for c2 in [eq_col - 2, eq_col + 1] if c2 >= 1]:
                        for r in range(3, min((ws.max_row or 40) + 1, 50)):
                            try:
                                v = ws.cell(r, alt_c).value
                            except Exception:
                                continue
                            if v and str(v).strip().isdigit():
                                u_col = alt_c; has_u = True; break
                        if has_u: break
            if has_u:
                location = ''
                for r in [1, 2]:
                    for cc in [eq_col, eq_col - 1, eq_col + 1, u_col]:
                        if cc < 1: continue
                        try:
                            v = ws.cell(r, cc).value
                        except Exception:
                            continue
                        if v:
                            vs = str(v).strip()
                            if vs and not ('통신' in vs and ('랙' in vs or '렉' in vs)) and '실장도' not in vs and len(vs) > 2:
                                if any(k in vs for k in ['서버', '전산', '층', 'EPS']):
                                    location = vs; break
                    if location: break
                rack_positions.append((u_col, eq_col, rname, location))
    else:
        for r in range(3, 15):
            for c in [1, 2]:
                try:
                    v = ws.cell(r, c).value
                except Exception:
                    continue
                if v and str(v).strip() == '1':
                    eq_c = c + 1
                    loc = ''
                    for rr in range(1, r):
                        for cc in range(1, 15):
                            try:
                                vv = ws.cell(rr, cc).value
                            except Exception:
                                continue
                            if vv:
                                vvs = str(vv).strip()
                                if any(k in vvs for k in ['서버', '전산', '층', 'EPS']) and '실장도' not in vvs:
                                    loc = vvs
                    rname = '통신랙'
                    for rr in range(1, r):
                        try:
                            vv = ws.cell(rr, eq_c).value
                        except Exception:
                            continue
                        if vv and ('렉' in str(vv) or '랙' in str(vv)) and '실장도' not in str(vv):
                            rname = str(vv).strip()
                    rack_positions.append((c, eq_c, rname, loc))
                    for alt_c in range(c + 6, min((ws.max_column or 30) + 1, 40), 6):
                        try:
                            alt_v = ws.cell(r, alt_c).value
                        except Exception:
                            continue
                        if alt_v and str(alt_v).strip() == '1':
                            rack_positions.append((alt_c, alt_c + 1, '통신랙', ''))
                    break
            if rack_positions: break

    if not rack_positions:
        # U번호 없는 단순 목록 형식 처리
        # 통신랙 이름이 있는 열에서 장비를 직접 수집
        if rack_name_cols:
            racks = []
            for eq_col, rname in sorted(rack_name_cols.items()):
                # 랙 이름 행 찾기
                rack_row = None
                for r in range(1, 8):
                    try:
                        v = ws.cell(r, eq_col).value
                    except Exception:
                        continue
                    if v and str(v).strip() == rname:
                        rack_row = r; break
                if not rack_row:
                    rack_row = 3

                # 위치 찾기 (랙 이름 위 행)
                location = ''
                for r in range(1, rack_row):
                    for cc in range(max(1, eq_col - 1), eq_col + 2):
                        try:
                            v = ws.cell(r, cc).value
                        except Exception:
                            continue
                        if v:
                            vs = str(v).strip()
                            if vs and '실장도' not in vs and not ('통신' in vs and ('랙' in vs or '렉' in vs)):
                                if any(k in vs for k in ['서버', '전산', '층', 'EPS', '교무', '컴퓨터']):
                                    location = vs

                # 장비 수집 (랙 이름 아래 행)
                items = []
                for r in range(rack_row + 1, (ws.max_row or 40) + 1):
                    try:
                        v = ws.cell(r, eq_col).value
                    except Exception:
                        continue
                    if v:
                        name = str(v).strip()
                        if name and len(name) > 1 and not is_location(name) and name != rname:
                            dev_id, dev_name = split_id(name)
                            items.append({'u': 0, 'name': name, 'device_id': dev_id, 'model': dev_name, 'type': detect_type(name)})

                if items:
                    racks.append({
                        'rack_name': rname.replace('렉', '랙'),
                        'location': location,
                        'items': items,
                        'source': 'nas',
                    })
            wb.close()
            return racks if racks else None

        wb.close()
        return None

    racks = []
    for u_col, eq_col, rack_name, location in rack_positions:
        # U#1 행 또는 첫 번째 U번호 행 찾기
        u1_row = None
        first_u_row = None
        for r in range(3, min((ws.max_row or 40) + 1, 50)):
            try:
                v = ws.cell(r, u_col).value
            except Exception:
                continue
            if v and str(v).strip().isdigit():
                if first_u_row is None:
                    first_u_row = r
                if str(v).strip() == '1':
                    u1_row = r; break
        if not u1_row:
            u1_row = first_u_row
        if not u1_row: continue

        items = []
        for r in range(max(3, u1_row - 5), u1_row):
            try:
                v = ws.cell(r, eq_col).value
                u_v = ws.cell(r, u_col).value
            except Exception:
                continue
            if v and not u_v:
                vs = str(v).strip()
                if not vs or len(vs) <= 1 or is_location(vs) or vs == rack_name or vs == location:
                    continue
                dev_id, dev_name = split_id(vs)
                items.append({'u': 0, 'name': vs, 'device_id': dev_id, 'model': dev_name, 'type': detect_type(vs)})

        current_u = 0
        for r in range(u1_row, (ws.max_row or 40) + 1):
            try:
                u_v = ws.cell(r, u_col).value
                eq_v = ws.cell(r, eq_col).value
            except Exception:
                continue
            if u_v and str(u_v).strip().isdigit():
                current_u = int(u_v)
            if eq_v:
                name = str(eq_v).strip()
                if name and name != rack_name and name != location and not is_location(name):
                    dev_id, dev_name = split_id(name)
                    items.append({'u': current_u, 'name': name, 'device_id': dev_id, 'model': dev_name, 'type': detect_type(name)})

        if items:
            racks.append({
                'rack_name': rack_name.replace('렉', '랙'),
                'location': location,
                'items': items,
                'source': 'nas',
            })

    wb.close()
    return racks if racks else None


@celery_app.task
def sync_nas_rack():
    """NAS 랙실장도 파일 사전 파싱 → School.rack_data에 저장

    - mtime 비교로 변경된 파일만 재파싱
    - 매일 새벽 자동 실행 (CELERY_BEAT_SCHEDULE)
    """
    from apps.schools.models import School

    NAS_BASE = os.environ.get('NAS_ARTIFACT_ROOT', '/app/nas/media/npms/산출물')
    rack_folder = os.path.join(NAS_BASE, '2025년 테크센터', '2025년 테크센터-네트워크 통신랙실장도')
    if not os.path.isdir(rack_folder):
        logger.warning(f'랙실장도 폴더 없음: {rack_folder}')
        return {'error': '폴더 없음'}

    school_map = {s.name: s for s in School.objects.all()}
    stats = {'total': 0, 'parsed': 0, 'skipped': 0, 'failed': 0}

    files = [f for f in os.listdir(rack_folder)
             if (f.endswith('.xlsx') or f.endswith('.xlsm')) and not f.startswith('~')]

    for fname in files:
        stats['total'] += 1
        filepath = os.path.join(rack_folder, fname)

        school_name = fname.rsplit('.', 1)[0]
        if '_' in school_name:
            school_name = school_name.rsplit('_', 1)[-1]
        school_name = school_name.strip()

        school = school_map.get(school_name)
        if not school:
            for variant in (school_name, '서울' + school_name, school_name.replace('서울', '', 1)):
                if variant in school_map:
                    school = school_map[variant]; break
        if not school:
            stats['skipped'] += 1
            continue

        # mtime 비교
        file_mtime = os.path.getmtime(filepath)
        if school.rack_data and isinstance(school.rack_data, list) and len(school.rack_data) > 0:
            saved_mtime = school.rack_data[0].get('_file_mtime', 0) if isinstance(school.rack_data[0], dict) else 0
            if abs(saved_mtime - file_mtime) < 1:
                stats['skipped'] += 1
                continue

        try:
            racks = _parse_rack_file(filepath)
        except Exception as e:
            logger.error(f'[{school_name}] 랙실장도 파싱 오류: {e}')
            stats['failed'] += 1
            gc.collect()
            continue

        if not racks:
            stats['skipped'] += 1
            continue

        # mtime 메타 추가
        if racks[0]:
            racks[0]['_file_mtime'] = file_mtime

        school.rack_data = racks
        school.save(update_fields=['rack_data'])
        stats['parsed'] += 1

        if (stats['parsed'] + stats['skipped'] + stats['failed']) % 100 == 0:
            gc.collect()
            logger.info(f'랙실장도 파싱 진행: {stats["parsed"] + stats["skipped"] + stats["failed"]}/{stats["total"]}')

    logger.info(f'랙실장도 사전 파싱 완료: 총 {stats["total"]}, 파싱 {stats["parsed"]}, 스킵 {stats["skipped"]}, 실패 {stats["failed"]}')
    return stats


# ══════════════════════════════════════════════════════════
# 구성도 PPTX 네트워크 장비 추출기
# ══════════════════════════════════════════════════════════
import re as _re
import tempfile
import zipfile
from collections import defaultdict as _defaultdict
from xml.etree import ElementTree as _ET

_NS = {
    'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
    'p': 'http://schemas.openxmlformats.org/presentationml/2006/main',
}
_EMU = 914400
_DEVICE_ID_RE = _re.compile(r'^(Secui|TrusGuard|[KHMGPi]#)')
_PORT_RE = _re.compile(r'^\((\d+)\s+(\d+)\)$')
_CABLE_COLOR_MAP = {
    'FF0000': '광', '00AFEF': 'Cat6', '00B0F0': 'Cat6',
    '00B050': 'Cat5', 'FF66CC': 'Cat5e',
}


def _pptx_get_xfrm(spPr):
    xfrm = spPr.find('a:xfrm', _NS)
    if xfrm is None:
        return None
    off = xfrm.find('a:off', _NS)
    ext = xfrm.find('a:ext', _NS)
    chOff = xfrm.find('a:chOff', _NS)
    chExt = xfrm.find('a:chExt', _NS)
    return {
        'x': int(off.get('x')) if off is not None else 0,
        'y': int(off.get('y')) if off is not None else 0,
        'w': int(ext.get('cx')) if ext is not None else 0,
        'h': int(ext.get('cy')) if ext is not None else 0,
        'chX': int(chOff.get('x')) if chOff is not None else 0,
        'chY': int(chOff.get('y')) if chOff is not None else 0,
        'chW': int(chExt.get('cx')) if chExt is not None else 0,
        'chH': int(chExt.get('cy')) if chExt is not None else 0,
        'flipH': xfrm.get('flipH', '0'),
        'flipV': xfrm.get('flipV', '0'),
    }


def _pptx_walk_shapes(slide_root):
    text_shapes, lines = [], []

    def _walk(elem, ax, ay, sx, sy):
        for child in elem:
            tag = child.tag.split('}')[-1]
            if tag == 'grpSp':
                grpSpPr = child.find('p:grpSpPr', _NS)
                if grpSpPr is None:
                    _walk(child, ax, ay, sx, sy)
                    continue
                xf = _pptx_get_xfrm(grpSpPr)
                if xf is None or not xf['chW'] or not xf['chH']:
                    _walk(child, ax, ay, sx, sy)
                    continue
                new_sx = sx * (xf['w'] / xf['chW'])
                new_sy = sy * (xf['h'] / xf['chH'])
                new_ax = ax + xf['x'] * sx - xf['chX'] * new_sx
                new_ay = ay + xf['y'] * sy - xf['chY'] * new_sy
                _walk(child, new_ax, new_ay, new_sx, new_sy)
            elif tag in ('sp', 'cxnSp'):
                spPr = child.find('p:spPr', _NS)
                if spPr is None:
                    continue
                xf = _pptx_get_xfrm(spPr)
                if xf is None:
                    continue
                abs_x = (ax + xf['x'] * sx) / _EMU
                abs_y = (ay + xf['y'] * sy) / _EMU
                abs_w = xf['w'] * sx / _EMU
                abs_h = xf['h'] * sy / _EMU
                ln = spPr.find('a:ln', _NS)
                color, dash = None, 'solid'
                if ln is not None:
                    srgb = ln.find('.//a:srgbClr', _NS)
                    if srgb is not None:
                        color = srgb.get('val')
                    d = ln.find('a:prstDash', _NS)
                    if d is not None:
                        dash = d.get('val')
                prst = spPr.find('a:prstGeom', _NS)
                cust = spPr.find('a:custGeom', _NS)
                geom = (prst.get('prst') if prst is not None else 'custom' if cust is not None else 'unknown')
                paras = []
                txBody = child.find('p:txBody', _NS)
                if txBody is not None:
                    for p in txBody.findall('a:p', _NS):
                        t = ''.join(r.text or '' for r in p.findall('.//a:t', _NS)).strip()
                        if t:
                            paras.append(t)
                is_line = (tag == 'cxnSp' or geom == 'line'
                           or (geom == 'custom' and color is not None and ln is not None))
                entry = {
                    'x': abs_x, 'y': abs_y, 'w': abs_w, 'h': abs_h,
                    'cx': abs_x + abs_w / 2, 'cy': abs_y + abs_h / 2,
                    'flipH': xf['flipH'], 'flipV': xf['flipV'],
                    'color': color, 'dash': dash, 'geom': geom, 'paras': paras,
                }
                if is_line and cust is not None:
                    pathEl = cust.find('.//a:path', _NS)
                    if pathEl is not None:
                        pw = int(pathEl.get('w', 0))
                        ph = int(pathEl.get('h', 0))
                        pts = []
                        for c2 in pathEl:
                            ttag = c2.tag.split('}')[-1]
                            for pt in c2.findall('a:pt', _NS):
                                pts.append((ttag, int(pt.get('x')), int(pt.get('y'))))
                        entry['pts_local'] = pts
                        entry['pw'] = pw
                        entry['ph'] = ph
                if is_line:
                    lines.append(entry)
                elif paras:
                    text_shapes.append(entry)

    spTree = slide_root.find('.//p:cSld/p:spTree', _NS)
    if spTree is not None:
        _walk(spTree, 0, 0, 1.0, 1.0)
    return text_shapes, lines


def _pptx_extract_school_name(pptx_path):
    from pathlib import Path
    p = Path(pptx_path)
    for cand in [p.stem, p.parent.name]:
        m = _re.search(r'([가-힣]{2,}?(?:초등학교|중학교|고등학교|특수학교|학교|유치원))', cand)
        if m:
            return m.group(1)
    for cand in [p.stem, p.parent.name]:
        parts = _re.split(r'[_\-\s\.]+', cand)
        for token in reversed(parts):
            if _re.search(r'[가-힣]{2,}', token):
                return token
    return p.stem


def _pptx_net_from_id(sid):
    if sid.startswith('K#'): return '교사망'
    if sid.startswith('H#'): return '학생망'
    if sid.startswith('M#') or sid.startswith('P#'): return '무선망'
    if sid.startswith('G#'): return '기타망'
    return '보안(공통)'


def _pptx_tier_from_y(cy):
    if cy < 1.6: return '상단(보안)'
    if cy < 2.9: return '1계위'
    if cy < 5.9: return '2계위'
    return '3계위'


def _pptx_extract_short_id(text):
    m = _re.match(r'(Secui\s*\d+ED\s*#\d+|TrusGuard\s*\w+|[KHMGPi]#\s*[A-Z0-9]+)',
                  _re.sub(r'\s+', ' ', text).strip())
    return _re.sub(r'\s+', '', m.group(1)) if m else text[:10]


def _pptx_parse_meta(d):
    paras = d.get('paras_clean', [])
    p0 = paras[0] if paras else ''
    p1 = paras[1] if len(paras) > 1 else ''
    short = d['short']
    if short.startswith('Secui'):
        return 'Secui 800ED', '본관', '1F', '방화벽'
    if short.startswith('TrusGuard'):
        return 'TrusGuard 200E', '본관', '1F', '10G 방화벽'
    model = p0.replace(short, '', 1).strip()
    building = floor = location = ''
    if p1:
        bm = _re.search(r'(본관|정보관|정보동|체육관|별관|학생관|학생동)', p1)
        if bm:
            building = bm.group(1)
        fm = _re.search(r'(\d+)\s*[F층]', p1)
        if fm:
            floor = f"{fm.group(1)}F"
        rest = p1
        if bm: rest = rest.replace(bm.group(0), '', 1)
        if fm: rest = rest.replace(fm.group(0), '', 1)
        location = _re.sub(r'\s+', ' ', rest).strip()
    return model, building, floor, location


def extract_pptx_network(pptx_path):
    """단일 PPTX 파일에서 네트워크 장비 정보 추출 → 행 리스트 반환"""
    import math
    school = _pptx_extract_school_name(pptx_path)
    rows = []
    with tempfile.TemporaryDirectory() as tmp:
        with zipfile.ZipFile(pptx_path) as z:
            z.extractall(tmp)
        slides_dir = os.path.join(tmp, 'ppt', 'slides')
        if not os.path.isdir(slides_dir):
            return []
        slide_files = sorted(
            [f for f in os.listdir(slides_dir) if f.startswith('slide') and f.endswith('.xml')],
            key=lambda f: int(_re.search(r'\d+', f).group())
        )
        for slide_file in slide_files:
            slide_num = int(_re.search(r'\d+', slide_file).group())
            try:
                tree = _ET.parse(os.path.join(slides_dir, slide_file))
            except _ET.ParseError:
                continue
            text_shapes, line_shapes = _pptx_walk_shapes(tree.getroot())
            # 장비/포트 분리
            devices, port_labels = [], []
            for s in text_shapes:
                paras = s['paras']
                if not paras:
                    continue
                if len(paras) == 1 and _PORT_RE.match(paras[0]):
                    m = _PORT_RE.match(paras[0])
                    port_labels.append({'p1': int(m.group(1)), 'p2': int(m.group(2)), 'cx': s['cx'], 'cy': s['cy']})
                elif _DEVICE_ID_RE.match(paras[0]):
                    paras_clean = paras[:]
                    if len(paras) >= 3 and _PORT_RE.match(paras[-1]):
                        m = _PORT_RE.match(paras[-1])
                        port_labels.append({'p1': int(m.group(1)), 'p2': int(m.group(2)), 'cx': s['cx'], 'cy': s['y'] + s['h'] - 0.05})
                        paras_clean = paras[:-1]
                    devices.append({
                        'x': s['x'], 'y': s['y'], 'w': s['w'], 'h': s['h'],
                        'cx': s['cx'], 'cy': s['cy'], 'paras_clean': paras_clean,
                        'short': _pptx_extract_short_id(paras_clean[0]),
                    })
            if not devices:
                continue
            # 포트 매칭
            used = set()
            for d in sorted(devices, key=lambda x: (x['cy'], x['cx'])):
                best, best_score = None, 999
                for i, pl in enumerate(port_labels):
                    if i in used or pl['cy'] >= d['y']:
                        continue
                    dx = abs(pl['cx'] - d['cx'])
                    dy = d['y'] - pl['cy']
                    if dx > d['w'] / 2 + 0.3 or dy > 0.8:
                        continue
                    score = dy + dx * 0.4
                    if score < best_score:
                        best_score, best = score, i
                if best is not None:
                    pl = port_labels[best]
                    d['port'] = (pl['p1'], pl['p2'])
                    used.add(best)
                else:
                    d['port'] = None
            # 케이블 엣지
            edge_cable = {}
            for l in line_shapes:
                if not l.get('color'):
                    continue
                cable = _CABLE_COLOR_MAP.get(l['color'])
                if not cable:
                    continue
                if l.get('pts_local') and l.get('pw') and l.get('ph'):
                    first, last = l['pts_local'][0], l['pts_local'][-1]
                    x1 = l['x'] + (first[1] / l['pw']) * l['w']
                    y1 = l['y'] + (first[2] / l['ph']) * l['h']
                    x2 = l['x'] + (last[1] / l['pw']) * l['w']
                    y2 = l['y'] + (last[2] / l['ph']) * l['h']
                else:
                    x1, y1, x2, y2 = l['x'], l['y'], l['x'] + l['w'], l['y'] + l['h']
                    if l['flipH'] == '1': x1, x2 = x2, x1
                    if l['flipV'] == '1': y1, y2 = y2, y1

                def nearest(x, y):
                    bd, br = 1.2, None
                    for dd in devices:
                        dx2 = max(0, abs(x - dd['cx']) - dd['w'] / 2)
                        dy2 = max(0, abs(y - dd['cy']) - dd['h'] / 2)
                        dist = math.hypot(dx2, dy2)
                        if dist < bd:
                            bd, br = dist, dd
                    return br

                a = nearest(x1, y1)
                b = nearest(x2, y2)
                if a and b and a['short'] != b['short']:
                    edge_cable[(a['short'], b['short'])] = cable
                    edge_cable[(b['short'], a['short'])] = cable
            # 행 생성
            for d in devices:
                short = d['short']
                network = _pptx_net_from_id(short)
                tier = _pptx_tier_from_y(d['cy'])
                model, building, floor, location = _pptx_parse_meta(d)
                p = d.get('port')
                port_str = f"({p[0]}→{p[1]})" if p else ''
                # 상위 연결 찾기
                upper_short = ''
                cable = ''
                for (a, b), c in edge_cable.items():
                    if a == short:
                        other = next((dd for dd in devices if dd['short'] == b and dd['cy'] < d['cy'] - 0.3), None)
                        if other:
                            upper_short = b
                            cable = c
                            break
                cable_pair = f"{upper_short} → {short}" if upper_short else ''
                rows.append({
                    '학교명': school, '슬라이드': slide_num, '망': network,
                    '망ID': short, '모델명': model, '건물명': building,
                    '층': floor, '위치': location, '계위': tier,
                    '포트(다운→업)': port_str, '케이블 업→다운': cable_pair,
                    '업링크 케이블': cable, '케이블 출처': '추출' if cable else '',
                })
    return rows

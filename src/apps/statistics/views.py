from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from rest_framework import viewsets, status
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from django.utils import timezone
from django.db.models import Avg, Count, Q, Sum
from datetime import timedelta, date

@login_required
def statistics_view(request):
    return render(request, 'statistics/index.html')


from .models import (
    StatisticsDaily, StatisticsMonthly, SLARecord,
    SatisfactionSurvey, PerformanceMetric
)
from .serializers import (
    StatisticsDailySerializer, StatisticsMonthlySerializer,
    SLARecordSerializer, SatisfactionSurveySerializer, SurveyResponseSerializer
)
from core.pagination import StandardPagination


class StatisticsDailyViewSet(viewsets.ReadOnlyModelViewSet):
    """일별 통계"""
    serializer_class = StatisticsDailySerializer
    permission_classes = [IsAuthenticated]
    pagination_class = StandardPagination

    def get_queryset(self):
        qs = StatisticsDaily.objects.all()
        start = self.request.query_params.get('start')
        end   = self.request.query_params.get('end')
        if start:
            qs = qs.filter(stat_date__gte=start)
        if end:
            qs = qs.filter(stat_date__lte=end)
        return qs

    @action(detail=False, methods=['get'])
    def trend(self, request):
        """최근 30일 트렌드 (5분 캐시)"""
        from django.core.cache import cache
        cache_key = 'stats_trend_30d'
        cached = cache.get(cache_key)
        if cached is not None:
            return Response(cached)

        since = timezone.localdate() - timedelta(days=30)
        qs = StatisticsDaily.objects.filter(stat_date__gte=since).order_by('stat_date')
        data = {
            'dates':      [str(s.stat_date) for s in qs],
            'incidents':  [s.total_incidents for s in qs],
            'completed':  [s.completed_incidents for s in qs],
            'sla_arrival':[round(s.sla_arrival_ok / s.total_incidents * 100, 1) if s.total_incidents else 0 for s in qs],
        }
        cache.set(cache_key, data, 300)
        return Response(data)

    @action(detail=False, methods=['get'])
    def today(self, request):
        """오늘 통계"""
        today = timezone.localdate()
        stat = StatisticsDaily.objects.filter(stat_date=today).first()
        if not stat:
            return Response({'stat_date': str(today), 'total_incidents': 0, 'completed_incidents': 0})
        return Response(StatisticsDailySerializer(stat).data)

    @action(detail=False, methods=['post'], url_path='aggregate')
    def aggregate(self, request):
        """일별 통계 집계 트리거 (admin 전용)"""
        from apps.accounts.models import User
        if not (request.user.role in ('admin', 'superadmin')):
            return Response({'error': 'Forbidden'}, status=status.HTTP_403_FORBIDDEN)
        from .services import aggregate_daily
        from datetime import date as date_type
        date_str = request.data.get('date')
        if date_str:
            try:
                target = date_type.fromisoformat(date_str)
            except ValueError:
                return Response({'error': '날짜 형식 오류 (YYYY-MM-DD)'}, status=400)
        else:
            target = timezone.localdate()
        aggregate_daily(target)
        return Response({'status': 'ok', 'date': str(target)})


class StatisticsMonthlyViewSet(viewsets.ReadOnlyModelViewSet):
    """월별 통계"""
    serializer_class = StatisticsMonthlySerializer
    permission_classes = [IsAuthenticated]
    pagination_class = StandardPagination

    def get_queryset(self):
        qs = StatisticsMonthly.objects.select_related('support_center')
        year = self.request.query_params.get('year')
        if year:
            qs = qs.filter(year=year)
        month = self.request.query_params.get('month')
        if month:
            qs = qs.filter(month=month)
        center_id = self.request.query_params.get('center_id')
        if center_id:
            qs = qs.filter(support_center_id=center_id)
        return qs

    @action(detail=False, methods=['get'])
    def yearly_summary(self, request):
        """연도별 요약"""
        year = request.query_params.get('year', timezone.now().year)
        qs = StatisticsMonthly.objects.filter(year=year, support_center__isnull=True).order_by('month')
        return Response({
            'year': year,
            'months': StatisticsMonthlySerializer(qs, many=True).data
        })

    @action(detail=False, methods=['get'])
    def center_comparison(self, request):
        """지원청별 비교"""
        year  = request.query_params.get('year', timezone.now().year)
        month = request.query_params.get('month', timezone.now().month)
        qs = StatisticsMonthly.objects.filter(
            year=year, month=month, support_center__isnull=False
        ).select_related('support_center')
        return Response(StatisticsMonthlySerializer(qs, many=True).data)


class SLARecordViewSet(viewsets.ReadOnlyModelViewSet):
    """SLA 기록 조회"""
    serializer_class = SLARecordSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = StandardPagination

    def get_queryset(self):
        qs = SLARecord.objects.select_related('incident', 'incident__school')
        arrival_ok = self.request.query_params.get('arrival_ok')
        if arrival_ok == '0':
            qs = qs.filter(arrival_ok=False)
        elif arrival_ok == '1':
            qs = qs.filter(arrival_ok=True)
        resolve_ok = self.request.query_params.get('resolve_ok')
        if resolve_ok == '0':
            qs = qs.filter(resolve_ok=False)
        elif resolve_ok == '1':
            qs = qs.filter(resolve_ok=True)
        start = self.request.query_params.get('start')
        if start:
            qs = qs.filter(created_at__date__gte=start)
        end = self.request.query_params.get('end')
        if end:
            qs = qs.filter(created_at__date__lte=end)
        return qs

    @action(detail=False, methods=['get'])
    def summary(self, request):
        """SLA 전체 요약"""
        qs = SLARecord.objects.all()
        total = qs.count()
        arrival_ok = qs.filter(arrival_ok=True).count()
        resolve_ok = qs.filter(resolve_ok=True).count()
        avgs = qs.aggregate(
            avg_arrival=Avg('arrival_actual_min'),
            avg_resolve=Avg('resolve_actual_min'),
        )
        return Response({
            'total':             total,
            'arrival_ok_count':  arrival_ok,
            'resolve_ok_count':  resolve_ok,
            'arrival_ok_rate':   round(arrival_ok / total * 100, 1) if total else 0,
            'resolve_ok_rate':   round(resolve_ok / total * 100, 1) if total else 0,
            'avg_arrival_min':   round(avgs['avg_arrival'] or 0, 1),
            'avg_resolve_min':   round(avgs['avg_resolve'] or 0, 1),
        })


class SatisfactionSurveyViewSet(viewsets.ReadOnlyModelViewSet):
    """만족도 조사 결과"""
    serializer_class = SatisfactionSurveySerializer
    permission_classes = [IsAuthenticated]
    pagination_class = StandardPagination

    def get_queryset(self):
        qs = SatisfactionSurvey.objects.select_related('incident', 'incident__school')
        st = self.request.query_params.get('status')
        if st:
            qs = qs.filter(status=st)
        start = self.request.query_params.get('start')
        if start:
            qs = qs.filter(sent_at__date__gte=start)
        end = self.request.query_params.get('end')
        if end:
            qs = qs.filter(sent_at__date__lte=end)
        return qs

    @action(detail=False, methods=['get'])
    def summary(self, request):
        """만족도 요약"""
        qs = SatisfactionSurvey.objects.filter(status='responded')
        total     = SatisfactionSurvey.objects.count()
        responded = qs.count()
        avg_score = qs.aggregate(avg=Avg('score'))['avg']
        by_score  = qs.values('score').annotate(cnt=Count('id')).order_by('score')
        return Response({
            'total':          total,
            'responded':      responded,
            'response_rate':  round(responded / total * 100, 1) if total else 0,
            'avg_score':      round(avg_score or 0, 2),
            'by_score':       {str(s['score']): s['cnt'] for s in by_score},
        })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def comprehensive_stats_api(request):
    """통계 종합 데이터 API — GET ?year=2025&month=4(optional) (5분 캐시)"""
    from django.core.cache import cache
    import calendar
    from datetime import date as dt_date
    from django.db.models import Count, Avg, Q, F
    from apps.incidents.models import Incident
    from apps.schools.models import SupportCenter, SchoolType
    from apps.workforce.models import AttendanceLog

    year  = int(request.query_params.get('year',  timezone.now().year))
    month = request.query_params.get('month')
    month = int(month) if month else None

    cache_key = f'comprehensive_stats_{year}_{month}'
    cached = cache.get(cache_key)
    if cached is not None:
        return Response(cached)

    if month:
        date_from = dt_date(year, month, 1)
        date_to   = dt_date(year, month, calendar.monthrange(year, month)[1])
        label = f'{year}년 {month}월'
    else:
        date_from = dt_date(year, 1, 1)
        date_to   = dt_date(year, 12, 31)
        label = f'{year}년 전체'

    inc_qs = Incident.objects.filter(
        received_at__date__gte=date_from,
        received_at__date__lte=date_to,
    )

    # KPI
    total_inc   = inc_qs.count()
    completed   = inc_qs.filter(status='completed').count()
    completed_rate = round(completed / total_inc * 100, 1) if total_inc else 0

    # 장애 유형별
    by_type = list(
        inc_qs.values('fault_type').annotate(cnt=Count('id')).order_by('-cnt')
    )
    FAULT_LABELS = {
        'service_stop': '서비스 중단', 'speed_slow': '속도 저하',
        'intermittent': '간헐적 장애', 'hardware': '하드웨어',
        'software': '소프트웨어', 'configuration': '설정 오류',
        'cable': '케이블', 'power': '전원 장애', 'security': '보안',
        'user_error': '사용자 실수', 'other': '기타',
    }
    for row in by_type:
        row['label'] = FAULT_LABELS.get(row['fault_type'], row['fault_type'])

    # 월별 추이 (연간 조회 시 12개월, 월별 조회 시 일별)
    if month:
        from django.db.models.functions import TruncDay
        monthly_trend = list(
            inc_qs.annotate(d=TruncDay('received_at'))
            .values('d').annotate(total=Count('id'), done=Count('id', filter=Q(status='completed')))
            .order_by('d')
        )
        trend_labels = [r['d'].strftime('%d일') for r in monthly_trend]
        trend_total  = [r['total'] for r in monthly_trend]
        trend_done   = [r['done']  for r in monthly_trend]
    else:
        from django.db.models.functions import TruncMonth
        monthly_trend = list(
            inc_qs.annotate(m=TruncMonth('received_at'))
            .values('m').annotate(total=Count('id'), done=Count('id', filter=Q(status='completed')))
            .order_by('m')
        )
        trend_labels = [r['m'].strftime('%m월') for r in monthly_trend]
        trend_total  = [r['total'] for r in monthly_trend]
        trend_done   = [r['done']  for r in monthly_trend]

    # 교육청 고정 순서 (차트/표 공통)
    CTR_ORDER = ['동부','서부','남부','북부','중부',
                 '강동송파','강서양천','강남서초','동작관악','성동광진','성북강북']

    # 지원청별 — 0건 포함 고정 순서
    _ctr_raw = {
        r['school__support_center__name']: r
        for r in inc_qs.values('school__support_center__name')
            .annotate(cnt=Count('id'), done=Count('id', filter=Q(status='completed')))
    }
    # DB에 있는 모든 교육청을 고정 순서로
    from apps.schools.models import SupportCenter as _SC
    _all_ctr_names = list(_SC.objects.values_list('name', flat=True))
    _ordered_names = [c for c in CTR_ORDER if c in _all_ctr_names] + \
                     [c for c in _all_ctr_names if c not in CTR_ORDER]
    by_center = []
    for name in _ordered_names:
        raw = _ctr_raw.get(name, {})
        cnt  = raw.get('cnt', 0)
        done = raw.get('done', 0)
        by_center.append({
            'label': name,
            'cnt':   cnt,
            'done':  done,
            'rate':  round(done / cnt * 100, 1) if cnt else 0,
        })

    # 학제별 — 0건 포함 고정 순서
    from apps.schools.models import SchoolType as _ST
    _st_raw = {
        r['school__school_type__name']: r['cnt']
        for r in inc_qs.values('school__school_type__name').annotate(cnt=Count('id'))
    }
    by_school_type = []
    for st in _ST.objects.order_by('order', 'id'):
        by_school_type.append({
            'label': st.name,
            'cnt':   _st_raw.get(st.name, 0),
        })

    # 우선순위별
    by_priority = list(
        inc_qs.values('priority').annotate(cnt=Count('id')).order_by('-cnt')
    )
    P_LABELS = {'critical': '긴급', 'high': '높음', 'medium': '보통', 'low': '낮음'}
    for r in by_priority:
        r['label'] = P_LABELS.get(r['priority'], r['priority'])

    # 상위 장애 학교
    top_schools = list(
        inc_qs.values('school__name', 'school__support_center__name')
        .annotate(cnt=Count('id'))
        .order_by('-cnt')[:10]
    )
    for r in top_schools:
        r['school']  = r.pop('school__name')
        r['center']  = r.pop('school__support_center__name') or ''

    # SLA 월별 (SLAMonthly 사용)
    from apps.incidents.models import SLAMonthly
    sla_qs = SLAMonthly.objects.filter(year=year)
    if month:
        sla_qs = sla_qs.filter(month=month)
    sla_months = list(sla_qs.order_by('month').values(
        'month', 'total_score', 'grade', 'uptime_pct',
        'fault_count', 'avg_fault_min', 'satisfaction_pct',
    ))

    # 근태 요약
    att_qs = AttendanceLog.objects.filter(
        work_date__gte=date_from, work_date__lte=date_to
    )
    att_total   = att_qs.count()
    att_normal  = att_qs.filter(status='normal').count()
    att_late    = att_qs.filter(status='late').count()
    att_absent  = att_qs.filter(status='absent').count()

    # 만족도
    from .models import SatisfactionSurvey
    sur_qs = SatisfactionSurvey.objects.filter(
        sent_at__date__gte=date_from, sent_at__date__lte=date_to
    )
    sur_total    = sur_qs.count()
    sur_resp     = sur_qs.filter(status='responded').count()
    sur_avg      = sur_qs.filter(status='responded').aggregate(a=Avg('score'))['a']
    sur_by_score = list(
        sur_qs.filter(status='responded').values('score').annotate(cnt=Count('id')).order_by('score')
    )

    # ── 지원청 × 학제 피벗 ─────────────────────────────────────
    # 고정 순서 (대시보드 CENTER_ORDER 와 동일)
    CENTER_ORDER = ['동부','서부','남부','북부','중부',
                    '강동송파','강서양천','강남서초','동작관악','성동광진','성북강북']
    TYPE_ORDER   = ['유치원','초등학교','중학교','고등학교','특수학교','각종학교','고등기술학교']

    from apps.schools.models import SupportCenter, SchoolType
    # DB에 있는 항목만 유효 목록으로 (순서 유지)
    db_centers = set(SupportCenter.objects.values_list('name', flat=True))
    db_types   = set(SchoolType.objects.values_list('name', flat=True))
    all_centers = [c for c in CENTER_ORDER if c in db_centers]
    all_types   = [t for t in TYPE_ORDER   if t in db_types]
    # DB에는 있지만 순서 목록에 없는 경우 뒤에 추가
    for c in sorted(db_centers):
        if c not in all_centers:
            all_centers.append(c)
    for t in sorted(db_types):
        if t not in all_types:
            all_types.append(t)

    pivot_raw = list(
        inc_qs.values(
            'school__support_center__name',
            'school__school_type__name',
        ).annotate(cnt=Count('id'))
    )
    pivot_map = {}
    for r in pivot_raw:
        c = r['school__support_center__name'] or '미분류'
        t = r['school__school_type__name']    or '미분류'
        pivot_map.setdefault(c, {})[t] = r['cnt']

    grand_total = sum(r['cnt'] for r in pivot_raw) or 1

    # 행 = 학제, 열 = 교육청
    pivot_rows = []
    for t in all_types:
        row_total = sum(pivot_map.get(c, {}).get(t, 0) for c in all_centers)
        cells = []
        for c in all_centers:
            cnt = pivot_map.get(c, {}).get(t, 0)
            col_total = sum(pivot_map.get(c2, {}).get(t, 0) for c2 in all_centers)
            ctr_total = sum(pivot_map.get(c, {}).get(t2, 0) for t2 in all_types)
            cells.append({
                'cnt':       cnt,
                'row_pct':   round(cnt / row_total  * 100, 1) if row_total  else 0,
                'col_pct':   round(cnt / ctr_total   * 100, 1) if ctr_total  else 0,
            })
        pivot_rows.append({
            'type':       t,
            'row_total':  row_total,
            'row_pct':    round(row_total / grand_total * 100, 1),
            'cells':      cells,
        })

    # 합계 행 (교육청별)
    center_totals = []
    for c in all_centers:
        ct = sum(pivot_map.get(c, {}).get(t, 0) for t in all_types)
        center_totals.append({
            'cnt':  ct,
            'pct':  round(ct / grand_total * 100, 1),
        })

    pivot_data = {
        'centers':       all_centers,
        'types':         all_types,
        'rows':          pivot_rows,
        'center_totals': center_totals,
        'grand_total':   grand_total,
    }

    # ── 장애 유형 상세 분석 (category × subcategory) ──────────
    from apps.incidents.models import IncidentCategory, IncidentSubcategory

    # 대분류별 집계 + 소분류별 집계
    cat_counts = {
        r['category_id']: r['cnt']
        for r in inc_qs.values('category_id').annotate(cnt=Count('id'))
    }
    cat_completed = {
        r['category_id']: r['cnt']
        for r in inc_qs.filter(status='completed').values('category_id').annotate(cnt=Count('id'))
    }
    sub_counts = {
        (r['category_id'], r['subcategory_id']): r['cnt']
        for r in inc_qs.values('category_id', 'subcategory_id').annotate(cnt=Count('id'))
    }

    # 대분류 × 교육청 피벗
    cat_center_raw = list(
        inc_qs.values('category_id', 'school__support_center__name')
        .annotate(cnt=Count('id'))
    )
    cat_center_map = {}
    for r in cat_center_raw:
        cat_center_map.setdefault(r['category_id'], {})[
            r['school__support_center__name'] or '미분류'
        ] = r['cnt']

    categories_qs = IncidentCategory.objects.prefetch_related('subcategories').order_by('order', 'id')
    total_inc2 = inc_qs.count() or 1

    fault_categories = []
    for cat in categories_qs:
        cat_total = cat_counts.get(cat.id, 0)
        cat_done  = cat_completed.get(cat.id, 0)

        # 소분류 목록 (DB 정의 순서 + 소분류없음)
        subs = []
        for sub in cat.subcategories.order_by('order', 'id'):
            sc = sub_counts.get((cat.id, sub.id), 0)
            subs.append({
                'name':    sub.name,
                'cnt':     sc,
                'cat_pct': round(sc / cat_total * 100, 1) if cat_total else 0,
                'total_pct': round(sc / total_inc2 * 100, 1),
            })
        # 소분류 미지정 건
        no_sub = sub_counts.get((cat.id, None), 0)
        if no_sub:
            subs.append({
                'name': '소분류 미지정',
                'cnt':  no_sub,
                'cat_pct': round(no_sub / cat_total * 100, 1) if cat_total else 0,
                'total_pct': round(no_sub / total_inc2 * 100, 1),
            })

        # 교육청별 건수 (고정 순서)
        center_cells = []
        for c in all_centers:
            cc = cat_center_map.get(cat.id, {}).get(c, 0)
            center_cells.append({'cnt': cc, 'pct': round(cc / cat_total * 100, 1) if cat_total else 0})

        fault_categories.append({
            'id':             cat.id,
            'name':           cat.name,
            'total':          cat_total,
            'pct':            round(cat_total / total_inc2 * 100, 1),
            'completed':      cat_done,
            'completed_rate': round(cat_done / cat_total * 100, 1) if cat_total else 0,
            'subcategories':  subs,
            'center_cells':   center_cells,
        })

    fault_analysis_data = {
        'categories': fault_categories,
        'centers':    all_centers,
        'total':      total_inc2 if total_inc2 > 1 else 0,
    }

    # ── 점검 진척 (InspectionPlan / SchoolInspection) ──────────
    from apps.progress.models import InspectionPlan, SchoolInspection, WorkerArea
    from apps.schools.models import School as _School
    PLAN_TYPE_LABELS = {
        'regular': '정기점검', 'special': '특별점검',
        'quarterly': '분기점검', 'project': '사업점검',
        'survey': '실태조사', 'followup': '사후점검',
    }

    # 지원청별 기초 자원 현황 (고정 순서)
    _worker_map = {
        r['support_center__name']: r['cnt']
        for r in WorkerArea.objects.filter(is_primary=True)
            .values('support_center__name')
            .annotate(cnt=Count('worker', distinct=True))
    }
    _school_map = {
        r['support_center__name']: r['cnt']
        for r in _School.objects.filter(is_active=True)
            .values('support_center__name').annotate(cnt=Count('id'))
    }
    ctr_resource = []
    for cname in all_centers:
        wc = _worker_map.get(cname, 0)
        sc = _school_map.get(cname, 0)
        ctr_resource.append({
            'center':        cname,
            'workers':       wc,
            'schools':       sc,
            'schools_per_worker': round(sc / wc, 1) if wc else None,
        })

    plans_qs = InspectionPlan.objects.filter(year=year)
    inspection_plans = []
    for p in plans_qs.order_by('plan_type', 'start_date'):
        total_schools = p.school_inspections.count()
        done_schools  = p.school_inspections.filter(status='completed').count()
        rate = round(done_schools / total_schools * 100, 1) if total_schools else 0

        # 지원청별 진척
        sub_raw = list(
            p.school_inspections.values('school__support_center__name')
            .annotate(
                total=Count('id'),
                done=Count('id', filter=Q(status='completed')),
                scheduled=Count('id', filter=Q(status='scheduled')),
            )
        )
        sub_map = {r['school__support_center__name']: r for r in sub_raw}
        center_progress = []
        for cname in all_centers:
            r = sub_map.get(cname, {})
            t = r.get('total', 0)
            d = r.get('done', 0)
            wc = _worker_map.get(cname, 0)
            center_progress.append({
                'center':    cname,
                'total':     t,
                'done':      d,
                'scheduled': r.get('scheduled', 0),
                'rate':      round(d / t * 100, 1) if t else 0,
                'per_worker': round(t / wc, 1) if wc else None,
            })

        inspection_plans.append({
            'id':               p.id,
            'name':             p.name,
            'plan_type':        p.plan_type,
            'type_label':       PLAN_TYPE_LABELS.get(p.plan_type, p.plan_type),
            'status':           p.status,
            'start_date':       p.start_date.isoformat(),
            'end_date':         p.end_date.isoformat(),
            'total':            total_schools,
            'done':             done_schools,
            'rate':             rate,
            'center_progress':  center_progress,
        })

    # 유형별 집계
    by_plan_type = {}
    for p in inspection_plans:
        t = p['type_label']
        if t not in by_plan_type:
            by_plan_type[t] = {'total': 0, 'done': 0, 'plans': 0}
        by_plan_type[t]['total'] += p['total']
        by_plan_type[t]['done']  += p['done']
        by_plan_type[t]['plans'] += 1
    for v in by_plan_type.values():
        v['rate'] = round(v['done'] / v['total'] * 100, 1) if v['total'] else 0
    inspection_by_type = [
        {'type_label': k, **v} for k, v in by_plan_type.items()
    ]

    # ── 인력별 업무 부하 분석 ──────────────────────────────
    from apps.accounts.models import User as _User
    from apps.incidents.models import IncidentAssignment

    # 장애 배정 건수 (기간 필터)
    _inc_by_worker = {
        r['worker_id']: {'total': r['total'], 'done': r['done']}
        for r in IncidentAssignment.objects.filter(
            incident__received_at__date__gte=date_from,
            incident__received_at__date__lte=date_to,
        ).values('worker_id').annotate(
            total=Count('id'),
            done=Count('id', filter=Q(incident__status='completed')),
        )
    }

    # 점검 배정 건수 (연도 기준, plan_type별)
    _insp_by_worker = {}
    for r in SchoolInspection.objects.filter(
        plan__year=year,
        assigned_worker__isnull=False,
    ).values('assigned_worker_id', 'plan__plan_type').annotate(
        total=Count('id'),
        done=Count('id', filter=Q(status='completed')),
    ):
        wid   = r['assigned_worker_id']
        ptype = r['plan__plan_type']
        _insp_by_worker.setdefault(wid, {})[ptype] = {
            'total': r['total'], 'done': r['done'],
        }

    # 보고서 작성 건수 (report_type별)
    from apps.reports.models import Report as _Report
    _rpt_by_worker = {}
    for r in _Report.objects.filter(
        created_at__date__gte=date_from,
        created_at__date__lte=date_to,
        created_by__isnull=False,
    ).values('created_by_id', 'template__report_type').annotate(cnt=Count('id')):
        wid   = r['created_by_id']
        rtype = r['template__report_type'] or 'other'
        _rpt_by_worker.setdefault(wid, {})[rtype] = r['cnt']

    RPT_TYPE_SHORT = {
        'incident':      '장애확인서',
        'regular':       '정기점검서',
        'cable':         '케이블공사',
        'switch_install':'스위치설치',
        'quarterly':     '분기점검서',
        'other':         '기타보고서',
    }
    PLAN_TYPE_SHORT = {
        'regular':   '정기점검',
        'special':   '특별점검',
        'quarterly': '분기점검',
        'project':   '사업점검',
        'survey':    '실태조사',
        'followup':  '사후점검',
    }

    # 지원청 배정 인력만 (support_center 없는 인력 제외)
    _workers_raw = list(
        _User.objects.filter(
            role__in=['worker', 'resident'],
            support_center__isnull=False,
        ).select_related('support_center').order_by('name')
    )
    # CENTER_ORDER 기준으로 정렬 (동부→서부→남부→…)
    _center_rank = {c: i for i, c in enumerate(CENTER_ORDER)}
    _workers_raw.sort(key=lambda w: (
        _center_rank.get(w.support_center.name, 99),
        w.name
    ))

    worker_workload = []
    for w in _workers_raw:
        cname      = w.support_center.name
        inc_data   = _inc_by_worker.get(w.id, {'total': 0, 'done': 0})
        insp_data  = _insp_by_worker.get(w.id, {})
        rpt_data   = _rpt_by_worker.get(w.id, {})

        insp_total = sum(v['total'] for v in insp_data.values())
        insp_done  = sum(v['done']  for v in insp_data.values())
        rpt_total  = sum(rpt_data.values())
        total_work = inc_data['total'] + insp_total + rpt_total

        # 점검 상세 (유형별)
        insp_detail = {
            PLAN_TYPE_SHORT.get(pt, pt): v
            for pt, v in insp_data.items()
        }
        # 보고서 상세 (유형별)
        rpt_detail = {
            RPT_TYPE_SHORT.get(rt, rt): cnt
            for rt, cnt in rpt_data.items()
        }

        worker_workload.append({
            'id':            w.id,
            'name':          w.name,
            'center':        cname,
            'incidents':     inc_data['total'],
            'inc_done':      inc_data['done'],
            'insp_total':    insp_total,
            'insp_done':     insp_done,
            'insp_detail':   insp_detail,
            'rpt_total':     rpt_total,
            'rpt_detail':    rpt_detail,
            'total_work':    total_work,
        })

    response_data = {
        'period': {'year': year, 'month': month, 'label': label,
                   'date_from': date_from.isoformat(), 'date_to': date_to.isoformat()},
        'kpi': {
            'total_incidents':  total_inc,
            'completed_count':  completed,
            'completed_rate':   completed_rate,
            'schools_served':   inc_qs.values('school').distinct().count(),
            'avg_satisfaction': round(float(sur_avg), 2) if sur_avg else None,
            'avg_sla_score': round(
                sum(s['total_score'] for s in sla_months if s['total_score']) /
                len([s for s in sla_months if s['total_score']]), 1
            ) if any(s['total_score'] for s in sla_months) else None,
        },
        'trend': {'labels': trend_labels, 'total': trend_total, 'done': trend_done},
        'by_type':          by_type,
        'by_priority':      by_priority,
        'by_center':        by_center,
        'by_school_type':   by_school_type,
        'top_schools':      top_schools,
        'sla_months':       sla_months,
        'attendance': {
            'total': att_total, 'normal': att_normal,
            'late': att_late, 'absent': att_absent,
        },
        'satisfaction': {
            'total': sur_total, 'responded': sur_resp,
            'response_rate': round(sur_resp / sur_total * 100, 1) if sur_total else 0,
            'avg_score': round(float(sur_avg), 2) if sur_avg else None,
            'by_score': {str(r['score']): r['cnt'] for r in sur_by_score},
        },
        'inspection': {
            'plans':        inspection_plans,
            'by_type':      inspection_by_type,
            'centers':      all_centers,
            'ctr_resource': ctr_resource,
        },
        'pivot':           pivot_data,
        'fault_analysis':  fault_analysis_data,
        'worker_workload': worker_workload,
    }
    cache.set(cache_key, response_data, 300)
    return Response(response_data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def comprehensive_stats_excel(request):
    """통계 종합 Excel 내보내기 — GET ?year=2025&month=4(optional)"""
    try:
        import io, calendar as cal_mod
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from apps.incidents.models import Incident
        from django.db.models import Count, Q
        from datetime import date as dt_date
    except ImportError:
        from rest_framework.response import Response
        return Response({'error': 'openpyxl 미설치'}, status=500)

    year  = int(request.query_params.get('year',  timezone.now().year))
    month_str = request.query_params.get('month')
    month = int(month_str) if month_str else None

    if month:
        date_from = dt_date(year, month, 1)
        date_to   = dt_date(year, month, cal_mod.monthrange(year, month)[1])
        label = f'{year}년 {month}월'
    else:
        date_from = dt_date(year, 1, 1)
        date_to   = dt_date(year, 12, 31)
        label = f'{year}년 전체'

    inc_qs = Incident.objects.filter(
        received_at__date__gte=date_from,
        received_at__date__lte=date_to,
    )

    wb = openpyxl.Workbook()
    hdr_fill = PatternFill('solid', fgColor='1E3A5F')
    hdr_font = Font(color='FFFFFF', bold=True, size=10)
    hdr_align = Alignment(horizontal='center', vertical='center')
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    def _write_header(ws, row, cols):
        for c_idx, col_name in enumerate(cols, 1):
            cell = ws.cell(row=row, column=c_idx, value=col_name)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = hdr_align
            cell.border = thin

    # ── 시트 1: 장애 현황 ──────────────────────
    ws1 = wb.active
    ws1.title = '장애현황'
    ws1.append([f'분석기간: {label}'])
    ws1.cell(1, 1).font = Font(bold=True, size=12)
    ws1.append([])
    cols1 = ['접수번호', '학교명', '지원청', '분류', '상태', '긴급도',
             '접수일시', '도착일시', '완료일시', '소요시간(분)', '처리유형']
    _write_header(ws1, 3, cols1)
    STATUS_LBL = {'received':'접수','assigned':'배정','moving':'이동','arrived':'도착',
                  'processing':'처리중','completed':'완료','cancelled':'취소'}
    PRI_LBL = {'critical':'긴급','high':'높음','medium':'중간','low':'낮음'}
    for r_idx, inc in enumerate(inc_qs.select_related('school','school__support_center','category'), 1):
        ws1.append([
            inc.incident_number,
            inc.school.name if inc.school else '',
            inc.school.support_center.name if inc.school and inc.school.support_center else '',
            inc.category.name if inc.category else '',
            STATUS_LBL.get(inc.status, inc.status),
            PRI_LBL.get(inc.priority, inc.priority),
            inc.received_at.strftime('%Y-%m-%d %H:%M') if inc.received_at else '',
            inc.arrived_at.strftime('%Y-%m-%d %H:%M') if inc.arrived_at else '',
            inc.completed_at.strftime('%Y-%m-%d %H:%M') if inc.completed_at else '',
            inc.get_elapsed_minutes() if hasattr(inc, 'get_elapsed_minutes') else '',
            getattr(inc, 'resolution_type', ''),
        ])

    # ── 시트 2: 지원청별 집계 ──────────────────
    ws2 = wb.create_sheet('지원청별집계')
    by_center = list(
        inc_qs.values('school__support_center__name')
        .annotate(total=Count('id'), done=Count('id', filter=Q(status='completed')))
        .order_by('-total')
    )
    _write_header(ws2, 1, ['지원청', '총건수', '완료', '완료율(%)'])
    for row in by_center:
        total = row['total']
        done  = row['done']
        ws2.append([
            row['school__support_center__name'] or '미분류',
            total, done,
            round(done / total * 100, 1) if total else 0,
        ])

    # ── 시트 3: 분류별 집계 ──────────────────────
    ws3 = wb.create_sheet('분류별집계')
    by_cat = list(
        inc_qs.values('category__name')
        .annotate(cnt=Count('id'))
        .order_by('-cnt')
    )
    _write_header(ws3, 1, ['분류', '건수'])
    for row in by_cat:
        ws3.append([row['category__name'] or '미분류', row['cnt']])

    for ws in [ws1, ws2, ws3]:
        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 35)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from django.http import HttpResponse
    fname = f'statistics_{year}{"_"+str(month).zfill(2) if month else ""}.xlsx'
    resp = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp


@api_view(['POST'])
@permission_classes([AllowAny])
def survey_respond(request):
    """만족도 응답 (토큰 기반, 인증 불필요)"""
    serializer = SurveyResponseSerializer(data=request.data)
    if not serializer.is_valid():
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    data = serializer.validated_data
    try:
        survey = SatisfactionSurvey.objects.get(token=data['token'], status='sent')
    except SatisfactionSurvey.DoesNotExist:
        return Response({'error': '유효하지 않거나 이미 응답된 토큰입니다.'}, status=status.HTTP_400_BAD_REQUEST)
    survey.score = data['score']
    survey.comment = data.get('comment', '')
    survey.status = 'responded'
    survey.responded_at = timezone.now()
    survey.save(update_fields=['score', 'comment', 'status', 'responded_at'])
    return Response({'message': '응답해 주셔서 감사합니다.'})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def pattern_api(request):
    """장애 패턴 분석 API

    Query params:
        center (int): 지원청 PK (선택)
        year   (int): 연도 (기본: 올해)
        month  (int): 월 (기본: 이번 달)
    """
    from apps.incidents.models import Incident
    from apps.schools.models import SupportCenter
    from .services import IncidentPatternAnalyzer

    now = timezone.now()
    year = int(request.query_params.get('year', now.year))
    month = int(request.query_params.get('month', now.month))

    qs = Incident.objects.all()
    center_id = request.query_params.get('center')
    center = None
    if center_id:
        try:
            center = SupportCenter.objects.get(pk=center_id)
            qs = qs.filter(school__support_center=center)
        except SupportCenter.DoesNotExist:
            return Response({'error': '지원청을 찾을 수 없습니다.'}, status=404)

    return Response({
        'hourly':         IncidentPatternAnalyzer.hourly_distribution(qs),
        'weekday':        IncidentPatternAnalyzer.weekday_distribution(qs),
        'category_trend': IncidentPatternAnalyzer.category_trend(year, month),
        'hotspots':       IncidentPatternAnalyzer.recurrence_hotspots(
                              center=center or SupportCenter.objects.first(),
                              top_n=10
                          ) if SupportCenter.objects.exists() else [],
    })

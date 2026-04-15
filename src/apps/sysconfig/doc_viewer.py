"""
산출물 통합 조회 API — 모든 서류를 자동 탐지, 상세 데이터를 표 형태로 제공
기존 exports.py의 상세 컬럼 로직을 재활용
"""
import io
import json
import csv
import urllib.parse
from django.apps import apps
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.utils import timezone


def _admin_required(view_func):
    from functools import wraps
    @wraps(view_func)
    def _wrapped(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return JsonResponse({'error': '로그인 필요'}, status=401)
        if request.user.role not in ('superadmin', 'admin'):
            return JsonResponse({'error': '권한 없음'}, status=403)
        return view_func(request, *args, **kwargs)
    return _wrapped


# ══════════════════════════════════════════════════════
# 서류 유형 레지스트리 (자동 탐지 + 수동 등록 혼합)
# ══════════════════════════════════════════════════════

def _build_catalog():
    """카탈로그 자동 생성 — ReportTemplate + 기타 서류 자동 탐지"""
    items = []

    # ── 1) Report 템플릿 기반 (자동) ─────────────────
    try:
        from apps.reports.models import ReportTemplate, Report
        for tmpl in ReportTemplate.objects.all():
            items.append({
                'id': f'report_{tmpl.id}',
                'name': tmpl.name,
                'count': Report.objects.filter(template=tmpl).count(),
                'category': '업무보고서',
            })
    except Exception:
        pass

    # ── 2) 고정 서류 유형 ────────────────────────────
    FIXED = [
        ('incidents',      '장애처리보고서',    '장애관리',   'apps.incidents.models', 'Incident', None),
        ('sla',            'SLA 성과',        '장애관리',   'apps.incidents.models', 'IncidentSLA', None),
        ('assets',         '장비목록',         '장비관리',   'apps.assets.models', 'Asset', None),
        ('assets_inbound', '장비입고증',       '장비관리',   'apps.assets.models', 'AssetInbound', None),
        ('assets_outbound','장비출고증',       '장비관리',   'apps.assets.models', 'AssetOutbound', None),
        ('mat_inbound',    '자재입고내역',     '자재관리',   'apps.materials.models', 'MaterialInbound', None),
        ('mat_outbound',   '자재출고내역',     '자재관리',   'apps.materials.models', 'MaterialOutbound', None),
        ('mat_inventory',  '자재현재고',       '자재관리',   'apps.materials.models', 'WarehouseInventory', None),
        ('photos',         '사진목록',         '현장사진',   'apps.photos.models', 'Photo', None),
        ('attendance',     '출퇴근기록',       '인력관리',   'apps.workforce.models', 'AttendanceLog', None),
        ('progress',       '점검현황',         '진척관리',   'apps.progress.models', 'SchoolInspection', None),
        ('education',      '교육이수증',       '교육관리',   'apps.education.models', 'EducationCompletion', None),
        ('users',          '인력목록',         '인력관리',   'apps.accounts.models', 'User', "role__in=['worker','resident']"),
        ('schools',        '학교목록',         '학교정보',   'apps.schools.models', 'School', None),
        ('buildings',      '건물정보',         '학교정보',   'apps.schools.models', 'SchoolRoom', None),
        ('contacts',       '학교담당자',       '학교정보',   'apps.schools.models', 'SchoolContact', None),
    ]
    for doc_id, name, cat, mod_path, cls_name, extra_filter in FIXED:
        try:
            parts = mod_path.rsplit('.', 1)
            import importlib
            mod = importlib.import_module(mod_path)
            model = getattr(mod, cls_name)
            qs = model.objects.all()
            if extra_filter:
                qs = qs.filter(**dict([extra_filter.split('=')])) if '=' in extra_filter else qs
            items.append({
                'id': doc_id,
                'name': name,
                'count': qs.count(),
                'category': cat,
            })
        except Exception:
            items.append({'id': doc_id, 'name': name, 'count': 0, 'category': cat})

    return items


@login_required
@_admin_required
def doc_catalog(request):
    """산출물 카탈로그"""
    items = _build_catalog()
    tree = {}
    for item in items:
        cat = item['category']
        if cat not in tree:
            tree[cat] = []
        tree[cat].append(item)
    return JsonResponse({'tree': tree, 'items': items})


# ══════════════════════════════════════════════════════
# 각 서류별 상세 데이터 생성기
# ══════════════════════════════════════════════════════

def _paginate(all_rows, page, page_size):
    total = len(all_rows)
    offset = (page - 1) * page_size
    return {
        'rows': all_rows[offset:offset + page_size],
        'total': total,
        'page': page,
        'page_size': page_size,
        'total_pages': (total + page_size - 1) // page_size if total else 0,
    }


def _dt(val):
    if not val:
        return ''
    return timezone.localtime(val).strftime('%Y-%m-%d %H:%M')


def _d(val):
    return str(val) if val else ''


def _get_incidents(q):
    from apps.incidents.models import Incident
    qs = Incident.objects.select_related(
        'school__support_center', 'school__school_type',
        'category', 'subcategory', 'received_by'
    ).order_by('-received_at')
    if q:
        qs = qs.filter(Q(incident_number__icontains=q) | Q(school__name__icontains=q) | Q(resolution__icontains=q))
    headers = ['접수번호', '접수일시', '교육지원청', '학제', '학교명',
               '장애분류', '소분류', '상태', '긴급도', '접수방법',
               '접수자', '요청자', '연락처', '요청내용',
               '장애위치(건물)', '장애위치(층)', '장애위치(교실)',
               '현장도착', '처리완료', '소요시간(분)',
               '처리유형', '처리내용', '장애유형',
               'SLA도착준수', 'SLA처리준수', '만족도점수']
    rows = []
    for i in qs:
        elapsed = ''
        if i.received_at and i.completed_at:
            elapsed = int((i.completed_at - i.received_at).total_seconds() / 60)
        rows.append({h: v for h, v in zip(headers, [
            i.incident_number, _dt(i.received_at),
            i.school.support_center.name if i.school and i.school.support_center else '',
            i.school.school_type.name if i.school and i.school.school_type else '',
            i.school.name if i.school else '',
            i.category.name if i.category else '',
            i.subcategory.name if i.subcategory else '',
            i.get_status_display(), i.get_priority_display(),
            i.get_contact_method_display() if hasattr(i, 'get_contact_method_display') else i.contact_method,
            i.received_by.name if i.received_by else '',
            i.requester_name, i.requester_phone, i.description or '',
            str(i.location_building) if i.location_building else '',
            str(i.location_floor) if i.location_floor else '',
            str(i.location_room) if i.location_room else '',
            _dt(i.arrived_at), _dt(i.completed_at), elapsed,
            i.get_resolution_type_display() if hasattr(i, 'get_resolution_type_display') else (i.resolution_type or ''),
            i.resolution or '', i.get_fault_type_display() if hasattr(i, 'get_fault_type_display') else (i.fault_type or ''),
            'Y' if i.sla_arrival_ok else ('N' if i.sla_arrival_ok is False else ''),
            'Y' if i.sla_resolve_ok else ('N' if i.sla_resolve_ok is False else ''),
            i.satisfaction_score or '',
        ])})
    return headers, rows


def _get_sla(q):
    from apps.incidents.models import IncidentSLA
    qs = IncidentSLA.objects.select_related(
        'incident__school__support_center', 'incident__school__school_type', 'incident__category'
    ).order_by('-incident__received_at')
    if q:
        qs = qs.filter(Q(incident__incident_number__icontains=q) | Q(incident__school__name__icontains=q))
    headers = ['접수번호', '접수일시', '교육지원청', '학제', '학교명', '장애분류',
               '도착목표', '도착실제', '도착준수', '도착차이(분)',
               '처리목표', '처리실제', '처리준수', '처리차이(분)']
    rows = []
    for s in qs:
        i = s.incident
        rows.append({h: v for h, v in zip(headers, [
            i.incident_number, _dt(i.received_at),
            i.school.support_center.name if i.school and i.school.support_center else '',
            i.school.school_type.name if i.school and i.school.school_type else '',
            i.school.name if i.school else '',
            i.category.name if i.category else '',
            _dt(s.arrival_target), _dt(s.arrival_actual),
            'Y' if s.arrival_ok else ('N' if s.arrival_ok is False else ''), s.arrival_diff_min or '',
            _dt(s.resolve_target), _dt(s.resolve_actual),
            'Y' if s.resolve_ok else ('N' if s.resolve_ok is False else ''), s.resolve_diff_min or '',
        ])})
    return headers, rows


def _get_assets(q):
    from apps.assets.models import Asset
    qs = Asset.objects.select_related('asset_model__category', 'current_school__support_center', 'current_school__school_type').order_by('serial_number')
    if q:
        qs = qs.filter(Q(serial_number__icontains=q) | Q(current_school__name__icontains=q) | Q(asset_model__model_name__icontains=q))
    headers = ['분류', '제조번호(S/N)', '관리번호', '모델명', '제조사', '상태',
               '교육지원청', '학제', '학교명', '설치위치', '설치일', '보증만료일', '사업명']
    rows = []
    for a in qs:
        rows.append({h: v for h, v in zip(headers, [
            a.asset_model.category.name if a.asset_model and a.asset_model.category else '',
            a.serial_number or '', a.asset_tag or '',
            a.asset_model.model_name if a.asset_model else '',
            a.asset_model.manufacturer if a.asset_model else '',
            a.get_status_display() if hasattr(a, 'get_status_display') else a.status,
            a.current_school.support_center.name if a.current_school and a.current_school.support_center else '',
            a.current_school.school_type.name if a.current_school and a.current_school.school_type else '',
            a.current_school.name if a.current_school else '',
            a.install_location or '', _d(a.installed_at), _d(getattr(a, 'warranty_expire', '')),
            a.project_name or '',
        ])})
    return headers, rows


def _get_model_simple(doc_id, q):
    """장비입출고, 자재입출고 등 — 기존 exports 로직 재활용"""
    from apps.sysconfig.exports import (
        export_assets, export_materials, export_photos,
        export_attendance, export_progress, export_sla, export_reports,
    )
    # 이 함수들은 HttpResponse를 반환하므로 여기서는 직접 쿼리
    if doc_id == 'assets_inbound':
        return _get_assets_inbound(q)
    elif doc_id == 'assets_outbound':
        return _get_assets_outbound(q)
    elif doc_id == 'mat_inbound':
        return _get_materials_sub('inbound', q)
    elif doc_id == 'mat_outbound':
        return _get_materials_sub('outbound', q)
    elif doc_id == 'mat_inventory':
        return _get_materials_sub('inventory', q)
    elif doc_id == 'photos':
        return _get_photos(q)
    elif doc_id == 'attendance':
        return _get_attendance(q)
    elif doc_id == 'progress':
        return _get_progress(q)
    elif doc_id == 'education':
        return _get_education(q)
    elif doc_id == 'users':
        return _get_users(q)
    elif doc_id == 'schools':
        return _get_schools(q)
    elif doc_id == 'buildings':
        return _get_buildings(q)
    elif doc_id == 'contacts':
        return _get_contacts(q)
    return [], []


def _get_assets_inbound(q):
    from apps.assets.models import AssetInbound
    qs = AssetInbound.objects.select_related('asset__asset_model__category', 'asset__current_school__support_center', 'received_by').order_by('-created_at')
    if q: qs = qs.filter(Q(asset__serial_number__icontains=q) | Q(inbound_number__icontains=q))
    headers = ['입고번호', '입고일', '분류', '제조번호(S/N)', '모델명', '제조사', '출발지', '도착지', '인수자', '비고']
    rows = []
    for r in qs:
        rows.append({h: v for h, v in zip(headers, [
            r.inbound_number, _d(r.created_at.date() if r.created_at else ''),
            r.asset.asset_model.category.name if r.asset and r.asset.asset_model and r.asset.asset_model.category else '',
            r.asset.serial_number if r.asset else '', r.asset.asset_model.model_name if r.asset and r.asset.asset_model else '',
            r.asset.asset_model.manufacturer if r.asset and r.asset.asset_model else '',
            r.from_location_name or '', r.to_location_name if hasattr(r, 'to_location_name') else '',
            r.received_by.name if r.received_by else '', r.note or '',
        ])})
    return headers, rows


def _get_assets_outbound(q):
    from apps.assets.models import AssetOutbound
    qs = AssetOutbound.objects.select_related('asset__asset_model__category', 'issued_by').order_by('-created_at')
    if q: qs = qs.filter(Q(asset__serial_number__icontains=q) | Q(outbound_number__icontains=q))
    headers = ['출고번호', '출고일', '분류', '제조번호(S/N)', '모델명', '제조사', '출발지', '도착지', '출고자', '비고']
    rows = []
    for r in qs:
        rows.append({h: v for h, v in zip(headers, [
            r.outbound_number, _d(r.created_at.date() if r.created_at else ''),
            r.asset.asset_model.category.name if r.asset and r.asset.asset_model and r.asset.asset_model.category else '',
            r.asset.serial_number if r.asset else '', r.asset.asset_model.model_name if r.asset and r.asset.asset_model else '',
            r.asset.asset_model.manufacturer if r.asset and r.asset.asset_model else '',
            '', '', r.issued_by.name if r.issued_by else '', r.note or '',
        ])})
    return headers, rows


def _get_materials_sub(sub, q):
    if sub == 'inbound':
        from apps.materials.models import MaterialInbound
        qs = MaterialInbound.objects.select_related('material__category').order_by('-inbound_date')
        if q: qs = qs.filter(Q(material__name__icontains=q))
        headers = ['입고일', '분류', '자재코드', '자재명', '규격', '단위', '수량', '단가', '공급업체', '인계자', '인수자', '비고']
        rows = [{h: v for h, v in zip(headers, [
            _d(r.inbound_date), r.material.category.name if r.material and r.material.category else '',
            r.material.code if r.material else '', r.material.name if r.material else '',
            r.material.spec if r.material else '', r.material.unit if r.material else '',
            r.quantity, r.unit_price or '', r.supplier or '',
            r.handover_person or '', r.receiver_person or '', r.note or '',
        ])} for r in qs]
    elif sub == 'outbound':
        from apps.materials.models import MaterialOutbound
        qs = MaterialOutbound.objects.select_related('material__category', 'to_center').order_by('-outbound_date')
        if q: qs = qs.filter(Q(material__name__icontains=q))
        headers = ['출고일', '분류', '자재코드', '자재명', '규격', '단위', '수량', '출고지원청', '인계자', '인수자', '비고']
        rows = [{h: v for h, v in zip(headers, [
            _d(r.outbound_date), r.material.category.name if r.material and r.material.category else '',
            r.material.code if r.material else '', r.material.name if r.material else '',
            r.material.spec if r.material else '', r.material.unit if r.material else '',
            r.quantity, r.to_center.name if r.to_center else '',
            r.handover_person or '', r.receiver_person or '', r.note or '',
        ])} for r in qs]
    else:
        from apps.materials.models import WarehouseInventory
        qs = WarehouseInventory.objects.select_related('material__category').order_by('material__category__name')
        if q: qs = qs.filter(Q(material__name__icontains=q))
        headers = ['분류', '자재코드', '자재명', '규격', '단위', '현재고', '최소재고', '공급업체']
        rows = [{h: v for h, v in zip(headers, [
            r.material.category.name if r.material and r.material.category else '',
            r.material.code if r.material else '', r.material.name if r.material else '',
            r.material.spec if r.material else '', r.material.unit if r.material else '',
            r.quantity, r.material.min_stock if r.material else '', r.material.supplier if r.material else '',
        ])} for r in qs]
    return headers, rows


def _get_photos(q):
    from apps.photos.models import Photo
    qs = Photo.objects.select_related('school__support_center', 'work_type', 'taken_by', 'building', 'floor').order_by('-taken_at')
    if q: qs = qs.filter(Q(school__name__icontains=q) | Q(file_name__icontains=q))
    headers = ['촬영일시', '교육지원청', '학교명', '건물', '층', '작업명', '단계', '파일명', '파일크기(KB)', '촬영자', 'GPS위도', 'GPS경도', '관련장애번호']
    rows = [{h: v for h, v in zip(headers, [
        _dt(p.taken_at),
        p.school.support_center.name if p.school and p.school.support_center else '',
        p.school.name if p.school else '', p.building.name if p.building else '',
        p.floor.floor_name if p.floor else '',
        p.work_type.name if p.work_type else getattr(p, 'work_type_etc', ''),
        p.get_photo_stage_display() if hasattr(p, 'get_photo_stage_display') else '',
        p.file_name, round(p.file_size / 1024, 1) if p.file_size else 0,
        p.taken_by.name if p.taken_by else '',
        str(p.gps_lat) if p.gps_lat else '', str(p.gps_lng) if p.gps_lng else '',
        p.incident.incident_number if hasattr(p, 'incident') and p.incident else '',
    ])} for p in qs]
    return headers, rows


def _get_attendance(q):
    from apps.workforce.models import AttendanceLog
    qs = AttendanceLog.objects.select_related('worker__support_center').order_by('-work_date')
    if q: qs = qs.filter(Q(worker__name__icontains=q))
    headers = ['근무일', '인력명', '소속지원청', '출근시각', '퇴근시각', '근무시간(h)', '상태', '비고']
    rows = [{h: v for h, v in zip(headers, [
        _d(r.work_date), r.worker.name if r.worker else '',
        r.worker.support_center.name if r.worker and r.worker.support_center else '',
        timezone.localtime(r.check_in_at).strftime('%H:%M') if r.check_in_at else '',
        timezone.localtime(r.check_out_at).strftime('%H:%M') if r.check_out_at else '',
        r.get_work_hours() if hasattr(r, 'get_work_hours') else '',
        r.get_status_display() if hasattr(r, 'get_status_display') else r.status,
        r.note or '',
    ])} for r in qs]
    return headers, rows


def _get_progress(q):
    from apps.progress.models import SchoolInspection
    qs = SchoolInspection.objects.select_related('plan', 'school__support_center', 'school__school_type', 'assigned_worker').order_by('school__support_center__name', 'school__name')
    if q: qs = qs.filter(Q(school__name__icontains=q) | Q(assigned_worker__name__icontains=q))
    headers = ['점검계획', '교육지원청', '학제', '학교명', '예정일', '완료일', '상태', '담당기사', '비고']
    rows = [{h: v for h, v in zip(headers, [
        r.plan.name if r.plan else '', r.school.support_center.name if r.school and r.school.support_center else '',
        r.school.school_type.name if r.school and r.school.school_type else '',
        r.school.name if r.school else '', _d(r.scheduled_date),
        _d(r.completed_date) if hasattr(r, 'completed_date') else '',
        r.get_status_display() if hasattr(r, 'get_status_display') else r.status,
        r.assigned_worker.name if r.assigned_worker else '',
        r.note if hasattr(r, 'note') else '',
    ])} for r in qs]
    return headers, rows


def _get_education(q):
    from apps.education.models import EducationCompletion
    qs = EducationCompletion.objects.select_related('user', 'course__category').order_by('-completed_at')
    if q: qs = qs.filter(Q(user__name__icontains=q) | Q(course__title__icontains=q))
    headers = ['이수증번호', '성명', '교육과정', '분류', '이수점수', '이수일']
    rows = [{h: v for h, v in zip(headers, [
        c.certificate_no or '', c.user.name if c.user else '',
        c.course.title if c.course else '',
        c.course.category.name if c.course and c.course.category else '',
        c.score, _d(c.completed_at),
    ])} for c in qs]
    return headers, rows


def _get_users(q):
    from apps.accounts.models import User
    qs = User.objects.filter(role__in=['worker', 'resident']).select_related('support_center').order_by('name')
    if q: qs = qs.filter(Q(name__icontains=q) | Q(username__icontains=q))
    ROLE_KO = {'worker': '현장기사', 'resident': '상주인력', 'admin': '관리자', 'superadmin': '슈퍼관리자', 'customer': '고객'}
    headers = ['아이디', '이름', '역할', '전화번호', '이메일', '소속지원청', '자택주소', '서비스만료일', '활성']
    rows = [{h: v for h, v in zip(headers, [
        u.username, u.name, ROLE_KO.get(u.role, u.role), u.phone, u.email,
        u.support_center.name if u.support_center else '', u.home_address or '',
        _d(u.service_expiry), 'Y' if u.is_active else 'N',
    ])} for u in qs]
    return headers, rows


def _get_schools(q):
    from apps.schools.models import School
    qs = School.objects.select_related('support_center', 'school_type').order_by('support_center__name', 'name')
    if q: qs = qs.filter(Q(name__icontains=q) | Q(support_center__name__icontains=q))
    headers = ['교육지원청', '학제', '학교명', '주소', '전화번호', '위도', '경도']
    rows = [{h: v for h, v in zip(headers, [
        s.support_center.name if s.support_center else '', s.school_type.name if s.school_type else '',
        s.name, s.address or '', s.phone if hasattr(s, 'phone') else '',
        str(s.lat) if s.lat else '', str(s.lng) if s.lng else '',
    ])} for s in qs]
    return headers, rows


def _get_buildings(q):
    from apps.schools.models import SchoolRoom
    qs = SchoolRoom.objects.select_related('floor__building__school__support_center', 'floor__building__school__school_type', 'floor__building').order_by(
        'floor__building__school__support_center__name', 'floor__building__school__name', 'floor__building__name', 'floor__floor_num')
    if q: qs = qs.filter(Q(floor__building__school__name__icontains=q) | Q(name__icontains=q))
    headers = ['교육지원청', '학제', '학교명', '건물명', '층명', '층번호', '호실번호', '호실명', '유형', '면적(㎡)']
    rows = [{h: v for h, v in zip(headers, [
        r.floor.building.school.support_center.name if r.floor and r.floor.building and r.floor.building.school and r.floor.building.school.support_center else '',
        r.floor.building.school.school_type.name if r.floor and r.floor.building and r.floor.building.school and r.floor.building.school.school_type else '',
        r.floor.building.school.name if r.floor and r.floor.building and r.floor.building.school else '',
        r.floor.building.name if r.floor and r.floor.building else '',
        r.floor.floor_name if r.floor else '', r.floor.floor_num if r.floor else '',
        r.room_number or '', r.name or '',
        r.get_room_type_display() if hasattr(r, 'get_room_type_display') else (r.room_type or ''),
        float(r.area_m2) if r.area_m2 else '',
    ])} for r in qs[:10000]]
    return headers, rows


def _get_contacts(q):
    from apps.schools.models import SchoolContact
    qs = SchoolContact.objects.select_related('school__support_center', 'school__school_type').order_by('school__support_center__name', 'school__name')
    if q: qs = qs.filter(Q(name__icontains=q) | Q(school__name__icontains=q))
    headers = ['교육지원청', '학제', '학교명', '성명', '직위', '전화번호', '이메일', '주담당']
    rows = [{h: v for h, v in zip(headers, [
        c.school.support_center.name if c.school and c.school.support_center else '',
        c.school.school_type.name if c.school and c.school.school_type else '',
        c.school.name if c.school else '',
        c.name, c.position or '', c.phone or '', c.email or '',
        'Y' if c.is_primary else '',
    ])} for c in qs]
    return headers, rows


def _get_report_template(template_id, q):
    """Report.data JSON 펼치기 — 기존 reports export 로직 재활용"""
    from apps.reports.models import Report, ReportTemplate
    tmpl = ReportTemplate.objects.get(id=template_id)
    qs = Report.objects.filter(template=tmpl).select_related('school__support_center', 'school__school_type', 'created_by').order_by('-created_at')
    if q: qs = qs.filter(Q(title__icontains=q) | Q(school__name__icontains=q))

    sample = qs.first()
    if not sample or not sample.data:
        headers = ['보고서ID', '교육지원청', '학제', '학교명', '상태', '작성일']
        rows = [{h: v for h, v in zip(headers, [
            r.id, r.school.support_center.name if r.school and r.school.support_center else '',
            r.school.school_type.name if r.school and r.school.school_type else '',
            r.school.name if r.school else '', r.get_status_display(), _d(r.created_at),
        ])} for r in qs]
        return headers, rows

    # JSON 구조 분석
    data = sample.data
    array_key = None
    array_fields = []
    top_fields = []
    SKIP = {'signature', 'photo', 'data:image'}

    for k, v in data.items():
        if any(s in k for s in SKIP):
            continue
        if isinstance(v, list) and v and isinstance(v[0], dict):
            array_key = k
            for ak, av in v[0].items():
                if any(s in ak for s in SKIP):
                    continue
                if isinstance(av, dict):
                    for akk in av.keys():
                        array_fields.append((f'{ak}_{akk}', ak, akk))
                else:
                    array_fields.append((ak, None, None))
        elif isinstance(v, (str, int, float, bool)):
            top_fields.append(k)

    # 한글 라벨
    LBL = {
        'notes': '특이사항', 'doc_type': '문서유형', 'quantity': '수량',
        'install_date': '설치일', 'work_date': '작업일',
        'floor': '층', 'building': '건물', 'location': '설치장소',
        'model_name': '모델명', 'manufacturer': '제조사', 'serial_number': 'S/N',
        'asset_id': '자산번호', 'category': '분류',
        'network_type': '망종류', 'network_type_label': '망종류명',
        'prev_model': '교체전모델', 'prev_manufacturer': '교체전제조사',
        'cable_type': '케이블종류', 'cable_length': '길이(m)',
        'work_types': '작업유형', 'work_label': '작업내용',
        'port': '포트', 'room': '교실', 'rack': '랙',
    }
    def lbl(k):
        if k in LBL: return LBL[k]
        return ' '.join(LBL.get(p, p) for p in k.split('_'))

    headers = ['보고서ID', '교육지원청', '학제', '학교명', '상태']
    for tf in top_fields:
        headers.append(lbl(tf))
    for af_key, parent, child in array_fields:
        headers.append(lbl(af_key))
    headers.append('작성일')

    rows = []
    for report in qs:
        base = [
            report.id,
            report.school.support_center.name if report.school and report.school.support_center else '',
            report.school.school_type.name if report.school and report.school.school_type else '',
            report.school.name if report.school else '',
            report.get_status_display(),
        ]
        top_vals = [report.data.get(tf, '') for tf in top_fields]

        arr = report.data.get(array_key, []) if array_key else [None]
        if not arr:
            arr = [None]
        for item in arr:
            arr_vals = []
            for af_key, parent, child in array_fields:
                if item is None:
                    arr_vals.append('')
                elif parent and child:
                    arr_vals.append((item.get(parent) or {}).get(child, '') if isinstance(item.get(parent), dict) else '')
                else:
                    val = item.get(af_key, '')
                    if isinstance(val, list):
                        val = ', '.join(str(x) for x in val)
                    arr_vals.append(val)
            row_vals = base + top_vals + arr_vals + [_d(report.created_at)]
            rows.append({h: v for h, v in zip(headers, row_vals)})

    return headers, rows


# ══════════════════════════════════════════════════════
# 메인 API
# ══════════════════════════════════════════════════════

# 서류 ID → (이름, 데이터 함수) 매핑
_DOC_NAME = {
    'incidents': '장애처리보고서', 'sla': 'SLA성과', 'assets': '장비목록',
    'assets_inbound': '장비입고증', 'assets_outbound': '장비출고증',
    'mat_inbound': '자재입고내역', 'mat_outbound': '자재출고내역', 'mat_inventory': '자재현재고',
    'photos': '사진목록', 'attendance': '출퇴근기록', 'progress': '점검현황',
    'education': '교육이수증', 'users': '인력목록',
    'schools': '학교목록', 'buildings': '건물정보', 'contacts': '학교담당자',
}


def _get_doc_data(doc_id, q):
    """doc_id에 따라 (name, headers, rows) 반환"""
    if doc_id.startswith('report_'):
        tid = int(doc_id.replace('report_', ''))
        from apps.reports.models import ReportTemplate
        tmpl = ReportTemplate.objects.get(id=tid)
        headers, rows = _get_report_template(tid, q)
        return tmpl.name, headers, rows
    elif doc_id == 'incidents':
        headers, rows = _get_incidents(q)
        return '장애처리보고서', headers, rows
    elif doc_id == 'sla':
        headers, rows = _get_sla(q)
        return 'SLA성과', headers, rows
    elif doc_id == 'assets':
        headers, rows = _get_assets(q)
        return '장비목록', headers, rows
    else:
        name = _DOC_NAME.get(doc_id, doc_id)
        headers, rows = _get_model_simple(doc_id, q)
        return name, headers, rows


@login_required
@_admin_required
def doc_data(request, doc_id):
    """서류 상세 데이터 — JSON 응답"""
    page = max(1, int(request.GET.get('page', 1)))
    page_size = min(100, int(request.GET.get('page_size', 50)))
    q = request.GET.get('q', '').strip()

    try:
        name, headers, all_rows = _get_doc_data(doc_id, q)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

    columns = [{'key': h, 'label': h} for h in headers]
    result = _paginate(all_rows, page, page_size)
    result['name'] = name
    result['columns'] = columns
    return JsonResponse(result)


@login_required
@_admin_required
def doc_export(request, doc_id):
    """서류 데이터 Excel 다운로드 — 파일명은 서류명.xlsx"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    q = request.GET.get('q', '').strip()
    try:
        name, headers, all_rows = _get_doc_data(doc_id, q)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = name[:31]

    # 헤더
    ws.append(headers)
    hdr_fill = PatternFill('solid', fgColor='1F497D')
    hdr_font = Font(bold=True, color='FFFFFF', size=10)
    for cell in ws[1]:
        cell.font = hdr_font
        cell.fill = hdr_fill

    # 데이터
    for row in all_rows[:10000]:
        ws.append([row.get(h, '') for h in headers])

    # 컬럼 너비
    for ci, h in enumerate(headers, 1):
        max_len = len(h)
        for row in ws.iter_rows(min_row=2, min_col=ci, max_col=ci, max_row=min(50, ws.max_row)):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)[:30]))
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = min(max_len + 3, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f'{name}.xlsx'
    encoded = urllib.parse.quote(fname)
    resp = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f"attachment; filename*=UTF-8''{encoded}"
    return resp

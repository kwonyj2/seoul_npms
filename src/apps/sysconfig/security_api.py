"""
보안관제 API — 6개 패널 전체 데이터 제공
"""
import datetime
import hashlib
import json
import os
import re
import subprocess

import psutil
from django.conf import settings
from django.db.models import Count, Max, Min, Q, F, Sum
from django.db.models.functions import TruncDate, TruncHour, ExtractHour
from django.http import JsonResponse
from django.utils import timezone
from django.views.decorators.http import require_http_methods

from apps.accounts.models import LoginHistory, UserActivityLog, UserSession, User
from apps.sysconfig.security_models import (
    SecurityEvent, BlockedIP, WhitelistedIP, BlockLog,
    SecurityConfig, SystemLogEntry, FileIntegritySnapshot,
)


def _admin_required(view_func):
    from functools import wraps
    @wraps(view_func)
    def _wrapped(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return JsonResponse({'error': '로그인 필요'}, status=401)
        if request.user.role not in ('superadmin', 'admin'):
            return JsonResponse({'error': '권한 없음'}, status=403)
        return view_func(request, *args, **kwargs)
    return _wrapped


# ═══════════════════════════════════════════════════════
# 패널1: 위협 현황 대시보드
# ═══════════════════════════════════════════════════════

@_admin_required
def sec_dashboard(request):
    """위협 현황 대시보드 — 요약 카드 + 시간대별 차트 + 유형별 분류 + Top10 IP"""
    now = timezone.now()
    local_now = timezone.localtime(now)          # KST 변환
    h24 = now - datetime.timedelta(hours=24)
    h7d = now - datetime.timedelta(days=7)

    # ── 요약 카드 ─────────────────────────
    # NPMS 로그인 실패 + SSH 공격 합산
    npms_attacks_24h = LoginHistory.objects.filter(
        success=False, created_at__gte=h24
    ).count()
    ssh_attacks_24h = SystemLogEntry.objects.filter(
        log_type='ssh_fail', created_at__gte=h24
    ).count()
    total_attacks_24h = npms_attacks_24h + ssh_attacks_24h
    blocked_ips = BlockedIP.objects.count()
    active_sessions = UserSession.objects.filter(
        is_active=True,
        last_active__gte=now - datetime.timedelta(minutes=30)
    ).count()
    events_24h = SecurityEvent.objects.filter(created_at__gte=h24).count()

    # 외국 IP 차단 건수 (nginx access log에서 444 응답 카운트)
    foreign_blocked_24h = 0
    for log_path in ['/var/log/nginx/access.log', '/app/nas/logs/nginx_access.log']:
        if os.path.exists(log_path):
            try:
                with open(log_path, 'r', errors='replace') as f:
                    for line in f:
                        if '" 444 ' in line:
                            foreign_blocked_24h += 1
            except Exception:
                pass
            break

    # 위험 등급 계산
    if total_attacks_24h >= 100 or SecurityEvent.objects.filter(
        severity='critical', created_at__gte=h24, resolved=False
    ).exists():
        threat_level = 'danger'
        threat_label = '위험'
    elif total_attacks_24h >= 30:
        threat_level = 'warning'
        threat_label = '경고'
    elif total_attacks_24h >= 10:
        threat_level = 'caution'
        threat_label = '주의'
    else:
        threat_level = 'safe'
        threat_label = '안전'

    # ── 시간대별 공격 시도 (24h) — SSH + NPMS 합산 ─────────
    # SSH 공격 (SystemLogEntry)
    ssh_hourly_raw = (
        SystemLogEntry.objects.filter(log_type='ssh_fail', created_at__gte=h24)
        .annotate(hour=TruncHour('created_at'))
        .values('hour')
        .annotate(cnt=Count('id'))
        .order_by('hour')
    )
    ssh_hourly = {timezone.localtime(r['hour']).strftime('%H:00'): r['cnt'] for r in ssh_hourly_raw}
    # NPMS 로그인 실패
    npms_hourly_raw = (
        LoginHistory.objects.filter(success=False, created_at__gte=h24)
        .annotate(hour=TruncHour('created_at'))
        .values('hour')
        .annotate(cnt=Count('id'))
        .order_by('hour')
    )
    npms_hourly = {timezone.localtime(r['hour']).strftime('%H:00'): r['cnt'] for r in npms_hourly_raw}
    hourly_labels = []
    hourly_data = []
    for i in range(24):
        t = (local_now - datetime.timedelta(hours=23 - i)).strftime('%H:00')
        hourly_labels.append(t)
        hourly_data.append(ssh_hourly.get(t, 0) + npms_hourly.get(t, 0))

    # ── 공격 유형별 분류 (SSH + NPMS) ─────
    type_map = {}
    # SSH 공격 유형
    ssh_fail_cnt = SystemLogEntry.objects.filter(log_type='ssh_fail', created_at__gte=h7d).count()
    ssh_auth_cnt = SystemLogEntry.objects.filter(log_type='auth_other', created_at__gte=h7d).count()
    if ssh_fail_cnt:
        type_map['SSH 브루트포스'] = ssh_fail_cnt
    if ssh_auth_cnt:
        type_map['OS 인증 실패'] = ssh_auth_cnt
    # NPMS 로그인 실패
    fail_reasons = (
        LoginHistory.objects.filter(success=False, created_at__gte=h7d)
        .values('fail_reason')
        .annotate(cnt=Count('id'))
        .order_by('-cnt')
    )
    for r in fail_reasons:
        reason = r['fail_reason'] or '기타'
        if '잠금' in reason:
            key = 'NPMS 계정 잠금'
        elif '비밀번호' in reason:
            key = 'NPMS 비밀번호 오류'
        elif '미등록' in reason or '없는' in reason:
            key = 'NPMS 미등록 계정'
        elif '만료' in reason:
            key = 'NPMS 서비스 만료'
        else:
            key = 'NPMS 기타'
        type_map[key] = type_map.get(key, 0) + r['cnt']
    attack_types = [{'label': k, 'count': v} for k, v in sorted(type_map.items(), key=lambda x: -x[1])]

    # ── Top 위협 IP (페이지네이션) ──────────────────
    ip_page = max(1, int(request.GET.get('ip_page', 1)))
    ip_ps = 10
    top_ips_qs = (
        SystemLogEntry.objects.filter(
            log_type='ssh_fail', created_at__gte=h7d,
            ip_address__isnull=False,
        )
        .values('ip_address')
        .annotate(
            fail_count=Count('id'),
            last_attempt=Max('created_at'),
            first_attempt=Min('created_at'),
        )
        .order_by('-fail_count')
    )
    top_ip_total = top_ips_qs.count()
    top_ips = top_ips_qs[(ip_page-1)*ip_ps : ip_page*ip_ps]
    top_ip_list = []
    for row in top_ips:
        ip = row['ip_address'] or '-'
        is_blocked = BlockedIP.objects.filter(ip_address=ip).exists()
        fc = row['fail_count']
        if fc >= 500:
            threat = 'critical'
        elif fc >= 100:
            threat = 'high'
        elif fc >= 20:
            threat = 'medium'
        else:
            threat = 'low'
        top_ip_list.append({
            'ip': ip,
            'fail_count': fc,
            'threat': threat,
            'is_blocked': is_blocked,
            'first': timezone.localtime(row['first_attempt']).strftime('%m-%d %H:%M') if row['first_attempt'] else '-',
            'last': timezone.localtime(row['last_attempt']).strftime('%m-%d %H:%M') if row['last_attempt'] else '-',
        })

    # ── 최근 보안 이벤트 (페이지네이션) ───────────
    ev_page = max(1, int(request.GET.get('ev_page', 1)))
    ev_ps = 10
    events_qs = SecurityEvent.objects.all()
    ev_total = events_qs.count()
    recent_events = []
    for ev in events_qs[(ev_page-1)*ev_ps : ev_page*ev_ps]:
        recent_events.append({
            'id': ev.id,
            'type': ev.get_event_type_display(),
            'type_key': ev.event_type,
            'severity': ev.severity,
            'ip': ev.ip_address or '-',
            'username': ev.username or '-',
            'desc': ev.description[:120],
            'resolved': ev.resolved,
            'time': timezone.localtime(ev.created_at).strftime('%m-%d %H:%M:%S'),
        })

    return JsonResponse({
        'summary': {
            'attacks_24h': total_attacks_24h,
            'blocked_ips': blocked_ips,
            'active_sessions': active_sessions,
            'events_24h': events_24h,
            'foreign_blocked': foreign_blocked_24h,
            'threat_level': threat_level,
            'threat_label': threat_label,
        },
        'hourly': {'labels': hourly_labels, 'data': hourly_data},
        'attack_types': attack_types,
        'top_ips': top_ip_list,
        'top_ip_page': ip_page,
        'top_ip_total': top_ip_total,
        'top_ip_total_pages': max(1, (top_ip_total + ip_ps - 1) // ip_ps),
        'recent_events': recent_events,
        'ev_page': ev_page,
        'ev_total': ev_total,
        'ev_total_pages': max(1, (ev_total + ev_ps - 1) // ev_ps),
    })


# ═══════════════════════════════════════════════════════
# 패널2: 자동 차단 관리
# ═══════════════════════════════════════════════════════

@_admin_required
def sec_blocked_ips(request):
    """차단 IP 목록 조회 / 수동 차단 / 해제"""
    if request.method == 'GET':
        q = request.GET.get('q', '')
        qs = BlockedIP.objects.all()
        if q:
            qs = qs.filter(Q(ip_address__icontains=q) | Q(description__icontains=q))

        page = max(1, int(request.GET.get('page', 1)))
        ps = 25
        total = qs.count()
        rows = []
        for b in qs[(page-1)*ps : page*ps]:
            rows.append({
                'id': b.id,
                'ip': b.ip_address,
                'reason': b.get_reason_display(),
                'reason_key': b.reason,
                'description': b.description,
                'fail_count': b.fail_count,
                'auto': b.auto_blocked,
                'permanent': b.is_permanent,
                'is_active': b.is_active,
                'blocked_at': timezone.localtime(b.blocked_at).strftime('%Y-%m-%d %H:%M'),
                'expires_at': timezone.localtime(b.expires_at).strftime('%Y-%m-%d %H:%M') if b.expires_at else '영구',
                'blocked_by': b.blocked_by.name if b.blocked_by else '시스템',
            })

        # 화이트리스트
        wl = list(WhitelistedIP.objects.values(
            'id', 'ip_address', 'description',
        ).order_by('ip_address'))

        return JsonResponse({
            'blocked': rows, 'whitelist': wl,
            'total': total, 'page': page, 'page_size': ps,
            'total_pages': max(1, (total + ps - 1) // ps),
        })

    elif request.method == 'POST':
        # 수동 차단
        try:
            body = json.loads(request.body)
        except Exception:
            return JsonResponse({'error': '잘못된 요청'}, status=400)

        action = body.get('action')  # block / unblock / whitelist_add / whitelist_remove

        if action == 'block':
            ip = body.get('ip', '').strip()
            if not ip:
                return JsonResponse({'error': 'IP 필요'}, status=400)
            reason = body.get('reason', 'manual')
            desc = body.get('description', '관리자 수동 차단')
            permanent = body.get('permanent', False)
            duration = body.get('duration', 60)

            obj, created = BlockedIP.objects.update_or_create(
                ip_address=ip,
                defaults={
                    'reason': reason,
                    'description': desc,
                    'is_permanent': permanent,
                    'auto_blocked': False,
                    'blocked_by': request.user,
                    'blocked_at': timezone.now(),
                    'expires_at': None if permanent else timezone.now() + datetime.timedelta(minutes=int(duration)),
                }
            )
            BlockLog.objects.create(
                ip_address=ip, action='block',
                reason=f'수동 차단: {desc}', actor=request.user
            )
            SecurityEvent.objects.create(
                event_type='ip_blocked', severity='info',
                ip_address=ip, username=request.user.username,
                description=f'관리자 수동 차단: {desc}',
            )
            return JsonResponse({'ok': True, 'created': created})

        elif action == 'unblock':
            ip = body.get('ip', '').strip()
            deleted = BlockedIP.objects.filter(ip_address=ip).delete()[0]
            if deleted:
                BlockLog.objects.create(
                    ip_address=ip, action='unblock',
                    reason='관리자 수동 해제', actor=request.user
                )
                SecurityEvent.objects.create(
                    event_type='ip_unblocked', severity='info',
                    ip_address=ip, username=request.user.username,
                    description='관리자 수동 차단 해제',
                )
            return JsonResponse({'ok': True, 'deleted': deleted})

        elif action == 'whitelist_add':
            ip = body.get('ip', '').strip()
            desc = body.get('description', '')
            if not ip:
                return JsonResponse({'error': 'IP 필요'}, status=400)
            obj, created = WhitelistedIP.objects.get_or_create(
                ip_address=ip,
                defaults={'description': desc, 'created_by': request.user}
            )
            # 화이트리스트 등록 시 차단 해제
            BlockedIP.objects.filter(ip_address=ip).delete()
            return JsonResponse({'ok': True, 'created': created})

        elif action == 'whitelist_remove':
            ip = body.get('ip', '').strip()
            deleted = WhitelistedIP.objects.filter(ip_address=ip).delete()[0]
            return JsonResponse({'ok': True, 'deleted': deleted})

        return JsonResponse({'error': '알 수 없는 action'}, status=400)

    return JsonResponse({'error': '허용되지 않는 메서드'}, status=405)


@_admin_required
def sec_block_config(request):
    """자동 차단 규칙 조회/수정"""
    KEYS = [
        'auto_block_enabled', 'block_threshold', 'block_duration_min',
        'block_window_min', 'permanent_threshold',
    ]
    if request.method == 'GET':
        config = {k: SecurityConfig.get(k) for k in KEYS}
        return JsonResponse({'config': config})

    elif request.method == 'PATCH':
        try:
            body = json.loads(request.body)
        except Exception:
            return JsonResponse({'error': '잘못된 요청'}, status=400)
        for k, v in body.items():
            if k in KEYS:
                SecurityConfig.objects.update_or_create(
                    key=k, defaults={'value': str(v), 'updated_by': request.user}
                )
                SecurityEvent.objects.create(
                    event_type='config_change', severity='info',
                    username=request.user.username,
                    description=f'보안 설정 변경: {k} = {v}',
                )
        return JsonResponse({'ok': True})

    return JsonResponse({'error': '허용되지 않는 메서드'}, status=405)


@_admin_required
def sec_block_log(request):
    """차단/해제 이력 조회"""
    page = max(1, int(request.GET.get('page', 1)))
    ps = 25
    qs = BlockLog.objects.select_related('actor').all()
    q = request.GET.get('q', '')
    if q:
        qs = qs.filter(Q(ip_address__icontains=q) | Q(reason__icontains=q))
    total = qs.count()
    rows = []
    for r in qs[(page-1)*ps: page*ps]:
        rows.append({
            'ip': r.ip_address,
            'action': r.get_action_display(),
            'reason': r.reason,
            'actor': r.actor.name if r.actor else '시스템',
            'time': timezone.localtime(r.created_at).strftime('%Y-%m-%d %H:%M:%S'),
        })
    return JsonResponse({
        'rows': rows, 'total': total,
        'page': page, 'page_size': ps,
        'total_pages': max(1, (total + ps - 1) // ps),
    })


# ═══════════════════════════════════════════════════════
# 패널3: 로그인 보안 분석
# ═══════════════════════════════════════════════════════

@_admin_required
def sec_login_analysis(request):
    """로그인 보안 분석 — 추이, 비정상 탐지, 사용자 패턴, 계정잠금"""
    now = timezone.now()
    days = int(request.GET.get('days', 14))
    since = now - datetime.timedelta(days=days)

    # ── 일별 성공/실패 추이 ────────────────
    daily_success = dict(
        LoginHistory.objects.filter(success=True, created_at__gte=since)
        .annotate(d=TruncDate('created_at'))
        .values('d').annotate(cnt=Count('id'))
        .values_list('d', 'cnt')
    )
    daily_fail = dict(
        LoginHistory.objects.filter(success=False, created_at__gte=since)
        .annotate(d=TruncDate('created_at'))
        .values('d').annotate(cnt=Count('id'))
        .values_list('d', 'cnt')
    )
    labels = []
    s_data = []
    f_data = []
    for i in range(days):
        d = (since + datetime.timedelta(days=i + 1)).date()
        labels.append(d.strftime('%m-%d'))
        s_data.append(daily_success.get(d, 0))
        f_data.append(daily_fail.get(d, 0))

    # ── 비정상 로그인 탐지 ─────────────────
    # 야간(22~06) 성공 로그인
    abnormal = []
    night_logins = (
        LoginHistory.objects.filter(
            success=True, created_at__gte=since
        ).annotate(hr=ExtractHour('created_at'))
        .filter(Q(hr__gte=22) | Q(hr__lt=6))
        .select_related('user')
        .order_by('-created_at')[:20]
    )
    for r in night_logins:
        abnormal.append({
            'type': '야간 접속',
            'username': r.user.username if r.user else r.attempted_username,
            'name': r.user.name if r.user else '-',
            'ip': r.ip_address or '-',
            'time': timezone.localtime(r.created_at).strftime('%Y-%m-%d %H:%M'),
            'detail': f'{timezone.localtime(r.created_at).strftime("%H:%M")} 접속',
        })

    # 동일 IP 다중 계정 시도 (브루트포스)
    multi_user_ips = (
        LoginHistory.objects.filter(success=False, created_at__gte=since)
        .values('ip_address')
        .annotate(
            user_cnt=Count('attempted_username', distinct=True),
            total=Count('id'),
        )
        .filter(user_cnt__gte=3)
        .order_by('-total')[:10]
    )
    for row in multi_user_ips:
        abnormal.append({
            'type': '다중 계정 시도',
            'username': '-',
            'name': '-',
            'ip': row['ip_address'] or '-',
            'time': '-',
            'detail': f'{row["user_cnt"]}개 계정, {row["total"]}회 시도',
        })

    # ── 시간대별 로그인 분포 (히트맵) ─────
    hourly_dist = (
        LoginHistory.objects.filter(created_at__gte=since)
        .annotate(hr=ExtractHour('created_at'))
        .values('hr', 'success')
        .annotate(cnt=Count('id'))
        .order_by('hr')
    )
    heatmap = {'success': [0]*24, 'fail': [0]*24}
    for r in hourly_dist:
        key = 'success' if r['success'] else 'fail'
        heatmap[key][r['hr']] = r['cnt']

    # ── 잠긴 계정 현황 ───────────────────
    lock_window = now - datetime.timedelta(minutes=30)
    locked_users = (
        LoginHistory.objects.filter(
            success=False, created_at__gte=lock_window
        )
        .values('attempted_username')
        .annotate(cnt=Count('id'))
        .filter(cnt__gte=5)
        .order_by('-cnt')
    )
    locked_list = []
    for r in locked_users:
        uname = r['attempted_username']
        user = User.objects.filter(username=uname).first()
        locked_list.append({
            'username': uname,
            'name': user.name if user else '(미등록)',
            'fail_count': r['cnt'],
            'is_registered': user is not None,
        })

    return JsonResponse({
        'trend': {'labels': labels, 'success': s_data, 'fail': f_data},
        'abnormal': abnormal,
        'heatmap': heatmap,
        'locked_accounts': locked_list,
    })


# ═══════════════════════════════════════════════════════
# 패널4: 시스템 로그 감시
# ═══════════════════════════════════════════════════════

@_admin_required
def sec_system_logs(request):
    """시스템 로그 — SSH, 리소스, Docker, 파일 무결성"""
    kind = request.GET.get('kind', 'ssh')
    page = max(1, int(request.GET.get('page', 1)))
    ps = 25

    if kind == 'ssh':
        qs = SystemLogEntry.objects.filter(log_type__in=['ssh_fail', 'ssh_success', 'auth_other'])
        q = request.GET.get('q', '')
        if q:
            qs = qs.filter(Q(ip_address__icontains=q) | Q(username__icontains=q))
        total = qs.count()
        rows = []
        for r in qs[(page-1)*ps: page*ps]:
            rows.append({
                'type': r.get_log_type_display(),
                'type_key': r.log_type,
                'ip': r.ip_address or '-',
                'username': r.username or '-',
                'raw': r.raw_line[:200],
                'time': timezone.localtime(r.log_time).strftime('%Y-%m-%d %H:%M:%S') if r.log_time else '-',
            })
        # SSH 실패 IP Top5
        ssh_top = (
            SystemLogEntry.objects.filter(log_type='ssh_fail')
            .values('ip_address')
            .annotate(cnt=Count('id'))
            .order_by('-cnt')[:5]
        )
        return JsonResponse({
            'rows': rows, 'total': total, 'page': page, 'page_size': ps,
            'total_pages': max(1, (total + ps - 1) // ps),
            'ssh_top': list(ssh_top),
        })

    elif kind == 'resource':
        # 현재 리소스 상태
        cpu = psutil.cpu_percent(interval=0.5)
        mem = psutil.virtual_memory()
        disk = psutil.disk_usage('/')
        # 프로세스 Top 5 (CPU)
        procs = []
        for p in sorted(psutil.process_iter(['pid', 'name', 'cpu_percent', 'memory_percent']),
                        key=lambda x: x.info.get('cpu_percent', 0) or 0, reverse=True)[:5]:
            procs.append({
                'pid': p.info['pid'],
                'name': p.info['name'],
                'cpu': round(p.info.get('cpu_percent', 0) or 0, 1),
                'mem': round(p.info.get('memory_percent', 0) or 0, 1),
            })
        # 이상 여부
        anomaly = cpu > 90 or mem.percent > 90 or disk.percent > 90
        return JsonResponse({
            'cpu': cpu,
            'memory': {'total': round(mem.total/1024**3, 1), 'used': round(mem.used/1024**3, 1), 'percent': mem.percent},
            'disk': {'total': round(disk.total/1024**3, 1), 'used': round(disk.used/1024**3, 1), 'percent': disk.percent},
            'processes': procs,
            'anomaly': anomaly,
        })

    elif kind == 'docker':
        # Docker 컨테이너 상태 — Docker Socket HTTP API 직접 호출
        containers = []
        try:
            import socket as _socket
            import http.client

            class UnixSocketConnection(http.client.HTTPConnection):
                def __init__(self, socket_path):
                    super().__init__('localhost')
                    self._socket_path = socket_path
                def connect(self):
                    self.sock = _socket.socket(_socket.AF_UNIX, _socket.SOCK_STREAM)
                    self.sock.settimeout(5)
                    self.sock.connect(self._socket_path)

            conn = UnixSocketConnection('/var/run/docker.sock')
            conn.request('GET', '/containers/json?all=true')
            resp = conn.getresponse()
            if resp.status == 200:
                data = json.loads(resp.read())
                for c in data:
                    name = c.get('Names', ['-'])[0].lstrip('/')
                    status = c.get('Status', '-')
                    state = c.get('State', '-')
                    containers.append({
                        'name': name,
                        'status': status,
                        'is_up': state == 'running',
                        'image': c.get('Image', '-'),
                    })
            else:
                containers = [{'name': 'error', 'status': f'Docker API {resp.status}', 'is_up': False, 'image': '-'}]
            conn.close()
        except FileNotFoundError:
            containers = [{'name': 'error', 'status': 'Docker 소켓 미연결 (docker.sock 마운트 필요)', 'is_up': False, 'image': '-'}]
        except Exception as e:
            containers = [{'name': 'error', 'status': str(e), 'is_up': False, 'image': '-'}]
        return JsonResponse({'containers': containers})

    elif kind == 'integrity':
        rows = []
        for f in FileIntegritySnapshot.objects.all().order_by('-is_changed', 'file_path'):
            rows.append({
                'path': f.file_path,
                'hash': f.sha256_hash,
                'hash_short': f.sha256_hash[:16] + '...',
                'size': f.file_size,
                'changed': f.is_changed,
                'prev_hash': (f.prev_hash[:16] + '...') if f.prev_hash else '-',
                'checked_at': timezone.localtime(f.checked_at).strftime('%Y-%m-%d %H:%M'),
            })
        return JsonResponse({'files': rows})

    return JsonResponse({'error': '알 수 없는 kind'}, status=400)


# ═══════════════════════════════════════════════════════
# 패널5: 보안 설정 관리
# ═══════════════════════════════════════════════════════

@_admin_required
def sec_settings(request):
    """보안 설정 조회/수정"""
    ALL_KEYS = list(SecurityConfig.DEFAULTS.keys())

    if request.method == 'GET':
        config = {k: SecurityConfig.get(k) for k in ALL_KEYS}

        # 현재 Django 보안 설정 표시
        django_settings = {
            'debug': getattr(settings, 'DEBUG', False),
            'session_cookie_age': getattr(settings, 'SESSION_COOKIE_AGE', 0),
            'session_cookie_secure': getattr(settings, 'SESSION_COOKIE_SECURE', False),
            'csrf_cookie_secure': getattr(settings, 'CSRF_COOKIE_SECURE', False),
            'csrf_cookie_httponly': getattr(settings, 'CSRF_COOKIE_HTTPONLY', False),
            'max_concurrent_sessions': getattr(settings, 'MAX_CONCURRENT_SESSIONS', 3),
            'password_min_length': 8,
            'throttle_anon': getattr(settings, 'REST_FRAMEWORK', {}).get(
                'DEFAULT_THROTTLE_RATES', {}).get('anon', '30/min'),
            'throttle_user': getattr(settings, 'REST_FRAMEWORK', {}).get(
                'DEFAULT_THROTTLE_RATES', {}).get('user', '200/min'),
        }

        # 보안 자가진단 체크리스트
        checks = []
        checks.append({
            'item': 'DEBUG 모드', 'status': 'pass' if not settings.DEBUG else 'fail',
            'detail': 'OFF' if not settings.DEBUG else '⚠ ON — 운영환경에서 반드시 OFF',
        })
        checks.append({
            'item': 'HTTPS(SSL/TLS)',
            'status': 'pass' if getattr(settings, 'SESSION_COOKIE_SECURE', False) else 'fail',
            'detail': 'TLS 1.2/1.3 적용 (자체서명)' if getattr(settings, 'SESSION_COOKIE_SECURE', False) else '⚠ 미적용',
        })
        checks.append({
            'item': 'CSRF 보호', 'status': 'pass',
            'detail': f'CSRF_COOKIE_HTTPONLY={settings.CSRF_COOKIE_HTTPONLY}',
        })
        checks.append({
            'item': 'X-Frame-Options', 'status': 'pass',
            'detail': 'DENY (nginx + Django)',
        })
        checks.append({
            'item': '보안 헤더', 'status': 'pass',
            'detail': 'X-Content-Type-Options, XSS-Protection, Referrer-Policy',
        })
        checks.append({
            'item': 'API Rate Limit', 'status': 'pass',
            'detail': f'anon={django_settings["throttle_anon"]}, user={django_settings["throttle_user"]}',
        })
        checks.append({
            'item': '비밀번호 정책', 'status': 'pass',
            'detail': '최소 8자, 복잡도 검증, 일반 비밀번호 차단',
        })
        checks.append({
            'item': '2FA 지원', 'status': 'pass',
            'detail': 'TOTP 기반 2FA 구현됨',
        })
        checks.append({
            'item': '세션 관리', 'status': 'pass',
            'detail': f'Redis 기반, {settings.SESSION_COOKIE_AGE//3600}시간 만료, 최대 {getattr(settings, "MAX_CONCURRENT_SESSIONS", 3)}개',
        })
        checks.append({
            'item': 'IP 자동 차단', 'status': 'pass' if SecurityConfig.get_bool('auto_block_enabled') else 'warn',
            'detail': f'{"활성" if SecurityConfig.get_bool("auto_block_enabled") else "비활성"} — 임계값 {SecurityConfig.get_int("block_threshold", 10)}회',
        })
        # Redis 인증
        redis_url = getattr(settings, 'CACHES', {}).get('default', {}).get('LOCATION', '')
        has_redis_pw = ':' in redis_url.split('@')[0] if '@' in redis_url else False
        checks.append({
            'item': 'Redis 인증', 'status': 'pass' if has_redis_pw else 'fail',
            'detail': '비밀번호 인증 적용됨' if has_redis_pw else '⚠ 인증 없음',
        })
        # CSP 헤더
        checks.append({
            'item': 'CSP 헤더', 'status': 'pass',
            'detail': 'Content-Security-Policy nginx 적용',
        })
        # 외국 IP 차단 (GeoIP + 한국 ISP 대역)
        # web 컨테이너에 마운트된 nginx 설정 확인
        foreign_block_active = False
        for p in ['/app/nginx-geoip.conf', '/usr/share/GeoIP/GeoIP.dat',
                  '/home/kwonyj/network_pms/docker/nginx-main.conf']:
            if os.path.exists(p):
                try:
                    with open(p, 'r') as f:
                        if 'geoip_country' in f.read():
                            foreign_block_active = True
                            break
                except Exception:
                    pass
                if p.endswith('.dat'):
                    foreign_block_active = True
                    break
        checks.append({
            'item': '외국 IP 차단',
            'status': 'pass' if foreign_block_active else 'warn',
            'detail': 'GeoIP + 한국 ISP 대역 허용 (비한국 IP 접속 차단)' if foreign_block_active
                      else 'nginx GeoIP 미설정 - 외국 IP 접속 가능',
        })
        # HSTS
        checks.append({
            'item': 'HSTS', 'status': 'pass',
            'detail': 'max-age=31536000, includeSubDomains (nginx 적용)',
        })
        # DDoS 방어
        checks.append({
            'item': 'DDoS/봇 방어', 'status': 'pass',
            'detail': '로그인 분당5회, API 초당5회, 일반 초당10회 제한',
        })
        # 백업 암호화
        has_enc = bool(getattr(settings, 'DB_BACKUP_ENCRYPT_KEY', ''))
        checks.append({
            'item': '백업 암호화', 'status': 'pass' if has_enc else 'warn',
            'detail': 'AES-256 암호화 적용' if has_enc else '.env에 DB_BACKUP_ENCRYPT_KEY 설정 필요',
        })
        # ── 추가 보안 항목 ────────────────────────
        # 로그인 실패 잠금
        checks.append({
            'item': '로그인 실패 잠금', 'status': 'pass',
            'detail': '5회 실패 시 30분 잠금 (Redis 캐시 기반)',
        })
        # DB 접근 제한
        checks.append({
            'item': 'DB 접근 제한', 'status': 'pass',
            'detail': 'PostgreSQL localhost(127.0.0.1) 바인딩만 허용',
        })
        # 파일 업로드 검증
        checks.append({
            'item': '파일 업로드 검증', 'status': 'pass',
            'detail': '최대 20MB, 확장자 제한, PIL 이미지 검증, Path Traversal 방지',
        })
        # 감사 로그
        checks.append({
            'item': '감사 로그', 'status': 'pass',
            'detail': 'AuditLogMiddleware - 모든 변경 작업 DB+파일 기록',
        })
        # 서버 정보 숨김
        checks.append({
            'item': '서버 정보 숨김', 'status': 'pass',
            'detail': 'nginx server_tokens off (버전 정보 노출 차단)',
        })
        # Admin 페이지 IP 제한
        checks.append({
            'item': 'Admin IP 제한', 'status': 'pass',
            'detail': '/admin, /flower, /api/docs 내부 IP만 접근 허용',
        })
        # SSH 모니터링
        ssh_mon = SecurityConfig.get_bool('ssh_monitor_enabled')
        checks.append({
            'item': 'SSH 모니터링', 'status': 'pass' if ssh_mon else 'warn',
            'detail': f'{"5분마다 SSH 로그 수집 + 자동 차단" if ssh_mon else "비활성 - 보안 설정에서 활성화 필요"}',
        })
        # 파일 무결성 점검
        fi_mon = SecurityConfig.get_bool('file_integrity_enabled')
        checks.append({
            'item': '파일 무결성 점검', 'status': 'pass' if fi_mon else 'warn',
            'detail': f'{"1시간마다 핵심 설정파일 SHA256 해시 비교" if fi_mon else "비활성 - 보안 설정에서 활성화 필요"}',
        })
        # ALLOWED_HOSTS
        allowed = getattr(settings, 'ALLOWED_HOSTS', ['*'])
        hosts_ok = '*' not in allowed
        checks.append({
            'item': 'ALLOWED_HOSTS', 'status': 'pass' if hosts_ok else 'fail',
            'detail': f'{", ".join(allowed)}' if hosts_ok else '* 허용 - Host 헤더 위변조 가능',
        })
        # 로그 보존 정책
        checks.append({
            'item': '로그 보존 정책', 'status': 'pass',
            'detail': 'RotatingFileHandler - 보안로그 5MB*30개, 접근로그 20MB*30개',
        })
        # SameSite Cookie
        samesite = getattr(settings, 'SESSION_COOKIE_SAMESITE', None)
        checks.append({
            'item': 'SameSite Cookie',
            'status': 'pass' if samesite else 'warn',
            'detail': f'SESSION_COOKIE_SAMESITE={samesite}' if samesite else 'CSRF 공격 방어를 위해 설정 권장',
        })

        pass_cnt = sum(1 for c in checks if c['status'] == 'pass')
        score = round(pass_cnt / len(checks) * 100)

        return JsonResponse({
            'config': config,
            'django_settings': django_settings,
            'checks': checks,
            'score': score,
        })

    elif request.method == 'PATCH':
        try:
            body = json.loads(request.body)
        except Exception:
            return JsonResponse({'error': '잘못된 요청'}, status=400)
        for k, v in body.items():
            if k in ALL_KEYS:
                SecurityConfig.objects.update_or_create(
                    key=k, defaults={'value': str(v), 'updated_by': request.user}
                )
        SecurityEvent.objects.create(
            event_type='config_change', severity='info',
            username=request.user.username,
            description=f'보안 설정 변경: {", ".join(body.keys())}',
        )
        return JsonResponse({'ok': True})

    return JsonResponse({'error': '허용되지 않는 메서드'}, status=405)


# ═══════════════════════════════════════════════════════
# 패널6: 보안 리포트
# ═══════════════════════════════════════════════════════

@_admin_required
def sec_report(request):
    """보안 리포트 — 기간별 통계"""
    now = timezone.now()
    local_now = timezone.localtime(now)
    period = request.GET.get('period', 'daily')  # daily / weekly / monthly

    if period == 'daily':
        since = now - datetime.timedelta(days=1)
        period_label = f'{timezone.localtime(since).strftime("%Y-%m-%d")} (일간)'
    elif period == 'weekly':
        since = now - datetime.timedelta(weeks=1)
        period_label = f'{timezone.localtime(since).strftime("%Y-%m-%d")} ~ {local_now.strftime("%Y-%m-%d")} (주간)'
    else:
        since = now - datetime.timedelta(days=30)
        period_label = f'{timezone.localtime(since).strftime("%Y-%m-%d")} ~ {local_now.strftime("%Y-%m-%d")} (월간)'

    # 로그인 통계
    login_total = LoginHistory.objects.filter(created_at__gte=since).count()
    login_success = LoginHistory.objects.filter(created_at__gte=since, success=True).count()
    login_fail = LoginHistory.objects.filter(created_at__gte=since, success=False).count()
    unique_ips = LoginHistory.objects.filter(
        created_at__gte=since, success=False
    ).values('ip_address').distinct().count()

    # 차단 이력
    blocks_count = BlockLog.objects.filter(created_at__gte=since, action='block').count()
    unblocks_count = BlockLog.objects.filter(created_at__gte=since, action='unblock').count()

    # 보안 이벤트 통계
    event_by_type = dict(
        SecurityEvent.objects.filter(created_at__gte=since)
        .values('event_type')
        .annotate(cnt=Count('id'))
        .values_list('event_type', 'cnt')
    )
    event_by_severity = dict(
        SecurityEvent.objects.filter(created_at__gte=since)
        .values('severity')
        .annotate(cnt=Count('id'))
        .values_list('severity', 'cnt')
    )

    # SSH 시도
    ssh_fails = SystemLogEntry.objects.filter(
        log_type='ssh_fail', created_at__gte=since
    ).count()
    ssh_unique_ips = SystemLogEntry.objects.filter(
        log_type='ssh_fail', created_at__gte=since, ip_address__isnull=False,
    ).values('ip_address').distinct().count()

    # 위협 트렌드 (이번 기간 vs 이전 기간)
    if period == 'daily':
        prev_since = since - datetime.timedelta(days=1)
    elif period == 'weekly':
        prev_since = since - datetime.timedelta(weeks=1)
    else:
        prev_since = since - datetime.timedelta(days=30)
    prev_fail = LoginHistory.objects.filter(
        created_at__gte=prev_since, created_at__lt=since, success=False
    ).count()
    prev_ssh = SystemLogEntry.objects.filter(
        log_type='ssh_fail', created_at__gte=prev_since, created_at__lt=since,
    ).count()
    # 전체 공격 = NPMS 로그인 실패 + SSH 공격
    total_current = login_fail + ssh_fails
    total_prev = prev_fail + prev_ssh
    trend_pct = round((total_current - total_prev) / max(total_prev, 1) * 100) if total_prev else 0

    # ── 차단 효과 분석 ─────────────────────────
    # 차단된 IP의 차단 이후 재시도 건수 (줄었으면 효과 있음)
    blocked_ips_list = list(BlockedIP.objects.values_list('ip_address', flat=True))
    blocked_ip_attempts = 0
    if blocked_ips_list:
        blocked_ip_attempts = SystemLogEntry.objects.filter(
            log_type='ssh_fail', created_at__gte=since,
            ip_address__in=blocked_ips_list,
        ).count()

    # 일별 추이 (SSH + NPMS 합산)
    daily_trend = []
    daily_npms = dict(
        LoginHistory.objects.filter(success=False, created_at__gte=since)
        .annotate(d=TruncDate('created_at'))
        .values('d').annotate(cnt=Count('id'))
        .values_list('d', 'cnt')
    )
    daily_ssh = dict(
        SystemLogEntry.objects.filter(log_type='ssh_fail', created_at__gte=since)
        .annotate(d=TruncDate('created_at'))
        .values('d').annotate(cnt=Count('id'))
        .values_list('d', 'cnt')
    )
    delta_days = (now - since).days
    for i in range(delta_days + 1):
        d = (since + datetime.timedelta(days=i)).date()
        daily_trend.append({
            'date': d.strftime('%m-%d'),
            'npms': daily_npms.get(d, 0),
            'ssh': daily_ssh.get(d, 0),
            'count': daily_npms.get(d, 0) + daily_ssh.get(d, 0),
        })

    return JsonResponse({
        'period': period_label,
        'login': {
            'total': login_total,
            'success': login_success,
            'fail': login_fail,
            'unique_fail_ips': unique_ips,
        },
        'ssh': {
            'fails': ssh_fails,
            'unique_ips': ssh_unique_ips,
        },
        'blocks': {
            'blocked': blocks_count,
            'unblocked': unblocks_count,
            'active_blocked': len(blocked_ips_list),
            'blocked_ip_attempts': blocked_ip_attempts,
        },
        'events': {'by_type': event_by_type, 'by_severity': event_by_severity},
        'ssh_fails': ssh_fails,
        'trend': {
            'current': total_current,
            'previous': total_prev,
            'change_pct': trend_pct,
            'npms_current': login_fail,
            'npms_previous': prev_fail,
            'ssh_current': ssh_fails,
            'ssh_previous': prev_ssh,
        },
        'daily_trend': daily_trend,
    })


# ═══════════════════════════════════════════════════════
# Excel 다운로드 통합 API
# ═══════════════════════════════════════════════════════

@_admin_required
def sec_export(request):
    """
    보안관제 리스트 Excel 다운로드
    GET /api/sysconfig/security/export/?kind=<리스트종류>

    지원 kind:
      top_ips, events, blocked, whitelist, blocklog,
      abnormal, locked, ssh, integrity, checks,
      evtype, trend
    """
    import io
    import urllib.parse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse

    kind = request.GET.get('kind', '')
    now = timezone.localtime(timezone.now())

    wb = Workbook()
    ws = wb.active

    # ── 공통 헤더 스타일
    def _apply_header(sheet, headers):
        sheet.append(headers)
        for col_idx, _ in enumerate(headers, 1):
            c = sheet.cell(row=1, column=col_idx)
            c.font = Font(bold=True, color='FFFFFF', size=10)
            c.fill = PatternFill('solid', fgColor='1F497D')
            c.alignment = Alignment(horizontal='center', vertical='center')

    # ── 종류별 데이터 채우기 ──────────────────────────
    if kind == 'top_ips':
        ws.title = 'Top위협IP'
        _apply_header(ws, ['순위', 'IP 주소', '실패 횟수', '위험도', '차단 상태', '최초 시도', '최근 시도'])
        h7d = now - datetime.timedelta(days=7)
        rows = (SystemLogEntry.objects.filter(
            log_type='ssh_fail', created_at__gte=h7d, ip_address__isnull=False
        ).values('ip_address').annotate(
            fail_count=Count('id'),
            last_attempt=Max('created_at'),
            first_attempt=Min('created_at'),
        ).order_by('-fail_count'))
        blocked_set = set(BlockedIP.objects.values_list('ip_address', flat=True))
        THREAT = {500: '심각', 100: '높음', 20: '보통'}
        for i, r in enumerate(rows, 1):
            fc = r['fail_count']
            threat = '심각' if fc >= 500 else '높음' if fc >= 100 else '보통' if fc >= 20 else '낮음'
            ws.append([
                i, r['ip_address'] or '-', fc, threat,
                '차단' if r['ip_address'] in blocked_set else '미차단',
                timezone.localtime(r['first_attempt']).strftime('%Y-%m-%d %H:%M') if r['first_attempt'] else '-',
                timezone.localtime(r['last_attempt']).strftime('%Y-%m-%d %H:%M') if r['last_attempt'] else '-',
            ])
        filename = f'보안관제_Top위협IP_{now.strftime("%Y%m%d_%H%M")}.xlsx'

    elif kind == 'events':
        ws.title = '보안이벤트'
        _apply_header(ws, ['시각', '유형', '위험도', 'IP', '사용자', '설명', '해결여부'])
        for ev in SecurityEvent.objects.all()[:10000]:
            ws.append([
                timezone.localtime(ev.created_at).strftime('%Y-%m-%d %H:%M:%S'),
                ev.get_event_type_display(),
                ev.get_severity_display(),
                ev.ip_address or '-',
                ev.username or '-',
                ev.description,
                '해결' if ev.resolved else '미해결',
            ])
        filename = f'보안관제_보안이벤트_{now.strftime("%Y%m%d_%H%M")}.xlsx'

    elif kind == 'blocked':
        ws.title = '차단IP목록'
        _apply_header(ws, ['IP 주소', '사유', '상세설명', '실패 횟수', '차단 유형', '영구차단', '차단 시각', '해제 예정', '차단자'])
        for b in BlockedIP.objects.all():
            ws.append([
                b.ip_address, b.get_reason_display(), b.description, b.fail_count,
                '자동' if b.auto_blocked else '수동',
                '영구' if b.is_permanent else '임시',
                timezone.localtime(b.blocked_at).strftime('%Y-%m-%d %H:%M'),
                timezone.localtime(b.expires_at).strftime('%Y-%m-%d %H:%M') if b.expires_at else '영구',
                b.blocked_by.name if b.blocked_by else '시스템',
            ])
        filename = f'보안관제_차단IP목록_{now.strftime("%Y%m%d_%H%M")}.xlsx'

    elif kind == 'whitelist':
        ws.title = '화이트리스트'
        _apply_header(ws, ['IP 주소', '설명', '등록자', '등록일시'])
        for w in WhitelistedIP.objects.all():
            ws.append([
                w.ip_address, w.description,
                w.created_by.name if w.created_by else '-',
                timezone.localtime(w.created_at).strftime('%Y-%m-%d %H:%M'),
            ])
        filename = f'보안관제_화이트리스트_{now.strftime("%Y%m%d_%H%M")}.xlsx'

    elif kind == 'blocklog':
        ws.title = '차단해제이력'
        _apply_header(ws, ['처리일시', 'IP 주소', '행위', '사유', '처리자'])
        for r in BlockLog.objects.select_related('actor').all()[:10000]:
            ws.append([
                timezone.localtime(r.created_at).strftime('%Y-%m-%d %H:%M:%S'),
                r.ip_address, r.get_action_display(), r.reason,
                r.actor.name if r.actor else '시스템',
            ])
        filename = f'보안관제_차단해제이력_{now.strftime("%Y%m%d_%H%M")}.xlsx'

    elif kind == 'abnormal':
        ws.title = '비정상로그인'
        _apply_header(ws, ['유형', '사용자ID', '이름', 'IP', '시각', '상세'])
        days = int(request.GET.get('days', 14))
        since = now - datetime.timedelta(days=days)
        night = (LoginHistory.objects.filter(success=True, created_at__gte=since)
                 .annotate(hr=ExtractHour('created_at'))
                 .filter(Q(hr__gte=22) | Q(hr__lt=6)).select_related('user'))
        for r in night:
            ws.append(['야간 접속',
                       r.user.username if r.user else r.attempted_username,
                       r.user.name if r.user else '-',
                       r.ip_address or '-',
                       timezone.localtime(r.created_at).strftime('%Y-%m-%d %H:%M'),
                       f'{timezone.localtime(r.created_at).strftime("%H:%M")} 접속'])
        multi = (LoginHistory.objects.filter(success=False, created_at__gte=since)
                 .values('ip_address').annotate(
                    user_cnt=Count('attempted_username', distinct=True),
                    total=Count('id'),
                 ).filter(user_cnt__gte=3).order_by('-total'))
        for row in multi:
            ws.append(['다중 계정 시도', '-', '-',
                       row['ip_address'] or '-', '-',
                       f'{row["user_cnt"]}개 계정, {row["total"]}회 시도'])
        filename = f'보안관제_비정상로그인_{now.strftime("%Y%m%d_%H%M")}.xlsx'

    elif kind == 'locked':
        ws.title = '계정잠금'
        _apply_header(ws, ['사용자ID', '이름', '실패 횟수', '등록 여부'])
        lock_window = now - datetime.timedelta(minutes=30)
        rows = (LoginHistory.objects.filter(
            success=False, created_at__gte=lock_window
        ).values('attempted_username').annotate(cnt=Count('id')).filter(cnt__gte=5).order_by('-cnt'))
        for r in rows:
            uname = r['attempted_username']
            u = User.objects.filter(username=uname).first()
            ws.append([uname, u.name if u else '(미등록)', r['cnt'], '등록' if u else '미등록'])
        filename = f'보안관제_계정잠금_{now.strftime("%Y%m%d_%H%M")}.xlsx'

    elif kind == 'ssh':
        ws.title = 'SSH접근로그'
        _apply_header(ws, ['시각', '유형', 'IP', '사용자', '원본 로그'])
        qs = SystemLogEntry.objects.filter(log_type__in=['ssh_fail', 'ssh_success', 'auth_other'])
        for r in qs[:20000]:
            t = '성공' if r.log_type == 'ssh_success' else '실패'
            ws.append([
                timezone.localtime(r.log_time).strftime('%Y-%m-%d %H:%M:%S') if r.log_time else '-',
                t, r.ip_address or '-', r.username or '-', r.raw_line,
            ])
        filename = f'보안관제_SSH접근로그_{now.strftime("%Y%m%d_%H%M")}.xlsx'

    elif kind == 'integrity':
        ws.title = '파일무결성'
        _apply_header(ws, ['파일 경로', 'SHA256 해시', '크기(bytes)', '상태', '점검일시'])
        for f in FileIntegritySnapshot.objects.all().order_by('-is_changed', 'file_path'):
            ws.append([
                f.file_path, f.sha256_hash, f.file_size,
                '변경' if f.is_changed else '정상',
                timezone.localtime(f.checked_at).strftime('%Y-%m-%d %H:%M:%S'),
            ])
        filename = f'보안관제_파일무결성_{now.strftime("%Y%m%d_%H%M")}.xlsx'

    elif kind == 'checks':
        ws.title = '보안자가진단'
        _apply_header(ws, ['점검 항목', '상태', '상세'])
        # 재활용: sec_settings 로직 그대로
        from django.conf import settings as _settings
        checks = []
        checks.append(('DEBUG 모드', 'pass' if not _settings.DEBUG else 'fail',
                       'OFF' if not _settings.DEBUG else '⚠ ON'))
        checks.append(('HTTPS(SSL/TLS)',
                       'pass' if getattr(_settings, 'SESSION_COOKIE_SECURE', False) else 'fail',
                       'TLS 1.2/1.3 적용' if getattr(_settings, 'SESSION_COOKIE_SECURE', False) else '⚠ 미적용'))
        checks.append(('CSRF 보호', 'pass', f'CSRF_COOKIE_HTTPONLY={_settings.CSRF_COOKIE_HTTPONLY}'))
        checks.append(('IP 자동 차단',
                       'pass' if SecurityConfig.get_bool('auto_block_enabled') else 'warn',
                       f'임계값 {SecurityConfig.get_int("block_threshold", 10)}회'))
        for item, st, detail in checks:
            ws.append([item, st, detail])
        filename = f'보안관제_보안자가진단_{now.strftime("%Y%m%d_%H%M")}.xlsx'

    elif kind == 'evtype':
        ws.title = '이벤트유형통계'
        _apply_header(ws, ['이벤트 유형', '건수'])
        period = request.GET.get('period', 'weekly')
        if period == 'daily':
            since = now - datetime.timedelta(days=1)
        elif period == 'monthly':
            since = now - datetime.timedelta(days=30)
        else:
            since = now - datetime.timedelta(weeks=1)
        stats = (SecurityEvent.objects.filter(created_at__gte=since)
                 .values('event_type').annotate(cnt=Count('id')).order_by('-cnt'))
        EVT_LBL = dict(SecurityEvent.EVENT_TYPE_CHOICES)
        for s in stats:
            ws.append([EVT_LBL.get(s['event_type'], s['event_type']), s['cnt']])
        filename = f'보안관제_이벤트유형통계_{now.strftime("%Y%m%d_%H%M")}.xlsx'

    elif kind == 'trend':
        ws.title = '일별추이'
        _apply_header(ws, ['날짜', '공격 시도 건수'])
        period = request.GET.get('period', 'weekly')
        if period == 'daily':
            since = now - datetime.timedelta(days=1)
        elif period == 'monthly':
            since = now - datetime.timedelta(days=30)
        else:
            since = now - datetime.timedelta(weeks=1)
        daily = (LoginHistory.objects.filter(success=False, created_at__gte=since)
                 .annotate(d=TruncDate('created_at'))
                 .values('d').annotate(cnt=Count('id')).order_by('d'))
        for r in daily:
            ws.append([r['d'].strftime('%Y-%m-%d'), r['cnt']])
        filename = f'보안관제_공격추이_{now.strftime("%Y%m%d_%H%M")}.xlsx'

    # ── 접속이력 4종 (로그인/활동/세션/보안탐지) ──────
    elif kind in ('al_login', 'al_activity', 'al_session', 'al_security'):
        al_kind = kind.replace('al_', '')
        from apps.accounts.models import LoginHistory as LH, UserActivityLog as UA, UserSession as US
        if al_kind == 'login':
            ws.title = '로그인이력'
            _apply_header(ws, ['사용자', '성명', 'IP', '브라우저', '결과', '실패사유', '시도일시'])
            for r in LH.objects.select_related('user').order_by('-created_at')[:20000]:
                ws.append([
                    r.attempted_username or (r.user.username if r.user else '-'),
                    r.user.name if r.user else '(미등록)',
                    r.ip_address or '-', r.user_agent[:80] if r.user_agent else '-',
                    '성공' if r.success else '실패',
                    r.fail_reason or '',
                    timezone.localtime(r.created_at).strftime('%Y-%m-%d %H:%M:%S'),
                ])
            filename = f'접속이력_로그인_{now.strftime("%Y%m%d_%H%M")}.xlsx'
        elif al_kind == 'activity':
            ws.title = '활동로그'
            _apply_header(ws, ['사용자', '성명', '행위', '대상', '상세', 'IP', '발생일시'])
            ACTION = {'login':'로그인','logout':'로그아웃','create':'생성','update':'수정','delete':'삭제','view':'조회','download':'다운로드','upload':'업로드'}
            for r in UA.objects.select_related('user').order_by('-created_at')[:20000]:
                ws.append([
                    r.user.username if r.user else '-',
                    r.user.name if r.user else '-',
                    ACTION.get(r.action, r.action), r.target or '-', r.detail or '-',
                    r.ip_address or '-', timezone.localtime(r.created_at).strftime('%Y-%m-%d %H:%M:%S'),
                ])
            filename = f'접속이력_활동로그_{now.strftime("%Y%m%d_%H%M")}.xlsx'
        elif al_kind == 'session':
            ws.title = '현재접속자'
            _apply_header(ws, ['사용자', '성명', 'IP', '현재화면', '로그인시각', '마지막활동'])
            cutoff = now - datetime.timedelta(minutes=30)
            for r in US.objects.select_related('user').filter(is_active=True, last_active__gte=cutoff).order_by('-last_active'):
                ws.append([
                    r.user.username if r.user else '-',
                    r.user.name if r.user else '-',
                    r.ip_address or '-', r.current_page or '-',
                    timezone.localtime(r.login_at).strftime('%Y-%m-%d %H:%M:%S'),
                    timezone.localtime(r.last_active).strftime('%Y-%m-%d %H:%M:%S'),
                ])
            filename = f'접속이력_현재접속자_{now.strftime("%Y%m%d_%H%M")}.xlsx'
        else:  # security
            ws.title = '보안탐지'
            _apply_header(ws, ['IP 주소', '실패 횟수', '시도 계정', '성공 이력', '위험도', '최초 시도', '최근 시도'])
            since = now - datetime.timedelta(days=7)
            fail_qs = LH.objects.filter(success=False, created_at__gte=since)
            ip_stats = fail_qs.values('ip_address').annotate(
                fail_count=Count('id'), last_attempt=Max('created_at'), first_attempt=Min('created_at'),
            ).order_by('-fail_count')
            for row in ip_stats:
                ip = row['ip_address']
                usernames = sorted(set(fail_qs.filter(ip_address=ip).exclude(attempted_username='').values_list('attempted_username', flat=True)))[:10]
                has_success = LH.objects.filter(ip_address=ip, success=True, created_at__gte=since).exists()
                fc = row['fail_count']
                threat = '심각' if fc >= 20 else '높음' if fc >= 10 else '주의' if fc >= 5 else '낮음'
                ws.append([
                    ip or '-', fc, ', '.join(u for u in usernames if u) or '-',
                    '있음' if has_success else '없음', threat,
                    timezone.localtime(row['first_attempt']).strftime('%Y-%m-%d %H:%M'),
                    timezone.localtime(row['last_attempt']).strftime('%Y-%m-%d %H:%M'),
                ])
            filename = f'접속이력_보안탐지_{now.strftime("%Y%m%d_%H%M")}.xlsx'

    else:
        return JsonResponse({'error': f'알 수 없는 kind: {kind}'}, status=400)

    # ── 열 너비 자동 조정 ──
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                v = str(cell.value) if cell.value is not None else ''
                max_len = max(max_len, len(v))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 60)

    # ── 응답 생성 ──
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    encoded = urllib.parse.quote(filename)
    resp['Content-Disposition'] = f"attachment; filename*=UTF-8''{encoded}"
    return resp

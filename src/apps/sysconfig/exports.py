"""
sysconfig 데이터 내보내기 API
GET /api/sysconfig/export/<module>/
  ?format=csv|xlsx
  &date_from=YYYY-MM-DD
  &date_to=YYYY-MM-DD
  &center=<지원청ID>
  &(모듈별 추가 파라미터)
"""
import csv
import io
from datetime import date as _date
import urllib.parse

from django.http import HttpResponse, JsonResponse
from django.contrib.auth.decorators import login_required
from django.utils import timezone


def _admin_required(view_func):
    from functools import wraps
    @wraps(view_func)
    def _wrapped(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return JsonResponse({'error': '로그인 필요'}, status=401)
        if request.user.role not in ('superadmin', 'admin'):
            return JsonResponse({'error': '권한 없음'}, status=403)
        return view_func(request, *args, **kwargs)
    return _wrapped


def _csv_response(filename):
    resp = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    encoded = urllib.parse.quote(filename)
    resp['Content-Disposition'] = f"attachment; filename*=UTF-8''{encoded}"
    return resp


def _xlsx_response(filename):
    resp = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    encoded = urllib.parse.quote(filename)
    resp['Content-Disposition'] = f"attachment; filename*=UTF-8''{encoded}"
    return resp


def _wb_header(ws, headers, *, fill_hex='1F497D'):
    """openpyxl 워크시트 헤더 스타일 적용"""
    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        hdr_fill = PatternFill('solid', fgColor=fill_hex)
        hdr_font = Font(bold=True, color='FFFFFF', size=10)
        ctr = Alignment(horizontal='center', vertical='center')
        thin = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'),  bottom=Side(style='thin'),
        )
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = hdr_font; c.fill = hdr_fill
            c.alignment = ctr; c.border = thin
    except Exception:
        pass


def _apply_date_filter(qs, field, date_from, date_to):
    if date_from:
        qs = qs.filter(**{f'{field}__date__gte': date_from})
    if date_to:
        qs = qs.filter(**{f'{field}__date__lte': date_to})
    return qs


def _today():
    return _date.today().strftime('%Y%m%d')


# ══════════════════════════════════════════════════════
# 1. 장애 목록
# ══════════════════════════════════════════════════════
def export_incidents(request, fmt, date_from, date_to, center):
    from apps.incidents.models import Incident
    qs = Incident.objects.select_related(
        'school__support_center', 'category', 'subcategory',
        'received_by'
    ).order_by('-received_at')
    qs = _apply_date_filter(qs, 'received_at', date_from, date_to)
    if center:
        qs = qs.filter(school__support_center_id=center)
    status_f = request.GET.get('status')
    if status_f:
        qs = qs.filter(status=status_f)

    headers = ['접수번호', '접수일시', '교육지원청', '학교명', '장애분류', '소분류',
               '상태', '긴급도', '접수자', '요청자', '연락처',
               '현장도착', '처리완료', '소요시간(분)', '처리내용']

    def row(i):
        return [
            i.incident_number,
            timezone.localtime(i.received_at).strftime('%Y-%m-%d %H:%M'),
            i.school.support_center.name if i.school and i.school.support_center else '',
            i.school.name if i.school else '',
            i.category.name if i.category else '',
            i.subcategory.name if i.subcategory else '',
            i.get_status_display(),
            i.get_priority_display(),
            i.received_by.name if i.received_by else '',
            i.requester_name, i.requester_phone,
            timezone.localtime(i.arrived_at).strftime('%Y-%m-%d %H:%M') if i.arrived_at else '',
            timezone.localtime(i.completed_at).strftime('%Y-%m-%d %H:%M') if i.completed_at else '',
            i.get_elapsed_minutes() if hasattr(i, 'get_elapsed_minutes') else '',
            i.description or '',
        ]

    fname = f'장애목록_{_today()}'
    if fmt == 'xlsx':
        import openpyxl
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = '장애목록'
        _wb_header(ws, headers)
        for ri, i in enumerate(qs, 2):
            for ci, v in enumerate(row(i), 1):
                ws.cell(row=ri, column=ci, value=str(v) if v is not None else '')
        resp = _xlsx_response(fname + '.xlsx')
        buf = io.BytesIO(); wb.save(buf); resp.content = buf.getvalue()
        return resp
    else:
        resp = _csv_response(fname + '.csv')
        w = csv.writer(resp); w.writerow(headers)
        for i in qs: w.writerow(row(i))
        return resp


# ══════════════════════════════════════════════════════
# 2. 사진 목록 (스위치 설치 확인서 포함)
# ══════════════════════════════════════════════════════
def export_photos(request, fmt, date_from, date_to, center):
    from apps.photos.models import Photo
    qs = Photo.objects.select_related(
        'school__support_center', 'work_type', 'taken_by', 'building', 'floor'
    ).order_by('-taken_at')
    qs = _apply_date_filter(qs, 'taken_at', date_from, date_to)
    if center:
        qs = qs.filter(school__support_center_id=center)
    work_type = request.GET.get('work_type')
    if work_type:
        qs = qs.filter(work_type_id=work_type)

    headers = ['촬영일시', '교육지원청', '학교명', '건물', '층', '작업명', '단계',
               '파일명', '파일크기(KB)', '촬영자', 'GPS 위도', 'GPS 경도', '관련장애번호']

    def row(p):
        return [
            timezone.localtime(p.taken_at).strftime('%Y-%m-%d %H:%M'),
            p.school.support_center.name if p.school and p.school.support_center else '',
            p.school.name if p.school else '',
            p.building.name if p.building else '',
            p.floor.floor_name if p.floor else '',
            p.work_type.name if p.work_type else p.work_type_etc,
            p.get_photo_stage_display(),
            p.file_name,
            round(p.file_size / 1024, 1) if p.file_size else 0,
            p.taken_by.name if p.taken_by else '',
            str(p.gps_lat) if p.gps_lat else '',
            str(p.gps_lng) if p.gps_lng else '',
            p.incident.incident_number if p.incident else '',
        ]

    fname = f'사진목록_{_today()}'
    if fmt == 'xlsx':
        import openpyxl
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = '사진목록'
        _wb_header(ws, headers)
        for ri, p in enumerate(qs, 2):
            for ci, v in enumerate(row(p), 1):
                ws.cell(row=ri, column=ci, value=v if v is not None else '')
        resp = _xlsx_response(fname + '.xlsx')
        buf = io.BytesIO(); wb.save(buf); resp.content = buf.getvalue()
        return resp
    else:
        resp = _csv_response(fname + '.csv')
        w = csv.writer(resp); w.writerow(headers)
        for p in qs: w.writerow(row(p))
        return resp


# ══════════════════════════════════════════════════════
# 3. 장비 전체 목록
# ══════════════════════════════════════════════════════
def export_assets(request, fmt, date_from, date_to, center):
    from apps.assets.models import Asset
    qs = Asset.objects.select_related(
        'asset_model__category', 'current_school__support_center'
    ).order_by('asset_model__category__name', 'serial_number')
    if center:
        qs = qs.filter(current_school__support_center_id=center)
    if date_from:
        qs = qs.filter(installed_at__gte=date_from)
    if date_to:
        qs = qs.filter(installed_at__lte=date_to)
    status_f = request.GET.get('status')
    if status_f:
        qs = qs.filter(status=status_f)

    headers = ['분류', '제조번호(S/N)', '관리번호', '모델명', '제조사',
               '상태', '설치학교', '교육지원청', '설치위치', '설치일',
               '보증만료일', '사업명']

    def row(a):
        return [
            a.asset_model.category.name if a.asset_model and a.asset_model.category else '',
            a.serial_number or '',
            a.asset_tag or '',
            a.asset_model.model_name if a.asset_model else '',
            a.asset_model.manufacturer if a.asset_model else '',
            a.get_status_display() if hasattr(a, 'get_status_display') else a.status,
            a.current_school.name if a.current_school else '',
            a.current_school.support_center.name if a.current_school and a.current_school.support_center else '',
            a.install_location or '',
            str(a.installed_at) if a.installed_at else '',
            str(a.warranty_expire) if a.warranty_expire else '',
            a.project_name or '',
        ]

    fname = f'장비목록_{_today()}'
    if fmt == 'xlsx':
        import openpyxl
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = '장비목록'
        _wb_header(ws, headers)
        for ri, a in enumerate(qs, 2):
            for ci, v in enumerate(row(a), 1):
                ws.cell(row=ri, column=ci, value=v if v is not None else '')
        resp = _xlsx_response(fname + '.xlsx')
        buf = io.BytesIO(); wb.save(buf); resp.content = buf.getvalue()
        return resp
    else:
        resp = _csv_response(fname + '.csv')
        w = csv.writer(resp); w.writerow(headers)
        for a in qs: w.writerow(row(a))
        return resp


# ══════════════════════════════════════════════════════
# 4. 자재 내역 (입고/출고/현재고)
# ══════════════════════════════════════════════════════
def export_materials(request, fmt, date_from, date_to, center):
    sub = request.GET.get('sub', 'inventory')  # inventory | inbound | outbound

    if sub == 'inbound':
        from apps.materials.models import MaterialInbound
        qs = MaterialInbound.objects.select_related('material__category').order_by('-inbound_date')
        if date_from: qs = qs.filter(inbound_date__gte=date_from)
        if date_to:   qs = qs.filter(inbound_date__lte=date_to)
        headers = ['입고일', '분류', '자재코드', '자재명', '규격', '단위', '수량', '단가', '공급업체', '인계자', '인수자', '비고']
        def row(r):
            return [str(r.inbound_date), r.material.category.name if r.material and r.material.category else '',
                    r.material.code if r.material else '', r.material.name if r.material else '',
                    r.material.spec if r.material else '', r.material.unit if r.material else '',
                    r.quantity, r.unit_price or '', r.supplier or '',
                    r.handover_person or '', r.receiver_person or '', r.note or '']
        fname = f'자재입고내역_{_today()}'
        rows_qs = qs

    elif sub == 'outbound':
        from apps.materials.models import MaterialOutbound
        qs = MaterialOutbound.objects.select_related('material__category', 'to_center').order_by('-outbound_date')
        if date_from: qs = qs.filter(outbound_date__gte=date_from)
        if date_to:   qs = qs.filter(outbound_date__lte=date_to)
        if center:    qs = qs.filter(to_center_id=center)
        headers = ['출고일', '분류', '자재코드', '자재명', '규격', '단위', '수량', '출고지원청', '인계자', '인수자', '비고']
        def row(r):
            return [str(r.outbound_date), r.material.category.name if r.material and r.material.category else '',
                    r.material.code if r.material else '', r.material.name if r.material else '',
                    r.material.spec if r.material else '', r.material.unit if r.material else '',
                    r.quantity, r.to_center.name if r.to_center else '',
                    r.handover_person or '', r.receiver_person or '', r.note or '']
        fname = f'자재출고내역_{_today()}'
        rows_qs = qs

    else:  # inventory
        from apps.materials.models import WarehouseInventory
        qs = WarehouseInventory.objects.select_related('material__category').order_by('material__category__name', 'material__name')
        headers = ['분류', '자재코드', '자재명', '규격', '단위', '현재고', '최소재고', '공급업체']
        def row(r):
            return [r.material.category.name if r.material and r.material.category else '',
                    r.material.code if r.material else '', r.material.name if r.material else '',
                    r.material.spec if r.material else '', r.material.unit if r.material else '',
                    r.quantity, r.material.min_stock if r.material else '', r.material.supplier if r.material else '']
        fname = f'자재현재고_{_today()}'
        rows_qs = qs

    if fmt == 'xlsx':
        import openpyxl
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = fname[:30]
        _wb_header(ws, headers)
        for ri, r in enumerate(rows_qs, 2):
            for ci, v in enumerate(row(r), 1):
                ws.cell(row=ri, column=ci, value=v if v is not None else '')
        resp = _xlsx_response(fname + '.xlsx')
        buf = io.BytesIO(); wb.save(buf); resp.content = buf.getvalue()
        return resp
    else:
        resp = _csv_response(fname + '.csv')
        w = csv.writer(resp); w.writerow(headers)
        for r in rows_qs: w.writerow(row(r))
        return resp


# ══════════════════════════════════════════════════════
# 5. 출퇴근 기록
# ══════════════════════════════════════════════════════
def export_attendance(request, fmt, date_from, date_to, center):
    from apps.workforce.models import AttendanceLog
    from django.utils import timezone as tz
    qs = AttendanceLog.objects.select_related(
        'worker__support_center'
    ).order_by('-work_date', 'worker__name')
    if date_from: qs = qs.filter(work_date__gte=date_from)
    if date_to:   qs = qs.filter(work_date__lte=date_to)
    if center:    qs = qs.filter(worker__support_center_id=center)

    headers = ['근무일', '인력명', '소속지원청', '출근시각', '퇴근시각', '근무시간(h)', '상태', '비고']

    def row(r):
        return [
            str(r.work_date),
            r.worker.name if r.worker else '',
            r.worker.support_center.name if r.worker and r.worker.support_center else '',
            tz.localtime(r.check_in_at).strftime('%H:%M') if r.check_in_at else '',
            tz.localtime(r.check_out_at).strftime('%H:%M') if r.check_out_at else '',
            r.get_work_hours() if hasattr(r, 'get_work_hours') else '',
            r.get_status_display() if hasattr(r, 'get_status_display') else r.status,
            r.note or '',
        ]

    fname = f'출퇴근기록_{_today()}'
    if fmt == 'xlsx':
        import openpyxl
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = '출퇴근기록'
        _wb_header(ws, headers)
        for ri, r in enumerate(qs, 2):
            for ci, v in enumerate(row(r), 1):
                ws.cell(row=ri, column=ci, value=v if v is not None else '')
        resp = _xlsx_response(fname + '.xlsx')
        buf = io.BytesIO(); wb.save(buf); resp.content = buf.getvalue()
        return resp
    else:
        resp = _csv_response(fname + '.csv')
        w = csv.writer(resp); w.writerow(headers)
        for r in qs: w.writerow(row(r))
        return resp


# ══════════════════════════════════════════════════════
# 6. 진척관리 (점검 결과)
# ══════════════════════════════════════════════════════
def export_progress(request, fmt, date_from, date_to, center):
    from apps.progress.models import SchoolInspection
    qs = SchoolInspection.objects.select_related(
        'plan', 'school__support_center', 'school__school_type', 'assigned_worker'
    ).order_by('plan__name', 'school__support_center__name', 'school__name')
    if center: qs = qs.filter(school__support_center_id=center)
    if date_from: qs = qs.filter(scheduled_date__gte=date_from)
    if date_to:   qs = qs.filter(scheduled_date__lte=date_to)

    headers = ['점검계획', '교육지원청', '학교명', '학제', '예정일', '완료일', '상태', '담당기사', '비고']

    def row(r):
        return [
            r.plan.name if r.plan else '',
            r.school.support_center.name if r.school and r.school.support_center else '',
            r.school.name if r.school else '',
            r.school.school_type.name if r.school and r.school.school_type else '',
            str(r.scheduled_date) if r.scheduled_date else '',
            str(r.completed_date) if hasattr(r, 'completed_date') and r.completed_date else '',
            r.get_status_display() if hasattr(r, 'get_status_display') else r.status,
            r.assigned_worker.name if r.assigned_worker else '',
            r.note if hasattr(r, 'note') else '',
        ]

    fname = f'진척관리_{_today()}'
    if fmt == 'xlsx':
        import openpyxl
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = '진척관리'
        _wb_header(ws, headers)
        for ri, r in enumerate(qs, 2):
            for ci, v in enumerate(row(r), 1):
                ws.cell(row=ri, column=ci, value=v if v is not None else '')
        resp = _xlsx_response(fname + '.xlsx')
        buf = io.BytesIO(); wb.save(buf); resp.content = buf.getvalue()
        return resp
    else:
        resp = _csv_response(fname + '.csv')
        w = csv.writer(resp); w.writerow(headers)
        for r in qs: w.writerow(row(r))
        return resp


# ══════════════════════════════════════════════════════
# 7. SLA 성과
# ══════════════════════════════════════════════════════
def export_sla(request, fmt, date_from, date_to, center):
    from apps.incidents.models import IncidentSLA, Incident
    qs = IncidentSLA.objects.select_related(
        'incident__school__support_center', 'incident__category'
    ).order_by('-incident__received_at')
    if center:    qs = qs.filter(incident__school__support_center_id=center)
    if date_from: qs = qs.filter(incident__received_at__date__gte=date_from)
    if date_to:   qs = qs.filter(incident__received_at__date__lte=date_to)

    headers = ['접수번호', '접수일시', '교육지원청', '학교명', '장애분류',
               '도착목표', '도착실제', '도착준수', '도착차이(분)',
               '처리목표', '처리실제', '처리준수', '처리차이(분)']

    def row(s):
        i = s.incident
        return [
            i.incident_number,
            timezone.localtime(i.received_at).strftime('%Y-%m-%d %H:%M') if i.received_at else '',
            i.school.support_center.name if i.school and i.school.support_center else '',
            i.school.name if i.school else '',
            i.category.name if i.category else '',
            str(s.arrival_target) if s.arrival_target else '',
            str(s.arrival_actual) if s.arrival_actual else '',
            'Y' if s.arrival_ok else ('N' if s.arrival_ok is False else ''),
            s.arrival_diff_min or '',
            str(s.resolve_target) if s.resolve_target else '',
            str(s.resolve_actual) if s.resolve_actual else '',
            'Y' if s.resolve_ok else ('N' if s.resolve_ok is False else ''),
            s.resolve_diff_min or '',
        ]

    fname = f'SLA성과_{_today()}'
    if fmt == 'xlsx':
        import openpyxl
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'SLA성과'
        _wb_header(ws, headers)
        for ri, s in enumerate(qs, 2):
            for ci, v in enumerate(row(s), 1):
                ws.cell(row=ri, column=ci, value=v if v is not None else '')
        resp = _xlsx_response(fname + '.xlsx')
        buf = io.BytesIO(); wb.save(buf); resp.content = buf.getvalue()
        return resp
    else:
        resp = _csv_response(fname + '.csv')
        w = csv.writer(resp); w.writerow(headers)
        for s in qs: w.writerow(row(s))
        return resp


# ══════════════════════════════════════════════════════
# 8. 업무보고서
# ══════════════════════════════════════════════════════
def export_reports(request, fmt, date_from, date_to, center):
    from apps.reports.models import Report
    qs = Report.objects.select_related(
        'template', 'school__support_center', 'incident', 'created_by'
    ).order_by('-created_at')
    qs = _apply_date_filter(qs, 'created_at', date_from, date_to)
    if center:
        qs = qs.filter(school__support_center_id=center)
    status_f = request.GET.get('status')
    if status_f:
        qs = qs.filter(status=status_f)

    headers = ['보고서ID', '유형', '제목', '학교명', '교육지원청', '상태', '작성자', '작성일시', '완료일시', '관련장애번호']

    def row(r):
        return [
            r.id,
            r.template.get_report_type_display() if r.template else '',
            r.title,
            r.school.name if r.school else '',
            r.school.support_center.name if r.school and r.school.support_center else '',
            r.get_status_display(),
            r.created_by.name if r.created_by else '',
            timezone.localtime(r.created_at).strftime('%Y-%m-%d %H:%M'),
            timezone.localtime(r.completed_at).strftime('%Y-%m-%d %H:%M') if r.completed_at else '',
            r.incident.incident_number if r.incident else '',
        ]

    fname = f'업무보고서_{_today()}'
    if fmt == 'xlsx':
        import openpyxl
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = '업무보고서'
        _wb_header(ws, headers)
        for ri, r in enumerate(qs, 2):
            for ci, v in enumerate(row(r), 1):
                ws.cell(row=ri, column=ci, value=v if v is not None else '')
        resp = _xlsx_response(fname + '.xlsx')
        buf = io.BytesIO(); wb.save(buf); resp.content = buf.getvalue()
        return resp
    else:
        resp = _csv_response(fname + '.csv')
        w = csv.writer(resp); w.writerow(headers)
        for r in qs: w.writerow(row(r))
        return resp


# ══════════════════════════════════════════════════════
# 메인 라우터
# ══════════════════════════════════════════════════════
EXPORT_MAP = {
    'incidents':  export_incidents,
    'photos':     export_photos,
    'assets':     export_assets,
    'materials':  export_materials,
    'attendance': export_attendance,
    'progress':   export_progress,
    'sla':        export_sla,
    'reports':    export_reports,
}


@login_required
@_admin_required
def export_view(request, module):
    if module not in EXPORT_MAP:
        return JsonResponse({'error': f'알 수 없는 모듈: {module}'}, status=400)
    fmt       = request.GET.get('format', 'csv').lower()
    date_from = request.GET.get('date_from', '')
    date_to   = request.GET.get('date_to', '')
    center    = request.GET.get('center', '')
    if fmt not in ('csv', 'xlsx'):
        fmt = 'csv'
    try:
        return EXPORT_MAP[module](request, fmt, date_from or None, date_to or None, center or None)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

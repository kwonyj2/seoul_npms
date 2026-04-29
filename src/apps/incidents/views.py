import csv
import logging
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.db.models import Q, Count

logger = logging.getLogger(__name__)
from rest_framework import viewsets, status, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from .models import (
    IncidentCategory, IncidentSubcategory, Incident, IncidentSLA,
    IncidentAssignment, IncidentComment, IncidentPhoto, SLARule, SLAMonthly,
    IncidentDelayReason
)
from .serializers import (
    IncidentCategorySerializer, IncidentListSerializer, IncidentDetailSerializer,
    IncidentCreateSerializer, IncidentUpdateSerializer, IncidentAssignmentSerializer,
    IncidentCommentSerializer, IncidentPhotoSerializer, SLARuleSerializer
)
from .services import ai_assign_worker, create_assignment, get_available_workers, get_best_worker
from core.permissions.roles import IsAdmin, IsWorker


# ─────────────────────────────────────────
# SLA 정적 보고서 생성 (API 키 없을 때 폴백)
# ─────────────────────────────────────────
def _static_report(indicator, sla, rows, year, month):
    """API 키 미설정 또는 AI 호출 실패 시 데이터 기반 정적 보고서 반환"""
    summary = detail = action = ''
    cnt = len(rows)

    if indicator == 'uptime':
        pct = sla.get('uptime_pct')
        fault_min = sla.get('uptime_fault_min', 0)
        total_min = sla.get('uptime_total_min', 0)
        achieved = pct is not None and pct >= 99.98
        summary = (
            f"{year}년 {month}월 장비 가동률은 {f'{pct:.4f}%' if pct is not None else '측정값 없음'}으로, "
            f"목표 수준(99.98%) 대비 {'달성' if achieved else '미달'}하였습니다."
        )
        detail = (
            f"월 총 가동시간 {total_min//60}시간 중 서비스 중단 손실 시간은 {fault_min}분이며, "
            f"서비스 중단 장애는 총 {cnt}건 발생하였습니다. "
            f"장애 발생 시간대·네트워크 구간별로 분류하여 원인을 분석하였으며, 조치 이력을 문서화하였습니다."
        )
        if pct is not None and pct < 99.85:
            action = ("가동률이 임계 수준 미만으로 하락하여 장애 로그 및 현장 점검 결과를 토대로 서비스 중단 원인을 분석하였습니다. "
                      "분석 결과를 바탕으로 취약 구간을 파악하고 집중 모니터링 체계를 강화하는 재발방지 대책을 수립하였습니다.")
        elif pct is not None and pct < 99.98:
            action = ("가동률이 목표 미달로, 월간 장애 이력을 구간별·장비별로 분석하여 취약 지점을 파악하였습니다. "
                      "취약 구간을 중점 관리 대상으로 지정하고 예방점검 주기를 조정하여 관리 체계를 강화하였습니다.")
        else:
            action = ("목표 가동률을 안정적으로 달성하였습니다. "
                      "월간 장애 데이터를 주기적으로 모니터링하여 이상 징후를 사전에 감지하는 체계를 유지하고 있습니다.")

    elif indicator == 'inspection':
        pct   = sla.get('inspection_pct')
        total = sla.get('inspection_total', 0)
        done  = sla.get('inspection_completed', 0)
        undone = total - done
        summary = (
            f"예방점검 준수율은 {f'{pct:.1f}%' if pct is not None else '측정값 없음'}으로, "
            f"총 {total}개교 중 {done}개교가 완료하였습니다."
        )
        if undone > 0:
            detail = (
                f"{undone}개교가 점검을 미완료하였습니다. "
                f"학교별 담당자에게 직접 연락하여 미완료 사유(방문 일정 조율 실패, 담당자 부재, 교내 행사 등)를 "
                f"항목별로 확인하였으며, 재방문 일정을 재조율하고 보완 점검을 추진하였습니다."
            )
        else:
            detail = ("모든 학교의 예방점검이 완료되었습니다. "
                      "점검 시 네트워크 장비 동작 상태·케이블 연결 상태·설정값 이상 여부를 항목별로 확인하고 "
                      "결과를 점검 대장에 기록하였습니다.")
        if pct is not None and pct < 97:
            action = ("준수율이 최소 기준(97%) 미달로, 미완료 학교를 긴급 방문하여 보완 점검을 실시하였습니다. "
                      "월초 사전 학교 통보 및 복수 예비 일정 확보 방식으로 재발방지 대책을 수립하였습니다.")
        elif pct is not None and pct < 100:
            action = ("미완료 학교에 대해 다음 월 내 만회 점검 일정을 학교 측과 협의하여 수립하고, "
                      "우선순위에 따라 방문 일정을 배정하였습니다.")
        else:
            action = ("모든 예방점검을 차질 없이 완료하였습니다. "
                      "점검 결과를 시스템에 즉시 등록하고 이상 항목에 대한 후속 조치를 완료하였습니다.")

    elif indicator in ('avg_fault', 'fault_count'):
        avg  = sla.get('avg_fault_min')
        fc   = sla.get('fault_count', cnt)
        over = sum(1 for r in rows if (r.get('biz_min') or 0) > 480)
        ft_cnt: dict = {}
        for r in rows:
            ft_cnt[r.get('fault_type','기타')] = ft_cnt.get(r.get('fault_type','기타'), 0) + 1
        ft_str = ', '.join(f"{k} {v}건" for k, v in ft_cnt.items()) or '없음'
        summary = (
            f"총 장애건수는 {fc}건, 평균 장애 조치시간은 {f'{avg:.0f}분' if avg is not None else '-'}입니다."
        )
        if over > 0:
            detail = (
                f"이 중 허용 기준(480분)을 초과한 장애 {over}건에 대해 접수~도착~조치~완료 단계별 소요 시간을 "
                f"추적하여 지연 구간을 특정하고 원인을 심층 분석하였습니다. "
                f"장애 유형별 현황({ft_str})을 토대로 각 건별 조치 이력을 문서화하였습니다."
            )
        else:
            detail = (
                f"모든 장애의 조치시간이 허용 기준(480분) 이내에 처리되었습니다. "
                f"장애 유형별 현황({ft_str})을 분석하여 조치 이력 및 원인 분류를 완료하였습니다."
            )
        if fc > 30:
            action = (f"장애건수가 허용 최소 기준(30건)을 초과하여 장애 유형별 근본 원인 분석을 수행하였습니다. "
                      f"빈발 장애 유형을 선별하고 현장 대응 절차를 보완하여 재발방지 대책을 수립하였습니다.")
        elif fc > 20:
            action = (f"장애건수가 목표 대비 다소 높아 취약 학교 및 발생 유형을 집중 분석하였습니다. "
                      f"취약 학교에 대한 예방적 유지보수 방문 횟수를 조정하고 대응 체계를 강화하였습니다.")
        else:
            action = (f"장애건수는 목표 범위 내에서 관리되었습니다. "
                      f"모든 장애에 대해 조치 이력 및 원인 분류를 완료하여 월간 관리 대장에 반영하였습니다.")

    elif indicator == 'overtime':
        summary = f"장애조치 최대 허용시간(480분) 초과 건수는 {cnt}건입니다."
        if cnt > 0:
            avg_over = sum(r.get('over_min', 0) for r in rows) / cnt
            max_over = max(r.get('over_min', 0) for r in rows)
            schools  = ', '.join({r.get('school','') for r in rows})
            detail = (
                f"초과 장애의 평균 초과시간은 {avg_over:.0f}분이며, 최대 {max_over:.0f}분까지 지연된 사례에 대해 "
                f"접수~도착~조치 완료 단계별 시간을 추적하여 지연 구간을 특정하고 원인을 분석하였습니다. "
                f"발생 학교({schools})별 지연 사유를 유형별로 분류하고 조치 이력을 문서화하였습니다."
            )
            action = (
                f"초과 장애 {cnt}건의 지연 원인 분석 결과를 바탕으로 유형별 재발방지 대책을 수립하였습니다. "
                f"조기 에스컬레이션 기준을 재정비하고 현장 대응 매뉴얼을 보완하였습니다."
            )
        else:
            detail = "모든 장애가 허용 시간 내에 처리되었습니다. 단계별 조치 시간을 검토한 결과 신속 대응 체계가 정상 유지되고 있음을 확인하였습니다."
            action = "최고 등급(A)을 달성하였습니다. 현 수준의 신속 대응 체계를 지속 유지하고 있습니다."

    elif indicator == 'human_error':
        total_fc = sla.get('fault_count', 1) or 1
        pct_str  = f"{cnt / total_fc * 100:.1f}%" if cnt else "0%"
        summary  = f"인적장애 발생 건수는 {cnt}건입니다."
        if cnt > 0:
            schools = ', '.join({r.get('school','') for r in rows})
            detail  = (
                f"인적장애는 전체 장애({total_fc}건)의 {pct_str}를 차지합니다. "
                f"작업 전·중·후 절차를 단계별로 재검토하여 오류 발생 단계와 원인을 파악하였으며, "
                f"담당자별 조치 내용을 기록하였습니다."
            )
            action  = (
                f"발생 원인 분석을 바탕으로 오류가 집중된 작업 단계에 이중 확인 절차를 적용하고, "
                f"관련 작업 절차를 보완하여 재발방지 대책을 수립하였습니다."
            )
        else:
            detail = "이번 달에는 인적장애가 발생하지 않았습니다. 작업 전 표준 절차 확인 및 이중 점검 체계가 잘 준수되고 있습니다."
            action = "인적장애 없음 — 작업 표준이 잘 준수되고 있으며, 정기 예방 교육의 효과가 유지되고 있습니다."

    elif indicator == 'recurrence':
        school_cnt: dict = {}
        for r in rows:
            school_cnt[r.get('school','')] = school_cnt.get(r.get('school',''), 0) + 1
        top_schools = ', '.join(f"{k}({v}회)" for k, v in sorted(school_cnt.items(), key=lambda x: -x[1])[:5])
        summary = f"반복장애 발생 건수는 {cnt}건입니다."
        if cnt > 0:
            detail = (
                f"반복장애는 동일 학교·동일 분류의 장애가 24시간 내 재발한 건입니다. "
                f"최초 조치 기록과 재발 기록을 비교하여 임시 처리에 그친 원인을 분석하였으며, "
                f"재발 학교({top_schools}) 현황을 파악하여 조치 완결성 미흡 사유를 건별로 기록하였습니다."
            )
            if cnt >= 6:
                action = (f"반복장애가 허용 기준(5건)을 초과하여 원인 유형별로 분류하고, "
                          f"항구적 해결 조치 계획을 수립하여 우선순위에 따라 시행하였으며, 결과를 교육청에 보고하였습니다.")
            else:
                action = (f"반복 발생 학교를 특정하여 조치 후 재확인 절차를 적용하고, "
                          f"초기 조치 완결성 향상을 위한 현장 점검 기준을 강화하였습니다.")
        else:
            detail = "이번 달에는 반복장애가 발생하지 않았습니다. 초기 조치 완결성 확인 절차가 현장에서 잘 지켜지고 있습니다."
            action = "조치 완료 후 재확인 절차를 지속 적용하여 초기 조치의 완결성을 유지하고 있습니다."

    elif indicator == 'security':
        sec_cnt  = sla.get('security_count', 0)
        sec_note = sla.get('security_note', '(내용 없음)')
        summary  = f"보안위규 발생 건수는 {sec_cnt}건입니다."
        if sec_cnt > 0:
            detail = (
                f"위규 내용: {sec_note}. "
                f"위반 발생 즉시 해당 담당자에게 통보하고, 위반 경위를 단계별로 확인하여 "
                f"정책 인지 여부·절차 준수 여부를 점검하는 방식으로 위반 원인을 분석하고 시정 조치를 완료하였습니다."
            )
            action = (
                f"위반 사례를 교육 자료로 정리하여 전 직원 보안 교육을 실시하고, "
                f"취약 절차를 식별하여 보완하는 재발방지 대책을 수립하였습니다."
            )
        else:
            detail = "이번 달에는 보안위규 사항이 발생하지 않았습니다. 보안 점검 항목을 정기적으로 확인하여 정책 준수 상태를 유지하고 있습니다."
            action = "정기 보안 점검을 통해 정책 준수 여부를 지속 확인하고 있습니다."

    elif indicator == 'satisfaction':
        pct  = sla.get('satisfaction_pct')
        sc   = sla.get('satisfaction_count', cnt)
        low  = sum(1 for r in rows if (r.get('score') or 5) <= 2)
        summary = f"서비스 만족도는 {f'{pct:.1f}%' if pct is not None else '측정값 없음'}(총 {sc}건 응답)입니다."
        if low > 0:
            detail = (
                f"낮은 평가(2점 이하) {low}건이 확인되었습니다. "
                f"해당 장애의 접수~처리~복구 완료 과정을 단계별로 재검토하고 피드백을 항목별로 분류하여 "
                f"불만 원인을 심층 분석하였으며, 건별 사후 조치를 시행하였습니다."
            )
        else:
            detail = ("대체로 높은 만족도를 유지하고 있습니다. "
                      "긍정 피드백의 공통 요인을 분석하여 우수 대응 사례를 전 담당자와 공유하였습니다.")
        if pct is not None and pct < 97:
            action = ("만족도가 최소 기준(97%) 미만으로, 불만 원인을 항목별로 분류하여 "
                      "응대 절차 개선 및 담당자 집중 교육 등 맞춤형 서비스 품질 개선 대책을 수립하였습니다.")
        else:
            action = ("서비스 만족도가 목표 수준 이상으로 유지되었습니다. "
                      "우수 대응 사례를 공유하고 현장 서비스 품질을 지속 유지하고 있습니다.")

    return {'summary': summary, 'detail': detail, 'action': action}


# ─────────────────────────────────────────
# 템플릿 뷰
# ─────────────────────────────────────────
@login_required
def incident_list_view(request):
    return render(request, 'incidents/list.html')


@login_required
def incident_detail_view(request, pk):
    incident = get_object_or_404(Incident, pk=pk)
    return render(request, 'incidents/detail.html', {'incident': incident})


@login_required
def incident_create_view(request):
    return render(request, 'incidents/create.html')


@login_required
def work_orders_view(request):
    return render(request, 'incidents/work_orders.html')


@login_required
def sla_view(request):
    """SLA 관리 화면"""
    from django.utils import timezone as tz
    now = tz.localdate()
    records = SLAMonthly.objects.order_by('-year', '-month')[:12]
    ctx = {
        'records':       records,
        'current_year':  now.year,
        'current_month': now.month,
        'indicators': [
            {'key': 'uptime',      'label': '장비 가동률'},
            {'key': 'inspection',  'label': '예방점검'},
            {'key': 'avg_fault',   'label': '평균 장애시간'},
            {'key': 'fault_count', 'label': '장애건수'},
            {'key': 'overtime',    'label': '허용시간 초과'},
            {'key': 'human_error', 'label': '인적장애'},
            {'key': 'recurrence',  'label': '반복장애'},
            {'key': 'security',    'label': '보안위규'},
            {'key': 'satisfaction','label': '서비스 만족도'},
        ],
    }
    return render(request, 'incidents/sla.html', ctx)


@login_required
def sla_calculate_api(request):
    """POST: SLA 월간 지표 자동 계산 & 저장"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST only'}, status=405)
    if request.user.role not in ('superadmin', 'admin'):
        return JsonResponse({'error': '권한 없음'}, status=403)
    import json
    from core.sla_calculator import save_monthly
    try:
        body = json.loads(request.body)
        year      = int(body.get('year',  timezone.localdate().year))
        month     = int(body.get('month', timezone.localdate().month))
        sec_count = int(body.get('security_count', 0))
        sec_note  = body.get('security_note', '')
        memo      = body.get('memo', '')
        obj, created = save_monthly(year, month, request.user,
                                    sec_count, sec_note, memo)
        return JsonResponse({
            'status':      'created' if created else 'updated',
            'year':        obj.year,
            'month':       obj.month,
            'total_score': obj.total_score,
            'grade':       obj.grade,
            'grade_label': dict(SLAMonthly.GRADE_CHOICES).get(obj.grade, ''),
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def sla_detail_api(request, year, month):
    """GET: 특정 월 SLA 상세 조회"""
    try:
        obj = SLAMonthly.objects.get(year=year, month=month)
    except SLAMonthly.DoesNotExist:
        return JsonResponse({'error': '데이터 없음'}, status=404)
    gl = dict(SLAMonthly.GRADE_CHOICES)
    return JsonResponse({
        'year': obj.year, 'month': obj.month,
        'uptime_pct': obj.uptime_pct,       'uptime_score': obj.uptime_score,
        'uptime_total_min': obj.uptime_total_min,
        'uptime_fault_min': obj.uptime_fault_min,
        'uptime_maint_min': obj.uptime_maint_min,
        'inspection_pct': obj.inspection_pct,   'inspection_score': obj.inspection_score,
        'inspection_total': obj.inspection_total, 'inspection_completed': obj.inspection_completed,
        'avg_fault_min': obj.avg_fault_min,   'avg_fault_score': obj.avg_fault_score,
        'fault_count': obj.fault_count,       'fault_count_score': obj.fault_count_score,
        'overtime_count': obj.overtime_count,  'overtime_score': obj.overtime_score,
        'human_error_count': obj.human_error_count, 'human_error_score': obj.human_error_score,
        'recurrence_count': obj.recurrence_count,  'recurrence_score': obj.recurrence_score,
        'security_count': obj.security_count,  'security_score': obj.security_score,
        'security_note': obj.security_note,
        'satisfaction_pct': obj.satisfaction_pct, 'satisfaction_score': obj.satisfaction_score,
        'satisfaction_count': obj.satisfaction_count,
        'total_score': obj.total_score, 'grade': obj.grade,
        'grade_label': gl.get(obj.grade, ''),
        'memo': obj.memo,
        'calculated_at': timezone.localtime(obj.calculated_at).strftime('%Y-%m-%d %H:%M') if obj.calculated_at else None,
    })


@login_required
def sla_delete_api(request, year, month):
    """DELETE: 특정 월 SLA 데이터 삭제 (관리자 전용)"""
    if request.method != 'DELETE':
        return JsonResponse({'error': 'DELETE only'}, status=405)
    if request.user.role not in ('superadmin', 'admin'):
        return JsonResponse({'error': '권한 없음'}, status=403)
    try:
        obj = SLAMonthly.objects.get(year=year, month=month)
        obj.delete()
        return JsonResponse({'status': 'deleted', 'year': year, 'month': month})
    except SLAMonthly.DoesNotExist:
        return JsonResponse({'error': '데이터 없음'}, status=404)


@login_required
def sla_indicator_api(request, year, month, indicator):
    """GET: 특정 월 특정 지표의 원본 데이터 조회"""
    import calendar
    from datetime import date, datetime
    from django.db.models import Avg

    first_day = date(year, month, 1)
    last_day  = date(year, month, calendar.monthrange(year, month)[1])
    month_start = datetime(year, month, 1, 0, 0, 0, tzinfo=timezone.get_current_timezone())
    month_end   = datetime(year, month, last_day.day, 23, 59, 59, tzinfo=timezone.get_current_timezone())

    # 기본 완료 장애 QuerySet (completed_at 기준)
    base_qs = Incident.objects.filter(
        completed_at__year=year,
        completed_at__month=month,
        status='completed',
    ).exclude(fault_type__in=['', 'other']).select_related('school', 'category')

    FAULT_TYPE = dict(Incident.FAULT_TYPE_CHOICES)

    def _inc_list(qs):
        result = []
        for inc in qs.order_by('completed_at'):
            biz_min = None
            if inc.received_at and inc.completed_at:
                try:
                    from core.sla_utils import business_hours_elapsed_minutes
                    biz_min = round(business_hours_elapsed_minutes(inc.received_at, inc.completed_at), 1)
                except Exception:
                    biz_min = round((inc.completed_at - inc.received_at).total_seconds() / 60, 1)
            result.append({
                'id':             inc.id,
                'number':         inc.incident_number,
                'school':         inc.school.name if inc.school else '-',
                'fault_type':     FAULT_TYPE.get(inc.fault_type, inc.fault_type),
                'received_at':    timezone.localtime(inc.received_at).strftime('%m-%d %H:%M') if inc.received_at else '-',
                'completed_at':   timezone.localtime(inc.completed_at).strftime('%m-%d %H:%M') if inc.completed_at else '-',
                'biz_min':        biz_min,
                'is_human_error': inc.is_human_error,
                'is_recurrence':  inc.is_recurrence,
                'satisfaction':   inc.satisfaction_score,
            })
        return result

    # ── 지표별 분기 ──────────────────────────────────────────────────
    if indicator == 'uptime':
        from apps.incidents.models import IncidentSLA
        sla_qs = IncidentSLA.objects.filter(
            resolve_actual__year=year,
            resolve_actual__month=month,
            incident__fault_type='service_stop',
        ).select_related('incident__school')
        rows = []
        for sla in sla_qs.order_by('resolve_actual'):
            inc = sla.incident
            if inc.received_at and sla.resolve_actual:
                dur = round((sla.resolve_actual - inc.received_at).total_seconds() / 60, 1)
            else:
                dur = None
            rows.append({
                'id':           inc.id,
                'number':       inc.incident_number,
                'school':       inc.school.name if inc.school else '-',
                'received_at':  timezone.localtime(inc.received_at).strftime('%m-%d %H:%M') if inc.received_at else '-',
                'resolved_at':  timezone.localtime(sla.resolve_actual).strftime('%m-%d %H:%M') if sla.resolve_actual else '-',
                'duration_min': dur,
            })
        return JsonResponse({'indicator': 'uptime', 'rows': rows,
                             'columns': ['장애번호', '학교', '접수일시', '복구일시', '중단시간(분)']})

    elif indicator == 'inspection':
        from apps.progress.models import InspectionPlan, SchoolInspection
        plans = InspectionPlan.objects.filter(
            start_date__lte=last_day, end_date__gte=first_day
        )
        rows = []
        for plan in plans:
            for si in SchoolInspection.objects.filter(plan=plan).select_related('school'):
                rows.append({
                    'plan':       plan.name if hasattr(plan, 'name') else str(plan),
                    'school':     si.school.name if si.school else '-',
                    'scheduled':  si.scheduled_date.strftime('%m-%d') if hasattr(si, 'scheduled_date') and si.scheduled_date else '-',
                    'status':     si.get_status_display() if hasattr(si, 'get_status_display') else si.status,
                    'done':       si.status == 'completed',
                })
        return JsonResponse({'indicator': 'inspection', 'rows': rows,
                             'columns': ['점검계획', '학교', '예정일', '상태']})

    elif indicator in ('avg_fault', 'fault_count'):
        rows = _inc_list(base_qs)
        return JsonResponse({'indicator': indicator, 'rows': rows,
                             'columns': ['장애번호', '학교', '장애유형', '접수', '완료', '업무시간(분)']})

    elif indicator == 'overtime':
        overtime_rows = []
        for inc in base_qs.order_by('completed_at'):
            if inc.received_at and inc.completed_at:
                try:
                    from core.sla_utils import business_hours_elapsed_minutes
                    biz_min = round(business_hours_elapsed_minutes(inc.received_at, inc.completed_at), 1)
                except Exception:
                    biz_min = round((inc.completed_at - inc.received_at).total_seconds() / 60, 1)
                if biz_min > 480:
                    overtime_rows.append({
                        'id': inc.id, 'number': inc.incident_number,
                        'school': inc.school.name if inc.school else '-',
                        'fault_type': FAULT_TYPE.get(inc.fault_type, inc.fault_type),
                        'received_at': timezone.localtime(inc.received_at).strftime('%m-%d %H:%M'),
                        'completed_at': timezone.localtime(inc.completed_at).strftime('%m-%d %H:%M'),
                        'biz_min': biz_min,
                        'over_min': round(biz_min - 480, 1),
                    })
        return JsonResponse({'indicator': 'overtime', 'rows': overtime_rows,
                             'columns': ['장애번호', '학교', '장애유형', '접수', '완료', '업무시간(분)', '초과(분)']})

    elif indicator == 'human_error':
        rows = _inc_list(base_qs.filter(is_human_error=True))
        return JsonResponse({'indicator': 'human_error', 'rows': rows,
                             'columns': ['장애번호', '학교', '장애유형', '접수', '완료', '업무시간(분)']})

    elif indicator == 'recurrence':
        rows = _inc_list(base_qs.filter(is_recurrence=True))
        return JsonResponse({'indicator': 'recurrence', 'rows': rows,
                             'columns': ['장애번호', '학교', '장애유형', '접수', '완료', '업무시간(분)']})

    elif indicator == 'security':
        try:
            obj = SLAMonthly.objects.get(year=year, month=month)
            return JsonResponse({'indicator': 'security',
                                 'count': obj.security_count,
                                 'note':  obj.security_note or '(내용 없음)',
                                 'rows':  [], 'columns': []})
        except SLAMonthly.DoesNotExist:
            return JsonResponse({'error': '데이터 없음'}, status=404)

    elif indicator == 'satisfaction':
        sat_qs = Incident.objects.filter(
            completed_at__year=year,
            completed_at__month=month,
            satisfaction_score__isnull=False,
        ).select_related('school')
        rows = []
        for inc in sat_qs.order_by('completed_at'):
            rows.append({
                'id':           inc.id,
                'number':       inc.incident_number,
                'school':       inc.school.name if inc.school else '-',
                'completed_at': inc.completed_at.strftime('%m-%d %H:%M') if inc.completed_at else '-',
                'score':        inc.satisfaction_score,
                'score_pct':    round(inc.satisfaction_score / 5 * 100, 1) if inc.satisfaction_score else None,
            })
        return JsonResponse({'indicator': 'satisfaction', 'rows': rows,
                             'columns': ['장애번호', '학교', '완료일시', '만족도(점)', '환산(%)']})

    return JsonResponse({'error': '알 수 없는 지표'}, status=400)


@login_required
def sla_report_api(request):
    """POST: Claude AI를 활용한 SLA 지표별 분석 보고서 생성"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST only'}, status=405)

    import json
    from django.conf import settings

    try:
        body = json.loads(request.body)
        indicator = body.get('indicator', '')
        sla       = body.get('sla', {})
        rows      = body.get('rows', [])
        year      = body.get('year')
        month     = body.get('month')
    except Exception:
        return JsonResponse({'error': '잘못된 요청'}, status=400)

    api_key = getattr(settings, 'ANTHROPIC_API_KEY', '') or ''
    if not api_key:
        return JsonResponse(_static_report(indicator, sla, rows, year, month))

    # ── 지표 메타 ────────────────────────────────────────────────
    IND_META = {
        'uptime':      {'label': '장비 가동률',        'unit': '%',  'target': '99.98%', 'min': '99.85%'},
        'inspection':  {'label': '예방점검 준수율',     'unit': '%',  'target': '100%',  'min': '97%'},
        'avg_fault':   {'label': '평균 장애 조치시간',  'unit': '분', 'target': '240분', 'min': '480분'},
        'fault_count': {'label': '월간 장애건수',       'unit': '건', 'target': '20건',  'min': '30건'},
        'overtime':    {'label': '허용시간 초과 건수',  'unit': '건', 'target': '0건',   'min': '1건'},
        'human_error': {'label': '인적장애 건수',       'unit': '건', 'target': '0건',   'min': '1건'},
        'recurrence':  {'label': '반복장애 건수',       'unit': '건', 'target': '0건',   'min': '3건'},
        'security':    {'label': '보안위규 건수',        'unit': '건', 'target': '0건',   'min': '1건'},
        'satisfaction':{'label': '서비스 만족도',       'unit': '%',  'target': '100%',  'min': '97%'},
    }
    meta = IND_META.get(indicator, {'label': indicator, 'unit': '', 'target': '-', 'min': '-'})

    # ── 데이터 요약 텍스트 구성 ──────────────────────────────────
    score_val = sla.get(f'{indicator}_score') or sla.get('security_score') or sla.get('satisfaction_score')

    data_summary_lines = [
        f"- 지표명: {meta['label']}",
        f"- 분석 기간: {year}년 {month}월",
        f"- 목표 기준: {meta['target']}, 허용 최소: {meta['min']}",
    ]

    if indicator == 'uptime':
        data_summary_lines += [
            f"- 측정 가동률: {sla.get('uptime_pct', '-')}%",
            f"- 서비스 중단 손실시간: {sla.get('uptime_fault_min', 0)}분",
            f"- 월 총 가동시간: {sla.get('uptime_total_min', 0)}분",
            f"- 서비스 중단 장애 건수: {len(rows)}건",
        ]
    elif indicator == 'inspection':
        total = sla.get('inspection_total', 0)
        done  = sla.get('inspection_completed', 0)
        data_summary_lines += [
            f"- 예방점검 준수율: {sla.get('inspection_pct', '-')}%",
            f"- 전체 대상: {total}개교, 완료: {done}개교, 미완료: {total - done}개교",
        ]
        if rows:
            undone = [r for r in rows if not r.get('done')]
            data_summary_lines.append(f"- 미완료 학교 목록: {', '.join(r.get('school','') for r in undone[:10])}")
    elif indicator in ('avg_fault', 'fault_count'):
        over480 = [r for r in rows if (r.get('biz_min') or 0) > 480]
        schools = list({r.get('school','') for r in rows})
        fault_types = {}
        for r in rows:
            ft = r.get('fault_type','기타')
            fault_types[ft] = fault_types.get(ft, 0) + 1
        data_summary_lines += [
            f"- 총 장애건수: {sla.get('fault_count', len(rows))}건",
            f"- 평균 조치시간(업무시간 기준): {sla.get('avg_fault_min', '-')}분",
            f"- 480분 초과 건수: {len(over480)}건",
            f"- 장애 유형별 현황: {', '.join(f'{k} {v}건' for k,v in fault_types.items())}",
            f"- 발생 학교 수: {len(schools)}개교",
        ]
    elif indicator == 'overtime':
        if rows:
            avg_over = sum(r.get('over_min', 0) for r in rows) / len(rows)
            max_over = max(r.get('over_min', 0) for r in rows)
            data_summary_lines += [
                f"- 초과 장애 건수: {len(rows)}건",
                f"- 평균 초과시간: {avg_over:.0f}분, 최대 초과시간: {max_over:.0f}분",
                f"- 초과 학교 목록: {', '.join({r.get('school','') for r in rows})}",
                f"- 장애 유형: {', '.join(set(r.get('fault_type','') for r in rows))}",
            ]
        else:
            data_summary_lines.append("- 초과 장애 건수: 0건 (전건 허용 시간 내 처리)")
    elif indicator == 'human_error':
        fault_types = {}
        for r in rows:
            ft = r.get('fault_type','기타')
            fault_types[ft] = fault_types.get(ft, 0) + 1
        data_summary_lines += [
            f"- 인적장애 건수: {len(rows)}건",
            f"- 전체 장애 대비 비율: {sla.get('fault_count', 1)}건 중 {len(rows)}건",
            f"- 인적장애 유형: {', '.join(f'{k} {v}건' for k,v in fault_types.items()) if fault_types else '없음'}",
            f"- 발생 학교: {', '.join({r.get('school','') for r in rows}) if rows else '없음'}",
        ]
    elif indicator == 'recurrence':
        school_cnt = {}
        for r in rows:
            s = r.get('school','')
            school_cnt[s] = school_cnt.get(s, 0) + 1
        data_summary_lines += [
            f"- 반복장애 건수: {len(rows)}건",
            f"- 반복 발생 학교: {', '.join(f'{k}({v}회)' for k,v in sorted(school_cnt.items(), key=lambda x:-x[1])[:5])}",
        ]
    elif indicator == 'security':
        data_summary_lines += [
            f"- 보안위규 건수: {sla.get('security_count', 0)}건",
            f"- 위규 내용: {sla.get('security_note', '(내용 없음)')}",
        ]
    elif indicator == 'satisfaction':
        low = [r for r in rows if (r.get('score') or 5) <= 2]
        score_dist = {}
        for r in rows:
            s = str(r.get('score',''))
            score_dist[s] = score_dist.get(s, 0) + 1
        data_summary_lines += [
            f"- 서비스 만족도: {sla.get('satisfaction_pct', '-')}%",
            f"- 응답 건수: {sla.get('satisfaction_count', len(rows))}건",
            f"- 낮은 평가(2점 이하) 건수: {len(low)}건",
            f"- 점수 분포: {', '.join(f'{k}점 {v}건' for k,v in sorted(score_dist.items()))}",
        ]

    data_summary = '\n'.join(data_summary_lines)

    # ── Claude 프롬프트 ──────────────────────────────────────────
    system_prompt = """당신은 학교 네트워크 유지보수 용역 사업의 SLA 보고서 작성 전문가입니다.

[사업 배경]
- 서울시 교육청 산하 학교를 대상으로 네트워크 장비 유지보수 용역을 수행하는 사업입니다.
- 점검 항목은 계약에 의해 사전에 확정되어 있으며, 수행사가 임의로 변경할 수 없습니다.
- 이 사업은 장비 부품 교체나 신규 구매를 담당하는 사업이 아닙니다. 유지보수·장애대응·예방점검이 핵심입니다.
- 보고 대상은 교육청 담당 공무원입니다. 공식 행정 보고서 형식으로 작성합니다.

[보고서 작성 원칙]
1. 보고서는 과거 완료형으로 서술합니다. (~하였습니다, ~을 분석하였습니다, ~대책을 수립하였습니다)
2. "~이 필요합니다", "~하십시오", "~이 요구됩니다" 같은 지시형 문장은 절대 사용하지 않습니다.
3. 분석 항목에서는 어떻게 분석하였는지 방법을 구체적으로 서술합니다.
4. 조치 및 대책 항목에서는 어떤 방법으로 조치하고 어떤 대책을 수립하였는지를 서술합니다.
5. 부품 교체, 장비 구매, 점검 항목 변경 등 계약 범위 밖의 내용은 절대 포함하지 않습니다.
6. 각 항목은 2~3문장으로 간결하게 서술합니다.

[출력 형식]
다음 JSON 형식으로만 응답하세요. 마크다운, 코드블록, 추가 설명 없이 순수 JSON만 출력하세요:
{
  "summary": "총평 (1~2문장, 측정값과 목표 대비 결과 서술)",
  "detail": "세부 분석 (2~3문장, 어떻게 분석하였는지 방법 포함)",
  "action": "조치 및 대책 (2~3문장, 어떤 방법으로 조치하고 대책을 수립하였는지 포함)"
}"""

    user_prompt = f"""다음 SLA 지표 데이터를 분석하여 교육청 보고용 분석 보고서를 작성해주세요.

{data_summary}

위 데이터를 바탕으로 총평, 세부 분석(어떻게 분석하였는지), 조치 및 대책(어떻게 조치하였는지)을 JSON으로 작성하세요."""

    try:
        import anthropic, re
        client = anthropic.Anthropic(api_key=api_key)
        message = client.messages.create(
            model='claude-haiku-4-5-20251001',
            max_tokens=1024,
            system=system_prompt,
            messages=[{'role': 'user', 'content': user_prompt}],
        )
        raw = message.content[0].text.strip()
        json_match = re.search(r'\{[\s\S]*\}', raw)
        if json_match:
            report = json.loads(json_match.group())
        else:
            raise ValueError('JSON 응답 없음')

        return JsonResponse({
            'summary': report.get('summary', ''),
            'detail':  report.get('detail', ''),
            'action':  report.get('action', ''),
        })

    except Exception:
        # AI 호출 실패 시 정적 보고서로 폴백
        return JsonResponse(_static_report(indicator, sla, rows, year, month))


# ─────────────────────────────────────────
# REST API
# ─────────────────────────────────────────
class IncidentCategoryViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = IncidentCategory.objects.filter(is_active=True).prefetch_related('subcategories')
    serializer_class = IncidentCategorySerializer
    permission_classes = [permissions.IsAuthenticated]

    @action(detail=False, methods=['get'])
    def csv_template(self, request):
        """장애 분류 CSV 양식 다운로드"""
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = 'attachment; filename="incident_categories_template.csv"'
        writer = csv.writer(response)
        writer.writerow(['대분류코드', '대분류명', '소분류명', '정렬순서'])
        writer.writerow(['wired', '유선망', '단선', '1'])
        writer.writerow(['wired', '유선망', '접속불가', '2'])
        writer.writerow(['wireless', '무선망', 'AP 불량', '1'])
        return response

    @action(detail=False, methods=['post'], permission_classes=[IsAdmin])
    def csv_upload(self, request):
        """장애 분류 CSV 업로드 (대분류+소분류)"""
        import io
        f = request.FILES.get('file')
        mode = request.data.get('mode', 'add_update')
        if not f:
            return Response({'error': 'CSV 파일을 첨부해 주세요.'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            text = f.read().decode('utf-8-sig')
        except Exception:
            return Response({'error': 'UTF-8 인코딩 파일만 지원합니다.'}, status=status.HTTP_400_BAD_REQUEST)

        reader = csv.DictReader(io.StringIO(text))
        rows = list(reader)

        if mode == 'reset':
            IncidentSubcategory.objects.all().delete()
            IncidentCategory.objects.all().delete()

        created, updated, errors = 0, 0, []
        for row_num, row in enumerate(rows, 2):
            code = (row.get('대분류코드') or '').strip()
            cat_name = (row.get('대분류명') or '').strip()
            sub_name = (row.get('소분류명') or '').strip()
            order = int((row.get('정렬순서') or '0').strip() or 0)
            if not code or not cat_name:
                errors.append({'row': row_num, 'error': '대분류코드/대분류명 필수'})
                continue
            try:
                cat, cat_new = IncidentCategory.objects.get_or_create(
                    code=code, defaults={'name': cat_name, 'order': order}
                )
                if not cat_new:
                    cat.name = cat_name; cat.save()
                if sub_name:
                    _, sub_new = IncidentSubcategory.objects.get_or_create(
                        category=cat, name=sub_name, defaults={'order': order}
                    )
                    created += 1 if sub_new else 0
                    updated += 0 if sub_new else 1
                else:
                    created += 1 if cat_new else 0
                    updated += 0 if cat_new else 1
            except Exception as e:
                errors.append({'row': row_num, 'error': str(e)})
        return Response({'created': created, 'updated': updated, 'errors': errors})


class IncidentViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        qs = Incident.objects.select_related(
            'school', 'school__support_center', 'school__school_type',
            'category', 'subcategory', 'received_by', 'sla'
        ).prefetch_related('assignments__worker').order_by('-received_at')

        params = self.request.query_params
        center      = params.get('center')
        school      = params.get('school')
        school_type = params.get('school_type')
        status_     = params.get('status')
        priority    = params.get('priority')
        category    = params.get('category')
        date_from   = params.get('date_from')
        date_to     = params.get('date_to')
        q           = params.get('q')

        if center:
            qs = qs.filter(school__support_center__code=center)
        if school:
            qs = qs.filter(school_id=school)
        if school_type:
            qs = qs.filter(school__school_type_id=school_type)
        if status_:
            qs = qs.filter(status=status_)
        if priority:
            qs = qs.filter(priority=priority)
        if category:
            qs = qs.filter(category__code=category)
        if date_from:
            qs = qs.filter(received_at__date__gte=date_from)
        if date_to:
            qs = qs.filter(received_at__date__lte=date_to)
        if q:
            qs = qs.filter(
                Q(incident_number__icontains=q) |
                Q(school__name__icontains=q) |
                Q(requester_name__icontains=q) |
                Q(description__icontains=q)
            )
        # 현장기사는 자신에게 배정된 건만
        user = self.request.user
        if user.role == 'worker':
            qs = qs.filter(assignments__worker=user)
        return qs

    def get_serializer_class(self):
        if self.action == 'list':
            return IncidentListSerializer
        if self.action == 'create':
            return IncidentCreateSerializer
        if self.action in ('update', 'partial_update'):
            return IncidentUpdateSerializer
        return IncidentDetailSerializer

    def get_permissions(self):
        if self.action == 'destroy':
            return [IsAdmin()]
        return super().get_permissions()

    # ── 상태 변경 ──────────────────────────
    @action(detail=True, methods=['post'])
    def change_status(self, request, pk=None):
        incident = self.get_object()
        new_status = request.data.get('status')
        allowed = ['received', 'assigned', 'moving', 'arrived', 'processing', 'completed', 'cancelled']
        if new_status not in allowed:
            return Response({'error': '올바르지 않은 상태입니다.'}, status=status.HTTP_400_BAD_REQUEST)

        update_data = {'status': new_status}

        # 도착 처리 시 arrived_at 자동 기록
        if new_status == 'arrived' and not incident.arrived_at:
            update_data['arrived_at'] = timezone.now()

        # 완료 처리 시 추가 필드 저장
        if new_status == 'completed':
            if request.data.get('resolution'):
                update_data['resolution'] = request.data['resolution']
            if request.data.get('resolution_type'):
                update_data['resolution_type'] = request.data['resolution_type']
            if request.data.get('fault_type'):
                update_data['fault_type'] = request.data['fault_type']
            if 'is_human_error' in request.data:
                update_data['is_human_error'] = request.data['is_human_error']
            if request.data.get('arrived_at') and not incident.arrived_at:
                update_data['arrived_at'] = request.data['arrived_at']
            if request.data.get('completed_at'):
                update_data['completed_at'] = request.data['completed_at']
            else:
                update_data['completed_at'] = timezone.now()

        serializer = IncidentUpdateSerializer(
            incident, data=update_data, partial=True, context={'request': request}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(IncidentDetailSerializer(incident).data)

    # ── 고객 협의 방문 약속 (SLA 조정) ──────────────────────────
    @action(detail=True, methods=['post'])
    def set_appointment(self, request, pk=None):
        """고객 협의로 방문 약속시간 설정 → SLA 기준시간 조정"""
        from django.conf import settings as django_settings
        from core.sla_utils import add_business_hours

        incident = self.get_object()
        appointment_at   = request.data.get('appointment_at')
        customer_call_at = request.data.get('customer_call_at')
        customer_call_note = request.data.get('customer_call_note', '')

        if not appointment_at:
            return Response({'error': '방��� 약속시간을 입력하세요.'}, status=status.HTTP_400_BAD_REQUEST)
        if not customer_call_at:
            return Response({'error': '고객 통���시간을 입력하세요.'}, status=status.HTTP_400_BAD_REQUEST)

        # Incident 필드 저장
        incident.appointment_at = appointment_at
        incident.customer_call_at = customer_call_at
        incident.customer_call_note = customer_call_note
        incident.save(update_fields=['appointment_at', 'customer_call_at', 'customer_call_note'])

        # SLA 재계산: 약속시간 기준으로 도착 목표 = 약속시간, 처리 목표 = 약속시간 + 8h
        sla_resolve = getattr(django_settings, 'SLA_RESOLVE_HOURS', 8)
        try:
            sla = incident.sla
            sla.arrival_target = appointment_at
            sla.resolve_target = add_business_hours(appointment_at, sla_resolve)
            sla.is_adjusted = True
            update_fields = ['arrival_target', 'resolve_target', 'is_adjusted']
            # 이미 도착/완료 기록이 있으면 재판정
            if sla.arrival_actual:
                sla.arrival_ok = sla.arrival_actual <= sla.arrival_target
                sla.arrival_diff_min = int((sla.arrival_actual - sla.arrival_target).total_seconds() / 60)
                update_fields += ['arrival_ok', 'arrival_diff_min']
                incident.sla_arrival_ok = sla.arrival_ok
                incident.save(update_fields=['sla_arrival_ok'])
            if sla.resolve_actual:
                sla.resolve_ok = sla.resolve_actual <= sla.resolve_target
                sla.resolve_diff_min = int((sla.resolve_actual - sla.resolve_target).total_seconds() / 60)
                update_fields += ['resolve_ok', 'resolve_diff_min']
                incident.sla_resolve_ok = sla.resolve_ok
                incident.save(update_fields=['sla_resolve_ok'])
            sla.save(update_fields=update_fields)
        except IncidentSLA.DoesNotExist:
            IncidentSLA.objects.create(
                incident=incident,
                arrival_target=appointment_at,
                resolve_target=add_business_hours(appointment_at, sla_resolve),
                is_adjusted=True,
            )

        return Response(IncidentDetailSerializer(incident).data)

    # ── 인력 배정 ──────────────────────────
    @action(detail=True, methods=['post'])
    def assign(self, request, pk=None):
        incident = self.get_object()
        worker_id = request.data.get('worker_id')
        if not worker_id:
            return Response({'error': '인력을 선택하세요.'}, status=status.HTTP_400_BAD_REQUEST)
        from apps.accounts.models import User
        worker = get_object_or_404(User, id=worker_id)
        assign = create_assignment(incident, worker, request.user)
        return Response(IncidentAssignmentSerializer(assign).data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'])
    def ai_assign(self, request, pk=None):
        """AI 자동 인력 배정"""
        incident = self.get_object()
        result = ai_assign_worker(incident)
        if not result:
            # AI 서버 없을 시 거리 기반 자동 배정
            workers = get_available_workers(incident)
            if not workers.exists():
                return Response({'error': '배정 가능한 인력이 없습니다.'}, status=status.HTTP_404_NOT_FOUND)
            worker, dist_km = get_best_worker(incident, workers)
            assign = create_assignment(incident, worker, request.user, is_ai=True,
                                       distance_km=dist_km)
        else:
            from apps.accounts.models import User
            worker = get_object_or_404(User, id=result['worker_id'])
            assign = create_assignment(
                incident, worker, request.user, is_ai=True,
                distance_km=result.get('distance_km'),
                eta_minutes=result.get('eta_minutes')
            )
        return Response(IncidentAssignmentSerializer(assign).data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['get'])
    def available_workers(self, request, pk=None):
        """배정 가능 인력 목록"""
        incident = self.get_object()
        from apps.accounts.serializers import UserListSerializer
        workers = get_available_workers(incident)
        return Response(UserListSerializer(workers, many=True).data)

    # ── 댓글 ──────────────────────────────
    @action(detail=True, methods=['get', 'post'])
    def comments(self, request, pk=None):
        incident = self.get_object()
        if request.method == 'GET':
            comments = incident.comments.all()
            return Response(IncidentCommentSerializer(comments, many=True).data)
        serializer = IncidentCommentSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save(incident=incident, author=request.user)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    # ── 사진 ──────────────────────────────
    @action(detail=True, methods=['get', 'post'])
    def photos(self, request, pk=None):
        incident = self.get_object()
        if request.method == 'GET':
            return Response(IncidentPhotoSerializer(incident.photos.all(), many=True).data)
        serializer = IncidentPhotoSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save(incident=incident, uploaded_by=request.user)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['delete'], url_path='photos/(?P<photo_id>[0-9]+)')
    def photo_delete(self, request, pk=None, photo_id=None):
        incident = self.get_object()
        photo = incident.photos.filter(id=photo_id).first()
        if not photo:
            return Response({'error': '사진을 찾을 수 없습니다.'}, status=status.HTTP_404_NOT_FOUND)
        if photo.image:
            import os
            path = photo.image.path
            if os.path.exists(path):
                os.remove(path)
        photo.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    # ── 처리 이력(타임라인) ────────────────
    @action(detail=True, methods=['get'])
    def timeline(self, request, pk=None):
        incident = self.get_object()
        STATUS_LABELS = {
            'received':   '접수', 'assigned':   '배정',
            'moving':     '이동', 'arrived':    '도착',
            'processing': '처리중', 'completed':  '완료', 'cancelled': '취소',
        }
        history = incident.status_history.all().order_by('changed_at')
        items = []
        for h in history:
            items.append({
                'from_status': h.from_status,
                'to_status':   h.to_status,
                'to_label':    STATUS_LABELS.get(h.to_status, h.to_status),
                'changed_by':  h.changed_by.name if h.changed_by else '시스템',
                'note':        h.note,
                'changed_at':  h.changed_at.isoformat(),
            })
        # 최초 접수 이벤트 (status_history에 없을 수 있음)
        if incident.received_at and (not items or items[0]['to_status'] != 'received'):
            items.insert(0, {
                'from_status': '',
                'to_status':   'received',
                'to_label':    '접수',
                'changed_by':  incident.received_by.name if incident.received_by else '시스템',
                'note':        '',
                'changed_at':  incident.received_at.isoformat(),
            })
        return Response(items)

    # ── PDF 생성 ──────────────────────────
    @action(detail=True, methods=['post'])
    def generate_pdf(self, request, pk=None):
        incident = self.get_object()
        from .services import generate_incident_pdf
        try:
            generate_incident_pdf(incident.id)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        download_url = f'/npms/api/incidents/incidents/{incident.id}/download_pdf/'
        return Response({'message': 'PDF가 생성되었습니다.', 'download_url': download_url})

    @action(detail=True, methods=['get'])
    def download_pdf(self, request, pk=None):
        import os
        from django.http import FileResponse
        incident = self.get_object()
        path = incident.report_pdf_path
        if not path or not os.path.exists(path):
            return Response({'error': '보고서가 없습니다. 먼저 생성해주세요.'}, status=status.HTTP_404_NOT_FOUND)
        filename = os.path.basename(path)
        response = FileResponse(open(path, 'rb'), content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="{filename}"'
        return response

    # ── 지연처리사유서 ──────────────────────
    @action(detail=True, methods=['get', 'post'])
    def delay_reason(self, request, pk=None):
        incident = self.get_object()

        if request.method == 'GET':
            try:
                dr = incident.delay_reason
                return Response({
                    'id': dr.id,
                    'reason': dr.reason,
                    'sig_worker_org': dr.sig_worker_org,
                    'sig_worker_name': dr.sig_worker_name,
                    'sig_worker_phone': dr.sig_worker_phone,
                    'sig_school_org': dr.sig_school_org,
                    'sig_school_name': dr.sig_school_name,
                    'sig_school_phone': dr.sig_school_phone,
                    'pdf_path': dr.pdf_path,
                    'created_at': dr.created_at,
                })
            except IncidentDelayReason.DoesNotExist:
                return Response({'exists': False})

        # POST — 사유서 저장 + PDF 생성 + SLA 인정
        reason = request.data.get('reason', '').strip()
        if not reason:
            return Response({'error': '지연 사유를 입력하세요.'}, status=status.HTTP_400_BAD_REQUEST)

        sig_worker_data = request.data.get('sig_worker_data', '')
        sig_school_data = request.data.get('sig_school_data', '')
        if not sig_worker_data:
            return Response({'error': '처리자 서명이 필요합니다.'}, status=status.HTTP_400_BAD_REQUEST)
        if not sig_school_data:
            return Response({'error': '담당 선생님 서명이 필요합니다.'}, status=status.HTTP_400_BAD_REQUEST)

        dr, created = IncidentDelayReason.objects.update_or_create(
            incident=incident,
            defaults={
                'reason': reason,
                'sig_worker_org':   request.data.get('sig_worker_org', '세종아이티엘 컨소시엄'),
                'sig_worker_name':  request.data.get('sig_worker_name', ''),
                'sig_worker_phone': request.data.get('sig_worker_phone', ''),
                'sig_worker_data':  sig_worker_data,
                'sig_school_org':   request.data.get('sig_school_org', ''),
                'sig_school_name':  request.data.get('sig_school_name', ''),
                'sig_school_phone': request.data.get('sig_school_phone', ''),
                'sig_school_data':  sig_school_data,
                'created_by':       request.user,
            }
        )

        # PDF 생성
        from .services import generate_delay_reason_pdf
        try:
            generate_delay_reason_pdf(dr)
        except Exception as e:
            logger.error(f'지연처리사유서 PDF 생성 실패: {e}')

        # SLA 인정 처리 — 기준시간 내 처리로 갱신
        try:
            sla = incident.sla
            sla.resolve_ok = True
            sla.save(update_fields=['resolve_ok'])
            incident.sla_resolve_ok = True
            incident.save(update_fields=['sla_resolve_ok'])
        except IncidentSLA.DoesNotExist:
            pass

        download_url = f'/npms/api/incidents/incidents/{incident.id}/download_delay_pdf/'
        return Response({
            'message': '지연처리사유서가 저장되었습니다. SLA 준수로 인정됩니다.',
            'download_url': download_url,
        })

    @action(detail=True, methods=['get'])
    def download_delay_pdf(self, request, pk=None):
        import os
        from django.http import FileResponse
        incident = self.get_object()
        try:
            dr = incident.delay_reason
        except IncidentDelayReason.DoesNotExist:
            return Response({'error': '지연처리사유서가 없습니다.'}, status=status.HTTP_404_NOT_FOUND)
        path = dr.pdf_path
        if not path or not os.path.exists(path):
            return Response({'error': 'PDF 파일이 없습니다.'}, status=status.HTTP_404_NOT_FOUND)
        filename = os.path.basename(path)
        response = FileResponse(open(path, 'rb'), content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="{filename}"'
        return response

    # ── CSV 다운로드 ──────────────────────
    @action(detail=False, methods=['get'])
    def csv_download(self, request):
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = 'attachment; filename="incidents.csv"'
        writer = csv.writer(response)
        writer.writerow(['접수번호', '접수일시', '교육지원청', '학교명', '장애분류',
                         '소분류', '상태', '긴급도', '접수자', '요청자', '연락처',
                         '도착일시', '완료일시', '소요시간(분)'])
        for i in self.get_queryset():
            writer.writerow([
                i.incident_number, i.received_at.strftime('%Y-%m-%d %H:%M'),
                i.school.support_center.name, i.school.name,
                i.category.name, i.subcategory.name if i.subcategory else '',
                i.get_status_display(), i.get_priority_display(),
                i.received_by.name if i.received_by else '',
                i.requester_name, i.requester_phone,
                i.arrived_at.strftime('%Y-%m-%d %H:%M') if i.arrived_at else '',
                i.completed_at.strftime('%Y-%m-%d %H:%M') if i.completed_at else '',
                i.get_elapsed_minutes(),
            ])
        return response

    # ── Excel 다운로드 ─────────────────────
    @action(detail=False, methods=['get'])
    def excel_download(self, request):
        """장애 목록 Excel 다운로드 (현재 필터 반영)"""
        try:
            import io
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            return HttpResponse('openpyxl 패키지가 필요합니다.', status=500)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '장애 목록'

        hdr_fill = PatternFill('solid', fgColor='1F497D')
        hdr_font = Font(bold=True, color='FFFFFF', size=10)
        ctr = Alignment(horizontal='center', vertical='center', wrap_text=False)
        thin = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'),  bottom=Side(style='thin'))

        STATUS_MAP = {
            'received': '접수', 'assigned': '배정', 'moving': '이동',
            'arrived': '도착', 'processing': '처리중', 'completed': '완료',
        }
        PRIORITY_MAP = {
            'critical': '긴급', 'high': '높음', 'medium': '보통', 'low': '낮음',
        }

        headers = ['접수번호', '접수일시', '교육지원청', '학교명', '장애분류', '소분류',
                   '상태', '긴급도', '담당자', '요청자', '연락처',
                   '현장도착', '처리완료', '소요시간(분)', '처리내용']
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = ctr
            c.border = thin

        ws.row_dimensions[1].height = 20
        col_widths = [14,16,8,14,10,10,8,6,8,8,12,16,16,12,30]
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

        for r, i in enumerate(self.get_queryset(), 2):
            assigned = (i.assignments.filter(is_accepted=True)
                        .select_related('worker').first())
            worker_name = assigned.worker.name if assigned else ''
            vals = [
                i.incident_number,
                i.received_at.strftime('%Y-%m-%d %H:%M'),
                i.school.support_center.name if i.school.support_center else '',
                i.school.name if i.school else '',
                i.category.name if i.category else '',
                i.subcategory.name if i.subcategory else '',
                STATUS_MAP.get(i.status, i.status),
                PRIORITY_MAP.get(i.priority, i.priority),
                worker_name,
                i.requester_name or '',
                i.requester_phone or '',
                i.arrived_at.strftime('%Y-%m-%d %H:%M') if i.arrived_at else '',
                i.completed_at.strftime('%Y-%m-%d %H:%M') if i.completed_at else '',
                i.get_elapsed_minutes() or '',
                i.resolution or '',
            ]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=r, column=ci, value=v)
                c.alignment = ctr if ci in (7, 8) else Alignment(vertical='center')
                c.border = thin

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        from django.utils import timezone as tz
        fname = f'incidents_{tz.localdate().isoformat()}.xlsx'
        resp = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = f'attachment; filename="{fname}"'
        return resp

    # ── 작업지시서 ─────────────────────────
    @action(detail=True, methods=['get', 'post'], url_path='work_orders')
    def work_orders(self, request, pk=None):
        """장애에 연결된 작업지시서 목록 조회 / 생성"""
        from .models import WorkOrder
        from .serializers import WorkOrderSerializer
        incident = self.get_object()

        if request.method == 'GET':
            wos = incident.work_orders.select_related('assigned_to', 'created_by', 'confirmed_by').all()
            return Response(WorkOrderSerializer(wos, many=True).data)

        # POST — 작업지시서 생성
        data = request.data.copy()
        wo = WorkOrder.objects.create(
            incident=incident,
            school=incident.school,
            work_order_number=WorkOrder.generate_number(),
            title=data.get('title') or f'[{incident.category.name}] {incident.school.name} 장애처리',
            work_description=data.get('work_description') or incident.description,
            work_type=data.get('work_type', 'repair'),
            assigned_to_id=data.get('assigned_to') or (
                incident.assignments.filter(is_accepted=True).values_list('worker_id', flat=True).first()
            ),
            due_date=data.get('due_date'),
            required_parts=data.get('required_parts', ''),
            created_by=request.user,
        )
        # 발행 즉시 PDF 자동 생성
        try:
            from .services import generate_work_order_pdf
            generate_work_order_pdf(wo)
        except Exception as e:
            logger.error(f'작업지시서 PDF 생성 실패: {e}')
        return Response(WorkOrderSerializer(wo).data, status=status.HTTP_201_CREATED)

    # ── CSV 업로드 (기존 장애 일괄 등록) ──────
    @action(detail=False, methods=['post'])
    def csv_upload(self, request):
        """기존 시스템 장애 데이터 CSV 일괄 업로드"""
        if request.user.role not in ('superadmin', 'admin'):
            return Response({'error': '권한이 없습니다.'}, status=status.HTTP_403_FORBIDDEN)

        file = request.FILES.get('file')
        if not file:
            return Response({'error': '파일을 선택하세요.'}, status=status.HTTP_400_BAD_REQUEST)

        # 인코딩 자동 감지 (EUC-KR / UTF-8)
        raw = file.read()
        text = None
        for enc in ('utf-8-sig', 'euc-kr', 'cp949', 'utf-8'):
            try:
                text = raw.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        if text is None:
            return Response({'error': '파일 인코딩을 인식할 수 없습니다.'}, status=status.HTTP_400_BAD_REQUEST)

        import io
        import datetime
        reader = csv.DictReader(io.StringIO(text))

        # ── 캐시 사전 로딩 ──
        from apps.schools.models import School
        from apps.accounts.models import User as UserModel
        school_cache = {}
        for s in School.objects.select_related('support_center').all():
            school_cache[s.name] = s

        cat_cache = {c.name: c for c in IncidentCategory.objects.all()}
        subcat_cache = {}
        for sc in IncidentSubcategory.objects.select_related('category').all():
            subcat_cache[(sc.category_id, sc.name)] = sc

        # 처리자 이름 → User 캐시 (이름 기준)
        worker_cache = {}
        for u in UserModel.objects.filter(role__in=('worker', 'resident', 'admin', 'superadmin')):
            worker_cache[u.name] = u

        STATUS_MAP = {
            '완료': 'completed', '처리완료': 'completed',
            '처리중': 'processing', '이동중': 'moving', '이동': 'moving',
            '배정': 'assigned', '접수': 'received', '도착': 'arrived',
            '취소': 'cancelled',
        }
        CONTACT_MAP = {
            '전화': 'phone', '방문': 'visit', '시스템': 'system',
            '이메일': 'email', '자동감지': 'auto', '자동': 'auto',
        }

        def parse_dt(val):
            if not val or not str(val).strip():
                return None
            val = str(val).strip()
            for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y-%m-%d',
                        '%Y/%m/%d %H:%M:%S', '%Y/%m/%d %H:%M', '%Y/%m/%d',
                        '%Y.%m.%d %H:%M', '%Y.%m.%d'):
                try:
                    dt = datetime.datetime.strptime(val, fmt)
                    return timezone.make_aware(dt)
                except ValueError:
                    continue
            return None

        created = skipped = errors = 0
        error_rows = []

        for row_num, row in enumerate(reader, start=2):
            try:
                # 학교 찾기
                school_name = (row.get('학교') or '').strip()
                school = school_cache.get(school_name)
                if not school:
                    errors += 1
                    error_rows.append(f'행{row_num}: 학교 "{school_name}" 없음')
                    continue

                # 대분류
                cat_name = (row.get('분류') or '').strip()
                category = cat_cache.get(cat_name)
                if not category and cat_name:
                    for n, c in cat_cache.items():
                        if cat_name in n or n in cat_name:
                            category = c
                            break
                if not category:
                    errors += 1
                    error_rows.append(f'행{row_num}: 분류 "{cat_name}" 없음')
                    continue

                # 소분류
                subcat_name = (row.get('세부분류') or '').strip()
                subcategory = subcat_cache.get((category.id, subcat_name)) if subcat_name else None

                # 접수번호
                inc_num = (row.get('접수번호') or '').strip()
                received_at = parse_dt(row.get('등록일') or '') or timezone.now()
                if not inc_num:
                    inc_num = Incident.generate_number(received_at)

                # 중복 체크 (접수번호 기준)
                if Incident.objects.filter(incident_number=inc_num).exists():
                    skipped += 1
                    continue

                inc_status = STATUS_MAP.get((row.get('진행사항') or '완료').strip(), 'completed')
                contact_method = CONTACT_MAP.get((row.get('접수방법') or '전화').strip(), 'phone')

                arrived_at   = parse_dt(row.get('현장 도착시간') or '')
                completed_at = parse_dt(row.get('완료일') or '')

                # 처리자 → User 매핑
                worker_name = (row.get('처리자') or '').strip()
                matched_worker = worker_cache.get(worker_name) if worker_name else None

                # 처리내용 조합 (DB에 매핑된 처리자는 텍스트에서 제외)
                resolution_parts = []
                if worker_name and not matched_worker:
                    resolution_parts.append(f'[처리자] {worker_name}')
                if row.get('처리내용'):
                    resolution_parts.append(row['처리내용'].strip())
                if row.get('조치방법'):
                    resolution_parts.append(f'[조치방법] {row["조치방법"].strip()}')
                resolution = '\n'.join(filter(None, resolution_parts))

                # 장애내용 조합
                desc_parts = []
                if row.get('장애 증상'):
                    desc_parts.append(row['장애 증상'].strip())
                if row.get('장애내용'):
                    desc_parts.append(row['장애내용'].strip())
                description = '\n'.join(filter(None, desc_parts)) or '-'

                incident = Incident.objects.create(
                    incident_number  = inc_num,
                    school           = school,
                    category         = category,
                    subcategory      = subcategory,
                    status           = inc_status,
                    priority         = 'medium',
                    contact_method   = contact_method,
                    requester_name   = (row.get('요청자') or '-').strip(),
                    requester_phone  = '',
                    description      = description,
                    received_at      = received_at,
                    arrived_at       = arrived_at,
                    completed_at     = completed_at,
                    resolution       = resolution,
                    resolution_type  = (row.get('처리유형') or '').strip(),
                )

                # 처리자 → IncidentAssignment 생성
                if matched_worker:
                    IncidentAssignment.objects.create(
                        incident    = incident,
                        worker      = matched_worker,
                        assigned_by = request.user,
                        note        = 'CSV 업로드 일괄등록',
                        arrived_at  = arrived_at,
                        completed_at= completed_at,
                    )

                created += 1

            except Exception as e:
                errors += 1
                error_rows.append(f'행{row_num}: {str(e)}')

        return Response({
            'created': created,
            'skipped': skipped,
            'errors': errors,
            'error_rows': error_rows[:20],
        })

    # ── CSV 샘플 다운로드 ─────────────────────
    @action(detail=False, methods=['get'])
    def csv_sample(self, request):
        """업로드용 CSV 샘플 파일 다운로드"""
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = 'attachment; filename="incident_upload_sample.csv"'
        writer = csv.writer(response)
        writer.writerow([
            '접수번호', '진행사항', '등록일', '접수방법',
            '지원청', '학교', '분류', '세부분류',
            '장애 증상', '요청자', '접수자', '처리자',
            '장애내용', '조치방법', '처리유형', '처리내용',
            '현장 도착시간', '완료일', '장애처리 소요시간', '년', '월',
        ])
        writer.writerow([
            '20230301_001', '완료', '2023-03-01 09:00', '전화',
            '동부', '○○초등학교', '유선망', 'LAN포트 불량',
            '인터넷 안됨', '홍길동', '김철수', '이영희',
            '교실 스위치 포트 불량', '포트 교체', '부품교체', '스위치 포트 1번 교체 완료',
            '2023-03-01 10:30', '2023-03-01 11:00', '120', '2023', '3',
        ])
        return response

    # ── 실시간 카운트 (대시보드용) ──────────
    @action(detail=False, methods=['get'])
    def summary(self, request):
        qs = Incident.objects.all()
        return Response({
            'total':      qs.count(),
            'received':   qs.filter(status='received').count(),
            'assigned':   qs.filter(status='assigned').count(),
            'processing': qs.filter(status__in=['moving','arrived','processing']).count(),
            'completed':  qs.filter(status='completed').count(),
            'today':      qs.filter(received_at__date=timezone.localdate()).count(),
        })


class IncidentAssignmentViewSet(viewsets.ModelViewSet):
    queryset = IncidentAssignment.objects.select_related('incident', 'worker', 'assigned_by')
    serializer_class = IncidentAssignmentSerializer
    permission_classes = [permissions.IsAuthenticated]

    @action(detail=True, methods=['post'])
    def accept(self, request, pk=None):
        assign = self.get_object()
        assign.is_accepted = True
        assign.accepted_at = timezone.now()
        assign.save()
        assign.incident.status = 'moving'
        assign.incident.save(update_fields=['status'])
        return Response({'status': '수락 완료'})

    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        assign = self.get_object()
        assign.is_accepted = False
        assign.reject_reason = request.data.get('reason', '')
        assign.save()
        return Response({'status': '거부 완료'})

    @action(detail=True, methods=['post'])
    def arrived(self, request, pk=None):
        assign = self.get_object()
        now = timezone.now()
        assign.arrived_at = now
        assign.save()
        incident = assign.incident
        incident.arrived_at = now
        incident.status = 'arrived'
        incident.save(update_fields=['arrived_at', 'status'])
        # SLA 도착 갱신
        try:
            sla = incident.sla
            sla.arrival_actual = now
            sla.arrival_ok = now <= sla.arrival_target
            sla.arrival_diff_min = int((now - sla.arrival_target).total_seconds() / 60)
            sla.save()
        except Exception as e:
            logger.warning('SLA 도착 갱신 실패 incident=%s: %s', incident.pk, e)
        return Response({'status': '도착 등록 완료', 'arrived_at': str(now)})


class SLARuleViewSet(viewsets.ModelViewSet):
    queryset = SLARule.objects.all()
    serializer_class = SLARuleSerializer
    permission_classes = [IsAdmin]

    @action(detail=False, methods=['get'])
    def csv_template(self, request):
        """SLA 규칙 CSV 양식 다운로드"""
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = 'attachment; filename="sla_rules_template.csv"'
        writer = csv.writer(response)
        writer.writerow(['규칙명', '도착기준(시간)', '처리기준(시간)', '적용시작일', '적용종료일', '적용중'])
        writer.writerow(['기본 SLA', '2', '8', '2026-01-01', '', 'Y'])
        return response

    @action(detail=False, methods=['post'])
    def csv_upload(self, request):
        """SLA 규칙 CSV 업로드"""
        import io
        f = request.FILES.get('file')
        mode = request.data.get('mode', 'add_update')
        if not f:
            return Response({'error': 'CSV 파일을 첨부해 주세요.'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            text = f.read().decode('utf-8-sig')
        except Exception:
            return Response({'error': 'UTF-8 인코딩 파일만 지원합니다.'}, status=status.HTTP_400_BAD_REQUEST)

        reader = csv.DictReader(io.StringIO(text))
        rows = list(reader)

        if mode == 'reset':
            SLARule.objects.all().delete()

        created, updated, errors = 0, 0, []
        for row_num, row in enumerate(rows, 2):
            name = (row.get('규칙명') or '').strip()
            if not name:
                errors.append({'row': row_num, 'error': '규칙명 필수'})
                continue
            try:
                apply_from = (row.get('적용시작일') or '2026-01-01').strip()
                apply_to = (row.get('적용종료일') or '').strip() or None
                is_active = (row.get('적용중') or 'Y').strip().upper() == 'Y'
                obj, is_new = SLARule.objects.get_or_create(
                    name=name,
                    defaults={
                        'arrival_hours': int((row.get('도착기준(시간)') or '2').strip()),
                        'resolve_hours': int((row.get('처리기준(시간)') or '8').strip()),
                        'apply_from': apply_from, 'apply_to': apply_to,
                        'is_active': is_active, 'created_by': request.user,
                    }
                )
                if not is_new:
                    obj.arrival_hours = int((row.get('도착기준(시간)') or '2').strip())
                    obj.resolve_hours = int((row.get('처리기준(시간)') or '8').strip())
                    obj.apply_from = apply_from; obj.apply_to = apply_to
                    obj.is_active = is_active; obj.save()
                created += 1 if is_new else 0
                updated += 0 if is_new else 1
            except Exception as e:
                errors.append({'row': row_num, 'error': str(e)})
        return Response({'created': created, 'updated': updated, 'errors': errors})


class WorkOrderViewSet(viewsets.ModelViewSet):
    """작업지시서 CRUD + 워크플로우"""
    permission_classes = [permissions.IsAuthenticated]
    pagination_class   = None

    def get_serializer_class(self):
        from .serializers import WorkOrderSerializer
        return WorkOrderSerializer

    def get_queryset(self):
        from .models import WorkOrder
        qs = WorkOrder.objects.select_related(
            'incident', 'school', 'assigned_to', 'created_by', 'confirmed_by'
        )
        incident_id = self.request.query_params.get('incident')
        status_     = self.request.query_params.get('status')
        school_id   = self.request.query_params.get('school')
        assigned_id = self.request.query_params.get('assigned_to')
        if incident_id: qs = qs.filter(incident_id=incident_id)
        if status_:     qs = qs.filter(status=status_)
        if school_id:   qs = qs.filter(school_id=school_id)
        if assigned_id: qs = qs.filter(assigned_to_id=assigned_id)
        return qs

    def perform_create(self, serializer):
        from .models import WorkOrder
        serializer.save(
            work_order_number=WorkOrder.generate_number(),
            created_by=self.request.user,
        )

    @action(detail=True, methods=['post'])
    def start(self, request, pk=None):
        """발행 → 수행중"""
        from .models import WorkOrder
        wo = self.get_object()
        wo.status     = 'in_progress'
        wo.started_at = timezone.now()
        wo.save(update_fields=['status', 'started_at', 'updated_at'])
        from .serializers import WorkOrderSerializer
        return Response(WorkOrderSerializer(wo).data)

    @action(detail=True, methods=['post'])
    def complete(self, request, pk=None):
        """수행중 → 완료"""
        from .models import WorkOrder
        wo = self.get_object()
        wo.status        = 'completed'
        wo.completed_at  = timezone.now()
        wo.actual_work   = request.data.get('actual_work', wo.actual_work)
        wo.parts_used    = request.data.get('parts_used',  wo.parts_used)
        wo.work_note     = request.data.get('work_note',   wo.work_note)
        wo.save()
        from .serializers import WorkOrderSerializer
        return Response(WorkOrderSerializer(wo).data)

    @action(detail=True, methods=['post'])
    def confirm(self, request, pk=None):
        """완료 → 확인완료"""
        from .models import WorkOrder
        wo = self.get_object()
        if wo.status != 'completed':
            return Response({'error': '완료 상태에서만 확인 가능합니다'}, status=400)
        wo.status       = 'confirmed'
        wo.confirmed_at = timezone.now()
        wo.confirmed_by = request.user
        wo.save()
        from .serializers import WorkOrderSerializer
        return Response(WorkOrderSerializer(wo).data)

    @action(detail=True, methods=['post'])
    def generate_pdf(self, request, pk=None):
        """작업지시서 PDF 생성"""
        wo = self.get_object()
        from .services import generate_work_order_pdf
        try:
            generate_work_order_pdf(wo)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        download_url = f'/npms/api/incidents/work-orders/{wo.id}/download_pdf/'
        return Response({'message': 'PDF가 생성되었습니다.', 'download_url': download_url})

    @action(detail=True, methods=['get'])
    def download_pdf(self, request, pk=None):
        """작업지시서 PDF 다운로드"""
        import os
        from django.http import FileResponse
        wo = self.get_object()
        path = wo.pdf_path
        if not path or not os.path.exists(path):
            return Response({'error': 'PDF가 없습니다. 먼저 생성해주세요.'}, status=status.HTTP_404_NOT_FOUND)
        filename = os.path.basename(path)
        response = FileResponse(open(path, 'rb'), content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="{filename}"'
        return response
